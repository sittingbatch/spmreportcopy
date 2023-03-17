# HttpResponse is used to
# pass the information
# back to view
import time
import os
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
#import numpy as np
import io, base64, uuid
from io import BytesIO
from datetime import datetime, timedelta, timezone, tzinfo,date
from django.shortcuts import render
from django.http import HttpResponse
from django.template import loader

from django.contrib.auth.models import User
from django.contrib.auth import get_user_model
User = get_user_model()
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.decorators import login_required
from .forms import RegisterUserForm
from django.contrib.auth import login,authenticate

import openpyxl
from django.template import Context
from openpyxl import workbook,load_workbook
from plotly.offline import plot
import plotly.graph_objs as go

import json



@login_required(login_url = '/login')

def chartcreation (request) :
    print(request.method)
    if request.method == "GET":
        loading = "true"
        
        mydata = User.objects.filter(first_name='abc')
        superusers = User.objects.filter(is_superuser=True)
        sup = str(superusers[0])
        ranvalue = str(mydata[0])
        ranvalue = int(ranvalue)
        today = date.today()
        todayvalue = str(today).split('-')
        for i in range(len(todayvalue)):
            ranvalue = ((int(todayvalue[i]))*5 + ranvalue + (i*3) + (i*5))*5




        return render(request, "index.html",{"loading":loading, "sup": sup, "ranvalue" : ranvalue})
    data = request.POST
    slist = data.get("spdlist")
    dlist = data.get("distlist")
    plotststn = data.get("pltststn")
    plotststn = str(plotststn)
    plotendstn = data.get("pltendstn")
    route = data.get("routeid")
    #print(plotststn)
    nameoflp = data.get("nameoflp")
    trainno = data.get("trainno")
    locono = data.get("locono")
    startdate = data.get("startdate")
    enddate = data.get("enddate")

    if route == "TPTED":
        signalkm = ["0","0.5","0.8","3.8","4.2","5.2","6.4","8","8.6","9.8","10.5","11","11.7","12","15.5","15.9","17","17","20.8","21.4","22.4","23","24","24.3","25","28.3","29.5","30.8","32","32","32.6","37.2","38.3","39.3","40","40.5","40.8","44.9","45.6","46.6","47.3","47.9","48.3","52.4","53.2","54.2","54.3","58.7","59.5","60.6","60.7","62","62.2","63.4","64.1","65.1","66.4","67.2","69.6","70.6","71","71.8","72.2","79.7","80.5","81.5","82.24","82.6","83","88.2","89.8","90.8","91.55","92.1","92.5","95.3","96.7","97.9","98.74","99.2","99.6","103","103.4","104.5","106.4","107.2","108.3","109.15","110.2","111.1","112.52","112.84","113.34","113.74","114.54","115.54","116.94","117.04","117.54","119.44","120.24","121.24","122.64","123.54","124.44","125.44","126.54","127.54","127.74","128.14","130.14","130.94","131.94","135.14","136.44","137.44","138.44","138.54","139.04","140.94","141.94","142.94","144.74","145.74","147.04","148.24","150.24","151.14","152.14","152.94","153.24","153.54","155.44","156.44","157.74","158.54","159.04","159.94","161.34","162.14","163.14","163.44","164.94","165.94","166.74","167.14","167.44","169.74","170.44","171.44","172.24","172.84","173.94","174.74","175.04","175.44","176.94","177.34","177.44","178.44","179.64"]
        signalname = ["TPT","STARTER","LSS","IB GWB","IB DISTANT","IB HOME/GD","GSS","GWB","GD","GSS","HOME","KEY","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","KNNT","GWB","DISTANT","HOME","SLY","STARTER","LSS/GD","GSS","GWB","DISTANT","HOME","DST","STARTER","LSS","GWB","DISTANT","HOME","DPI","STARTER","LSS","GWB","DISTANT","HOME","MAP","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","THONGNUR NBS","GWB","DISTANT","HOME","BDY","STARTER","LSS","GWB","GD","G/IB D","IBH","GWB","DISTANT","HOME","BQI","STARTER","LSS","GWB","DISTANT","HOME","LCR","STARTER","LSS","GWB","DISTANT","HOME","DSPT","STARTER","LSS","GWB","DISTANT","HOME","TNT","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","KPPR","LSS/DISTANT","HOME","MGSJ","STARTER","LSS","GWB","DISTANT","HOME","SA","STARTER","LSS","GWB","GD","LC 115 GSS/IBD","IBS","GWB","GD","LC 115 A GSS/D","HOME","VRPD","STARTER","LSS","GWB","IBD","IBS","GWB","DISTANT","HOME","DC","STARTER","LSS","GWB","GD","LC 116 C GSS","GWB","IBD","IBS/GD","LC 116 E GSS","GWB","DISTANT","HOME","MVPM","STARTER","LSS","GWB","DISTANT","HOME","SGE","STARTER","LSS","GWB","GD","LC 118 GSS","GWB","DISTANT","HOME","ANU","STARTER","LSS","GWB","GD","LC 121 A GSS","GWB","DISTANT","HOME","CV","STARTER","LSS","GWB","GD","LC 121 C GSS/D","HOME","ED"]
        nstn = ["TPT","KEY","SLY","DST","DPI","MAP","BDY","BQI","LCR","DSPT","TNT","KPPR","MGSJ","SA","VRPD","DC","MVPM","SGE","ANU","CV","ED"]
        annot = ["TPT","KEY","SLY","DST","DPI","MAP","BDY","BQI","LCR","DSPT","TNT","KPPR","MGSJ","SA","VRPD","DC","MVPM","SGE","ANU","CV","ED"]
        dstn = [0,11,23,32,40,47.3,60.7,71,82.24,91.55,98.74,109.15,112.52,116.94,127.54,138.44,152.94,158.54,166.74,174.74,179.64]
        dstnannot = ["0","11","23","32","40","47.3","60.7","71","82.24","91.55","98.74","109.15","112.52","116.94","127.54","138.44","152.94","158.54","166.74","174.74","179.64"]


    if route == "EDTPT":
        signalkm = ["0","0.21","0.61","1.71","2.21","3.41","4.91","5.41","5.61","5.81","6.21","7.41","8.21","11.21","11.51","12.61","13.41","13.61","14.01","14.41","14.81","15.91","17.01","18.81","19.21","20.21","21.41","21.51","22.21","23.91","24.71","25.81","26.41","26.81","27.11","29.81","30.81","31.81","32.81","34.01","36.71","37.91","39.11","39.71","40.21","40.61","44.11","44.91","46.11","48.41","49.11","50.11","50.71","51.21","51.61","52.31","53.61","54.41","55.51","56.51","58.51","59.31","60.61","61.41","62.01","62.71","62.91","63.41","64.41","65.09","65.81","66.21","66.91","68.06","68.26","68.86","70.36","71.16","72.16","74.96","75.66","76.76","78.02","78.36","78.76","82.26","83.06","84.06","85.21","85.41","85.81","91.11","92.71","93.71","94.77","95.01","95.41","98.61","99.41","100.41","101.21","102.01","103.11","103.61","104.01","105.11","105.84","106.41","106.81","108.71","109.21","110.21","110.51","111.01","112.01","113.61","114.41","115.51","116.38","116.51","116.81","120.41","121.11","122.11","123.11","125.51","126.81","127.81","129.23","129.33","129.63","133.43","134.33","135.33","136.24","136.54","136.94","141.74","142.54","143.54","144.53","144.84","145.34","148.64","149.34","150.34","151.34","152.48","152.64","153.04","155.84","156.54","157.54","158.54","160.84","161.34","162.34","163.34","163.76","164.24","164.64","166.24","166.64","167.64","168.44","171.54","172.04","173.54","174.14"]
        signalname = ["ED","STARTER","LSS","GWB","DISTANT","HOME","CV","STARTER","LSS/GD","LC 120 B GSS","GD","LC 120 A GSS/GD","LCC 119 GSS","GWB","DISTANT","HOME","ANU","STARTER","LSS","GWB","GD","LC 118 C GSS/GD","LC 118 B GSS","GWB","DISTANT","HOME","SGE","STARTER","LSS","GWB","DISTANT","HOME","MVPM","STARTER","LSS","GWB","GD","LC 117 A GSS/IBD","IBS/GD","LC 117  GSS","GWB","DISTANT","HOME","DC","STARTER","LSS","GWB","IBD","IBS","GWB","DISTANT","HOME","VRPD","STARTER","LSS/GD","LC 115 A GSS","IB GWB","IBD","IBS/GD","LC 115 GSS","GWB","DISTANT","HOME","SA","STARTER","LSS","GWB","DISTANT","HOME","MGSJ","STARTER","LSS/DISTANT","HOME","KPPR","STARTER","LSS"," GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","TNT","STARTER","LSS","GWB","DISTANT","HOME","DSPT","STARTER","LSS","GWB","DISTANT","HOME","LCR","STARTER","LSS","GWB","GD","GSS      ","GWB    ","GD    ","GSS","GWB","DISTANT","HOME","BQI","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","GD","GSS","GWB","DISTANT","HOME","BDY","STARTER","LSS","GWB","GD","GSS/ IB DIS","IB HOME","GWB","DISTANT","HOME","MAP","STARTER","LSS","GWB","DISTANT","HOME","DPI","STARTER","LSS","GWB","DISTANT","HOME","DST","STARTER","LSS","GWB","GD","GSS/DISTANT","HOME","SLY","STARTER","LSS","GWB","GD","GSS/ IB D","IB HOME","GWB","DISTANT","INNER HOME","ROUTING HOME","KEY","STARTER","LSS","GWB","GD","GSS/ IB D","IB HOME","GWB","DISATANT","HOME","TPT"]
        nstn = ["ED","CV","ANU","SGE","MVPM","DC","VRPD","SA","MGSJ","KPPR","TNT","DSPT","LCR","BQI","BDY","MAP","DPI","DST","SLY","KEY","TPT"]
        annot = ["ED","CV","ANU","SGE","MVPM","DC","VRPD","SA","MGSJ","KPPR","TNT","DSPT","LCR","BQI","BDY","MAP","DPI","DST","SLY","KEY","TPT"]
        dstn = [0,4.91,13.41,21.41,26.41,39.71,50.71,61.41,65.09,68.06,78.02,85.21,94.77,105.84,116.38,129.23,136.24,144.53,152.48,163.76,174.14]
        dstnannot = ["0","4.91","13.41","21.41","26.41","39.71","50.71","61.41","65.09","68.06","78.02","85.21","94.77","105.84","116.38","129.23","136.24","144.53","152.48","163.76","174.14"]

    if route == "JTJCBE":
        signalkm = ["0","0.48","2.58","3.6","4.68","5.88","6.58","7.18","7.68","7.98","10.98","11.38","12.38","13.58","15.18","15.78","16.98","17.68","18.18","18.88","19.18","22.68","23.08","24.18","24.18","27.98","28.58","29.58","30.18","31.18","31.48","32.18","35.48","36.68","37.98","39.18","39.18","39.78","44.38","45.48","46.48","47.18","47.68","47.98","52.08","52.78","53.78","54.48","55.08","55.48","59.58","60.38","61.38","61.48","65.88","66.68","67.78","67.88","69.18","69.38","70.58","71.28","72.28","73.58","74.38","76.78","77.78","78.18","78.98","79.38","86.88","87.68","88.68","89.42","89.78","90.18","95.38","96.98","97.98","98.73","99.28","99.68","102.48","103.88","105.08","105.92","106.38","106.78","110.18","110.58","111.68","113.58","114.38","115.48","116.33","117.38","118.28","119.7","120.02","120.52","120.92","121.72","122.72","124.12","124.22","124.72","126.62","127.42","128.42","129.82","130.72","131.62","132.62","133.72","134.72","134.92","135.32","137.32","138.12","139.12","142.32","143.62","144.62","145.62","145.72","146.22","148.12","149.12","150.12","151.92","152.92","154.22","155.42","157.42","158.32","159.32","160.12","160.42","160.72","162.62","163.62","164.92","165.72","166.22","167.12","168.52","169.32","170.32","170.62","172.12","173.12","173.92","174.32","174.62","176.92","177.62","178.62","179.42","180.02","181.12","181.92","182.22","182.62","184.12","184.52","184.62","185.62","185.62","185.7","186.72","189.52","190.72","191.72","192.01","192.72","196.47","196.56","197.54","199.21","199.92","200.12","202.42","202.54","203.54","204.67","205.42","205.52","209.54","210.92","211.92","212.32","212.44","212.52","214.56","216.12","216.42","219.48","219.58","220.64","222.27","222.52","222.62","226.48","227.72","228.92","229.42","229.6","230.6","231.6","232.46","233.52","234.56","235.78","236.22","236.62","241.52","241.92","242.92","243.82","244.22","244.52","245.02","245.42","246.52","247.82","248.82","249.42","250.42","251.12","251.52","252.72","253.42","253.42","254.32","254.42","255.42","256.42","258.32","258.42","259.42","259.42","261.52","261.92","262.62","262.92","265.42","265.52","266.52","268.32","268.42","268.62","269.02","270.02","270.12","270.82","271.42","272.82","273.22","274.22","275.62","276.22","276.82","277.12","278.42","279.42","280.12","281.22","282.92","282.92","283.32","284.52","285.72"]
        signalname = ["JTJ","STARTER","INNER STARTER I","INNER STARTER II","LSS","GSS","HOME","TPT","STARTER","LSS","IB GWB","IB DISTANT","IB HOME/GD","GSS","GWB","GD","GSS","HOME","KEY","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","KNNT","GWB","DISTANT","HOME","SLY","STARTER","LSS/GD","GSS","GWB","DISTANT","HOME","DST","STARTER","LSS","GWB","DISTANT","HOME","DPI","STARTER","LSS","GWB","DISTANT","HOME","MAP","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","THONGNUR NBS","GWB","DISTANT","HOME","BDY","STARTER","LSS","GWB","GD","G/IB D","IBH","GWB","DISTANT","HOME","BQI","STARTER","LSS","GWB","DISTANT","HOME","LCR","STARTER","LSS","GWB","DISTANT","HOME","DSPT","STARTER","LSS","GWB","DISTANT","HOME","TNT","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","KPPR","LSS/DISTANT","HOME","MGSJ","STARTER","LSS","GWB","DISTANT","HOME","SA","STARTER","LSS","GWB","GD","LC 115 GSS/IBD","IBS","GWB","GD","LC 115 A GSS/D","HOME","VRPD","STARTER","LSS","GWB","IBD","IBS","GWB","DISTANT","HOME","DC","STARTER","LSS","GWB","GD","LC 116 C GSS","GWB","IBD","IBS/GD","LC 116 E GSS","GWB","DISTANT","HOME","MVPM","STARTER","LSS","GWB","DISTANT","HOME","SGE","STARTER","LSS","GWB","GD","LC 118 GSS","GWB","DISTANT","HOME","ANU","STARTER","LSS","GWB","GD","LC 121 A GSS","GWB","DISTANT","HOME","CV","STARTER","LSS","GWB","GD","LC 121 C GSS/D","HOME","ED","STARTER","LSS","GWB","DISTANT","HOME","TPM","STARTER","GWB","DISTANT","HOME","PY","STARTER","LSS","GWB","DISTANT","HOME","IGR","STARTER","LSS","GWB","DISTANT","HOME","VZ","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","UKL","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","GD","GSS","GWB","DISTANT","HOME","HOME 2","TUP","STARTER","LSS","GWB","DISTANT","HOME","VNJ","STARTER","LSS","GWB","GD","GSS/DISTANT","GSS","GWB","GD","GSS","GWB","DISTANT","HOME","SNO","STARTER","LSS/GD","GSS","GWB","GD","GSS","GWB","GD","G/D","HOME","SUU","STARTER","LSS","GWB","DISTANT","HOME","IGU","STARTER","LSS/DISTANT","GSS/GD","GSS/GD","SWB","GSS","SHI - H","GWB","DISTANT","GSS/DISTANT","HOME","PLMD","STARTER","LSS/ GD","GSS/GD","GSS/GD","GSS/DISTANT","HOME","CBF","STARTER","LSS/DISTANT","HOME","CBE"]
        nstn = ["JTJ","TPT","KEY","SLY","DST","DPI","MAP","BDY","BQI","LCR","DSPT","TNT","KPPR","MGSJ","SA","VRPD","DC","MVPM","SGE","ANU","CV","ED","TPM","PY","IGR","VZ","UKL","TUP","VNJ","SNO","SUU","IGU","PLMD","CBF","CBE"]
        annot = ["JTJ","TPT","KEY","SLY","DST","DPI","MAP","BDY","BQI","LCR","DSPT","TNT","KPPR","MGSJ","SA","VRPD","DC","MVPM","SGE","ANU","CV","ED","TPM","PY","IGR","VZ","UKL","TUP","VNJ","SNO","SUU","IGU","PLMD","CBF","CBE"]
        dstn = [0,7.18,18.18,30.18,39.18,47.18,54.48,67.88,78.18,89.42,98.73,105.92,116.33,119.7,124.12,134.72,145.62,160.12,165.72,173.92,181.92,185.62,192.01,199.21,204.67,212.32,222.27,235.78,243.82,253.42,261.92,268.32,276.22,282.92,285.72]
        dstnannot = ["0","7.18","18.18","30.18","39.18","47.18","54.48","67.88","78.18","89.42","98.73","105.92","116.33","119.7","124.12","134.72","145.62","160.12","165.72","173.92","181.92","185.62","192.01","199.21","204.67","212.32","222.27","235.78","243.82","253.42","261.92","268.32","276.22","282.92","285.72"]

    if route == "CBEJTJ":
        signalkm = ["0.00","0.63","1.10","2.30","2.80","3.30","4.00","4.50","5.70","6.80","8.40","9.50","9.60","9.90","11.00","12.40","12.80","13.10","13.90","14.30","15.00","16.10","17.10","17.90","18.40","19.20","21.70","22.20","23.20","24.40","24.50","24.80","26.20","27.40","27.90","29.00","29.40","29.90","30.90","31.90","33.00","33.20","33.50","33.80","34.40","35.50","36.60","37.10","38.10","39.50","40.10","40.70","41.70","42.60","42.80","43.30","48.20","48.60","49.60","50.70","50.80","51.30","52.10","53.10","53.70","54.70","55.70","56.30","57.50","60.90","61.40","63.30","64.15","64.70","65.10","71.90","72.20","73.30","73.78","74.50","74.90","78.50","79.70","80.80","81.75","82.00","82.30","85.00","85.40","86.30","87.21","87.60","88.00","92.00","92.60","93.60","94.71","94.80","95.00","96.50","97.10","98.20","99.20","100.10","100.79","101.00","101.40","102.50","103.00","104.20","105.70","106.20","106.40","107.10","107.60","108.10","108.90","111.90","112.20","113.30","114.10","114.20","114.30","114.30","114.70","115.80","116.90","119.30","119.70","120.70","121.90","121.90","122.60","124.50","125.10","126.30","126.90","127.30","127.60","130.40","131.00","132.00","133.00","134.10","136.30","137.60","138.80","139.40","139.90","140.30","143.90","144.20","145.30","147.90","148.50","149.50","150.10","150.60","151.00","151.70","152.90","153.30","154.40","155.40","157.40","157.90","159.20","160.00","160.60","161.30","161.50","161.90","162.90","163.58","164.30","164.70","165.60","166.75","166.90","167.50","169.40","170.00","171.00","174.00","174.40","175.50","176.76","177.10","177.50","181.00","181.60","182.60","183.75","183.80","184.20","189.40","191.00","192.00","193.06","193.30","193.70","196.70","197.20","198.20","199.00","199.40","200.50","201.00","201.40","202.50","203.23","203.80","204.20","206.10","206.60","207.60","207.90","208.30","209.30","210.90","211.30","212.40","213.27","213.40","213.70","217.30","217.70","218.70","219.70","222.10","223.40","224.40","225.82","225.90","226.20","230.00","230.90","231.90","232.81","233.10","233.50","238.50","239.10","240.10","241.09","241.40","241.90","245.60","246.30","247.30","248.30","249.44","249.60","250.00","253.50","253.90","254.90","255.90","258.50","258.90","259.90","260.90","261.32","261.80","262.20","264.60","265.00","266.00","266.60","270.20","270.60","272.10","272.70","273.20","273.70","275.40","276.90","278.80","280.09"]
        signalname = ["CBE","STARTER","LSS /DISTANT","HOME","CBF","STARTER","LSS/GD","GSS/ GD","GSS/GD","GSS/DISTANT","HOME","PLMD","STARTER","LSS/GD","GSS","GWB","DISTANT","SWB","GSS/DISTANT","SHI","GSS/GD","GSS/DISTANT","HOME","IGU","STARTER","LSS","GWB","DISTANT","HOME","SUU","STARTER","LSS/GD","LC142 GSS","GWB","G D","LC141 GSS","GWB","GD","LC140 GSS/DIST","HOME","SNO","STARTER","LSS","GWB","G D","LC137 GSS","GWB","GD","LC136 GSS/GD","LC 135 GSS","GWB","DISTANT","HOME","VNJ","STARTER","LSS","GWB","DISTANT","HOME","TUP","STARTER","LSS/G D","GSS","GWB","GD","LC 131 E GSS","IBS GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","UKL","STARTER","LSS ","GWB","DISTANT","HOME","VZ","STARTER","LSS ","GWB","DISTANT","HOME","IGR","STARTER","LSS","GWB","DISTANT","HOME","PY","STARTER","LSS","GWB","DISTANT","HOME","TPM","STARTER","LSS","GWB","GD","GSS/DISTANT","HOME","RT HOME","ED","STARTER","LSS","GWB","DISTANT","HOME","CV","STARTER","LSS/GD","LC 120 B GSS","GD","LC 120 A GSS/GD","LCC 119 GSS","GWB","DISTANT","HOME","ANU","STARTER","LSS","GWB","GD","LC 118 C GSS/GD","LC 118 B GSS","GWB","DISTANT","HOME","SGE","STARTER","LSS","GWB","DISTANT","HOME","MVPM","STARTER","LSS","GWB","GD","LC 117 A GSS/IBD","IBS/GD","LC 117  GSS","GWB","DISTANT","HOME","DC","STARTER","LSS","GWB","IBD","IBS","GWB","DISTANT","HOME","VRPD","STARTER","LSS/GD","LC 115 A GSS","IB GWB","IBD","IBS/GD","LC 115 GSS","GWB","DISTANT","HOME","SA","STARTER","LSS","GWB","DISTANT","HOME","MGSJ","STARTER","LSS/DISTANT","HOME","KPPR","STARTER","LSS"," GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","TNT","STARTER","LSS","GWB","DISTANT","HOME","DSPT","STARTER","LSS","GWB","DISTANT","HOME","LCR","STARTER","LSS","GWB","GD","GSS      ","GWB    ","GD    ","GSS","GWB","DISTANT","HOME","BQI","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","GD","GSS","GWB","DISTANT","HOME","BDY","STARTER","LSS","GWB","GD","GSS/ IB DIS","IB HOME","GWB","DISTANT","HOME","MAP","STARTER","LSS","GWB","DISTANT","HOME","DPI","STARTER","LSS","GWB","DISTANT","HOME","DST","STARTER","LSS","GWB","GD","GSS/DISTANT","HOME","SLY","STARTER","LSS","GWB","GD","GSS/ IB D","IB HOME","GWB","DISTANT","INNER HOME","ROUTING HOME","KEY","STARTER","LSS","GWB","GD","GSS/ IB D","IB HOME","GWB","DISATANT","HOME","TPT","STARTER","LSS/D","HOME","HOME","HOME","JTJ"]
        nstn = ["CBE","CBF","PLMD","IGU","SUU","SNO","VNJ","TUP","UKL","VZ","IGR","PY","TPM","ED","CV","ANU","SGE","MVPM","DC","VRPD","SA","MGSJ","KPPR","TNT","DSPT","LCR","BQI","BDY","MAP","DPI","DST","SLY","KEY","TPT","JTJ"]
        annot = ["CBE","CBF","PLMD","IGU","SUU","SNO","VNJ","TUP","UKL","VZ","IGR","PY","TPM","ED","CV","ANU","SGE","MVPM","DC","VRPD","SA","MGSJ","KPPR","TNT","DSPT","LCR","BQI","BDY","MAP","DPI","DST","SLY","KEY","TPT","JTJ"]
        dstn = [0.00,2.80,9.50,17.90,24.40,33.00,42.60,50.70,64.15,73.78,81.75,87.21,94.71,100.79,105.70,114.10,121.90,126.90,139.40,150.10,160.00,163.58,166.75,176.76,183.75,193.06,203.23,213.27,225.82,232.81,241.09,249.44,261.32,272.70,280.09]
        dstnannot = ["0.00","2.80","9.50","17.90","24.40","33.00","42.60","50.70","64.15","73.78","81.75","87.21","94.71","100.79","105.70","114.10","121.90","126.90","139.40","150.10","160.00","163.58","166.75","176.76","183.75","193.06","203.23","213.27","225.82","232.81","241.09","249.44","261.32","272.70","280.09"]

    if route == "SACBE":
        signalkm = ["0","0.1","0.6","2.5","3.1","4.1","5.5","6.4","6.8","7.8","8.9","9.9","10.1","10.5","12.5","12.9","13.9","17.3","18.6","19.6","20.6","20.7","21.2","23.1","23.8","24.7","26.5","27","28.3","29.4","31.2","31.5","32.2","33.1","33.4","33.7","35.6","36.1","37.3","38.1","38.6","39.3","40.7","41.3","42.3","42.6","44.1","45.1","45.9","46.3","46.6","49.2","49.9","50.9","51.8","52.4","53.5","54.3","54.6","55","56.5","56.9","57","58","59.2","59.28","60.3","63.1","64.3","65.3","65.59","66.3","70.05","70.14","71.12","72.79","73.5","73.7","76","76.12","77.12","78.25","79","79.1","83.12","84.5","85.5","85.9","86.02","86.1","88.14","89.7","90","93.06","93.16","94.22","95.85","96.1","96.2","100.06","101.3","102.5","103","103.18","104.18","105.18","106.04","107.1","108.14","109.36","109.8","110.2","115.1","115.5","116.5","117.4","117.8","118.1","118.6","119","120.1","121.4","122.4","123","124","124.7","125.1","126.3","127","127","127.9","128","129","130","131.9","132","133","133","135.1","135.5","136.2","136.5","139","139.1","140.1","141.9","142","142.7","143.1","144.1","144.2","144.9","145.5","146.9","147.3","148.3","149.7","150.3","150.9","151.2","152.5","153.5","154.2","155.3","157","157","157.4","158.6","159.8"]
        signalname = ["SA","STARTER","LSS","GWB","GD","LC 115 GSS/IBD","IBS","GWB","GD","LC 115 A GSS/D","HOME","VRPD","STARTER","LSS","GWB","IBD","IBS","GWB","DISTANT","HOME","DC","STARTER","LSS","GWB","GD","LC 116 C GSS","GWB","IBD","IBS/GD","LC 116 E GSS","GWB","DISTANT","HOME","MVPM","STARTER","LSS","GWB","DISTANT","HOME","SGE","STARTER","LSS","GWB","GD","LC 118 GSS","GWB","DISTANT","HOME","ANU","STARTER","LSS","GWB","GD","LC 121 A GSS","GWB","DISTANT","HOME","CV","STARTER","LSS","GWB","GD","LC 121 C GSS/D","HOME","ED","STARTER","LSS","GWB","DISTANT","HOME","TPM","STARTER","GWB","DISTANT","HOME","PY","STARTER","LSS","GWB","DISTANT","HOME","IGR","STARTER","LSS","GWB","DISTANT","HOME","VZ","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","UKL","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","GD","GSS","GWB","DISTANT","HOME","HOME 2","TUP","STARTER","LSS","GWB","DISTANT","HOME","VNJ","STARTER","LSS","GWB","GD","GSS/DISTANT","GSS","GWB","GD","GSS","GWB","DISTANT","HOME","SNO","STARTER","LSS/GD","GSS","GWB","GD","GSS","GWB","GD","G/D","HOME","SUU","STARTER","LSS","GWB","DISTANT","HOME","IGU","STARTER","LSS/DISTANT","GSS/GD","GSS/GD","SWB","GSS","SHI - H","GWB","DISTANT","GSS/DISTANT","HOME","PLMD","STARTER","LSS/ GD","GSS/GD","GSS/GD","GSS/DISTANT","HOME","CBF","STARTER","LSS/DISTANT","HOME","CBE"]
        nstn = ["SA","VRPD","DC","MVPM","SGE","ANU","CV","ED","TPM","PY","IGR","VZ","UKL","TUP","VNJ","SNO","SUU","IGU","PLMD","CBF","CBE"]
        annot = ["SA","VRPD","DC","MVPM","SGE","ANU","CV","ED","TPM","PY","IGR","VZ","UKL","TUP","VNJ","SNO","SUU","IGU","PLMD","CBF","CBE"]
        dstn = [0,9.9,20.6,33.1,38.1,45.9,54.3,59.2,65.59,72.79,78.25,85.9,95.85,109.36,117.4,127,135.5,141.9,150.3,157,159.8]
        dstnannot = ["0","9.9","20.6","33.1","38.1","45.9","54.3","59.2","65.59","72.79","78.25","85.9","95.85","109.36","117.4","127","135.5","141.9","150.3","157","159.8"]

    if route == "CBESA":
        signalkm = ["0.00","0.63","1.10","2.30","2.80","3.30","4.00","4.50","5.70","6.80","8.40","9.50","9.60","9.90","11.00","12.40","12.80","13.10","13.90","14.30","15.00","16.10","17.10","17.90","18.40","19.20","21.70","22.20","23.20","24.40","24.50","24.80","26.20","27.40","27.90","29.00","29.40","29.90","30.90","31.90","33.00","33.20","33.50","33.80","34.40","35.50","36.60","37.10","38.10","39.50","40.10","40.70","41.70","42.60","42.80","43.30","48.20","48.60","49.60","50.70","50.80","51.30","52.10","53.10","53.70","54.70","55.70","56.30","57.50","60.90","61.40","63.30","64.15","64.70","65.10","71.90","72.20","73.30","73.78","74.50","74.90","78.50","79.70","80.80","81.75","82.00","82.30","85.00","85.40","86.30","87.21","87.60","88.00","92.00","92.60","93.60","94.71","94.80","95.00","96.50","97.10","98.20","99.20","100.10","100.79","101.00","101.40","102.50","103.00","104.20","105.70","106.20","106.40","107.10","107.60","108.10","108.90","111.90","112.20","113.30","114.10","114.30","114.30","114.40","114.70","115.80","116.90","119.30","119.70","120.70","121.90","121.90","122.60","124.50","125.10","126.30","126.90","127.30","127.60","130.40","131.00","132.00","133.00","134.10","136.30","137.60","138.80","139.40","139.90","140.30","143.90","144.20","145.30","147.90","148.50","149.50","150.10","150.60","151.00","151.70","152.90","153.30","154.40","155.40","157.40","157.90","159.20","160.00"]
        signalname = ["CBE","STARTER","LSS /DISTANT","HOME","CBF","STARTER","LSS/GD","GSS/ GD","GSS/GD","GSS/DISTANT","HOME","PLMD","STARTER","LSS/GD","GSS","GWB","DISTANT","SWB","GSS/DISTANT","SHI","GSS/GD","GSS/DISTANT","HOME","IGU","STARTER","LSS","GWB","DISTANT","HOME","SUU","STARTER","LSS/GD","LC142 GSS","GWB","G D","LC141 GSS","GWB","GD","LC140 GSS/DIST","HOME","SNO","STARTER","LSS","GWB","G D","LC137 GSS","GWB","GD","LC136 GSS/GD","LC 135 GSS","GWB","DISTANT","HOME","VNJ","STARTER","LSS","GWB","DISTANT","HOME","TUP","STARTER","LSS/G D","GSS","GWB","GD","LC 131 E GSS","IBS GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","UKL","STARTER","LSS ","GWB","DISTANT","HOME","VZ","STARTER","LSS ","GWB","DISTANT","HOME","IGR","STARTER","LSS","GWB","DISTANT","HOME","PY","STARTER","LSS","GWB","DISTANT","HOME","TPM","STARTER","LSS","GWB","GD","GSS/DISTANT","HOME","RT HOME","ED","STARTER","LSS","GWB","DISTANT","HOME","CV","STARTER","LSS/GD","LC 120 B GSS","GD","LC 120 A GSS/GD","LCC 119 GSS","GWB","DISTANT","HOME","ANU","STARTER","LSS","GWB","GD","LC 118 C GSS/GD","LC 118 B GSS","GWB","DISTANT","HOME","SGE","STARTER","LSS","GWB","DISTANT","HOME","MVPM","STARTER","LSS","GWB","GD","LC 117 A GSS/IBD","IBS/GD","LC 117  GSS","GWB","DISTANT","HOME","DC","STARTER","LSS","GWB","IBD","IBS","GWB","DISTANT","HOME","VRPD","STARTER","LSS/GD","LC 115 A GSS","IB GWB","IBD","IBS/GD","LC 115 GSS","GWB","DISTANT","HOME","SA"]
        nstn = ["CBE","CBF","PLMD","IGU","SUU","SNO","VNJ","TUP","UKL","VZ","IGR","PY","TPM","ED","CV","ANU","SGE","MVPM","DC","VRPD","SA"]
        annot = ["CBE","CBF","PLMD","IGU","SUU","SNO","VNJ","TUP","UKL","VZ","IGR","PY","TPM","ED","CV","ANU","SGE","MVPM","DC","VRPD","SA"]
        dstn = [0.00,2.80,9.50,17.90,24.40,33.00,42.60,50.70,64.15,73.78,81.75,87.21,94.71,100.79,105.70,114.10,121.90,126.90,139.40,150.10,160.00]
        dstnannot = ["0.00","2.80","9.50","17.90","24.40","33.00","42.60","50.70","64.15","73.78","81.75","87.21","94.71","100.79","105.70","114.10","121.90","126.90","139.40","150.10","160.00"]


    if request.method == "POST" and plotststn != "None":
        """speedlist = data.get("spdlist")
                                speedlist1 = speedlist.splitlines()
                                speedlist = list(map(float,speedlist1))
                                distancelist = data.get("distlist")
                                distancelist1 = distancelist.splitlines()
                                distancelist = list(map(float,distancelist1))"""

        #signal_sample = request.FILES["spmfile"]
     
        wb1 = load_workbook(filename = request.FILES['spmfile'].file,data_only=True)

        sh1 = wb1['Sheet1'] 

        rowcount = sh1.max_row
        colcount = sh1.max_column
        colcount = colcount+1

        for c in range(colcount):
            if c==0:
                c=c+1
            for r in range(rowcount):
                if r==0:
                    r=r+1
                if (type(sh1.cell(r,c).value).__name__) == "str":
                    if "distance" in sh1.cell(r,c).value.lower() or sh1.cell(r,c).value == "DISTANCE" or sh1.cell(r,c).value == "distance" or sh1.cell(r,c).value == "Distance" or sh1.cell(r,c).value == "DIST MTRS":
                        distancerow = r
                        distancecolumn = c
                    if "speed"  in sh1.cell(r,c).value.lower() or sh1.cell(r,c).value == "SPEED" or sh1.cell(r,c).value == "speed" or sh1.cell(r,c).value == "Speed" or sh1.cell(r,c).value == "INST.  KMPH":
                        speedrow = r
                        speedcolumn = c

        speedlist1 = []
        distancelist1 = []
        distancerow = distancerow+2
        speedrow = speedrow +2
        rowcount = rowcount

        for i in range(distancerow,rowcount):
            if i == 0:
                i = 1
            if (type(sh1.cell(i,distancecolumn).value).__name__) != "NoneType":
                distancelist1.append(sh1.cell(i,distancecolumn).value)

        for s in range(speedrow,rowcount):
            if s == 0:
                s = 1
            if (type(sh1.cell(s,speedcolumn).value).__name__) != "NoneType":
                speedlist1.append(sh1.cell(s,speedcolumn).value)
        #print(len(speedlist))
        #print(len(distancelist))
        speedlist = list(map(float,speedlist1))
        distancelist = list(map(float,distancelist1))

        #spmtype = data.get("spmtype")
        #print(spmtype)
        spmcount = 0
        for i in range(len(distancelist1)):
            if distancelist1[i] == 1 or distancelist1[i] == 2:
                spmcount = spmcount+1
        print(spmcount)

        if spmcount>=20:
            spmtype = "telpro"
        else:
            spmtype = ""

        if spmtype == "telpro":
            for i in range(len(distancelist)):
                if i != 0 or i != len(distancelist):
                    distancelist[i] = distancelist[i] + distancelist[i-1]
            for index in range(len(distancelist)):
                distancelist[index] = (distancelist[index]/1000)
                distancelist[index] = round(distancelist[index],2)
        today = datetime.now()

        if route == "JTJED":
            """
            signalkm = ["213.02","213.5","215.6","216.62","217.7","218.9","219.6","220.2","220.7","221","224","224.4","225.4","226.6","228.2","228.8","230","230.7","231.2","231.9","232.2","235.7","236.1","237.2","237.2","241","241.6","242.6","243.2","244.2","244.5","245.2","248.5","249.7","251","252.2","252.2","252.8","257.4","258.5","259.5","260.2","260.7","261","265.1","265.5","266.5","267.2","267.8","268.2","272.3","272.8","273.8","273.2","277.6","278","279.1","279.2","280.5","280.7","281.9","282.3","283.3","284.6","285.4","287.8","288.8","289.2","290","290.4","297.9","298.3","299.3","300.04","300.4","300.8","306","307.6","308.6","309.35","309.9","310.3","312.9","314.3","315.5","316.34","316.8","317.2","320.6","321","322.1","324","324.4","325.5","326.35","327.4","328.3","329.72","329.4","329.9","330.1","330.7","331.7","333.1","333.2","333.7","335.6","336.2","337.2","338.6","339.5","339.9","340.9","342","343","343.2","343.6","345.6","346","347","350.4","351.7","352.7","353.7","353.8","354.3","356.2","356.9","357.8","359.6","360.1","361.4","362.5","364.6","364.3","365.3","366.2","366.5","366.8","368.7","369.2","370.4","371.2","371.7","372.4","373.8","374.4","375.4","375.7","377.2","378.2","379","379.4","379.7","382.3","383","384","384.9","385.5","386.6","387.4","387.7","388.1","389.6","390","390.1","391.1","392.3","392.3","392.8","393.7"]
            signalname = ["JTJ","STARTER","INNER STARTER I","INNER STARTER II","LSS","GSS","HOME","TPT","STARTER","LSS","IB GWB","IB DISTANT","IB HOME/GD","GSS","GWB","GD","GSS","HOME","KEY","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","KNNT","GWB","DISTANT","HOME","SLY","STARTER","LSS/GD","GSS","GWB","DISTANT","HOME","DST","STARTER","LSS","GWB","DISTANT","HOME","DPI","STARTER","LSS","GWB","DISTANT","HOME","MAP","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","R","GWB","DISTANT","HOME","BDY","STARTER","LSS","GWB","GD","G/IB D","IBH","GWB","DISTANT","HOME","BQI","STARTER","LSS","GWB","DISTANT","HOME","LCR","STARTER","LSS","GWB","DISTANT","HOME","DSPT","STARTER","LSS","GWB","DISTANT","HOME","TNT","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","KPPR","LSS/DISTANT","HOME","MGSJ","STARTER","LSS","GWB","DISTANT","HOME","SA","STARTER","LSS","GWB","GD","LC 115 GSS/IBD","IBS","GWB","GD","LC 115 A GSS/D","HOME","VRPD","STARTER","LSS","GWB","IBD","IBS","GWB","DISTANT","HOME","DC","STARTER","LSS","GWB","GD","LC 116 C GSS","GWB","IBD","IBS/GD","LC 116 E GSS","GWB","DISTANT","HOME","MVPM","STARTER","LSS","GWB","DISTANT","HOME","SGE","STARTER","LSS","GWB","GD","LC 118 GSS","GWB","DISTANT","HOME","ANU","STARTER","LSS","GWB","GD","LC 121 A GSS","GWB","DISTANT","HOME","CV","STARTER","LSS","GWB","GD","LC 121 C GSS/D","HOME","ED","ED","STARTER","LSS"]
            nstn = ["JTJ","TPT","KEY","SLY","DST","DPI","MAP","BDY","BQI","LCR","DSPT","TNT","KPPR","MGSJ","SA","VRPD","DC","MVPM","SGE","ANU","CV","ED"]
            #input data for annotations
            annot = ["JTJ","TPT","KEY","SLY","DST","DPI","MAP","BDY","BQI","LCR","DSPT","TNT","KPPR","MGSJ","SA","VRPD","DC","MVPM","SGE","ANU","CV","ED"]
            dstn = [0.00,7.18,18.18,24.18,30.18,39.18,46.48,54.18,66.18,76.18,96.33,103.32,113.33,116.7,120.08,129.98,140.68,153.18,158.18,165.98,174.38,179.28]
            dstnannot = ["0","4.91","13.31","21.11","26.11","38.61","49.31","59.21","62.79","65.96","75.97","82.96","92.27","102.44","112.48","125.03","132.02","140.3","148.65","160.53","171.91","179.29"]
            
            signalkm = ["0","0.48","2.58","3.6","4.68","5.88","6.58","7.18","7.68","7.98","10.98","11.38","12.38","13.58","15.18","15.78","16.98","17.68","18.18","18.88","19.18","22.68","23.08","24.18","24.18","27.98","28.58","29.58","30.18","31.18","31.48","32.18","35.48","36.68","37.98","39.18","39.18","39.78","44.38","45.48","46.48","47.18","47.68","47.98","52.08","52.48","53.48","54.18","54.78","55.18","59.28","59.78","60.78","60.88","65.28","65.68","66.78","66.88","68.18","68.38","69.58","69.98","70.98","72.28","73.08","75.48","76.48","76.88","77.68","78.08","85.58","86.38","87.38","88.12","88.48","88.88","94.08","95.68","96.68","97.43","97.98","98.38","100.98","102.38","103.58","104.42","104.88","105.28","108.68","109.08","110.18","112.08","112.48","113.58","114.43","115.48","116.38","117.8","118.12","118.62","118.82","119.62","120.62","122.02","122.12","122.62","124.52","125.32","126.32","127.72","128.62","129.42","130.42","131.52","132.52","132.72","133.12","135.12","135.92","136.92","140.32","141.62","142.62","143.62","143.72","144.22","146.12","147.12","148.12","149.92","150.72","152.02","153.12","155.22","156.02","157.02","157.92","158.22","158.52","160.42","161.42","162.62","163.42","163.92","164.82","166.22","167.12","168.12","168.42","169.92","170.92","171.72","172.12","172.42","175.02","175.72","176.72","177.62","178.22","179.32","180.12","180.42","180.82","182.32","182.72","182.82","183.82","185.02","185.02","185.52","186.42"]
            signalname = ["JTJ","STARTER","INNER STARTER I","INNER STARTER II","LSS","GSS","HOME","TPT","STARTER","LSS","IB GWB","IB DISTANT","IB HOME/GD","GSS","GWB","GD","GSS","HOME","KEY","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","KNNT","GWB","DISTANT","HOME","SLY","STARTER","LSS/GD","GSS","GWB","DISTANT","HOME","DST","STARTER","LSS","GWB","DISTANT","HOME","DPI","STARTER","LSS","GWB","DISTANT","HOME","MAP","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","THONGNUR NBS","GWB","DISTANT","HOME","BDY","STARTER","LSS","GWB","GD","G/IB D","IBH","GWB","DISTANT","HOME","BQI","STARTER","LSS","GWB","DISTANT","HOME","LCR","STARTER","LSS","GWB","DISTANT","HOME","DSPT","STARTER","LSS","GWB","DISTANT","HOME","TNT","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","KPPR","LSS/DISTANT","HOME","MGSJ","STARTER","LSS","GWB","DISTANT","HOME","SA","STARTER","LSS","GWB","GD","LC 115 GSS/IBD","IBS","GWB","GD","LC 115 A GSS/D","HOME","VRPD","STARTER","LSS","GWB","IBD","IBS","GWB","DISTANT","HOME","DC","STARTER","LSS","GWB","GD","LC 116 C GSS","GWB","IBD","IBS/GD","LC 116 E GSS","GWB","DISTANT","HOME","MVPM","STARTER","LSS","GWB","DISTANT","HOME","SGE","STARTER","LSS","GWB","GD","LC 118 GSS","GWB","DISTANT","HOME","ANU","STARTER","LSS","GWB","GD","LC 121 A GSS","GWB","DISTANT","HOME","CV","STARTER","LSS","GWB","GD","LC 121 C GSS/D","HOME","ED","ED","STARTER","LSS"]
            nstn = ["JTJ","TPT","KEY","SLY","DST","DPI","MAP","BDY","BQI","LCR","DSPT","TNT","KPPR","MGSJ","SA","VRPD","DC","MVPM","SGE","ANU","CV","ED"]
            annot = ["JTJ","TPT","KEY","SLY","DST","DPI","MAP","BDY","BQI","LCR","DSPT","TNT","KPPR","MGSJ","SA","VRPD","DC","MVPM","SGE","ANU","CV","ED"]
            dstn = [0,7.18,18.18,30.18,39.18,47.18,54.18,66.88,76.88,88.12,97.43,104.42,114.43,117.8,122.02,132.52,143.62,157.92,163.42,171.72,180.12,185.02]
            dstnannot = ["0","7.18","18.18","30.18","39.18","47.18","54.18","66.88","76.88","88.12","97.43","104.42","114.43","117.8","122.02","132.52","143.62","157.92","163.42","171.72","180.12","185.02"]


            signalkm = ["0","0.48","2.58","3.6","4.68","5.88","6.58","7.18","7.68","7.98","10.98","11.38","12.38","13.58","15.18","15.78","16.98","17.68","18.18","18.88","19.18","22.68","23.08","24.18","24.18","27.98","28.58","29.58","30.18","31.18","31.48","32.18","35.48","36.68","37.98","39.18","39.18","39.78","44.38","45.48","46.48","47.18","47.68","47.98","52.08","52.78","53.78","54.48","55.08","55.48","59.58","60.38","61.38","61.48","65.88","66.68","67.78","67.88","69.18","69.38","70.58","71.28","72.28","73.58","74.38","76.78","77.78","78.18","78.98","79.38","86.88","87.68","88.68","89.42","89.78","90.18","95.38","96.98","97.98","98.73","99.28","99.68","102.48","103.88","105.08","105.92","106.38","106.78","110.18","110.58","111.68","113.58","114.38","115.48","116.33","117.38","118.28","119.7","120.02","120.52","120.92","121.72","122.72","124.12","124.22","124.72","126.62","127.42","128.42","129.82","130.72","131.62","132.62","133.72","134.72","134.92","135.32","137.32","138.12","139.12","142.32","143.62","144.62","145.62","145.72","146.22","148.12","149.12","150.12","151.92","152.92","154.22","155.42","157.42","158.32","159.32","160.12","160.42","160.72","162.62","163.62","164.92","165.72","166.22","167.12","168.52","169.32","170.32","170.62","172.12","173.12","173.92","174.32","174.62","176.92","177.62","178.62","179.42","180.02","181.12","181.92","182.22","182.62","184.12","184.52","184.62","185.62","186.82","186.82","187.32"]
            signalname = ["JTJ","STARTER","INNER STARTER I","INNER STARTER II","LSS","GSS","HOME","TPT","STARTER","LSS","IB GWB","IB DISTANT","IB HOME/GD","GSS","GWB","GD","GSS","HOME","KEY","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","KNNT","GWB","DISTANT","HOME","SLY","STARTER","LSS/GD","GSS","GWB","DISTANT","HOME","DST","STARTER","LSS","GWB","DISTANT","HOME","DPI","STARTER","LSS","GWB","DISTANT","HOME","MAP","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","THONGNUR NBS","GWB","DISTANT","HOME","BDY","STARTER","LSS","GWB","GD","G/IB D","IBH","GWB","DISTANT","HOME","BQI","STARTER","LSS","GWB","DISTANT","HOME","LCR","STARTER","LSS","GWB","DISTANT","HOME","DSPT","STARTER","LSS","GWB","DISTANT","HOME","TNT","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","KPPR","LSS/DISTANT","HOME","MGSJ","STARTER","LSS","GWB","DISTANT","HOME","SA","STARTER","LSS","GWB","GD","LC 115 GSS/IBD","IBS","GWB","GD","LC 115 A GSS/D","HOME","VRPD","STARTER","LSS","GWB","IBD","IBS","GWB","DISTANT","HOME","DC","STARTER","LSS","GWB","GD","LC 116 C GSS","GWB","IBD","IBS/GD","LC 116 E GSS","GWB","DISTANT","HOME","MVPM","STARTER","LSS","GWB","DISTANT","HOME","SGE","STARTER","LSS","GWB","GD","LC 118 GSS","GWB","DISTANT","HOME","ANU","STARTER","LSS","GWB","GD","LC 121 A GSS","GWB","DISTANT","HOME","CV","STARTER","LSS","GWB","GD","LC 121 C GSS/D","HOME","ED","ED","STARTER"]
            nstn = ["JTJ","TPT","KEY","SLY","DST","DPI","MAP","BDY","BQI","LCR","DSPT","TNT","KPPR","MGSJ","SA","VRPD","DC","MVPM","SGE","ANU","CV","ED"]
            annot = ["JTJ","TPT","KEY","SLY","DST","DPI","MAP","BDY","BQI","LCR","DSPT","TNT","KPPR","MGSJ","SA","VRPD","DC","MVPM","SGE","ANU","CV","ED"]
            dstn = [0,7.18,18.18,30.18,39.18,47.18,54.48,67.88,78.18,89.42,98.73,105.92,116.33,119.7,124.12,134.72,145.62,160.12,165.72,173.92,181.92,186.82]
            dstnannot = ["0","7.18","18.18","30.18","39.18","47.18","54.48","67.88","78.18","89.42","98.73","105.92","116.33","119.7","124.12","134.72","145.62","160.12","165.72","173.92","181.92","186.82"]
            """

            signalkm = ["0","0.48","2.58","3.6","4.68","5.88","6.58","7.18","7.68","7.98","10.98","11.38","12.38","13.58","15.18","15.78","16.98","17.68","18.18","18.88","19.18","22.68","23.08","24.18","24.18","27.98","28.58","29.58","30.18","31.18","31.48","32.18","35.48","36.68","37.98","39.18","39.18","39.78","44.38","45.48","46.48","47.18","47.68","47.98","52.08","52.78","53.78","54.48","55.08","55.48","59.58","60.38","61.38","61.48","65.88","66.68","67.78","67.88","69.18","69.38","70.58","71.28","72.28","73.58","74.38","76.78","77.78","78.18","78.98","79.38","86.88","87.68","88.68","89.42","89.78","90.18","95.38","96.98","97.98","98.73","99.28","99.68","102.48","103.88","105.08","105.92","106.38","106.78","110.18","110.58","111.68","113.58","114.38","115.48","116.33","117.38","118.28","119.7","120.02","120.52","120.92","121.72","122.72","124.12","124.22","124.72","126.62","127.42","128.42","129.82","130.72","131.62","132.62","133.72","134.72","134.92","135.32","137.32","138.12","139.12","142.32","143.62","144.62","145.62","145.72","146.22","148.12","149.12","150.12","151.92","152.92","154.22","155.42","157.42","158.32","159.32","160.12","160.42","160.72","162.62","163.62","164.92","165.72","166.22","167.12","168.52","169.32","170.32","170.62","172.12","173.12","173.92","174.32","174.62","176.92","177.62","178.62","179.42","180.02","181.12","181.92","182.22","182.62","184.12","184.52","184.62","185.62","186.82","186.82","187.32","188.22"]
            signalname = ["JTJ","STARTER","INNER STARTER I","INNER STARTER II","LSS","GSS","HOME","TPT","STARTER","LSS","IB GWB","IB DISTANT","IB HOME/GD","GSS","GWB","GD","GSS","HOME","KEY","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","KNNT","GWB","DISTANT","HOME","SLY","STARTER","LSS/GD","GSS","GWB","DISTANT","HOME","DST","STARTER","LSS","GWB","DISTANT","HOME","DPI","STARTER","LSS","GWB","DISTANT","HOME","MAP","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","THONGNUR NBS","GWB","DISTANT","HOME","BDY","STARTER","LSS","GWB","GD","G/IB D","IBH","GWB","DISTANT","HOME","BQI","STARTER","LSS","GWB","DISTANT","HOME","LCR","STARTER","LSS","GWB","DISTANT","HOME","DSPT","STARTER","LSS","GWB","DISTANT","HOME","TNT","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","KPPR","LSS/DISTANT","HOME","MGSJ","STARTER","LSS","GWB","DISTANT","HOME","SA","STARTER","LSS","GWB","GD","LC 115 GSS/IBD","IBS","GWB","GD","LC 115 A GSS/D","HOME","VRPD","STARTER","LSS","GWB","IBD","IBS","GWB","DISTANT","HOME","DC","STARTER","LSS","GWB","GD","LC 116 C GSS","GWB","IBD","IBS/GD","LC 116 E GSS","GWB","DISTANT","HOME","MVPM","STARTER","LSS","GWB","DISTANT","HOME","SGE","STARTER","LSS","GWB","GD","LC 118 GSS","GWB","DISTANT","HOME","ANU","STARTER","LSS","GWB","GD","LC 121 A GSS","GWB","DISTANT","HOME","CV","STARTER","LSS","GWB","GD","LC 121 C GSS/D","HOME","ED","ED","STARTER"]
            nstn = ["JTJ","TPT","KEY","SLY","DST","DPI","MAP","BDY","BQI","LCR","DSPT","TNT","KPPR","MGSJ","SA","VRPD","DC","MVPM","SGE","ANU","CV","ED"]
            annot = ["JTJ","TPT","KEY","SLY","DST","DPI","MAP","BDY","BQI","LCR","DSPT","TNT","KPPR","MGSJ","SA","VRPD","DC","MVPM","SGE","ANU","CV","ED"]
            dstn = [0,7.18,18.18,30.18,39.18,47.18,54.48,67.88,78.18,89.42,98.73,105.92,116.33,119.7,124.12,134.72,145.62,160.12,165.72,173.92,181.92,186.82]
            dstnannot = ["0","7.18","18.18","30.18","39.18","47.18","54.48","67.88","78.18","89.42","98.73","105.92","116.33","119.7","124.12","134.72","145.62","160.12","165.72","173.92","181.92","186.82"]


        if route == "EDJTJ":
            """
            signalkm = ["0","0.11","0.61","1.71","2.21","3.41","4.91","5.41","5.61","5.41","5.81","7.31","8.11","11.11","11.41","12.51","13.31","14.21","13.51","13.51","13.91","15.01","16.11","18.51","18.91","19.91","21.11","21.11","21.81","23.71","24.31","25.51","26.11","26.51","26.81","29.71","30.11","31.11","32.11","33.11","35.51","36.81","38.01","38.61","39.11","39.51","43.11","43.41","44.51","47.01","47.61","48.71","49.31","49.81","50.21","50.91","52.11","52.51","53.61","54.61","56.61","57.11","58.41","59.21","59.81","60.51","60.71","61.11","62.11","62.79","63.51","63.91","64.81","65.96","66.11","66.71","68.61","69.21","70.21","73.21","73.61","74.71","75.97","76.31","76.71","80.21","80.81","81.81","82.96","83.01","83.41","88.61","90.21","91.21","92.27","92.51","92.91","95.91","96.41","97.41","98.21","98.61","99.71","100.21","100.61","101.71","102.44","103.01","103.41","105.31","105.81","106.81","107.11","107.51","108.51","110.11","110.51","111.61","112.48","112.61","112.91","116.51","116.91","117.91","118.91","121.31","122.61","123.61","125.03","125.11","125.41","129.21","130.11","131.11","132.02","132.31","132.71","137.71","138.31","139.31","140.3","140.61","141.11","144.81","145.51","146.51","147.51","148.65","148.81","149.21","152.71","153.11","154.11","155.11","157.71","158.11","159.11","160.11","160.53","161.01","161.41","163.81","164.21","165.21","165.81","169.41","169.81","171.21","171.91","172.41","172.91","174.61","176.11","177.91","179.29","179.91"]
            signalname = ["ED","STARTER","LSS","GWB","DISTANT","HOME","CV","STARTER","LSS/GD","LC 120 B GSS","GD","LC 120 A GSS/GD","LCC 119 GSS","GWB","DISTANT","HOME","ANU","STARTER","LSS","GWB","GD","LC 118 C GSS/GD","LC 118 B GSS","GWB","DISTANT","HOME","SGE","STARTER","LSS","GWB","DISTANT","HOME","MVPM","STARTER","LSS","GWB","GD","LC 117 A GSS/IBD","IBS/GD","LC 117  GSS","GWB","DISTANT","HOME","DC","STARTER","LSS","GWB","IBD","IBS","GWB","DISTANT","HOME","VRPD","STARTER","LSS/GD","LC 115 A GSS","IB GWB","IBD","IBS/GD","LC 115 GSS","GWB","DISTANT","HOME","SA","STARTER","LSS","GWB","DISTANT","HOME","MGSJ","STARTER","LSS/DISTANT","HOME","KPPR","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","TNT","STARTER","LSS","GWB","DISTANT","HOME","DSPT","STARTER","LSS","GWB","DISTANT","HOME","LCR","STARTER","LSS","GWB","GD","GSS      ","GWB    ","GD    ","GSS","GWB","DISTANT","HOME","BQI","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","GD","GSS","GWB","DISTANT","HOME","BDY","STARTER","LSS","GWB","GD","GSS/ IB DIS","IB HOME","GWB","DISTANT","HOME","MAP","STARTER","LSS","GWB","DISTANT","HOME","DPI","STARTER","LSS","GWB","DISTANT","HOME","DST","STARTER","LSS","GWB","GD","GSS/DISTANT","HOME","SLY","STARTER","LSS","GWB","GD","GSS/ IB D","IB HOME","GWB","DISTANT","INNER HOME","ROUTING HOM","KEY","STARTER","LSS","GWB","GD","GSS/ IB D","IB HOME","GWB","DISATANT","HOME","TPT","STARTER","LSS/D","HOME","HOME","HOME","JTJ","STARTER"]
            nstn = ["ED","CV","ANU","SGE","MVPM","DC","VRPD","SA","MGSJ","KPPR","TNT","DSPT","LCR","BQI","BDY","MAP","DPI","DST","SLY","KEY","TPT","JTJ"]
            annot = ["ED","CV","ANU","SGE","MVPM","DC","VRPD","SA","MGSJ","KPPR","TNT","DSPT","LCR","BQI","BDY","MAP","DPI","DST","SLY","KEY","TPT","JTJ"]
            dstn = [0,4.91,13.31,21.11,26.11,38.61,49.31,59.21,62.79,65.96,75.97,82.96,92.27,102.44,112.48,125.03,132.02,140.3,148.65,160.53,171.91,179.29]
            dstnannot = ["0","4.91","13.31","21.11","26.11","38.61","49.31","59.21","62.79","65.96","75.97","82.96","92.27","102.44","112.48","125.03","132.02","140.3","148.65","160.53","171.91","179.29"]
            
            signalkm = ["0","0.21","0.61","1.71","2.21","3.41","4.91","5.41","5.61","5.41","5.81","7.31","8.11","11.11","11.41","12.51","13.31","14.21","13.51","13.51","13.91","15.01","16.11","18.51","18.91","19.91","21.11","21.11","21.81","23.71","24.31","25.51","26.11","26.51","26.81","29.61","30.41","31.41","32.41","33.61","36.11","37.51","38.81","39.41","39.91","40.31","44.21","45.01","46.11","49.01","49.81","50.81","51.41","51.91","52.31","53.01","54.31","55.11","56.21","57.21","59.21","60.01","61.31","62.11","62.71","63.41","63.61","64.11","65.11","65.79","66.51","66.91","67.81","68.96","69.11","69.71","71.61","72.41","73.41","76.51","77.21","78.31","79.57","79.91","80.31","83.81","84.61","85.61","86.76","86.81","87.21","92.51","94.11","95.11","96.17","96.41","96.81","100.01","100.81","101.81","102.61","103.41","104.51","105.01","105.41","106.51","107.24","107.81","108.21","110.11","110.61","111.61","111.91","112.31","113.31","114.91","115.71","116.81","117.68","117.81","118.11","121.71","122.41","123.41","124.41","126.81","128.11","129.11","130.53","130.61","130.91","134.71","135.61","136.61","137.52","137.81","138.21","143.21","144.01","145.01","146","146.31","146.81","150.51","151.21","152.21","153.21","154.35","154.51","154.91","158.41","159.11","160.11","161.11","163.71","164.11","165.11","166.11","166.53","167.01","167.41","169.81","170.21","171.21","171.81","175.41","175.81","177.31","177.91","178.41","178.91","180.61","182.11","184.01","185.3","186.01"]
            signalname = ["ED","STARTER","LSS","GWB","DISTANT","HOME","CV","STARTER","LSS/GD","LC 120 B GSS","GD","LC 120 A GSS/GD","LCC 119 GSS","GWB","DISTANT","HOME","ANU","STARTER","LSS","GWB","GD","LC 118 C GSS/GD","LC 118 B GSS","GWB","DISTANT","HOME","SGE","STARTER","LSS","GWB","DISTANT","HOME","MVPM","STARTER","LSS","GWB","GD","LC 117 A GSS/IBD","IBS/GD","LC 117  GSS","GWB","DISTANT","HOME","DC","STARTER","LSS","GWB","IBD","IBS","GWB","DISTANT","HOME","VRPD","STARTER","LSS/GD","LC 115 A GSS","IB GWB","IBD","IBS/GD","LC 115 GSS","GWB","DISTANT","HOME","SA","STARTER","LSS","GWB","DISTANT","HOME","MGSJ","STARTER","LSS/DISTANT","HOME","KPPR","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","TNT","STARTER","LSS","GWB","DISTANT","HOME","DSPT","STARTER","LSS","GWB","DISTANT","HOME","LCR","STARTER","LSS","GWB","GD","GSS      ","GWB    ","GD    ","GSS","GWB","DISTANT","HOME","BQI","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","GD","GSS","GWB","DISTANT","HOME","BDY","STARTER","LSS","GWB","GD","GSS/ IB DIS","IB HOME","GWB","DISTANT","HOME","MAP","STARTER","LSS","GWB","DISTANT","HOME","DPI","STARTER","LSS","GWB","DISTANT","HOME","DST","STARTER","LSS","GWB","GD","GSS/DISTANT","HOME","SLY","STARTER","LSS","GWB","GD","GSS/ IB D","IB HOME","GWB","DISTANT","INNER HOME","ROUTING HOME","KEY","STARTER","LSS","GWB","GD","GSS/ IB D","IB HOME","GWB","DISATANT","HOME","TPT","STARTER","LSS/D","HOME","HOME","HOME","JTJ","STARTER"]
            nstn = ["ED","CV","ANU","SGE","MVPM","DC","VRPD","SA","MGSJ","KPPR","TNT","DSPT","LCR","BQI","BDY","MAP","DPI","DST","SLY","KEY","TPT","JTJ"]
            annot = ["ED","CV","ANU","SGE","MVPM","DC","VRPD","SA","MGSJ","KPPR","TNT","DSPT","LCR","BQI","BDY","MAP","DPI","DST","SLY","KEY","TPT","JTJ"]
            dstn = [0,4.91,13.31,21.11,26.11,39.41,51.41,62.11,65.79,68.96,79.57,86.76,96.17,107.24,117.68,130.53,137.52,146,154.35,166.53,177.91,185.3]
            dstnannot = ["0","4.91","13.31","21.11","26.11","39.41","51.41","62.11","65.79","68.96","79.57","86.76","96.17","107.24","117.68","130.53","137.52","146","154.35","166.53","177.91","185.3"]

            signalkm = ["0","0.21","0.61","1.71","2.21","3.41","4.91","5.41","5.61","5.81","6.21","7.41","8.21","11.21","11.51","12.61","13.41","13.61","14.01","14.41","14.81","15.91","17.01","18.81","19.21","20.21","21.41","21.51","22.21","23.91","24.71","25.81","26.41","26.81","27.11","29.81","30.81","31.81","32.81","34.01","36.71","37.91","39.11","39.71","40.21","40.61","44.11","44.91","46.11","48.41","49.11","50.11","50.71","51.21","51.61","52.31","53.61","54.41","55.51","56.51","58.51","59.31","60.61","61.41","62.01","62.71","62.91","63.41","64.41","65.09","65.81","66.21","66.91","68.06","68.26","68.86","70.36","71.16","72.16","74.96","75.66","76.76","78.02","78.36","78.76","82.26","83.06","84.06","85.21","85.41","85.81","91.11","92.71","93.71","94.77","95.01","95.41","98.61","99.41","100.41","101.21","102.01","103.11","103.61","104.01","105.11","105.84","106.41","106.81","108.71","109.21","110.21","110.51","111.01","112.01","113.61","114.41","115.51","116.38","116.51","116.81","120.41","121.11","122.11","123.11","125.51","126.81","127.81","129.23","129.33","129.63","133.43","134.33","135.33","136.24","136.54","136.94","141.74","142.54","143.54","144.53","144.84","145.34","148.64","149.34","150.34","151.34","152.48","152.64","153.04","155.84","156.54","157.54","158.54","160.84","161.34","162.34","163.34","163.76","164.24","164.64","166.24","166.64","167.64","168.44","171.54","172.04","173.54","174.14","174.64","175.14","176.84","178.34","180.24","181.53","182.24"]
            signalname = ["ED","STARTER","LSS","GWB","DISTANT","HOME","CV","STARTER","LSS/GD","LC 120 B GSS","GD","LC 120 A GSS/GD","LCC 119 GSS","GWB","DISTANT","HOME","ANU","STARTER","LSS","GWB","GD","LC 118 C GSS/GD","LC 118 B GSS","GWB","DISTANT","HOME","SGE","STARTER","LSS","GWB","DISTANT","HOME","MVPM","STARTER","LSS","GWB","GD","LC 117 A GSS/IBD","IBS/GD","LC 117  GSS","GWB","DISTANT","HOME","DC","STARTER","LSS","GWB","IBD","IBS","GWB","DISTANT","HOME","VRPD","STARTER","LSS/GD","LC 115 A GSS","IB GWB","IBD","IBS/GD","LC 115 GSS","GWB","DISTANT","HOME","SA","STARTER","LSS","GWB","DISTANT","HOME","MGSJ","STARTER","LSS/DISTANT","HOME","KPPR","STARTER","LSS"," GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","TNT","STARTER","LSS","GWB","DISTANT","HOME","DSPT","STARTER","LSS","GWB","DISTANT","HOME","LCR","STARTER","LSS","GWB","GD","GSS      ","GWB    ","GD    ","GSS","GWB","DISTANT","HOME","BQI","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","GD","GSS","GWB","DISTANT","HOME","BDY","STARTER","LSS","GWB","GD","GSS/ IB DIS","IB HOME","GWB","DISTANT","HOME","MAP","STARTER","LSS","GWB","DISTANT","HOME","DPI","STARTER","LSS","GWB","DISTANT","HOME","DST","STARTER","LSS","GWB","GD","GSS/DISTANT","HOME","SLY","STARTER","LSS","GWB","GD","GSS/ IB D","IB HOME","GWB","DISTANT","INNER HOME","ROUTING HOME","KEY","STARTER","LSS","GWB","GD","GSS/ IB D","IB HOME","GWB","DISATANT","HOME","TPT","STARTER","LSS/D","HOME","HOME","HOME","JTJ","STARTER"]
            nstn = ["ED","CV","ANU","SGE","MVPM","DC","VRPD","SA","MGSJ","KPPR","TNT","DSPT","LCR","BQI","BDY","MAP","DPI","DST","SLY","KEY","TPT","JTJ"]
            annot = ["ED","CV","ANU","SGE","MVPM","DC","VRPD","SA","MGSJ","KPPR","TNT","DSPT","LCR","BQI","BDY","MAP","DPI","DST","SLY","KEY","TPT","JTJ"]
            dstn = [0,4.91,13.41,21.41,26.41,39.71,50.71,61.41,65.09,68.06,78.02,85.21,94.77,105.84,116.38,129.23,136.24,144.53,152.48,163.76,174.14,181.53]
            dstnannot = ["0","4.91","13.41","21.41","26.41","39.71","50.71","61.41","65.09","68.06","78.02","85.21","94.77","105.84","116.38","129.23","136.24","144.53","152.48","163.76","174.14","181.53"]
            """

            signalkm = ["0","0.21","0.61","1.71","2.21","3.41","4.91","5.41","5.61","5.81","6.21","7.41","8.21","11.21","11.51","12.61","13.41","13.61","14.01","14.41","14.81","15.91","17.01","18.81","19.21","20.21","21.41","21.51","22.21","23.91","24.71","25.81","26.41","26.81","27.11","29.81","30.81","31.81","32.81","34.01","36.71","37.91","39.11","39.71","40.21","40.61","44.11","44.91","46.11","48.41","49.11","50.11","50.71","51.21","51.61","52.31","53.61","54.41","55.51","56.51","58.51","59.31","60.61","61.41","62.01","62.71","62.91","63.41","64.41","65.09","65.81","66.21","66.91","68.06","68.26","68.86","70.36","71.16","72.16","74.96","75.66","76.76","78.02","78.36","78.76","82.26","83.06","84.06","85.21","85.41","85.81","91.11","92.71","93.71","94.77","95.01","95.41","98.61","99.41","100.41","101.21","102.01","103.11","103.61","104.01","105.11","105.84","106.41","106.81","108.71","109.21","110.21","110.51","111.01","112.01","113.61","114.41","115.51","116.38","116.51","116.81","120.41","121.11","122.11","123.11","125.51","126.81","127.81","129.23","129.33","129.63","133.43","134.33","135.33","136.24","136.54","136.94","141.74","142.54","143.54","144.53","144.84","145.34","148.64","149.34","150.34","151.34","152.48","152.64","153.04","155.84","156.54","157.54","158.54","160.84","161.34","162.34","163.34","163.76","164.24","164.64","166.24","166.64","167.64","168.44","171.54","172.04","173.54","174.14","174.64","175.14","176.84","178.34","180.24","181.53","182.24"]
            signalname = ["ED","STARTER","LSS","GWB","DISTANT","HOME","CV","STARTER","LSS/GD","LC 120 B GSS","GD","LC 120 A GSS/GD","LCC 119 GSS","GWB","DISTANT","HOME","ANU","STARTER","LSS","GWB","GD","LC 118 C GSS/GD","LC 118 B GSS","GWB","DISTANT","HOME","SGE","STARTER","LSS","GWB","DISTANT","HOME","MVPM","STARTER","LSS","GWB","GD","LC 117 A GSS/IBD","IBS/GD","LC 117  GSS","GWB","DISTANT","HOME","DC","STARTER","LSS","GWB","IBD","IBS","GWB","DISTANT","HOME","VRPD","STARTER","LSS/GD","LC 115 A GSS","IB GWB","IBD","IBS/GD","LC 115 GSS","GWB","DISTANT","HOME","SA","STARTER","LSS","GWB","DISTANT","HOME","MGSJ","STARTER","LSS/DISTANT","HOME","KPPR","STARTER","LSS"," GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","TNT","STARTER","LSS","GWB","DISTANT","HOME","DSPT","STARTER","LSS","GWB","DISTANT","HOME","LCR","STARTER","LSS","GWB","GD","GSS      ","GWB    ","GD    ","GSS","GWB","DISTANT","HOME","BQI","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","GD","GSS","GWB","DISTANT","HOME","BDY","STARTER","LSS","GWB","GD","GSS/ IB DIS","IB HOME","GWB","DISTANT","HOME","MAP","STARTER","LSS","GWB","DISTANT","HOME","DPI","STARTER","LSS","GWB","DISTANT","HOME","DST","STARTER","LSS","GWB","GD","GSS/DISTANT","HOME","SLY","STARTER","LSS","GWB","GD","GSS/ IB D","IB HOME","GWB","DISTANT","INNER HOME","ROUTING HOME","KEY","STARTER","LSS","GWB","GD","GSS/ IB D","IB HOME","GWB","DISATANT","HOME","TPT","STARTER","LSS/D","HOME","HOME","HOME","JTJ","STARTER"]
            nstn = ["ED","CV","ANU","SGE","MVPM","DC","VRPD","SA","MGSJ","KPPR","TNT","DSPT","LCR","BQI","BDY","MAP","DPI","DST","SLY","KEY","TPT","JTJ"]
            annot = ["ED","CV","ANU","SGE","MVPM","DC","VRPD","SA","MGSJ","KPPR","TNT","DSPT","LCR","BQI","BDY","MAP","DPI","DST","SLY","KEY","TPT","JTJ"]
            dstn = [0,4.91,13.41,21.41,26.41,39.71,50.71,61.41,65.09,68.06,78.02,85.21,94.77,105.84,116.38,129.23,136.24,144.53,152.48,163.76,174.14,181.53]
            dstnannot = ["0","4.91","13.41","21.41","26.41","39.71","50.71","61.41","65.09","68.06","78.02","85.21","94.77","105.84","116.38","129.23","136.24","144.53","152.48","163.76","174.14","181.53"]


        if route == "SATPJ":
            signalkm = ["0","0","0.09","10.05","11.01","12.02","13.3","13.4","13.8","14.01","23.06","24.02","25.02","25.87","26.4","26.8","32.02","33.5","36.07","37.05","38.05","39.6","39.8","40.01","48.08","49.04","50.04","51.4","51.8","52.01","56.01","57.8","67","67.06","68.06","69.5","69.9","70.02","72.05","78.5","81.06","82.02","83.02","83.08","85","85.2","85.6","87.6","88.9","90.4","90.8","91.8","92","92.4","93.4","94.52","94.6","94.9","95.2","95.6","96.6","98.9","99.4","100.4","101","101.5","102.5","103.48","103.7","104","107.6","107.64","107.8","108.5","109.5","110.41","110.7","111","112.5","112.8","113.9","114.2","114.8","115.2","116.3","116.3","116.7","117.8","118.49","118.7","119.2","120.2","120.6","122.1","122.2","123","123.5","124.3","124.7","125","125.7","126.1","126.5","127.6","128.3","128.7","129.8","130.4","130.9","131.9","133.19","133.5","133.8","134.6","136","136.4","137.4","138.4","139.17","140.37","140.57","141.67","143.57","144.09","144.67","144.87","145.77","146.07","147.17","149.06","150.27","150.77","151.77","153.47","154.82","154.97","155.42","156.02","156.62","157.72","158.72","159.02","159.72","160.72","162.02","162.58","163.68","163.98","164.98","164.98","165.38","167.08"]
            signalname = ["SA","STARTER","LSS","GWB","DISTANT","HOME","MALR","STARTER","LSS/GD","GSS","GWB","DISTANT","HOME","RASP","STARTER","LSS","SWB","PCTM - H","GWB","DISTANT","HOME","KLGN","STARTER","LSS","GWB","DISTANT","HOME","NMKL","STARTER","LSS","SWB","LDVD - H","GWB","DISTANT","HOME","MONR","STARTER","LSS","SWB","VNGL - H","GWB","GD","G/D","HOME","KRR","STARTER","LSS/GD","LC39 GSS/GD","LC40GSS","GWB","GD","LC41/GSS","GWB","DISTANT","HOME","VRQ","STARTER","LSS","GWB","GD","LC43/GSS","GWB","GD","LC44/GSS","GWB","DISTANT","HOME","MYU","STARTER","LSS","SITHALAVAI","SEV","GWB","DISTANT","HOME","MMH","STARTER","LSS","GWB","DISTANT","HOME","LP","STARTER","LSS/GD","GWB","GSS LC49","GD","LC50 GSS","TIC","GWB","GD","GSS LC49","GWB","D/GD","LC53 GSS/GD","LC54 GSS","HOME","KLT","STARTER","LSS/GD","LC55 GSS","GWB","G/D","LC NO 57 GSS","GWB","GD","LC59 GSS","GWB","DISTANT","HOME","PLI","STARTER","LSS/GD","LC63 GSS","GWB","GD","LC64 GSS/D","HOME","PGN","STARTER","LSS/GD","LC67/GSS/D","HOME","EL","STARTER","LSS","GWB","GD","LC73 GSS","JPM","GWB","GD/D","LC75 GSS","HOME","MTNL","STARTER","LSS/GD","LC 78 GSS","GWB","GD","LCNO 80 GSS","GWB","GD","LC 83 GSS/D","HOME","TP","STARTER","LSS/D","PALAKARAI","TPE","HOME","TPJ"]
            nstn = ["SA","MALR","RASP","KLGN","NMKL","MONR","KRR","VRQ","MYU","MMH","LP","KLT","PLI","PGN","EL","MTNL","TP","TPJ"]
            annot = ["SA","MALR","RASP","KLGN","NMKL","MONR","KRR","VRQ","MYU","MMH","LP","KLT","PLI","PGN","EL","MTNL","TP","TPJ"]
            dstn = [0,13.3,25.87,39.6,51.4,69.5,85,94.52,103.48,110.41,114.2,124.3,133.19,139.17,144.09,154.82,162.58,167.08]
            dstnannot = ["0","13.3","25.87","39.6","51.4","69.5","85","94.52","103.48","110.41","114.2","124.3","133.19","139.17","144.09","154.82","162.58","167.08"]



        if route == "TPJSA":
            signalkm = ["0","0.4","0.6","1.9","2.5","3.2","4.34","4.4","4.6","5.7","5.9","6.4","7.5","8","8.6","9.6","10.3","11.05","11.5","11.8","12.6","14","14.6","15.6","15.81","16.7","17.1","18.1","19.48","19.5","19.8","20.3","21.6","22.53","22.8","23","23.8","25.2","25.6","26.6","27.6","28.51","28.8","29","30","30.4","31.4","32.3","32.6","33.6","34","34.4","35.4","36.4","37.4","37.6","37.9","38.3","39","39.6","40","41","42","42.5","43.4","43.9","46.2","46.8","47.3","47.5","49","49.4","50.4","51.19","51.6","51.9","54.06","55.8","56.3","57.3","58.22","58.7","58.9","59.2","59.7","60.8","63.2","63.6","64.6","64.8","65.5","66.5","67.18","67.7","68","69.4","70.8","71.2","71.9","72.2","73.6","75.6","76.7","77.3","78.2","79.15","87.21","89.21","90.17","92.17","92.72","93.19","94.13","103.22","104.42","108.16","108.21","110.18","110.82","111.21","112.14","120.18","121.14","122.18","122.62","123.21","124.14","128.14","128.72","134.15","134.21","135.21","136.35","137.14","137.17","146.19","147.15","148.16","148.21","148.92","150.14","150.17","160.15","160.19","161.2","162.22"]
            signalname = ["TPJ","STARTER","LSS/D","PALAKARAI","TPE","HOME","TP","STARTER","LSS/GD","LC 83 GSS","GWB","GD","LC 82 GSS","GWB","D/GD","LC 78 GSS","HOME","MTNL","STARTER","LSS/GD","LC75 GSS","GWB","GD","LC73,72,71 GSS","JPM HALT","GWB","DISTANT","HOME","EL","STARTER","LSS/GD","LC 67 GSS/D","HOME","PGN","STARTER","LSS/GD","LC64 GSS","GWB","GD","LC 63 GSS/D","HOME","PLI","STARTER","LSS","GWB","GD","LC 59 GSS","GWB","GD","LC57 GSS","GWB","GD","LC 55 GSS/D","HOME","KLT","STARTER","LSS/GD","LC54 GSS/GD","LC53 GSS","GWB","GD","GSS LC52","GWB","GD","GSS/GWB","GD","HOME","LP","STARTER","LSS","GWB","DISTANT","HOME","MMH","STARTER","LSS","SEV HALT","GWB","DISTANT","HOME","MYU","STARTER","LSS","GWB","GD","LC44 GSS","GWB","GD","LC43 GSS","GWB","DISTANT","HOME","VRQ","STARTER","LSS/GD","LC41 GSS","GWB","GD","49 GSS/D","LC40 GSS/GD","LC39 GSS/D","HOME","KRR","STARTER","LSS/GD","GSS","SWB","GWB","DISTANT","HOME","MONR","STARTER","LSS","SWB","LDVD - H","GWB","DISTANT","HOME","NMKL","STARTER","LSS","GWB","DISTANT","HOME","KLGN","STARTER","LSS","SWB","PCTM","GWB","DISTANT","HOME","RASP","STARTER","LSS","GWB","DISTANT","GSS/DISTANT","HOME","MALR","STARTER","LSS","GWB","DISTANT","HOME","SA"]
            nstn = ["TPJ","TP","MTNL","EL","PGN","PLI","KLT","LP","MMH","MYU","VRQ","KRR","MONR","NMKL","KLGN","RASP","MALR","SA"]
            annot = ["TPJ","TP","MTNL","EL","PGN","PLI","KLT","LP","MMH","MYU","VRQ","KRR","MONR","NMKL","KLGN","RASP","MALR","SA"]
            dstn = [0,4.34,11.05,19.48,22.53,28.51,37.4,46.8,51.19,58.22,67.18,76.7,92.72,110.82,122.62,136.35,148.92,162.22]
            dstnannot = ["0","4.34","11.05","19.48","22.53","28.51","37.4","46.8","51.19","58.22","67.18","76.7","92.72","110.82","122.62","136.35","148.92","162.22"]



        if route == "EDTPJ":
            signalkm = ["0.00","0.00","0.70","1.13","1.18","2.05","3.06","4.06","5.06","6.07","7.19","8.15","8.21","9.08","10.09","11.00","11.14","11.22","13.19","14.06","15.06","15.11","15.19","16.20","18.07","19.00","19.11","19.20","20.08","21.08","22.12","22.19","23.07","24.07","24.20","26.01","27.09","27.18","28.17","29.18","30.15","31.05","32.00","32.08","32.14","33.05","33.14","34.15","35.15","36.02","37.03","38.00","38.15","38.24","39.05","39.13","40.12","41.15","42.05","43.06","48.04","48.13","49.14","50.00","50.19","51.01","52.19","53.07","54.00","54.09","55.11","55.17","56.03","57.07","61.16","62.08","63.10","64.80","65.00","65.40","67.40","68.70","70.20","70.60","71.60","71.80","72.20","73.20","74.32","74.40","74.70","75.00","75.40","76.40","78.70","79.20","80.20","80.80","81.30","82.30","83.28","83.50","83.80","87.4","87.44","87.60","88.30","89.30","90.21","90.50","90.80","92.30","92.60","93.70","94.60","94.76","95.00","96.10","96.10","96.50","97.60","98.29","98.50","99.00","100.00","100.40","101.90","102.00","102.80","103.30","104.10","104.50","104.80","105.50","105.90","106.30","107.40","108.10","108.50","109.60","110.20","110.70","111.70","112.99","113.30","113.60","114.40","115.80","116.20","117.20","118.20","118.97","119.40","119.60","120.70","121.50","122.02","122.02","122.60","122.80","123.50","123.80","124.90","125.69","126.90","127.40","128.40","129.10","130.45","130.60","130.90","131.50","132.10","132.60","133.60","133.90","134.30","135.30","136.60","137.16","137.70","138.00","139.00","139.00","139.40","141.10"]
            signalname = ["ED","STARTER","LSS/GD","LC 12D GSS","GWB","GD","LC 3 GSS/GD","LC 4 GSS/GD","LC 5 GSS/GD","LC 6 GSS/GD","LC 8 GSS/GD","LC 9 GSS","GWB","DISTANT","HOME","CVD","STARTER","LSS","GWB","GD","LC13 GSS","GWB","GD","LC14 GSS/D","HOME","PAS","STARTER","LC18 LSS/GD","LC18 GSS/GD","LC19 GSS/GD","LC20 GSS","GWB","GD","LC22 GSS/GD","LC23 GSS/GD","LC24 GSS","GWB","GD","LC25 GSS/GD","LC26 GSS/GD/D","LC27 GSS","HOME","URL","STARTER","LSS","GWB","GD","LC28A GSS","GWB","DISTANT","HOME","KMD","STARTER","LSS","GWB","GD","LC31A GSS ","GWB","GD","LC32 GSS ","GWB","DISTANT","HOME","PGR","STARTER","LSS","GWB","DISTANT","HOME","MPLM","STARTER","LSS/GD","LC34 GSS/GD","LC35 GSS","GWB","DISTANT","HOME","KRR","STARTER","LSS/GD","LC39 GSS/GD","LC40GSS","GWB","GD","LC41/GSS","GWB","DISTANT","HOME","VRQ","STARTER","LSS","GWB","GD","LC43/GSS","GWB","GD","LC44/GSS","GWB","DISTANT","HOME","MYU","STARTER","LSS","SITHALAVAI","SEV","GWB","DISTANT","HOME","MMH","STARTER","LSS","GWB","DISTANT","HOME","LP","STARTER","LSS/GD","GWB","GSS LC49","GD","LC50 GSS","TIC","GWB","GD","GSS LC49","GWB","D/GD","LC53 GSS/GD","LC54 GSS","HOME","KLT","STARTER","LSS/GD","LC55 GSS","GWB","G/D","LC NO 57 GSS","GWB","GD","LC59 GSS","GWB","DISTANT","HOME","PLI","STARTER","LSS/GD","LC63 GSS","GWB","GD","LC64 GSS/D","HOME","PGN","STARTER","LSS/GD","LC67/GSS/D","HOME","EL","ELAMANUR","STARTER","LSS","GWB","GD","LC73 GSS","JPM","GWB","GD/D","LC75 GSS","HOME","MTNL","STARTER","LSS/GD","LC 78 GSS","GWB","GD","LCNO 80 GSS","GWB","GD","LC 83 GSS/D","HOME","TP","STARTER","LSS/D","PALAKARAI","TPE","HOME","TPJ"]
            nstn = ["ED","CVD","PAS","URL","KMD","PGR","MPLM","KRR","VRQ","MYU","MMH","LP","KLT","PLI","PGN","EL","MTNL","TP","TPJ"]
            annot = ["ED","CVD","PAS","URL","KMD","PGR","MPLM","KRR","VRQ","MYU","MMH","LP","KLT","PLI","PGN","EL","MTNL","TP","TPJ"]
            dstn = [0.00,11.00,19.00,32.00,38.00,50.00,54.09,64.80,74.32,83.28,90.21,94.60,104.10,112.99,118.97,122.02,130.45,137.16,141.10]
            dstnannot = ["0.00","11.00","19.00","32.00","38.00","50.00","54.09","64.80","74.32","83.28","90.21","94.60","104.10","112.99","118.97","122.02","130.45","137.16","141.10"]

        if route == "TPJED":
            signalkm = ["0","0.4","0.6","1.9","2.5","3.2","4.34","4.4","4.6","5.7","5.9","6.4","7.5","8","8.6","9.6","10.3","11.05","11.5","11.8","12.6","14","14.6","15.6","15.81","16.7","17.1","18.1","19.48","19.5","19.8","20.3","21.6","22.53","22.8","23","23.8","25.2","25.6","26.6","27.6","28.51","28.8","29","30","30.4","31.4","32.3","32.6","33.6","34","34.4","35.4","36.4","37.4","37.6","37.9","38.3","39","39.6","40","41","42","42.5","43.4","43.9","46.2","46.8","47.3","47.5","49","49.4","50.4","51.19","51.6","51.9","54.06","55.8","56.3","57.3","58.22","58.7","58.9","59.2","59.7","60.8","63.2","63.6","64.6","64.8","65.5","66.5","67.18","67.7","68","69.4","70.8","71.2","71.9","72.2","73.6","75.6","76.7","77.4","78.4","82.5","83.4","84.4","85.4","86.3","86.3","87.3","87.4","89.4","89.5","90.5","91.06","91.5","92.4","96.5","97.4","98.4","99.5","100.3","101.3","101.4","101.5","102.5","103.3","104.3","104.4","105.4","105.5","106.5","107.5","108.4","109.22","109.4","110.4","110.4","110.5","111.5","112.5","114.3","114.4","115.4","116.5","117.4","117.5","117.5","119.3","120.4","120.5","121.5","122.58","123.4","123.4","124.5","124.5","125.4","126.4","127.5","128.4","129.5","130.17","131.3","131.4","131.4","131.5","132.5","133.5","135.4","136.4","137.4","138.4","138.5","139.5","140.5","141.5"]
            signalname = ["TPJ","STARTER","LSS/D","PALAKARAI","TPE","HOME","TP","STARTER","LSS/GD","LC 83 GSS","GWB","GD","LC 82 GSS","GWB","D/GD","LC 78 GSS","HOME","MTNL","STARTER","LSS/GD","LC75 GSS","GWB","GD","LC73,72,71 GSS","JPM HALT","GWB","DISTANT","HOME","EL","STARTER","LSS/GD","LC 67 GSS/D","HOME","PGN","STARTER","LSS/GD","LC64 GSS","GWB","GD","LC 63 GSS/D","HOME","PLI","STARTER","LSS","GWB","GD","LC 59 GSS","GWB","GD","LC57 GSS","GWB","GD","LC 55 GSS/D","HOME","KLT","STARTER","LSS/GD","LC54 GSS/GD","LC53 GSS","GWB","GD","GSS LC52","GWB","GD","GSS/GWB","GD","HOME","LP","STARTER","LSS","GWB","DISTANT","HOME","MMH","STARTER","LSS","SEV HALT","GWB","DISTANT","HOME","MYU","STARTER","LSS","GWB","GD","LC44 GSS","GWB","GD","LC43 GSS","GWB","DISTANT","HOME","VRQ","STARTER","LSS/GD","LC41 GSS","GWB","GD","49 GSS/D","LC40 GSS/GD","LC39 GSS/D","HOME","KRR","STARTER","LSS","GWB","GD","LC35 GSS/GD","LC34 GSS/D","HOME","MPLM","STARTER","LSS","GWB","DISTANT","HOME","PGR","STARTER","LSS","GWB","GD","LC32 GSS","GWB","GD","LC31A GSS","GWB","DISTANT","HOME","KMD","STARTER","LSS","GWB","GD","LC28A GSS","GWB","DISTANT","URL","HOME","STARTER","LSS/GD","LC27 GSS/GD","LC26 GSS/GD","LC25 GSS","GWB","GD","LC24 GSS/GD","LC23 GSS/GD","LC22 GSS","GWB","GD","LC20 GSS/GD","LC19 GSS/GD","LC18 GSS/D","HOME","PAS","STARTER","LSS/GD","LC14 GSS","GWB","GD","LC13 GSS","GWB","GD","HOME","CVD","STARTER","LSS","GWB","GD","LC9 GSS/GD","LC8 GSS/GD","LC6 GSS/GD","LC5 GSS/GD","LC4 GSS/GD","GWB","GD","LC121D GSS/D","HOME","ED"]
            nstn = ["TPJ","TP","MTNL","EL","PGN","PLI","KLT","LP","MMH","MYU","VRQ","KRR","MPLM","PGR","KMD","URL","PAS","CVD","ED"]
            annot = ["TPJ","TP","MTNL","EL","PGN","PLI","KLT","LP","MMH","MYU","VRQ","KRR","MPLM","PGR","KMD","URL","PAS","CVD","ED"]
            dstn = [0,4.34,11.05,19.48,22.53,28.51,37.4,46.8,51.19,58.22,67.18,76.7,86.3,91.06,103.3,109.22,122.58,130.17,141.5]
            dstnannot = ["0","4.34","11.05","19.48","22.53","28.51","37.4","46.8","51.19","58.22","67.18","76.7","86.3","91.06","103.3","109.22","122.58","130.17","141.5"]



        if route == "EDIGU":
            signalkm = ["0","0.08","1.1","3.9","5.1","6.1","6.39","7.1","10.85","10.94","11.92","13.59","14.3","14.5","16.8","16.92","17.92","19.05","19.8","19.9","23.92","25.3","26.3","26.7","26.82","26.9","28.94","30.5","30.8","33.86","33.96","35.02","36.65","36.9","37","40.86","42.1","43.3","43.8","43.98","44.98","45.98","46.84","47.9","48.94","50.16","50.6","51","55.9","56.3","57.3","58.2","58.6","58.9","59.4","59.8","60.9","62.2","63.2","63.8","64.8","65.5","65.9","67.1","67.8","67.8","68.7","68.8","69.8","70.8","72.7","72.8","73.8","73.8","75.9","76.3","77","77.3","79.8","79.9","80.9","82.8"]
            signalname = ["ED","STARTER","LSS","GWB","DISTANT","HOME","TPM","STARTER","GWB","DISTANT","HOME","PY","STARTER","LSS","GWB","DISTANT","HOME","IGR","STARTER","LSS","GWB","DISTANT","HOME","VZ","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","UKL","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","GD","GSS","GWB","DISTANT","HOME","HOME 2","TUP","STARTER","LSS","GWB","DISTANT","HOME","VNJ","STARTER","LSS","GWB","GD","GSS/DISTANT","GSS","GWB","GD","GSS","GWB","DISTANT","HOME","SNO","STARTER","LSS/GD","GSS","GWB","GD","GSS","GWB","GD","G/D","HOME","SUU","STARTER","LSS","GWB","DISTANT","HOME","IGU"]
            nstn = ["ED","TPM","PY","IGR","VZ","UKL","TUP","VNJ","SNO","SUU","IGU"]
            annot = ["ED","TPM","PY","IGR","VZ","UKL","TUP","VNJ","SNO","SUU","IGU"]
            dstn = [0,6.39,13.59,19.05,26.7,36.65,50.16,58.2,67.8,76.3,82.8]
            dstnannot = ["0","6.39","13.59","19.05","26.7","36.65","50.16","58.2","67.8","76.3","82.8"]


        if route == "IGUED":
            signalkm = ["0","0.5","1.3","3.8","4.3","5.3","6.6","6.5","6.9","8.3","9.5","10","11.1","11.5","12","13","14","15.1","15.2","15.6","15.9","16.5","17.6","18.7","19.2","20.2","21.6","22.2","22.8","23.8","24.7","24.9","25.4","30.3","30.7","31.7","32.8","32.9","33.4","34.2","35.2","35.8","36.8","37.8","38.4","39.6","43","43.5","45.4","46.3","46.8","47.2","54","54.3","55.4","55.9","56.6","57","60.6","61.8","62.9","63.9","64.1","64.4","67.1","67.5","68.4","69.8","69.7","70.1","74.1","74.7","75.7","76.9","76.9","77.1","78.6","79.2","80.9","81.3","82.2","82.9"]
            signalname = ["IGU","STARTER","LSS","GWB","DISTANT","HOME","SUU","STARTER","LSS/GD","LC142 GSS","GWB","G D","LC141 GSS","GWB","GD","LC140 GSS/DIST","HOME","SNO","STARTER","LSS","GWB","G D","LC137 GSS","GWB","GD","LC136 GSS/GD","LC 135 GSS","GWB","DISTANT","HOME","VNJ","STARTER","LSS","GWB","DISTANT","HOME","TUP","STARTER","LSS/G D","GSS","GWB","GD","LC 131 E GSS","IBS GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","UKL","STARTER","LSS ","GWB","DISTANT","HOME","VZ","STARTER","LSS ","GWB","DISTANT","HOME","IGR","STARTER","LSS","GWB","DISTANT","HOME","PY","STARTER","LSS","GWB","DISTANT","HOME","TPM","STARTER","LSS","GWB","GD","GSS/DISTANT","HOME","RT HOME","ED"]
            nstn = ["IGU","SUU","SNO","VNJ","TUP","UKL","VZ","IGR","PY","TPM","ED"]
            annot = ["IGU","SUU","SNO","VNJ","TUP","UKL","VZ","IGR","PY","TPM","ED"]
            dstn = [0,6.6,15.1,24.7,32.8,46.3,55.9,63.9,69.8,76.9,82.9]
            dstnannot = ["0","6.6","15.1","24.7","32.8","46.3","55.9","63.9","69.8","76.9","82.9"]

        if route == "KRRDG":
            signalkm = ["0","0.1","1","5.8","6.3","7.3","13","13.5","14.5","15.3","15.7","16","26.8","27.3","28.3","29.1","29.5","29.9","32.6","33","34","42.3","42.7","43.7","44.7","45.3","46.3","49.9","50.5","51.5","52.5","53.3","53.7","54","69.9","70.3","71.3","72.2","73.6","73.6"]
            signalname = ["KRR","Starter","LSS/GD","GSS","LC 04 GWB","GD","GSS","GWB","DISTANT","HOME","VEI","STARTER","LSS","GWB","DISTANT","HOME","PALM","STARTER","LSS","GWB","GD","LC 12 GSS","GWB","GD","LC 18 GSS","GWB","GD","LC 20 GSS","GWB","GD","LC 22 GSS/D","HOME","EDU","STARTER","LSS","GWB","GD/D","GSS","HOME","DG"]
            nstn = ["KRR","VEI","PALM","EDU","DG"]
            annot = ["KRR","VEI","PALM","EDU","DG"]
            dstn = [0,15.7,29.5,53.7,73.6]
            dstnannot = ["0","15.7","29.5","53.7","73.6"]

        if route == "DGKRR":
            signalkm = ["0","1","1.2","17.8","18.4","19.4","20.3","20.5","20.8","21.5","25.3","25.8","26.8","28.1","28.4","29.4","37.7","38.1","39.1","41.9","42.4","43.5","44.5","44.7","45","55.9","56.4","57.4","58.3","58.5","58.8","64.4","64.8","65.8","70.7","71.1","72.1","72.7","73.9"]
            signalname = ["DG","STARTER","LSS","GWB","DISTANT","HOME","EDU","STARTER","LSS/GD","LC 22  GSS","GWB","GD","LC 20 GSS","GWB","GD","LC 18 GSS","GWB","GD","LC 12 GSS","GWB","DISTANT","HOME","PALM","STARTER","LSS","GWB","DISTANT","HOME","VEI","STARTER","LSS","GWB","GD","LC 4 GSS","GWB","GD/D","GSS","HOME","KRR"]
            nstn = ["DG","EDU","PALM","VEI","KRR"]
            annot = ["DG","EDU","PALM","VEI","KRR"]
            dstn = [0,20.3,44.5,58.3,73.9]
            dstnannot = ["0","20.3","44.5","58.3","73.9"]

        if route == "EDPGTA":
            signalkm = ["0","0.08","1.1","3.9","5.1","6.1","6.39","7.1","10.85","10.94","11.92","13.59","14.3","14.5","16.8","16.92","17.92","19.05","19.8","19.9","23.92","25.3","26.3","26.7","26.82","26.9","28.94","30.5","30.8","33.86","33.96","35.02","36.65","36.9","37","40.86","42.1","43.3","43.8","43.98","44.98","45.98","46.84","47.9","48.94","50.16","50.6","51","55.9","56.3","57.3","58.2","58.6","58.9","59.4","59.8","60.9","62.2","63.2","63.8","64.8","65.5","65.9","67.1","67.8","67.8","68.7","68.8","69.8","70.8","72.7","72.8","73.8","73.8","75.9","76.3","77","77.3","79.8","79.9","80.9","82.8","83.1","83.6","88.8","88.9","89.9","90","90.8","91.8","93.2","93.6","93.9","100.5","100.8","102","102.8","103.7","103.9","105.6","105.8","107.7","107.7","114.6","114.8","115.8","116.6","117.1","117.8","122.6","121.8","122.7","122.7","122.9","123.9","125.6","125.9","127.7","128.5","128.8","129.72","130.86","138.6","138.9","139.9","140.9","141.7"]
            signalname = ["ED","STARTER","LSS","GWB","DISTANT","HOME","TPM","STARTER","GWB","DISTANT","HOME","PY","STARTER","LSS","GWB","DISTANT","HOME","IGR","STARTER","LSS","GWB","DISTANT","HOME","VZ","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","UKL","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","GD","GSS","GWB","DISTANT","HOME","HOME 2","TUP","STARTER","LSS","GWB","DISTANT","HOME","VNJ","STARTER","LSS","GWB","GD","GSS/DISTANT","GSS","GWB","GD","GSS","GWB","DISTANT","HOME","SNO","STARTER","LSS/GD","GSS","GWB","GD","GSS","GWB","GD","G/D","HOME","SUU","STARTER","LSS","GWB","DISTANT","HOME","IGU","STARTER","LSS","GWB","GD","GSS","GWB","DISTANT","HOME","PTJ","STARTER","LSS","GWB","DISTANT","HOME","MDKI","STARTER","LSS","GWB","DISTANT","HOME","ETTIMADAI","GWB","DISTANT","HOME","WAL","STARTER","LSS","GWB","DISTANT","HOME","CHULLIMADAI","GD","GSS","GWB","DISTANT","HOME","KJKD","STARTER","LSS/GD","GSS","GWB","DISTANT","INNER HOME","HOME","PGT"]
            nstn = ["ED","TPM","PY","IGR","VZ","UKL","TUP","VNJ","SNO","SUU","IGU","PTJ","MDKI","WAL","KJKD","PGT"]
            annot = ["ED","TPM","PY","IGR","VZ","UKL","TUP","VNJ","SNO","SUU","IGU","PTJ","MDKI","WAL","KJKD","PGT"]
            dstn = [0,6.39,13.59,19.05,26.7,36.65,50.16,58.2,67.8,76.3,82.8,93.2,102.8,116.6,128.5,141.7]
            dstnannot = ["0","6.39","13.59","19.05","26.7","36.65","50.16","58.2","67.8","76.3","82.8","93.2","102.8","116.6","128.5","141.7"]            



        if route == "PGTEDA":
            signalkm = ["0","0.55","0.94","4.2","4.9","5.9","5.9","8.1","9.7","10.68","11.9","12.98","13.56","13.84","17.2","17.98","19.92","19.92","22.1","22.82","23.8","23.98","24.92","25.7","31.1","31.88","32.92","33.98","35.1","36.66","37.64","38.24","38.78","38.98","45.1","45.94","46.86","48.5","48.8","49.1","49.3","49.7","50.7","52.5","52.9","53.9","56.4","56.8","57.8","58.8","59.3","60.1","62.6","63.1","64.1","65.1","65.3","65.7","67.1","68.3","68.8","69.9","70.3","70.8","71.8","72.8","73.9","74","74.4","74.7","75.3","76.4","77.5","78","79","80.4","81","81.6","82.6","83.5","83.7","84.2","89.1","89.5","90.5","91.6","91.7","92.2","93","94","94.6","95.6","96.6","97.2","98.4","101.8","102.3","104.2","105.1","105.6","106","112.8","113.1","114.2","114.7","115.4","115.8","119.4","120.6","121.7","122.7","122.9","123.2","125.9","126.3","127.2","128.1","128.5","128.9","132.9","133.5","134.5","135.7","135.7","135.9","137.4","139.7","140.1","141","141.7"]
            signalname = ["PGT","STARTER","LSS","GWB","DISTANT","HOME","KOTTEKAD","GWB","GD","GSS","HOME","KJKD","STARTER","LSS","GWB","DISTANT","HOME","CHULLIMADA","GWB","DISTANT","HOME","WAL","STARTER","LSS","GWB","DISTANT","HOME","ETTIMADAI","GWB","DISTANT","HOME","MDKI","STARTER","LSS","GWB","DISTANT","HOME","PTJ","STARTER","LSS","GWB","G D","LC 147 GSS","GWB","G D","LC 146 GSS","GWB","DISTANT","HOME","IGU","STARTER","LSS","GWB","DISTANT","HOME","SUU","STARTER","LSS/GD","LC142 GSS","GWB","G D","LC141 GSS","GWB","GD","LC140 GSS/DIST","HOME","SNO","STARTER","LSS","GWB","G D","LC137 GSS","GWB","GD","LC136 GSS/GD","LC 135 GSS","GWB","DISTANT","HOME","VNJ","STARTER","LSS","GWB","DISTANT","HOME","TUP","STARTER","LSS/G D","GSS","GWB","GD","LC 131 E GSS","IBS GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","UKL","STARTER","LSS ","GWB","DISTANT","HOME","VZ","STARTER","LSS ","GWB","DISTANT","HOME","IGR","STARTER","LSS","GWB","DISTANT","HOME","PY","STARTER","LSS","GWB","DISTANT","HOME","TPM","STARTER","LSS","GWB","DISTANT","HOME","RT HOME","ED"]
            nstn = ["PGT","KJKD","WAL","MDKI","PTJ","IGU","SUU","SNO","VNJ","TUP","UKL","VZ","IGR","PY","TPM","ED"]
            annot = ["PGT","KJKD","WAL","MDKI","PTJ","IGU","SUU","SNO","VNJ","TUP","UKL","VZ","IGR","PY","TPM","ED"]
            dstn = [0,12.98,23.98,38.24,48.5,58.8,65.1,73.9,83.5,91.6,105.1,114.7,122.7,128.1,135.7,141.7]
            dstnannot = ["0","12.98","23.98","38.24","48.5","58.8","65.1","73.9","83.5","91.6","105.1","114.7","122.7","128.1","135.7","141.7"]

        if route == "EDPGTB":
            signalkm = ["0","0.08","1.1","3.9","5.1","6.1","6.39","7.1","10.85","10.94","11.92","13.59","14.3","14.5","16.8","16.92","17.92","19.05","19.8","19.9","23.92","25.3","26.3","26.7","26.82","26.9","28.94","30.5","30.8","33.86","33.96","35.02","36.65","36.9","37","40.86","42.1","43.3","43.8","43.98","44.98","45.98","46.84","47.9","48.94","50.16","50.6","51","55.9","56.3","57.3","58.2","58.6","58.9","59.4","59.8","60.9","62.2","63.2","63.8","64.8","65.5","65.9","67.1","67.8","67.8","68.7","68.8","69.8","70.8","72.7","72.8","73.8","73.8","75.9","76.3","77","77.3","79.8","79.9","80.9","82.8","83.1","83.6","88.8","88.9","89.9","90","90.8","91.8","93.2","93.6","93.95","100.83","101.99","102.8","103.72","103.93","105.85","107.73","107.74","114.71","115.89","116.6","117.15","117.75","123","124.1","124.8","125.93","127.09","128.5","128.86","129.71","130.85","133.87","134.91","136.228","138.97","139.95","140.87","141.7"]
            signalname = ["ED","STARTER","LSS","GWB","DISTANT","HOME","TPM","STARTER","GWB","DISTANT","HOME","PY","STARTER","LSS","GWB","DISTANT","HOME","IGR","STARTER","LSS","GWB","DISTANT","HOME","VZ","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","UKL","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","GD","GSS","GWB","DISTANT","HOME","HOME 2","TUP","STARTER","LSS","GWB","DISTANT","HOME","VNJ","STARTER","LSS","GWB","GD","GSS/DISTANT","GSS","GWB","GD","GSS","GWB","DISTANT","HOME","SNO","STARTER","LSS/GD","GSS","GWB","GD","GSS","GWB","GD","G/D","HOME","SUU","STARTER","LSS","GWB","DISTANT","HOME","IGU","STARTER","LSS","GWB","GD","GSS","GWB","DISTANT","HOME","PTJ","STARTER","LSS","DISTANT","HOME","MDKI","STARTER","LSS","DISTANT","HOME","ETTIMADAI","DISTANT","HOME","WAL","STARTER","LSS","GWB","IBD","IBSS","DISTANT","HOME","KJKD","STARTER","LSS/GD","GSS","DISTANT","HOME","KOTTAKADU","DISTANT","HOME","INNER HOME","PGT"]
            nstn = ["ED","TPM","PY","IGR","VZ","UKL","TUP","VNJ","SNO","SUU","IGU","PTJ","MDKI","WAL","KJKD","PGT"]
            annot = ["ED","TPM","PY","IGR","VZ","UKL","TUP","VNJ","SNO","SUU","IGU","PTJ","MDKI","WAL","KJKD","PGT"]
            dstn = [0,6.39,13.59,19.05,26.7,36.65,50.16,58.2,67.8,76.3,82.8,93.2,102.8,116.6,128.5,141.7]
            dstnannot = ["0","6.39","13.59","19.05","26.7","36.65","50.16","58.2","67.8","76.3","82.8","93.2","102.8","116.6","128.5","141.7"]

        if route == "PGTEDB":
            signalkm = ["0","0.55","0.95","4.2","4.91","5.91","5.91","9.1","9.71","10.69","11.91","13.48","13.57","13.87","16.9","17.6","18.7","22.1","22.75","23.73","24.491","24.87","25.65","31.1","31.89","32.93","33.5","35.1","36.67","37.65","37.98","38.79","38.99","45.2","45.95","46.87","48.5","48.8","49.1","49.3","49.7","50.7","56.4","56.8","57.8","58.8","59.3","60.1","62.6","63.1","64.1","65.4","65.6","65.7","67.1","68.3","68.8","69.9","70.3","70.8","71.8","72.8","73.9","74","74.4","74.7","75.3","76.4","77.5","78","79","80.4","81","81.6","82.6","83.5","83.7","84.2","89.1","89.5","90.5","91.6","91.7","92.2","93","94","94.6","95.6","96.6","97.2","98.4","101.8","102.3","104.2","105.1","105.6","106","112.8","113.1","114.2","114.7","115.4","115.8","119.4","120.6","121.7","122.7","122.9","123.2","125.9","126.3","127.2","128.1","128.5","128.9","132.9","133.5","134.5","135.7","135.7","135.9","137.4","139.7","140.1","141","141.7"]
            signalname = ["PGT","STARTER","LSS","GWB","DISTANT","HOME","KOTTAIKADU","GWB","GD","GSS","HOME","KJKD","STARTER","LSS","GWB","IBD","IBSS","GWB","DISTANT","HOME","WAL","STARTER","LSS","GWB","DISTANT","HOME","ETTIMADAI","GWB","DISTANT","HOME","MDKI","STARTER","LSS","GWB","DISTANT","HOME","PTJ","STARTER","LSS","GWB","G D","LC 147 GSS","GWB","DISTANT","HOME","IGU","STARTER","LSS","GWB","DISTANT","HOME","SUU","STARTER","LSS/GD","LC142 GSS","GWB","G D","LC141 GSS","GWB","GD","LC140 GSS/DIST","HOME","SNO","STARTER","LSS","GWB","G D","LC137 GSS","GWB","GD","LC136 GSS/GD","LC 135 GSS","GWB","DISTANT","HOME","VNJ","STARTER","LSS","GWB","DISTANT","HOME","TUP","STARTER","LSS/G D","GSS","GWB","GD","LC 131 E GSS","IBS GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","UKL","STARTER","LSS ","GWB","DISTANT","HOME","VZ","STARTER","LSS ","GWB","DISTANT","HOME","IGR","STARTER","LSS","GWB","DISTANT","HOME","PY","STARTER","LSS","GWB","DISTANT","HOME","TPM","STARTER","LSS","GWB","DISTANT","HOME","RT HOME","ED"]
            nstn = ["PGT","KJKD","WAL","MDKI","PTJ","IGU","SUU","SNO","VNJ","TUP","UKL","VZ","IGR","PY","TPM","ED"]
            annot = ["PGT","KJKD","WAL","MDKI","PTJ","IGU","SUU","SNO","VNJ","TUP","UKL","VZ","IGR","PY","TPM","ED"]
            dstn = [0,13.48,24.491,37.98,48.5,58.8,65.4,73.9,83.5,91.6,105.1,114.7,122.7,128.1,135.7,141.7]
            dstnannot = ["0","13.48","24.491","37.98","48.5","58.8","65.4","73.9","83.5","91.6","105.1","114.7","122.7","128.1","135.7","141.7"]




        list1 = distancelist
        list3 = speedlist
        sstn = data.get("ststn")

        if nstn[0]  == sstn:
            start = 1
        else:
            start = 0
        count = len(nstn)
        signalkmfinal = list(map(float,signalkm))
        minvalue = signalkmfinal[0]
        for index in range(len(signalkmfinal)):
            signalkmfinal[index] = signalkmfinal[index]-minvalue
            signalkmfinal[index] = round(signalkmfinal[index],2)
        plotcountend = 0
        plotcountstart = 0
        for index in range(len(signalname)):
            if signalname[index] == plotststn:
                plotcountstart = index
            if signalname[index] == plotendstn:
                plotcountend = index
                #print(plotcountstart)
                #print(plotcountend)
        plotcount = plotcountend-plotcountstart+1
        plotsignal = [0]*plotcount
        plotsignalspeed = [0]*plotcount
        plotsignalname = [0]*plotcount
        for i in range(plotcountstart,plotcountend+1):
            j = i-plotcountstart
            plotsignal[j] = signalkmfinal[i]
            plotsignalname[j] = signalname[i]
        #print(plotsignal)
        list4 = list(map(float,list3))
        minvalue = list4[0]
        for index in range(len(list4)):
            #list4[index] = list4[index]-minvalue
            list4[index] = round(list4[index],2)
        list2 = list(map(float,list1))
        minvalue = list2[0]
        for index in range(len(list2)):
            list2[index] = list2[index]-minvalue
            list2[index] = round(list2[index],2)
        """
        for index in range(len(nstn)):
            if sstn == nstn[index]:
                for x in range(len(nstn)):
                    nstn[x] = list2[0]+dstn[x]-dstn[index]
        stn = [0]*count
        if start == 0:
            for neg in range(len(nstn)):
                if nstn[neg] > 0:
                    stn[neg] = nstn[neg]
            stn = list(dict.fromkeys(stn))
        if start == 1:
            for neg in range(len(nstn)):
                stn[neg] = nstn[neg]
            stn = list(stn)

        spd = [0]*(len(stn))
        for x in range(len(stn)):
            for index in range(len(list2)):
                if round(stn[x],0) == round(list2[index],0):
                    spd[x] = list4[index]
                    break
        """
        #print(plotsignal)

        #bwstnfound = plotsignal.copy()
        bwstnfound = [0] * len(plotsignal)
        for index in range(len(plotsignal)):
            for i in range(len(list4)):
                if plotsignal[index] == list2[i]:
                    #print(i)
                    bwstnfound[index] = 1
                    plotsignalspeed[index] = list4[i]

        for i in range(len(bwstnfound)):
            if bwstnfound[i] == 0:
                for k in range(15):
                    plotvalue = plotsignal[i]+(0.01*k)
                    for a in range(len(list4)):
                        if plotvalue == list2[a] and bwstnfound[i] == 0:
                            plotsignalspeed[i] = list4[a]
                            bwstnfound[i] = 1
        #print(bwstnfound)
        zcount = 0
        for i in range(len(list4)):
            if list4[i] == 0:
                zcount = zcount+1
        print(zcount)
        print(plotsignal)
        print(plotsignalspeed)
        print(plotsignalname)
        x=plotsignalname
        y=plotsignalspeed
        graphsize = (len(plotsignalname))/2
        if graphsize>100:
            graphsize = 50
        """
        highlightx = stn
        highlighty = spd
        if len(stn) != len(annot):
            print(annot)
            for i in range(len(annot)):
                annotlen = len(annot)
                stnlen = len(stn)
                if annotlen-stnlen !=0:
                    del annot[0]
        """
        """
        graphtitle = "Speedometer graph between "+plotststn+" and "+plotendstn
        plt.figure(figsize=(graphsize,6))
        plt.xlabel("Name of the Signal")
        plt.ylabel("Speed in (KMPH)")
        plt.title(graphtitle)
        plt.xticks(plotsignal,x)
        plt.xticks(rotation=90)

        plt.plot(plotsignal,y, color='red', marker='o')"""
        #plt.scatter(highlightx,highlighty, color = "r", marker = "o")
        """
        #loop for annotation
        for i, label in enumerate(annot):
            plt.text(highlightx[i],highlighty[i],label)
        #fig = plt.show()
        """
        """
        buffer = BytesIO()
        plt.savefig(buffer, format = 'png')
        buffer.seek(0)
        image_png = buffer.getvalue()
        #chart.clear()
        chart = base64.b64encode(image_png)
        chart = chart.decode('utf-8')
        buffer.close()
        plt.clf()"""

        graphtitle = "SPEEDOMETER GRAPH BETWEEN "+plotststn+" AND "+plotendstn


        fig = go.Figure()
        scatter = go.Scatter(x=plotsignal, y=y, mode='lines', name='test', opacity=0.8, marker_color='blue')
        fig.add_trace(scatter)
        #fig.update_xaxes(type='category')
        fig.update_layout(xaxis = dict(tickmode = 'array',tickvals = plotsignal, ticktext = x))
        arrow_list=[]
        
        for i in range(len(plotsignal)):
            for a in range(len(annot)):
                if plotsignalname[i] == annot[a]:
                    arrow=dict(x=plotsignal[i],y=y[i],text=annot[a],arrowhead = 2,
                               arrowwidth=1.5,
                               arrowcolor='rgb(255,51,0)')
                    arrow_list.append(arrow) 

        #print(arrow_list)
           
        fig.update_layout(annotations=arrow_list, xaxis_title="NAME OF THE SIGNAL", yaxis_title="SPEED (KMPH)", title={'text': graphtitle, 'y':0.9,'x':0.5,'xanchor': 'center','yanchor': 'top'})


        plot_div = plot(fig, output_type='div')

        return render(request, "index.html", {"minvalue" : minvalue, "today" : today, "plot_div" : plot_div, "sstn" : sstn, "slist" : slist, "dlist" : dlist})
    if request.method == "POST" and plotststn == "None":
        #buffer.flush()
        #chart.flush()
        #print(request.POST)
        #os.system('clear')
        #chart.clear()
        """
        speedlist = data.get("spdlist")
        #print(speedlist)
        speedlist1 = speedlist.splitlines()
        speedlist = list(map(float,speedlist1))
        #speedlist1 = list(map(float,speedlist))
        #print(speedlist1)
        #print(speedlist)
        distancelist = data.get("distlist")
        distancelist1 = distancelist.splitlines()
        distancelist = list(map(float,distancelist1))
        """
        wb1 = load_workbook(filename = request.FILES['spmfile'].file,data_only=True)

        sh1 = wb1['Sheet1'] 

        rowcount = sh1.max_row
        colcount = sh1.max_column
        colcount = colcount+1

       
        

        for c in range(colcount):
            if c==0:
                c=c+1
            for r in range(rowcount):
                if r==0:
                    r=r+1
                if (type(sh1.cell(r,c).value).__name__) == "str":
                    if "distance" in sh1.cell(r,c).value.lower() or sh1.cell(r,c).value == "DISTANCE" or sh1.cell(r,c).value == "distance" or sh1.cell(r,c).value == "Distance" or sh1.cell(r,c).value == "DIST MTRS":
                        distancerow = r
                        distancecolumn = c
                    if "speed"  in sh1.cell(r,c).value.lower() or sh1.cell(r,c).value == "SPEED" or sh1.cell(r,c).value == "speed" or sh1.cell(r,c).value == "Speed" or sh1.cell(r,c).value == "INST.  KMPH":
                        speedrow = r
                        speedcolumn = c
                if (type(sh1.cell(r,c).value).__name__) == "str":
                    if "time"  in sh1.cell(r,c).value.lower() or sh1.cell(r,c).value == "Time" or sh1.cell(r,c).value == "TIME" or ("time"  in sh1.cell(r,c).value.lower() and "date" in sh1.cell(r,c).value.lower()):
                        timerow = r
                        timecolumn = c
                    if c<=3 and ("date"  in sh1.cell(r,c).value.lower() or sh1.cell(r,c).value == "Date" or sh1.cell(r,c).value == "DATE" or ("time"  in sh1.cell(r,c).value.lower() and "date" in sh1.cell(r,c).value.lower())):
                        daterow = r
                        datecolumn = c
                """
                if sh1.cell(r,c).value == "Time" or sh1.cell(r,c).value == "TIME" or sh1.cell(r,c).value == "  DATE/Time" or sh1.cell(r,c).value == "Date       TIME":
                    timerow = r
                    timecolumn = c
                if sh1.cell(r,c).value == "Date" or sh1.cell(r,c).value == "DATE" or sh1.cell(r,c).value == "  DATE/Time" or sh1.cell(r,c).value == "Date       TIME":
                    daterow = r
                    datecolumn = c
                """


                


        speedlist1 = []
        distancelist1 = []
        timelist = []
        datelist = []
        print(daterow)
        print(datecolumn)
        timevalue = sh1.cell(timerow,timecolumn).value
       
        datevalue = sh1.cell(daterow,datecolumn).value
        #print(timevalue)
        distancerow = distancerow+2
        speedrow = speedrow+2
        timerow = timerow+2
        daterow = daterow+2
        rowcount = rowcount

        for i in range(daterow,rowcount):
            if i == 0:
                i = 1
            if (type(sh1.cell(i,datecolumn).value).__name__) != "NoneType":
                datelist.append(sh1.cell(i,datecolumn).value)
            else:
                print("found date")


        for i in range(timerow,rowcount):
            if i==0:
                i = 1
            if (type(sh1.cell(i,timecolumn).value).__name__) != "NoneType":
                timelist.append(sh1.cell(i,timecolumn).value)
            else:
                print("found time")
        #print(timelist)


        for i in range(distancerow,rowcount):
            if i == 0:
                i = 1
            if (type(sh1.cell(i,distancecolumn).value).__name__) != "NoneType":
                #print("found")
                distancelist1.append(sh1.cell(i,distancecolumn).value)
            else:
                print("found distance")
        #print(type(sh1.cell(28,2).value).__name__)
        #print(type(sh1.cell(29,2).value))

     


        for s in range(speedrow,rowcount):
            if s == 0:
                s = 1
            if (type(sh1.cell(s,speedcolumn).value).__name__) != "NoneType":
                #print("found speed")
                speedlist1.append(sh1.cell(s,speedcolumn).value)
            else:
                print("found speed")
        #print(len(speedlist))
        #print(speedlist1)
        #print(len(distancelist))
        speedlist = list(map(float,speedlist1))
        distancelist = list(map(float,distancelist1))




        #print(speedlist)
        #print(distancelist)
        #spmtype = data.get("spmtype")
        #print(spmtype)

        spmcount = 0
        for i in range(len(distancelist1)):
            if distancelist1[i] == 1 or distancelist1[i] == 2:
                spmcount = spmcount+1
        print(spmcount)
        #print(len(distancelist1))
        if spmcount>=20:
            spmtype = "telpro"
        else:
            spmtype = ""

        if spmtype == "telpro":
            for i in range(len(distancelist)):
                if i != 0 or i != len(distancelist):
                    distancelist[i] = distancelist[i] + distancelist[i-1]
            for index in range(len(distancelist)):
                distancelist[index] = (distancelist[index]/1000)
                distancelist[index] = round(distancelist[index],2)
        #print(distancelist)
        today = datetime.now()
        """
        list1 = ["381077.559","381077.56","381077.56","381077.561","381077.561","381077.562","381077.563","381077.564","381077.565","381077.566","381077.568","381077.569","381077.57","381077.572","381077.574","381077.575","381077.577","381077.579","381077.581","381077.584","381077.586","381077.589","381077.591","381077.594","381077.597","381077.6","381077.603","381077.606","381077.609","381077.613","381077.616","381077.62","381077.624","381077.628","381077.632","381077.636","381077.64","381077.645","381077.649","381077.653","381077.658","381077.662","381077.667","381077.671","381077.675","381077.68","381077.684","381077.688","381077.692","381077.695","381077.699","381077.702","381077.705","381077.708","381077.71","381077.712","381077.715","381077.717","381077.719","381077.72","381077.722","381077.724","381077.726","381077.728","381077.729","381077.731","381077.733","381077.735","381077.737","381077.739","381077.742","381077.744","381077.746","381077.748","381077.75","381077.753","381077.755","381077.757","381077.759","381077.762","381077.764","381077.767","381077.769","381077.772","381077.774","381077.777","381077.78","381077.782","381077.785","381077.788","381077.791","381077.793","381077.796","381077.799","381077.803","381077.806","381077.809","381077.813","381077.816","381077.82","381077.824","381077.827","381077.831","381077.835","381077.84","381077.844","381077.848","381077.852","381077.856","381077.86","381077.864","381077.869","381077.873","381077.877","381077.881","381077.885","381077.889","381077.893","381077.897","381077.9","381077.904","381077.908","381077.912","381077.916","381077.92","381077.924","381077.928","381077.931","381077.935","381077.939","381077.943","381077.947","381077.951","381077.955","381077.959","381077.963","381077.967","381077.971","381077.975","381077.98","381077.984","381077.988","381077.992","381077.996","381078","381078.004","381078.008","381078.012","381078.016","381078.02","381078.025","381078.029","381078.033","381078.037","381078.041","381078.045","381078.049","381078.054","381078.058","381078.062","381078.067","381078.071","381078.075","381078.079","381078.084","381078.088","381078.092","381078.096","381078.1","381078.105","381078.109","381078.113","381078.117","381078.121","381078.125","381078.13","381078.134","381078.138","381078.142","381078.146","381078.15","381078.154","381078.158","381078.162","381078.166","381078.17","381078.174","381078.178","381078.182","381078.186","381078.19","381078.194","381078.198","381078.202","381078.206","381078.211","381078.215","381078.219","381078.223","381078.227","381078.231","381078.235","381078.239","381078.243","381078.247","381078.251","381078.255","381078.259","381078.263","381078.267","381078.271","381078.275","381078.278","381078.282","381078.286","381078.29","381078.294","381078.298","381078.302","381078.307","381078.311","381078.315","381078.319","381078.323","381078.327","381078.332","381078.336","381078.34","381078.344","381078.348","381078.353","381078.357","381078.361","381078.366","381078.37","381078.375","381078.379","381078.384","381078.388","381078.392","381078.397","381078.401","381078.406","381078.41","381078.415","381078.419","381078.423","381078.428","381078.432","381078.437","381078.441","381078.446","381078.45","381078.455","381078.459","381078.463","381078.468","381078.472","381078.477","381078.482","381078.486","381078.491","381078.496","381078.5","381078.505","381078.51","381078.515","381078.52","381078.525","381078.531","381078.536","381078.541","381078.547","381078.552","381078.558","381078.564","381078.57","381078.576","381078.582","381078.588","381078.595","381078.601","381078.608","381078.615","381078.622","381078.63","381078.637","381078.645","381078.653","381078.661","381078.669","381078.678","381078.686","381078.695","381078.704","381078.713","381078.723","381078.733","381078.742","381078.752","381078.762","381078.773","381078.783","381078.794","381078.805","381078.816","381078.827","381078.839","381078.85","381078.862","381078.874","381078.886","381078.899","381078.911","381078.924","381078.937","381078.95","381078.963","381078.977","381078.991","381079.004","381079.019","381079.033","381079.047","381079.062","381079.077","381079.092","381079.107","381079.123","381079.138","381079.154","381079.17","381079.187","381079.203","381079.22","381079.237","381079.254","381079.271","381079.288","381079.306","381079.324","381079.342","381079.36","381079.379","381079.398","381079.416","381079.436","381079.455","381079.474","381079.494","381079.514","381079.534","381079.554","381079.574","381079.595","381079.615","381079.636","381079.656","381079.676","381079.697","381079.717","381079.737","381079.757","381079.777","381079.796","381079.816","381079.835","381079.854","381079.872","381079.89","381079.907","381079.925","381079.941","381079.957","381079.973","381079.988","381080.002","381080.016","381080.03","381080.043","381080.055","381080.067","381080.079","381080.09","381080.101","381080.112","381080.122","381080.133","381080.143","381080.153","381080.163","381080.173","381080.183","381080.192","381080.202","381080.212","381080.222","381080.232","381080.242","381080.252","381080.262","381080.272","381080.282","381080.292","381080.302","381080.313","381080.323","381080.334","381080.344","381080.355","381080.366","381080.377","381080.389","381080.4","381080.412","381080.423","381080.435","381080.447","381080.459","381080.471","381080.483","381080.496","381080.508","381080.521","381080.534","381080.547","381080.56","381080.573","381080.586","381080.6","381080.613","381080.627","381080.641","381080.655","381080.669","381080.683","381080.697","381080.712","381080.727","381080.741","381080.756","381080.771","381080.787","381080.802","381080.818","381080.833","381080.849","381080.865","381080.881","381080.898","381080.915","381080.931","381080.948","381080.965","381080.983","381081","381081.018","381081.036","381081.054","381081.072","381081.09","381081.109","381081.128","381081.147","381081.166","381081.185","381081.205","381081.225","381081.245","381081.265","381081.285","381081.306","381081.327","381081.347","381081.369","381081.39","381081.411","381081.433","381081.455","381081.477","381081.499","381081.521","381081.544","381081.566","381081.589","381081.612","381081.636","381081.659","381081.682","381081.706","381081.73","381081.754","381081.778","381081.802","381081.827","381081.852","381081.876","381081.901","381081.927","381081.952","381081.977","381082.003","381082.029","381082.054","381082.081","381082.107","381082.133","381082.159","381082.186","381082.213","381082.239","381082.266","381082.293","381082.321","381082.348","381082.375","381082.403","381082.431","381082.459","381082.487","381082.515","381082.542","381082.57","381082.598","381082.626","381082.654","381082.682","381082.71","381082.738","381082.765","381082.793","381082.821","381082.848","381082.876","381082.903","381082.931","381082.958","381082.986","381083.013","381083.04","381083.068","381083.095","381083.123","381083.15","381083.177","381083.205","381083.232","381083.259","381083.287","381083.314","381083.341","381083.369","381083.396","381083.424","381083.451","381083.478","381083.506","381083.533","381083.561","381083.588","381083.616","381083.643","381083.671","381083.698","381083.725","381083.753","381083.78","381083.808","381083.835","381083.863","381083.89","381083.918","381083.945","381083.973","381084","381084.028","381084.055","381084.083","381084.11","381084.137","381084.165","381084.192","381084.22","381084.247","381084.275","381084.302","381084.33","381084.357","381084.385","381084.412","381084.44","381084.467","381084.495","381084.522","381084.55","381084.578","381084.605","381084.633","381084.66","381084.688","381084.715","381084.743","381084.77","381084.798","381084.826","381084.853","381084.881","381084.909","381084.936","381084.964","381084.991","381085.019","381085.047","381085.074","381085.102","381085.13","381085.157","381085.185","381085.213","381085.24","381085.268","381085.296","381085.324","381085.351","381085.379","381085.407","381085.435","381085.462","381085.49","381085.518","381085.546","381085.574","381085.602","381085.629","381085.657","381085.685","381085.713","381085.741","381085.768","381085.796","381085.824","381085.851","381085.879","381085.906","381085.934","381085.961","381085.989","381086.016","381086.044","381086.071","381086.099","381086.126","381086.154","381086.182","381086.209","381086.237","381086.264","381086.292","381086.319","381086.347","381086.375","381086.402","381086.43","381086.457","381086.485","381086.513","381086.54","381086.568","381086.595","381086.623","381086.651","381086.678","381086.706","381086.733","381086.761","381086.789","381086.816","381086.844","381086.871","381086.899","381086.927","381086.954","381086.982","381087.009","381087.037","381087.064","381087.092","381087.12","381087.147","381087.175","381087.202","381087.23","381087.257","381087.285","381087.313","381087.34","381087.368","381087.395","381087.423","381087.45","381087.478","381087.505","381087.533","381087.56","381087.588","381087.616","381087.643","381087.671","381087.698","381087.726","381087.753","381087.781","381087.808","381087.836","381087.863","381087.891","381087.918","381087.946","381087.973","381088.001","381088.028","381088.055","381088.083","381088.11","381088.137","381088.165","381088.192","381088.219","381088.246","381088.274","381088.301","381088.328","381088.355","381088.383","381088.41","381088.437","381088.464","381088.491","381088.518","381088.545","381088.572","381088.6","381088.627","381088.654","381088.681","381088.708","381088.735","381088.762","381088.789","381088.816","381088.844","381088.871","381088.898","381088.925","381088.952","381088.98","381089.007","381089.034","381089.061","381089.088","381089.115","381089.143","381089.17","381089.197","381089.224","381089.251","381089.279","381089.306","381089.333","381089.36","381089.387","381089.414","381089.442","381089.469","381089.496","381089.523","381089.55","381089.578","381089.605","381089.632","381089.659","381089.686","381089.713","381089.741","381089.768","381089.795","381089.822","381089.849","381089.876","381089.904","381089.931","381089.958","381089.985","381090.013","381090.04","381090.067","381090.094","381090.122","381090.149","381090.176","381090.204","381090.231","381090.259","381090.286","381090.313","381090.341","381090.368","381090.396","381090.423","381090.451","381090.478","381090.506","381090.533","381090.561","381090.589","381090.616","381090.644","381090.671","381090.699","381090.727","381090.754","381090.782","381090.81","381090.837","381090.865","381090.892","381090.92","381090.947","381090.974","381091.002","381091.029","381091.057","381091.084","381091.111","381091.138","381091.166","381091.193","381091.22","381091.247","381091.274","381091.301","381091.328","381091.355","381091.382","381091.409","381091.436","381091.463","381091.49","381091.517","381091.544","381091.571","381091.598","381091.625","381091.652","381091.679","381091.706","381091.733","381091.76","381091.787","381091.814","381091.841","381091.868","381091.895","381091.922","381091.949","381091.977","381092.004","381092.031","381092.059","381092.086","381092.113","381092.141","381092.168","381092.195","381092.223","381092.25","381092.278","381092.305","381092.332","381092.36","381092.387","381092.414","381092.442","381092.469","381092.496","381092.524","381092.551","381092.578","381092.606","381092.633","381092.66","381092.687","381092.714","381092.742","381092.769","381092.796","381092.823","381092.85","381092.877","381092.904","381092.932","381092.959","381092.986","381093.013","381093.04","381093.067","381093.094","381093.122","381093.149","381093.176","381093.203","381093.23","381093.257","381093.285","381093.312","381093.339","381093.366","381093.393","381093.42","381093.447","381093.475","381093.502","381093.529","381093.556","381093.583","381093.61","381093.637","381093.664","381093.691","381093.718","381093.745","381093.772","381093.799","381093.826","381093.853","381093.88","381093.907","381093.934","381093.961","381093.988","381094.015","381094.042","381094.069","381094.096","381094.123","381094.15","381094.177","381094.204","381094.231","381094.258","381094.285","381094.312","381094.339","381094.366","381094.393","381094.42","381094.447","381094.474","381094.501","381094.528","381094.555","381094.582","381094.609","381094.636","381094.663","381094.691","381094.718","381094.745","381094.772","381094.8","381094.827","381094.854","381094.882","381094.909","381094.936","381094.964","381094.991","381095.019","381095.046","381095.074","381095.101","381095.129","381095.156","381095.184","381095.211","381095.239","381095.266","381095.294","381095.321","381095.348","381095.376","381095.403","381095.43","381095.458","381095.485","381095.512","381095.539","381095.566","381095.593","381095.62","381095.647","381095.674","381095.701","381095.728","381095.755","381095.782","381095.809","381095.835","381095.862","381095.889","381095.916","381095.943","381095.97","381095.997","381096.024","381096.051","381096.078","381096.105","381096.132","381096.159","381096.186","381096.214","381096.241","381096.268","381096.295","381096.323","381096.35","381096.378","381096.405","381096.433","381096.46","381096.487","381096.515","381096.542","381096.57","381096.598","381096.625","381096.653","381096.68","381096.708","381096.735","381096.763","381096.79","381096.818","381096.845","381096.872","381096.9","381096.927","381096.954","381096.981","381097.008","381097.036","381097.063","381097.09","381097.117","381097.143","381097.17","381097.197","381097.224","381097.25","381097.277","381097.303","381097.33","381097.356","381097.383","381097.409","381097.436","381097.462","381097.488","381097.514","381097.541","381097.567","381097.593","381097.619","381097.645","381097.671","381097.697","381097.723","381097.749","381097.775","381097.8","381097.826","381097.852","381097.877","381097.903","381097.929","381097.954","381097.98","381098.005","381098.03","381098.056","381098.081","381098.107","381098.132","381098.157","381098.183","381098.208","381098.234","381098.26","381098.285","381098.311","381098.337","381098.363","381098.389","381098.415","381098.441","381098.468","381098.494","381098.521","381098.548","381098.575","381098.602","381098.629","381098.657","381098.684","381098.711","381098.739","381098.767","381098.794","381098.822","381098.849","381098.877","381098.905","381098.932","381098.96","381098.988","381099.015","381099.043","381099.071","381099.099","381099.126","381099.154","381099.181","381099.209","381099.236","381099.264","381099.291","381099.319","381099.346","381099.374","381099.401","381099.428","381099.456","381099.483","381099.51","381099.537","381099.565","381099.592","381099.619","381099.646","381099.673","381099.7","381099.727","381099.754","381099.781","381099.808","381099.834","381099.861","381099.888","381099.915","381099.941","381099.968","381099.994","381100.021","381100.047","381100.074","381100.1","381100.126","381100.152","381100.179","381100.205","381100.231","381100.257","381100.283","381100.308","381100.334","381100.36","381100.386","381100.412","381100.438","381100.464","381100.49","381100.516","381100.543","381100.569","381100.596","381100.622","381100.649","381100.676","381100.703","381100.73","381100.757","381100.785","381100.812","381100.84","381100.868","381100.896","381100.924","381100.952","381100.98","381101.008","381101.036","381101.064","381101.092","381101.12","381101.148","381101.175","381101.203","381101.231","381101.259","381101.287","381101.314","381101.342","381101.369","381101.397","381101.424","381101.452","381101.479","381101.507","381101.534","381101.561","381101.589","381101.616","381101.643","381101.671","381101.698","381101.725","381101.753","381101.78","381101.808","381101.835","381101.862","381101.89","381101.917","381101.944","381101.972","381101.999","381102.027","381102.054","381102.081","381102.109","381102.136","381102.164","381102.191","381102.218","381102.246","381102.273","381102.3","381102.328","381102.355","381102.382","381102.409","381102.437","381102.464","381102.491","381102.519","381102.546","381102.573","381102.6","381102.628","381102.655","381102.682","381102.709","381102.736","381102.763","381102.791","381102.818","381102.845","381102.872","381102.899","381102.927","381102.954","381102.981","381103.008","381103.035","381103.062","381103.09","381103.117","381103.144","381103.171","381103.199","381103.226","381103.253","381103.28","381103.307","381103.335","381103.362","381103.389","381103.416","381103.444","381103.471","381103.498","381103.525","381103.552","381103.58","381103.607","381103.634","381103.661","381103.689","381103.716","381103.744","381103.771","381103.798","381103.826","381103.853","381103.881","381103.908","381103.936","381103.963","381103.991","381104.019","381104.046","381104.074","381104.102","381104.129","381104.157","381104.184","381104.212","381104.24","381104.267","381104.295","381104.323","381104.35","381104.378","381104.405","381104.433","381104.46","381104.488","381104.515","381104.543","381104.57","381104.598","381104.625","381104.653","381104.68","381104.708","381104.735","381104.762","381104.79","381104.817","381104.845","381104.872","381104.899","381104.927","381104.954","381104.981","381105.009","381105.036","381105.063","381105.091","381105.118","381105.145","381105.173","381105.2","381105.227","381105.255","381105.282","381105.309","381105.336","381105.364","381105.391","381105.418","381105.446","381105.473","381105.5","381105.527","381105.555","381105.582","381105.609","381105.637","381105.664","381105.691","381105.719","381105.746","381105.773","381105.8","381105.828","381105.855","381105.882","381105.909","381105.937","381105.964","381105.991","381106.018","381106.046","381106.073","381106.1","381106.127","381106.154","381106.182","381106.209","381106.236","381106.263","381106.29","381106.318","381106.345","381106.372","381106.399","381106.427","381106.454","381106.481","381106.508","381106.536","381106.563","381106.59","381106.618","381106.645","381106.672","381106.699","381106.726","381106.754","381106.781","381106.808","381106.835","381106.863","381106.89","381106.917","381106.944","381106.971","381106.998","381107.025","381107.053","381107.08","381107.107","381107.134","381107.161","381107.188","381107.215","381107.242","381107.269","381107.296","381107.324","381107.351","381107.378","381107.405","381107.432","381107.46","381107.487","381107.514","381107.541","381107.569","381107.596","381107.623","381107.65","381107.678","381107.705","381107.733","381107.76","381107.788","381107.815","381107.843","381107.87","381107.898","381107.926","381107.953","381107.981","381108.008","381108.036","381108.064","381108.092","381108.119","381108.147","381108.174","381108.202","381108.23","381108.257","381108.285","381108.313","381108.34","381108.368","381108.395","381108.423","381108.451","381108.478","381108.505","381108.533","381108.56","381108.588","381108.615","381108.643","381108.67","381108.698","381108.725","381108.752","381108.78","381108.807","381108.834","381108.862","381108.889","381108.916","381108.944","381108.971","381108.998","381109.026","381109.053","381109.08","381109.107","381109.135","381109.162","381109.189","381109.216","381109.244","381109.271","381109.298","381109.326","381109.353","381109.38","381109.408","381109.435","381109.463","381109.49","381109.517","381109.545","381109.572","381109.6","381109.627","381109.654","381109.682","381109.709","381109.737","381109.764","381109.792","381109.819","381109.846","381109.874","381109.901","381109.928","381109.956","381109.983","381110.01","381110.038","381110.065","381110.092","381110.119","381110.147","381110.174","381110.201","381110.228","381110.255","381110.283","381110.31","381110.337","381110.364","381110.391","381110.418","381110.445","381110.472","381110.499","381110.526","381110.553","381110.58","381110.608","381110.635","381110.662","381110.689","381110.716","381110.744","381110.771","381110.798","381110.825","381110.853","381110.88","381110.907","381110.934","381110.962","381110.989","381111.016","381111.043","381111.071","381111.098","381111.125","381111.152","381111.18","381111.207","381111.234","381111.261","381111.289","381111.316","381111.343","381111.37","381111.397","381111.424","381111.452","381111.479","381111.506","381111.533","381111.561","381111.588","381111.615","381111.642","381111.67","381111.697","381111.724","381111.752","381111.779","381111.806","381111.833","381111.861","381111.888","381111.916","381111.943","381111.971","381111.998","381112.025","381112.053","381112.08","381112.108","381112.135","381112.163","381112.19","381112.218","381112.245","381112.273","381112.3","381112.327","381112.355","381112.382","381112.41","381112.437","381112.465","381112.492","381112.52","381112.547","381112.574","381112.602","381112.629","381112.657","381112.684","381112.712","381112.739","381112.767","381112.794","381112.822","381112.849","381112.877","381112.904","381112.932","381112.959","381112.987","381113.014","381113.042","381113.069","381113.097","381113.124","381113.152","381113.179","381113.207","381113.234","381113.261","381113.289","381113.316","381113.344","381113.371","381113.399","381113.426","381113.454","381113.481","381113.508","381113.536","381113.563","381113.591","381113.618","381113.646","381113.673","381113.701","381113.728","381113.755","381113.783","381113.81","381113.838","381113.865","381113.893","381113.92","381113.947","381113.975","381114.002","381114.03","381114.057","381114.085","381114.112","381114.14","381114.167","381114.194","381114.222","381114.249","381114.277","381114.304","381114.332","381114.359","381114.386","381114.414","381114.441","381114.469","381114.496","381114.524","381114.551","381114.578","381114.606","381114.633","381114.661","381114.688","381114.716","381114.743","381114.77","381114.798","381114.825","381114.853","381114.88","381114.908","381114.935","381114.963","381114.99","381115.017","381115.045","381115.072","381115.1","381115.127","381115.155","381115.182","381115.209","381115.237","381115.264","381115.291","381115.319","381115.346","381115.374","381115.401","381115.428","381115.456","381115.483","381115.51","381115.538","381115.565","381115.593","381115.62","381115.647","381115.675","381115.702","381115.73","381115.757","381115.784","381115.812","381115.839","381115.867","381115.894","381115.922","381115.949","381115.977","381116.004","381116.031","381116.059","381116.086","381116.114","381116.141","381116.169","381116.196","381116.224","381116.251","381116.279","381116.306","381116.334","381116.361","381116.388","381116.416","381116.443","381116.471","381116.498","381116.526","381116.553","381116.581","381116.608","381116.636","381116.663","381116.691","381116.718","381116.746","381116.773","381116.801","381116.828","381116.856","381116.883","381116.911","381116.938","381116.966","381116.994","381117.021","381117.049","381117.076","381117.104","381117.131","381117.159","381117.186","381117.214","381117.241","381117.269","381117.296","381117.324","381117.351","381117.379","381117.406","381117.434","381117.461","381117.489","381117.516","381117.544","381117.571","381117.599","381117.626","381117.654","381117.681","381117.709","381117.736","381117.764","381117.791","381117.819","381117.846","381117.874","381117.901","381117.928","381117.956","381117.983","381118.011","381118.038","381118.065","381118.093","381118.12","381118.148","381118.175","381118.202","381118.23","381118.257","381118.285","381118.312","381118.339","381118.367","381118.394","381118.422","381118.449","381118.476","381118.504","381118.531","381118.559","381118.586","381118.614","381118.641","381118.668","381118.696","381118.723","381118.75","381118.778","381118.805","381118.832","381118.86","381118.887","381118.914","381118.942","381118.969","381118.996","381119.023","381119.051","381119.078","381119.105","381119.132","381119.159","381119.186","381119.213","381119.24","381119.266","381119.293","381119.319","381119.345","381119.371","381119.398","381119.423","381119.449","381119.475","381119.5","381119.526","381119.551","381119.576","381119.601","381119.626","381119.651","381119.676","381119.7","381119.725","381119.749","381119.773","381119.797","381119.821","381119.844","381119.867","381119.89","381119.913","381119.935","381119.957","381119.979","381120.001","381120.022","381120.043","381120.064","381120.085","381120.105","381120.125","381120.146","381120.166","381120.187","381120.207","381120.228","381120.248","381120.269","381120.289","381120.309","381120.33","381120.35","381120.371","381120.392","381120.412","381120.433","381120.453","381120.474","381120.494","381120.515","381120.535","381120.556","381120.577","381120.597","381120.618","381120.639","381120.659","381120.68","381120.701","381120.722","381120.742","381120.763","381120.784","381120.805","381120.825","381120.846","381120.867","381120.888","381120.908","381120.929","381120.95","381120.971","381120.991","381121.012","381121.033","381121.054","381121.074","381121.095","381121.116","381121.137","381121.158","381121.179","381121.2","381121.221","381121.242","381121.263","381121.284","381121.305","381121.327","381121.348","381121.37","381121.391","381121.413","381121.435","381121.457","381121.478","381121.5","381121.523","381121.545","381121.567","381121.59","381121.612","381121.635","381121.658","381121.681","381121.704","381121.727","381121.751","381121.774","381121.798","381121.822","381121.846","381121.87","381121.894","381121.919","381121.943","381121.968","381121.992","381122.017","381122.042","381122.067","381122.092","381122.118","381122.143","381122.169","381122.195","381122.221","381122.246","381122.273","381122.299","381122.325","381122.352","381122.378","381122.405","381122.432","381122.459","381122.486","381122.513","381122.54","381122.567","381122.594","381122.621","381122.648","381122.675","381122.702","381122.729","381122.756","381122.783","381122.811","381122.838","381122.865","381122.892","381122.919","381122.947","381122.974","381123.001","381123.029","381123.056","381123.084","381123.112","381123.139","381123.167","381123.195","381123.223","381123.25","381123.278","381123.306","381123.334","381123.361","381123.389","381123.417","381123.445","381123.472","381123.5","381123.528","381123.556","381123.584","381123.612","381123.639","381123.667","381123.695","381123.723","381123.751","381123.778","381123.806","381123.834","381123.862","381123.889","381123.917","381123.945","381123.973","381124","381124.028","381124.056","381124.084","381124.111","381124.139","381124.167","381124.194","381124.222","381124.25","381124.278","381124.305","381124.333","381124.361","381124.388","381124.416","381124.444","381124.471","381124.499","381124.527","381124.554","381124.582","381124.61","381124.637","381124.665","381124.693","381124.72","381124.748","381124.776","381124.803","381124.831","381124.859","381124.886","381124.914","381124.941","381124.969","381124.997","381125.024","381125.052","381125.08","381125.107","381125.135","381125.162","381125.19","381125.218","381125.245","381125.273","381125.3","381125.328","381125.356","381125.383","381125.411","381125.438","381125.466","381125.493","381125.521","381125.549","381125.576","381125.604","381125.631","381125.659","381125.686","381125.714","381125.742","381125.769","381125.797","381125.824","381125.852","381125.879","381125.907","381125.934","381125.961","381125.989","381126.016","381126.044","381126.071","381126.098","381126.126","381126.153","381126.18","381126.208","381126.235","381126.262","381126.289","381126.317","381126.344","381126.371","381126.399","381126.426","381126.453","381126.481","381126.508","381126.535","381126.563","381126.59","381126.618","381126.645","381126.673","381126.701","381126.728","381126.756","381126.783","381126.811","381126.838","381126.866","381126.894","381126.921","381126.949","381126.976","381127.004","381127.032","381127.059","381127.087","381127.114","381127.142","381127.17","381127.197","381127.225","381127.252","381127.28","381127.308","381127.335","381127.363","381127.39","381127.418","381127.446","381127.473","381127.501","381127.528","381127.556","381127.584","381127.611","381127.639","381127.666","381127.694","381127.721","381127.749","381127.777","381127.804","381127.832","381127.859","381127.887","381127.914","381127.942","381127.97","381127.997","381128.025","381128.052","381128.08","381128.107","381128.135","381128.162","381128.19","381128.218","381128.245","381128.273","381128.3","381128.328","381128.355","381128.383","381128.41","381128.438","381128.465","381128.493","381128.52","381128.548","381128.575","381128.603","381128.63","381128.658","381128.685","381128.712","381128.74","381128.767","381128.795","381128.822","381128.849","381128.877","381128.904","381128.931","381128.958","381128.986","381129.013","381129.04","381129.067","381129.095","381129.122","381129.149","381129.176","381129.204","381129.231","381129.258","381129.285","381129.313","381129.34","381129.367","381129.395","381129.422","381129.449","381129.477","381129.504","381129.531","381129.559","381129.586","381129.614","381129.641","381129.669","381129.696","381129.724","381129.751","381129.779","381129.807","381129.834","381129.862","381129.889","381129.917","381129.944","381129.972","381129.999","381130.027","381130.054","381130.082","381130.109","381130.137","381130.164","381130.191","381130.219","381130.246","381130.274","381130.301","381130.329","381130.356","381130.384","381130.411","381130.439","381130.466","381130.494","381130.521","381130.549","381130.576","381130.604","381130.631","381130.659","381130.686","381130.713","381130.741","381130.769","381130.796","381130.824","381130.851","381130.879","381130.906","381130.934","381130.961","381130.989","381131.017","381131.044","381131.072","381131.1","381131.127","381131.155","381131.182","381131.21","381131.238","381131.265","381131.293","381131.32","381131.348","381131.376","381131.403","381131.431","381131.459","381131.486","381131.514","381131.541","381131.569","381131.597","381131.624","381131.652","381131.679","381131.707","381131.735","381131.762","381131.79","381131.818","381131.845","381131.873","381131.9","381131.928","381131.956","381131.983","381132.011","381132.039","381132.067","381132.094","381132.122","381132.15","381132.177","381132.205","381132.233","381132.261","381132.289","381132.316","381132.344","381132.372","381132.4","381132.428","381132.456","381132.483","381132.511","381132.539","381132.567","381132.595","381132.623","381132.651","381132.679","381132.706","381132.735","381132.763","381132.791","381132.819","381132.846","381132.875","381132.902","381132.931","381132.959","381132.987","381133.015","381133.043","381133.071","381133.099","381133.127","381133.155","381133.183","381133.211","381133.238","381133.266","381133.293","381133.321","381133.348","381133.375","381133.402","381133.429","381133.456","381133.483","381133.51","381133.537","381133.564","381133.591","381133.618","381133.645","381133.672","381133.699","381133.726","381133.754","381133.781","381133.808","381133.835","381133.862","381133.89","381133.917","381133.944","381133.972","381133.999","381134.027","381134.054","381134.081","381134.109","381134.136","381134.164","381134.191","381134.219","381134.246","381134.274","381134.301","381134.329","381134.356","381134.383","381134.411","381134.438","381134.466","381134.493","381134.52","381134.548","381134.575","381134.602","381134.63","381134.657","381134.685","381134.712","381134.739","381134.766","381134.794","381134.821","381134.848","381134.875","381134.902","381134.929","381134.957","381134.984","381135.01","381135.037","381135.064","381135.091","381135.118","381135.145","381135.172","381135.198","381135.225","381135.252","381135.278","381135.305","381135.332","381135.358","381135.384","381135.411","381135.437","381135.464","381135.49","381135.516","381135.543","381135.569","381135.596","381135.622","381135.648","381135.675","381135.701","381135.728","381135.754","381135.781","381135.808","381135.834","381135.861","381135.887","381135.914","381135.941","381135.968","381135.995","381136.022","381136.049","381136.076","381136.103","381136.129","381136.156","381136.183","381136.21","381136.237","381136.264","381136.292","381136.319","381136.346","381136.373","381136.4","381136.427","381136.454","381136.482","381136.509","381136.536","381136.564","381136.591","381136.619","381136.646","381136.674","381136.701","381136.729","381136.757","381136.784","381136.812","381136.84","381136.868","381136.895","381136.923","381136.951","381136.978","381137.006","381137.034","381137.062","381137.089","381137.117","381137.145","381137.172","381137.2","381137.227","381137.255","381137.282","381137.31","381137.337","381137.365","381137.392","381137.42","381137.447","381137.474","381137.502","381137.529","381137.556","381137.584","381137.611","381137.638","381137.665","381137.693","381137.72","381137.747","381137.774","381137.802","381137.829","381137.856","381137.884","381137.911","381137.938","381137.966","381137.993","381138.021","381138.048","381138.076","381138.104","381138.131","381138.159","381138.187","381138.214","381138.242","381138.27","381138.297","381138.325","381138.353","381138.38","381138.408","381138.435","381138.463","381138.491","381138.518","381138.546","381138.573","381138.601","381138.628","381138.655","381138.683","381138.71","381138.738","381138.765","381138.792","381138.82","381138.847","381138.875","381138.902","381138.93","381138.957","381138.985","381139.013","381139.04","381139.068","381139.096","381139.124","381139.151","381139.179","381139.207","381139.235","381139.262","381139.29","381139.318","381139.346","381139.373","381139.401","381139.429","381139.457","381139.484","381139.512","381139.54","381139.567","381139.595","381139.623","381139.65","381139.678","381139.706","381139.734","381139.761","381139.789","381139.816","381139.844","381139.872","381139.899","381139.927","381139.954","381139.982","381140.01","381140.037","381140.065","381140.093","381140.12","381140.148","381140.175","381140.203","381140.231","381140.258","381140.286","381140.313","381140.341","381140.368","381140.396","381140.424","381140.451","381140.479","381140.506","381140.534","381140.561","381140.589","381140.616","381140.644","381140.671","381140.699","381140.726","381140.754","381140.782","381140.809","381140.837","381140.864","381140.892","381140.919","381140.947","381140.974","381141.002","381141.029","381141.057","381141.085","381141.112","381141.14","381141.167","381141.195","381141.222","381141.25","381141.277","381141.305","381141.332","381141.36","381141.387","381141.415","381141.442","381141.469","381141.497","381141.524","381141.551","381141.579","381141.606","381141.633","381141.661","381141.688","381141.715","381141.742","381141.77","381141.797","381141.824","381141.851","381141.879","381141.906","381141.933","381141.961","381141.988","381142.015","381142.042","381142.07","381142.097","381142.124","381142.151","381142.179","381142.206","381142.233","381142.261","381142.288","381142.315","381142.343","381142.37","381142.397","381142.425","381142.452","381142.479","381142.507","381142.534","381142.562","381142.589","381142.617","381142.644","381142.672","381142.699","381142.727","381142.754","381142.782","381142.809","381142.837","381142.865","381142.892","381142.92","381142.947","381142.975","381143.003","381143.03","381143.058","381143.085","381143.113","381143.141","381143.168","381143.196","381143.223","381143.251","381143.278","381143.306","381143.334","381143.361","381143.389","381143.416","381143.444","381143.471","381143.499","381143.527","381143.554","381143.582","381143.609","381143.637","381143.664","381143.691","381143.719","381143.746","381143.774","381143.801","381143.828","381143.856","381143.883","381143.91","381143.938","381143.965","381143.992","381144.019","381144.047","381144.074","381144.101","381144.128","381144.156","381144.183","381144.21","381144.237","381144.265","381144.292","381144.319","381144.347","381144.374","381144.401","381144.428","381144.456","381144.483","381144.51","381144.538","381144.565","381144.593","381144.62","381144.648","381144.675","381144.703","381144.73","381144.758","381144.785","381144.813","381144.841","381144.868","381144.896","381144.924","381144.951","381144.979","381145.007","381145.034","381145.062","381145.09","381145.117","381145.145","381145.173","381145.2","381145.228","381145.255","381145.283","381145.311","381145.338","381145.366","381145.393","381145.421","381145.448","381145.476","381145.504","381145.531","381145.559","381145.586","381145.614","381145.641","381145.669","381145.696","381145.723","381145.751","381145.778","381145.806","381145.833","381145.861","381145.888","381145.916","381145.943","381145.971","381145.998","381146.026","381146.053","381146.08","381146.108","381146.135","381146.163","381146.19","381146.218","381146.245","381146.273","381146.3","381146.328","381146.355","381146.383","381146.411","381146.438","381146.466","381146.493","381146.521","381146.548","381146.576","381146.603","381146.631","381146.658","381146.686","381146.713","381146.741","381146.768","381146.796","381146.824","381146.851","381146.879","381146.906","381146.934","381146.962","381146.989","381147.017","381147.045","381147.072","381147.1","381147.128","381147.156","381147.183","381147.211","381147.239","381147.266","381147.294","381147.321","381147.349","381147.376","381147.404","381147.431","381147.459","381147.486","381147.513","381147.54","381147.567","381147.594","381147.62","381147.647","381147.673","381147.7","381147.726","381147.752","381147.778","381147.804","381147.83","381147.856","381147.882","381147.907","381147.933","381147.958","381147.983","381148.008","381148.033","381148.058","381148.082","381148.107","381148.131","381148.155","381148.179","381148.203","381148.227","381148.251","381148.274","381148.297","381148.321","381148.344","381148.367","381148.39","381148.413","381148.435","381148.457","381148.48","381148.502","381148.524","381148.546","381148.567","381148.589","381148.61","381148.632","381148.653","381148.673","381148.694","381148.715","381148.735","381148.756","381148.776","381148.796","381148.816","381148.835","381148.855","381148.874","381148.893","381148.912","381148.931","381148.95","381148.969","381148.987","381149.005","381149.024","381149.042","381149.06","381149.077","381149.095","381149.112","381149.129","381149.147","381149.164","381149.18","381149.197","381149.213","381149.23","381149.245","381149.261","381149.276","381149.291","381149.305","381149.319","381149.333","381149.346","381149.358","381149.37","381149.382","381149.394","381149.406","381149.417","381149.428","381149.439","381149.45","381149.462","381149.473","381149.484","381149.495","381149.506","381149.517","381149.529","381149.54","381149.551","381149.562","381149.573","381149.585","381149.596","381149.607","381149.618","381149.629","381149.64","381149.652","381149.663","381149.674","381149.685","381149.696","381149.707","381149.718","381149.729","381149.74","381149.751","381149.762","381149.773","381149.784","381149.795","381149.806","381149.817","381149.828","381149.838","381149.849","381149.86","381149.871","381149.882","381149.892","381149.903","381149.913","381149.924","381149.934","381149.945","381149.955","381149.965","381149.975","381149.985","381149.994","381150.003","381150.013","381150.022","381150.031","381150.039","381150.048","381150.056","381150.065","381150.073","381150.081","381150.088","381150.096","381150.103","381150.111","381150.118","381150.125","381150.131","381150.138","381150.145","381150.151","381150.157","381150.163","381150.169","381150.175","381150.18","381150.185","381150.191","381150.196","381150.201","381150.205","381150.21","381150.215","381150.219","381150.223","381150.227","381150.231","381150.235","381150.239","381150.243","381150.247","381150.251","381150.255","381150.258","381150.262","381150.266","381150.27","381150.274","381150.278","381150.282","381150.286","381150.29","381150.294","381150.298","381150.302","381150.306","381150.31","381150.314","381150.318","381150.322","381150.326","381150.33","381150.334","381150.338","381150.342","381150.346","381150.35","381150.354","381150.358","381150.362","381150.366","381150.37","381150.374","381150.378","381150.381","381150.385","381150.389","381150.393","381150.397","381150.401","381150.405","381150.409","381150.413","381150.417","381150.421","381150.425","381150.429","381150.433","381150.437","381150.441","381150.445","381150.449","381150.453","381150.457","381150.461","381150.465","381150.469","381150.473","381150.477","381150.481","381150.485","381150.489","381150.493","381150.497","381150.501","381150.505","381150.508","381150.512","381150.516","381150.52","381150.524","381150.528","381150.532","381150.536","381150.54","381150.544","381150.548","381150.552","381150.556","381150.56","381150.564","381150.568","381150.571","381150.575","381150.579","381150.583","381150.587","381150.591","381150.595","381150.599","381150.603","381150.607","381150.61","381150.614","381150.618","381150.622","381150.626","381150.63","381150.634","381150.638","381150.642","381150.646","381150.649","381150.653","381150.657","381150.661","381150.665","381150.669","381150.673","381150.677","381150.68","381150.684","381150.688","381150.692","381150.696","381150.7","381150.704","381150.707","381150.711","381150.715","381150.719","381150.723","381150.726","381150.73","381150.734","381150.738","381150.742","381150.745","381150.749","381150.752","381150.756","381150.759","381150.762","381150.765","381150.768","381150.771","381150.774","381150.777","381150.779","381150.782","381150.784","381150.787","381150.789","381150.791","381150.794","381150.796","381150.799","381150.801","381150.804","381150.807","381150.809","381150.812","381150.815","381150.818","381150.821","381150.825","381150.828","381150.831","381150.835","381150.839","381150.843","381150.847","381150.85","381150.854","381150.858","381150.862","381150.866","381150.87","381150.874","381150.878","381150.882","381150.885","381150.889","381150.893","381150.897","381150.9","381150.904","381150.908","381150.912","381150.915","381150.919","381150.922","381150.926","381150.929","381150.933","381150.936","381150.94","381150.943","381150.946","381150.949","381150.952","381150.955","381150.958","381150.961","381150.964","381150.967","381150.969","381150.972","381150.975","381150.978","381150.98","381150.983","381150.986","381150.988","381150.991","381150.993","381150.995","381150.998","381151","381151.002","381151.004","381151.005","381151.007","381151.009","381151.01","381151.011","381151.013","381151.014","381151.015","381151.015","381151.016","381151.017","381151.017","381151.017","381151.017","381151.019","381151.02","381151.02","381151.021","381151.022","381151.022","381151.023","381151.024","381151.026","381151.027","381151.028","381151.029","381151.031","381151.032","381151.034","381151.036","381151.038","381151.039","381151.041","381151.044","381151.046","381151.048","381151.05","381151.053","381151.055","381151.058","381151.06","381151.063","381151.066","381151.069","381151.072","381151.075","381151.078","381151.081","381151.084","381151.088","381151.091","381151.095","381151.098","381151.102","381151.106","381151.11","381151.113","381151.117","381151.121","381151.125","381151.129","381151.133","381151.137","381151.141","381151.145","381151.149","381151.154","381151.158","381151.162","381151.166","381151.17","381151.174","381151.178","381151.182","381151.186","381151.19","381151.194","381151.199","381151.203","381151.207","381151.211","381151.215","381151.219","381151.223","381151.227","381151.231","381151.235","381151.239","381151.243","381151.247","381151.251","381151.255","381151.258","381151.262","381151.266","381151.27","381151.274","381151.278","381151.282","381151.286","381151.29","381151.294","381151.298","381151.302","381151.306","381151.31","381151.314","381151.318","381151.322","381151.326","381151.33","381151.334","381151.337","381151.341","381151.345","381151.349","381151.353","381151.357","381151.361","381151.365","381151.369","381151.373","381151.377","381151.381","381151.385","381151.389","381151.392","381151.396","381151.4","381151.404","381151.408","381151.412","381151.416","381151.42","381151.424","381151.428","381151.432","381151.436","381151.439","381151.443","381151.447","381151.451","381151.455","381151.459","381151.463","381151.467","381151.471","381151.475","381151.479","381151.483","381151.486","381151.49","381151.494","381151.498","381151.502","381151.506","381151.51","381151.514","381151.518","381151.522","381151.526","381151.53","381151.534","381151.537","381151.541","381151.545","381151.549","381151.553","381151.557","381151.561","381151.565","381151.569","381151.573","381151.577","381151.581","381151.585","381151.589","381151.592","381151.596","381151.6","381151.604","381151.608","381151.612","381151.616","381151.62","381151.624","381151.628","381151.632","381151.636","381151.64","381151.644","381151.648","381151.651","381151.655","381151.659","381151.663","381151.667","381151.671","381151.675","381151.679","381151.683","381151.687","381151.691","381151.695","381151.699","381151.703","381151.707","381151.71","381151.714","381151.718","381151.722","381151.726","381151.73","381151.734","381151.738","381151.742","381151.746","381151.75","381151.754","381151.758","381151.762","381151.766","381151.769","381151.773","381151.777","381151.781","381151.785","381151.789","381151.793","381151.797","381151.801","381151.805","381151.809","381151.813","381151.817","381151.821","381151.825","381151.828","381151.832","381151.836","381151.84","381151.844","381151.848","381151.852","381151.856","381151.86","381151.864","381151.868","381151.871","381151.875","381151.879","381151.883","381151.887","381151.891","381151.895","381151.899","381151.903","381151.907","381151.911","381151.914","381151.918","381151.922","381151.926","381151.93","381151.934","381151.938","381151.942","381151.946","381151.949","381151.953","381151.957","381151.961","381151.965","381151.969","381151.973","381151.977","381151.981","381151.985","381151.988","381151.992","381151.996","381152","381152.004","381152.008","381152.012","381152.016","381152.02","381152.024","381152.028","381152.031","381152.035","381152.039","381152.043","381152.047","381152.051","381152.055","381152.059","381152.063","381152.067","381152.07","381152.074","381152.078","381152.082","381152.086","381152.09","381152.094","381152.098","381152.102","381152.106","381152.11","381152.114","381152.118","381152.121","381152.125","381152.129","381152.134","381152.138","381152.142","381152.147","381152.151","381152.156","381152.161","381152.166","381152.171","381152.176","381152.181","381152.187","381152.192","381152.198","381152.204","381152.21","381152.216","381152.222","381152.229","381152.235","381152.242","381152.249","381152.256","381152.263","381152.27","381152.277","381152.284","381152.292","381152.3","381152.307","381152.315","381152.323","381152.331","381152.339","381152.348","381152.356","381152.365","381152.373","381152.382","381152.391","381152.4","381152.409","381152.418","381152.428","381152.437","381152.447","381152.457","381152.466","381152.476","381152.486","381152.496","381152.507","381152.517","381152.527","381152.538","381152.549","381152.559","381152.57","381152.581","381152.592","381152.604","381152.615","381152.626","381152.638","381152.649","381152.66","381152.672","381152.683","381152.695","381152.706","381152.717","381152.729","381152.74","381152.751","381152.762","381152.773","381152.784","381152.795","381152.806","381152.817","381152.828","381152.838","381152.849","381152.859","381152.87","381152.88","381152.89","381152.9","381152.91","381152.92","381152.93","381152.94","381152.949","381152.959","381152.969","381152.978","381152.987","381152.996","381153.005","381153.014","381153.023","381153.032","381153.041","381153.049","381153.058","381153.066","381153.075","381153.083","381153.091","381153.1","381153.108","381153.116","381153.124","381153.132","381153.14","381153.148","381153.155","381153.163","381153.171","381153.178","381153.186","381153.193","381153.201","381153.208","381153.215","381153.223","381153.23","381153.237","381153.244","381153.251","381153.258","381153.265","381153.272","381153.279","381153.286","381153.293","381153.3","381153.307","381153.314","381153.321","381153.328","381153.335","381153.342","381153.348","381153.355","381153.362","381153.369","381153.376","381153.383","381153.389","381153.396","381153.403","381153.411","381153.418","381153.425","381153.432","381153.439","381153.446","381153.454","381153.461","381153.469","381153.476","381153.484","381153.491","381153.499","381153.506","381153.514","381153.522","381153.53","381153.538","381153.546","381153.553","381153.561","381153.57","381153.578","381153.586","381153.594","381153.602","381153.61","381153.618","381153.626","381153.635","381153.643","381153.651","381153.659","381153.667","381153.675","381153.683","381153.691","381153.699","381153.707","381153.715","381153.723","381153.731","381153.739","381153.747","381153.755","381153.764","381153.772","381153.78","381153.788","381153.796","381153.804","381153.813","381153.821","381153.829","381153.838","381153.846","381153.855","381153.863","381153.872","381153.88","381153.889","381153.898","381153.907","381153.915","381153.925","381153.934","381153.943","381153.952","381153.961","381153.971","381153.981","381153.99","381154","381154.01","381154.02","381154.03","381154.04","381154.05","381154.061","381154.071","381154.082","381154.092","381154.103","381154.114","381154.125","381154.136","381154.147","381154.158","381154.17","381154.181","381154.193","381154.204","381154.216","381154.228","381154.24","381154.252","381154.264","381154.276","381154.288","381154.3","381154.313","381154.325","381154.337","381154.35","381154.362","381154.374","381154.386","381154.399","381154.411","381154.423","381154.435","381154.448","381154.46","381154.472","381154.484","381154.496","381154.509","381154.521","381154.533","381154.545","381154.557","381154.569","381154.581","381154.593","381154.606","381154.618","381154.63","381154.642","381154.654","381154.667","381154.679","381154.691","381154.704","381154.716","381154.728","381154.741","381154.753","381154.766","381154.778","381154.79","381154.803","381154.816","381154.828","381154.841","381154.853","381154.866","381154.879","381154.891","381154.904","381154.917","381154.929","381154.942","381154.955","381154.967","381154.98","381154.992","381155.004","381155.017","381155.029","381155.041","381155.053","381155.065","381155.077","381155.089","381155.1","381155.112","381155.124","381155.135","381155.147","381155.158","381155.17","381155.181","381155.192","381155.204","381155.215","381155.226","381155.238","381155.249","381155.26","381155.271","381155.282","381155.294","381155.305","381155.316","381155.327","381155.339","381155.35","381155.362","381155.373","381155.385","381155.397","381155.409","381155.42","381155.433","381155.445","381155.457","381155.47","381155.482","381155.495","381155.507","381155.52","381155.533","381155.546","381155.56","381155.573","381155.587","381155.6","381155.614","381155.628","381155.642","381155.656","381155.671","381155.685","381155.7","381155.715","381155.73","381155.745","381155.76","381155.776","381155.791","381155.807","381155.823","381155.839","381155.855","381155.872","381155.888","381155.905","381155.921","381155.938","381155.955","381155.972","381155.99","381156.007","381156.025","381156.042","381156.06","381156.078","381156.096","381156.114","381156.132","381156.151","381156.169","381156.188","381156.207","381156.226","381156.244","381156.264","381156.283","381156.302","381156.321","381156.341","381156.361","381156.38","381156.4","381156.42","381156.44","381156.46","381156.48","381156.501","381156.521","381156.542","381156.562","381156.583","381156.604","381156.624","381156.646","381156.667","381156.688","381156.709","381156.73","381156.752","381156.773","381156.795","381156.817","381156.839","381156.861","381156.883","381156.905","381156.927","381156.95","381156.972","381156.995","381157.017","381157.04","381157.063","381157.086","381157.109","381157.133","381157.156","381157.18","381157.203","381157.227","381157.251","381157.275","381157.299","381157.323","381157.348","381157.372","381157.397","381157.422","381157.447","381157.472","381157.497","381157.522","381157.547","381157.573","381157.598","381157.624","381157.65","381157.676","381157.702","381157.728","381157.754","381157.78","381157.807","381157.833","381157.86","381157.887","381157.914","381157.941","381157.968","381157.995","381158.022","381158.049","381158.076","381158.104","381158.131","381158.158","381158.185","381158.212","381158.24","381158.267","381158.294","381158.321","381158.349","381158.376","381158.403","381158.43","381158.457","381158.485","381158.512","381158.539","381158.567","381158.594","381158.621","381158.648","381158.676","381158.703","381158.73","381158.758","381158.785","381158.812","381158.84","381158.867","381158.894","381158.922","381158.949","381158.976","381159.004","381159.031","381159.059","381159.086","381159.113","381159.141","381159.168","381159.196","381159.223","381159.25","381159.278","381159.305","381159.333","381159.36","381159.388","381159.415","381159.442","381159.47","381159.498","381159.525","381159.553","381159.58","381159.608","381159.635","381159.663","381159.69","381159.718","381159.745","381159.773","381159.801","381159.828","381159.856","381159.883","381159.911","381159.939","381159.966","381159.994","381160.022","381160.049","381160.077","381160.105","381160.132","381160.16","381160.188","381160.215","381160.243","381160.271","381160.299","381160.326","381160.354","381160.382","381160.41","381160.437","381160.465","381160.493","381160.521","381160.549","381160.577","381160.604","381160.632","381160.66","381160.688","381160.716","381160.744","381160.772","381160.8","381160.827","381160.855","381160.883","381160.911","381160.939","381160.966","381160.994","381161.022","381161.049","381161.077","381161.104","381161.132","381161.16","381161.187","381161.215","381161.242","381161.27","381161.298","381161.325","381161.353","381161.381","381161.408","381161.436","381161.464","381161.491","381161.519","381161.547","381161.574","381161.602","381161.63","381161.657","381161.685","381161.713","381161.74","381161.768","381161.796","381161.824","381161.852","381161.879","381161.907","381161.935","381161.963","381161.991","381162.018","381162.046","381162.074","381162.102","381162.13","381162.157","381162.185","381162.213","381162.241","381162.269","381162.296","381162.324","381162.352","381162.379","381162.407","381162.435","381162.463","381162.49","381162.518","381162.546","381162.573","381162.601","381162.629","381162.656","381162.684","381162.711","381162.739","381162.767","381162.794","381162.822","381162.85","381162.877","381162.905","381162.932","381162.96","381162.988","381163.015","381163.043","381163.07","381163.098","381163.125","381163.153","381163.18","381163.207","381163.235","381163.262","381163.289","381163.316","381163.343","381163.37","381163.397","381163.424","381163.451","381163.477","381163.504","381163.531","381163.557","381163.584","381163.61","381163.636","381163.662","381163.688","381163.714","381163.74","381163.766","381163.792","381163.818","381163.843","381163.869","381163.895","381163.92","381163.946","381163.971","381163.996","381164.021","381164.046","381164.071","381164.095","381164.12","381164.144","381164.168","381164.191","381164.215","381164.238","381164.262","381164.285","381164.308","381164.33","381164.353","381164.375","381164.397","381164.419","381164.441","381164.462","381164.484","381164.505","381164.526","381164.547","381164.567","381164.588","381164.608","381164.628","381164.648","381164.667","381164.687","381164.706","381164.725","381164.744","381164.763","381164.781","381164.8","381164.818","381164.836","381164.854","381164.871","381164.889","381164.906","381164.923","381164.94","381164.956","381164.973","381164.989","381165.005","381165.021","381165.037","381165.053","381165.068","381165.084","381165.099","381165.114","381165.128","381165.143","381165.158","381165.172","381165.186","381165.201","381165.215","381165.23","381165.244","381165.258","381165.272","381165.286","381165.3","381165.314","381165.328","381165.342","381165.356","381165.37","381165.384","381165.397","381165.411","381165.424","381165.438","381165.451","381165.465","381165.478","381165.491","381165.504","381165.517","381165.53","381165.543","381165.556","381165.569","381165.582","381165.595","381165.608","381165.621","381165.634","381165.647","381165.659","381165.672","381165.685","381165.698","381165.71","381165.723","381165.736","381165.748","381165.761","381165.773","381165.786","381165.799","381165.811","381165.824","381165.836","381165.849","381165.861","381165.874","381165.886","381165.898","381165.911","381165.923","381165.935","381165.947","381165.959","381165.972","381165.984","381165.996","381166.007","381166.019","381166.031","381166.043","381166.054","381166.066","381166.077","381166.089","381166.1","381166.112","381166.123","381166.134","381166.145","381166.156","381166.167","381166.178","381166.189","381166.2","381166.211","381166.221","381166.232","381166.242","381166.252","381166.262","381166.272","381166.282","381166.292","381166.301","381166.311","381166.32","381166.329","381166.338","381166.347","381166.356","381166.364","381166.372","381166.38","381166.388","381166.396","381166.404","381166.411","381166.418","381166.425","381166.432","381166.439","381166.445","381166.452","381166.458","381166.465","381166.471","381166.477","381166.484","381166.49","381166.496","381166.503","381166.509","381166.515","381166.521","381166.527","381166.533","381166.539","381166.545","381166.551","381166.556","381166.562","381166.567","381166.572","381166.577","381166.582","381166.586","381166.59","381166.595","381166.599","381166.603","381166.607","381166.611","381166.614","381166.618","381166.621","381166.624","381166.626","381166.628","381166.63","381166.631","381166.632","381166.632","381166.632","381166.632","381166.634","381166.634","381166.635","381166.636","381166.637","381166.638","381166.639","381166.641","381166.642","381166.644","381166.646","381166.648","381166.65","381166.652","381166.654","381166.657","381166.66","381166.663","381166.666","381166.669","381166.673","381166.677","381166.68","381166.684","381166.688","381166.693","381166.697","381166.702","381166.707","381166.711","381166.717","381166.722","381166.727","381166.733","381166.739","381166.745","381166.751","381166.758","381166.764","381166.771","381166.778","381166.785","381166.792","381166.799","381166.807","381166.815","381166.823","381166.831","381166.839","381166.847","381166.856","381166.865","381166.874","381166.883","381166.892","381166.901","381166.911","381166.921","381166.931","381166.941","381166.951","381166.962","381166.973","381166.984","381166.995","381167.006","381167.017","381167.029","381167.04","381167.052","381167.064","381167.077","381167.089","381167.102","381167.114","381167.127","381167.14","381167.154","381167.167","381167.18","381167.194","381167.208","381167.222","381167.236","381167.251","381167.265","381167.28","381167.295","381167.31","381167.326","381167.341","381167.357","381167.372","381167.388","381167.404","381167.421","381167.437","381167.453","381167.47","381167.487","381167.504","381167.521","381167.538","381167.555","381167.573","381167.59","381167.608","381167.626","381167.644","381167.662","381167.68","381167.699","381167.717","381167.735","381167.754","381167.773","381167.792","381167.811","381167.83","381167.849","381167.868","381167.888","381167.907","381167.927","381167.946","381167.966","381167.986","381168.006","381168.026","381168.046","381168.067","381168.087","381168.108","381168.128","381168.149","381168.17","381168.191","381168.212","381168.233","381168.254","381168.275","381168.296","381168.318","381168.339","381168.361","381168.383","381168.404","381168.426","381168.448","381168.47","381168.492","381168.514","381168.537","381168.559","381168.582","381168.604","381168.627","381168.649","381168.672","381168.695","381168.718","381168.741","381168.764","381168.787","381168.81","381168.833","381168.857","381168.88","381168.903","381168.927","381168.95","381168.974","381168.998","381169.022","381169.045","381169.069","381169.093","381169.117","381169.141","381169.166","381169.19","381169.214","381169.239","381169.263","381169.288","381169.312","381169.337","381169.362","381169.386","381169.411","381169.436","381169.461","381169.486","381169.511","381169.536","381169.561","381169.587","381169.612","381169.637","381169.663","381169.688","381169.714","381169.739","381169.765","381169.791","381169.816","381169.842","381169.868","381169.894","381169.92","381169.946","381169.972","381169.998","381170.025","381170.051","381170.077","381170.104","381170.13","381170.157","381170.183","381170.21","381170.236","381170.263","381170.29","381170.317","381170.344","381170.37","381170.397","381170.424","381170.451","381170.478","381170.505","381170.532","381170.559","381170.586","381170.613","381170.64","381170.667","381170.694","381170.721","381170.748","381170.775","381170.802","381170.829","381170.855","381170.883","381170.91","381170.937","381170.964","381170.991","381171.018","381171.045","381171.073","381171.1","381171.127","381171.155","381171.182","381171.21","381171.237","381171.265","381171.293","381171.32","381171.348","381171.375","381171.403","381171.431","381171.458","381171.486","381171.514","381171.541","381171.569","381171.597","381171.624","381171.652","381171.68","381171.708","381171.735","381171.763","381171.791","381171.818","381171.846","381171.874","381171.901","381171.929","381171.957","381171.984","381172.012","381172.04","381172.067","381172.095","381172.123","381172.15","381172.178","381172.206","381172.233","381172.261","381172.289","381172.316","381172.344","381172.371","381172.399","381172.427","381172.454","381172.482","381172.509","381172.537","381172.565","381172.592","381172.62","381172.647","381172.675","381172.703","381172.73","381172.758","381172.785","381172.813","381172.84","381172.868","381172.896","381172.923","381172.951","381172.978","381173.006","381173.034","381173.061","381173.089","381173.116","381173.144","381173.172","381173.199","381173.227","381173.254","381173.282","381173.309","381173.337","381173.364","381173.392","381173.419","381173.447","381173.475","381173.502","381173.53","381173.557","381173.585","381173.612","381173.64","381173.667","381173.695","381173.722","381173.75","381173.777","381173.805","381173.832","381173.86","381173.887","381173.915","381173.942","381173.97","381173.997","381174.025","381174.053","381174.08","381174.108","381174.135","381174.163","381174.19","381174.218","381174.245","381174.273","381174.3","381174.328","381174.355","381174.383","381174.411","381174.438","381174.466","381174.493","381174.521","381174.548","381174.576","381174.603","381174.631","381174.659","381174.686","381174.714","381174.741","381174.769","381174.796","381174.824","381174.851","381174.879","381174.906","381174.934","381174.962","381174.989","381175.017","381175.044","381175.072","381175.099","381175.127","381175.154","381175.182","381175.209","381175.237","381175.264","381175.292","381175.319","381175.347","381175.375","381175.402","381175.43","381175.457","381175.485","381175.512","381175.54","381175.567","381175.595","381175.622","381175.65","381175.677","381175.705","381175.732","381175.76","381175.787","381175.815","381175.842","381175.87","381175.897","381175.925","381175.953","381175.98","381176.008","381176.035","381176.063","381176.09","381176.118","381176.146","381176.173","381176.201","381176.228","381176.256","381176.284","381176.311","381176.339","381176.366","381176.394","381176.421","381176.449","381176.476","381176.504","381176.532","381176.559","381176.587","381176.614","381176.642","381176.669","381176.697","381176.724","381176.752","381176.779","381176.807","381176.834","381176.862","381176.889","381176.917","381176.944","381176.972","381176.999","381177.027","381177.054","381177.082","381177.109","381177.137","381177.164","381177.192","381177.219","381177.247","381177.274","381177.302","381177.329","381177.357","381177.384","381177.412","381177.439","381177.467","381177.494","381177.522","381177.549","381177.577","381177.604","381177.632","381177.659","381177.687","381177.714","381177.742","381177.769","381177.797","381177.824","381177.852","381177.879","381177.907","381177.934","381177.962","381177.989","381178.017","381178.044","381178.072","381178.099","381178.127","381178.154","381178.182","381178.209","381178.237","381178.264","381178.292","381178.319","381178.347","381178.374","381178.402","381178.429","381178.457","381178.484","381178.512","381178.54","381178.567","381178.595","381178.622","381178.65","381178.677","381178.705","381178.732","381178.76","381178.787","381178.815","381178.842","381178.869","381178.897","381178.924","381178.952","381178.979","381179.007","381179.034","381179.061","381179.089","381179.116","381179.144","381179.171","381179.198","381179.226","381179.253","381179.281","381179.308","381179.335","381179.363","381179.39","381179.417","381179.445","381179.472","381179.5","381179.527","381179.555","381179.582","381179.61","381179.637","381179.665","381179.693","381179.72","381179.748","381179.775","381179.803","381179.831","381179.858","381179.886","381179.913","381179.941","381179.969","381179.996","381180.024","381180.052","381180.079","381180.107","381180.134","381180.162","381180.19","381180.217","381180.245","381180.272","381180.3","381180.328","381180.355","381180.383","381180.411","381180.438","381180.466","381180.493","381180.521","381180.549","381180.576","381180.604","381180.632","381180.659","381180.687","381180.714","381180.742","381180.77","381180.797","381180.825","381180.852","381180.88","381180.907","381180.935","381180.962","381180.989","381181.016","381181.044","381181.071","381181.098","381181.125","381181.151","381181.178","381181.205","381181.231","381181.257","381181.284","381181.31","381181.336","381181.363","381181.389","381181.415","381181.44","381181.466","381181.492","381181.517","381181.543","381181.568","381181.593","381181.617","381181.642","381181.666","381181.69","381181.713","381181.736","381181.759","381181.781","381181.803","381181.825","381181.845","381181.865","381181.885","381181.904","381181.922","381181.939","381181.956","381181.972","381181.987","381182.002","381182.016","381182.03","381182.043","381182.057","381182.069","381182.082","381182.094","381182.106","381182.119","381182.131","381182.142","381182.154","381182.166","381182.178","381182.189","381182.201","381182.212","381182.224","381182.235","381182.246","381182.257","381182.269","381182.28","381182.291","381182.301","381182.312","381182.323","381182.334","381182.345","381182.355","381182.366","381182.376","381182.387","381182.397","381182.408","381182.418","381182.428","381182.438","381182.448","381182.458","381182.469","381182.478","381182.488","381182.498","381182.508","381182.518","381182.528","381182.538","381182.548","381182.557","381182.567","381182.577","381182.586","381182.596","381182.605","381182.615","381182.624","381182.634","381182.643","381182.652","381182.662","381182.671","381182.68","381182.689","381182.699","381182.708","381182.717","381182.727","381182.736","381182.745","381182.754","381182.764","381182.773","381182.782","381182.791","381182.801","381182.81","381182.819","381182.828","381182.838","381182.847","381182.856","381182.865","381182.875","381182.884","381182.893","381182.903","381182.912","381182.921","381182.93","381182.939","381182.949","381182.958","381182.967","381182.976","381182.985","381182.994","381183.003","381183.012","381183.02","381183.029","381183.038","381183.047","381183.056","381183.064","381183.073","381183.081","381183.09","381183.098","381183.107","381183.115","381183.123","381183.132","381183.14","381183.148","381183.156","381183.164","381183.172","381183.18","381183.188","381183.196","381183.203","381183.211","381183.219","381183.226","381183.234","381183.241","381183.249","381183.256","381183.263","381183.27","381183.277","381183.284","381183.29","381183.297","381183.303","381183.309","381183.315","381183.321","381183.327","381183.333","381183.338","381183.344","381183.349","381183.355","381183.36","381183.365","381183.37","381183.375","381183.38","381183.385","381183.39","381183.395","381183.4","381183.405","381183.409","381183.414","381183.418","381183.423","381183.427","381183.431","381183.436","381183.44","381183.445","381183.449","381183.454","381183.458","381183.463","381183.468","381183.473","381183.477","381183.483","381183.488","381183.493","381183.498","381183.504","381183.51","381183.515","381183.521","381183.527","381183.533","381183.539","381183.546","381183.552","381183.559","381183.566","381183.573","381183.58","381183.588","381183.595","381183.603","381183.611","381183.619","381183.628","381183.636","381183.645","381183.654","381183.663","381183.672","381183.682","381183.691","381183.701","381183.711","381183.721","381183.731","381183.742","381183.752","381183.763","381183.773","381183.784","381183.795","381183.805","381183.816","381183.827","381183.837","381183.848","381183.859","381183.87","381183.881","381183.891","381183.902","381183.913","381183.924","381183.934","381183.945","381183.956","381183.967","381183.978","381183.988","381183.999","381184.01","381184.021","381184.032","381184.042","381184.053","381184.064","381184.075","381184.086","381184.096","381184.107","381184.118","381184.128","381184.139","381184.149","381184.159","381184.17","381184.18","381184.19","381184.2","381184.21","381184.219","381184.229","381184.239","381184.248","381184.258","381184.267","381184.277","381184.286","381184.296","381184.305","381184.315","381184.324","381184.334","381184.343","381184.352","381184.362","381184.371","381184.38","381184.39","381184.399","381184.409","381184.418","381184.427","381184.436","381184.445","381184.454","381184.463","381184.472","381184.481","381184.489","381184.497","381184.505","381184.513","381184.521","381184.529","381184.536","381184.543","381184.55","381184.557","381184.564","381184.571","381184.577","381184.584","381184.59","381184.597","381184.603","381184.609","381184.616","381184.622","381184.628","381184.634","381184.64","381184.645","381184.651","381184.656","381184.662","381184.667","381184.672","381184.676","381184.68","381184.684","381184.688","381184.691","381184.694","381184.697","381184.698","381184.7","381184.7","381184.7","381184.7","381184.702","381184.702","381184.703","381184.703","381184.704","381184.705","381184.706","381184.707","381184.709","381184.71","381184.712","381184.714","381184.716","381184.718","381184.721","381184.724","381184.726","381184.729","381184.733","381184.736","381184.74","381184.743","381184.747","381184.752","381184.756","381184.761","381184.765","381184.77","381184.776","381184.781","381184.787","381184.792","381184.799","381184.805","381184.811","381184.818","381184.825","381184.832","381184.839","381184.847","381184.854","381184.862","381184.87","381184.879","381184.887","381184.896","381184.904","381184.913","381184.923","381184.932","381184.942","381184.952","381184.962","381184.972","381184.983","381184.993","381185.004","381185.015","381185.027","381185.038","381185.05","381185.062","381185.074","381185.086","381185.099","381185.112","381185.125","381185.138","381185.151","381185.165","381185.179","381185.193","381185.207","381185.221","381185.236","381185.251","381185.266","381185.281","381185.296","381185.312","381185.328","381185.344","381185.36","381185.376","381185.393","381185.409","381185.426","381185.443","381185.46","381185.478","381185.495","381185.513","381185.531","381185.548","381185.566","381185.585","381185.603","381185.621","381185.64","381185.658","381185.677","381185.696","381185.715","381185.734","381185.754","381185.773","381185.792","381185.812","381185.832","381185.852","381185.872","381185.892","381185.912","381185.932","381185.953","381185.973","381185.994","381186.015","381186.035","381186.056","381186.077","381186.098","381186.119","381186.14","381186.161","381186.181","381186.202","381186.223","381186.243","381186.264","381186.285","381186.305","381186.326","381186.346","381186.367","381186.387","381186.408","381186.428","381186.449","381186.469","381186.49","381186.51","381186.53","381186.551","381186.571","381186.591","381186.611","381186.631","381186.652","381186.672","381186.692","381186.712","381186.732","381186.752","381186.772","381186.792","381186.812","381186.832","381186.852","381186.872","381186.893","381186.913","381186.934","381186.955","381186.976","381186.997","381187.018","381187.039","381187.06","381187.082","381187.104","381187.126","381187.147","381187.17","381187.192","381187.214","381187.237","381187.259","381187.282","381187.305","381187.328","381187.351","381187.374","381187.397","381187.421","381187.444","381187.468","381187.492","381187.515","381187.539","381187.564","381187.588","381187.612","381187.637","381187.661","381187.686","381187.711","381187.735","381187.76","381187.785","381187.81","381187.836","381187.861","381187.886","381187.912","381187.937","381187.963","381187.988","381188.014","381188.04","381188.065","381188.091","381188.117","381188.143","381188.169","381188.195","381188.222","381188.248","381188.274","381188.3","381188.327","381188.353","381188.38","381188.406","381188.433","381188.459","381188.486","381188.513","381188.54","381188.566","381188.593","381188.62","381188.647","381188.674","381188.701","381188.728","381188.756","381188.783","381188.81","381188.837","381188.865","381188.892","381188.919","381188.947","381188.974","381189.002","381189.03","381189.057","381189.085","381189.113","381189.141","381189.168","381189.196","381189.224","381189.252","381189.28","381189.308","381189.336","381189.364","381189.393","381189.421","381189.449","381189.477","381189.505","381189.533","381189.561","381189.589","381189.617","381189.645","381189.672","381189.7","381189.728","381189.756","381189.783","381189.81","381189.838","381189.865","381189.893","381189.92","381189.948","381189.975","381190.002","381190.03","381190.057","381190.085","381190.112","381190.139","381190.167","381190.194","381190.222","381190.249","381190.277","381190.304","381190.331","381190.359","381190.386","381190.414","381190.441","381190.469","381190.496","381190.524","381190.552","381190.579","381190.607","381190.634","381190.662","381190.689","381190.717","381190.745","381190.772","381190.8","381190.827","381190.855","381190.883","381190.91","381190.938","381190.965","381190.993","381191.021","381191.048","381191.076","381191.103","381191.131","381191.159","381191.186","381191.214","381191.241","381191.269","381191.296","381191.324","381191.351","381191.379","381191.407","381191.434","381191.462","381191.489","381191.517","381191.544","381191.572","381191.599","381191.627","381191.654","381191.681","381191.709","381191.736","381191.764","381191.791","381191.818","381191.846","381191.873","381191.9","381191.928","381191.955","381191.982","381192.01","381192.037","381192.065","381192.092","381192.119","381192.146","381192.174","381192.201","381192.229","381192.256","381192.283","381192.311","381192.338","381192.365","381192.393","381192.42","381192.448","381192.475","381192.503","381192.53","381192.557","381192.585","381192.612","381192.64","381192.667","381192.695","381192.722","381192.75","381192.777","381192.805","381192.832","381192.859","381192.887","381192.914","381192.942","381192.969","381192.997","381193.024","381193.052","381193.079","381193.107","381193.134","381193.162","381193.189","381193.217","381193.244","381193.272","381193.299","381193.327","381193.354","381193.382","381193.409","381193.436","381193.464","381193.491","381193.519","381193.546","381193.574","381193.601","381193.629","381193.656","381193.684","381193.711","381193.739","381193.766","381193.794","381193.821","381193.849","381193.876","381193.904","381193.931","381193.959","381193.986","381194.014","381194.041","381194.069","381194.096","381194.124","381194.151","381194.179","381194.207","381194.234","381194.262","381194.289","381194.317","381194.344","381194.372","381194.399","381194.427","381194.454","381194.482","381194.509","381194.537","381194.564","381194.592","381194.619","381194.647","381194.675","381194.702","381194.73","381194.757","381194.785","381194.812","381194.84","381194.867","381194.895","381194.922","381194.95","381194.978","381195.005","381195.033","381195.06","381195.088","381195.115","381195.143","381195.17","381195.198","381195.225","381195.253","381195.28","381195.308","381195.335","381195.363","381195.39","381195.418","381195.445","381195.473","381195.5","381195.527","381195.555","381195.582","381195.61","381195.637","381195.665","381195.692","381195.72","381195.747","381195.775","381195.803","381195.83","381195.858","381195.885","381195.913","381195.941","381195.968","381195.996","381196.023","381196.051","381196.079","381196.106","381196.134","381196.162","381196.19","381196.217","381196.245","381196.273","381196.301","381196.328","381196.356","381196.384","381196.411","381196.439","381196.467","381196.495","381196.522","381196.55","381196.578","381196.605","381196.633","381196.661","381196.689","381196.716","381196.744","381196.772","381196.799","381196.827","381196.854","381196.882","381196.91","381196.937","381196.965","381196.992","381197.02","381197.047","381197.075","381197.103","381197.13","381197.158","381197.185","381197.213","381197.24","381197.268","381197.295","381197.323","381197.35","381197.378","381197.406","381197.433","381197.461","381197.488","381197.516","381197.543","381197.571","381197.599","381197.626","381197.654","381197.681","381197.709","381197.737","381197.764","381197.792","381197.82","381197.847","381197.875","381197.902","381197.93","381197.958","381197.985","381198.013","381198.041","381198.069","381198.096","381198.124","381198.152","381198.179","381198.207","381198.235","381198.262","381198.29","381198.318","381198.345","381198.373","381198.4","381198.428","381198.456","381198.483","381198.511","381198.538","381198.566","381198.593","381198.621","381198.648","381198.675","381198.703","381198.73","381198.758","381198.785","381198.813","381198.84","381198.868","381198.895","381198.923","381198.95","381198.978","381199.005","381199.033","381199.061","381199.088","381199.116","381199.143","381199.171","381199.199","381199.226","381199.254","381199.282","381199.309","381199.337","381199.364","381199.392","381199.42","381199.448","381199.475","381199.503","381199.531","381199.558","381199.586","381199.613","381199.641","381199.669","381199.696","381199.724","381199.752","381199.779","381199.807","381199.835","381199.863","381199.89","381199.918","381199.946","381199.973","381200.001","381200.029","381200.056","381200.084","381200.112","381200.139","381200.167","381200.195","381200.222","381200.25","381200.278","381200.305","381200.333","381200.361","381200.388","381200.416","381200.443","381200.471","381200.498","381200.526","381200.553","381200.581","381200.608","381200.636","381200.663","381200.691","381200.718","381200.746","381200.773","381200.801","381200.828","381200.855","381200.883","381200.91","381200.938","381200.965","381200.993","381201.02","381201.048","381201.075","381201.103","381201.13","381201.158","381201.185","381201.213","381201.24","381201.268","381201.296","381201.323","381201.351","381201.379","381201.407","381201.434","381201.462","381201.49","381201.518","381201.545","381201.573","381201.601","381201.629","381201.657","381201.685","381201.712","381201.74","381201.768","381201.796","381201.824","381201.851","381201.879","381201.907","381201.935","381201.962","381201.99","381202.018","381202.045","381202.073","381202.101","381202.128","381202.156","381202.184","381202.211","381202.239","381202.266","381202.294","381202.321","381202.349","381202.376","381202.404","381202.431","381202.459","381202.486","381202.514","381202.541","381202.569","381202.596","381202.624","381202.651","381202.679","381202.706","381202.734","381202.761","381202.789","381202.816","381202.844","381202.871","381202.899","381202.926","381202.954","381202.981","381203.009","381203.036","381203.064","381203.092","381203.119","381203.147","381203.174","381203.202","381203.229","381203.257","381203.284","381203.312","381203.339","381203.367","381203.394","381203.422","381203.449","381203.476","381203.504","381203.531","381203.558","381203.586","381203.613","381203.64","381203.668","381203.695","381203.722","381203.75","381203.777","381203.805","381203.832","381203.859","381203.887","381203.914","381203.941","381203.969","381203.996","381204.024","381204.051","381204.078","381204.106","381204.133","381204.161","381204.188","381204.216","381204.243","381204.271","381204.298","381204.326","381204.353","381204.381","381204.408","381204.436","381204.464","381204.491","381204.519","381204.546","381204.574","381204.601","381204.629","381204.657","381204.684","381204.712","381204.739","381204.767","381204.795","381204.822","381204.85","381204.877","381204.905","381204.933","381204.96","381204.988","381205.015","381205.043","381205.071","381205.098","381205.126","381205.154","381205.181","381205.209","381205.236","381205.264","381205.292","381205.319","381205.347","381205.374","381205.402","381205.43","381205.457","381205.485","381205.513","381205.54","381205.568","381205.595","381205.623","381205.651","381205.678","381205.706","381205.734","381205.761","381205.789","381205.817","381205.844","381205.872","381205.9","381205.927","381205.955","381205.983","381206.01","381206.038","381206.066","381206.093","381206.121","381206.149","381206.176","381206.204","381206.232","381206.259","381206.287","381206.315","381206.342","381206.37","381206.398","381206.425","381206.453","381206.481","381206.508","381206.536","381206.563","381206.591","381206.619","381206.646","381206.674","381206.701","381206.729","381206.757","381206.784","381206.812","381206.839","381206.867","381206.895","381206.922","381206.95","381206.978","381207.005","381207.033","381207.061","381207.088","381207.116","381207.143","381207.171","381207.199","381207.226","381207.254","381207.281","381207.309","381207.336","381207.364","381207.391","381207.419","381207.446","381207.474","381207.501","381207.528","381207.556","381207.583","381207.61","381207.638","381207.665","381207.692","381207.72","381207.747","381207.775","381207.802","381207.829","381207.857","381207.884","381207.911","381207.939","381207.966","381207.993","381208.02","381208.047","381208.074","381208.1","381208.127","381208.153","381208.18","381208.205","381208.231","381208.257","381208.282","381208.307","381208.332","381208.357","381208.381","381208.406","381208.429","381208.453","381208.477","381208.5","381208.523","381208.546","381208.569","381208.591","381208.612","381208.634","381208.655","381208.675","381208.695","381208.714","381208.733","381208.751","381208.769","381208.786","381208.802","381208.818","381208.833","381208.848","381208.862","381208.875","381208.888","381208.901","381208.913","381208.925","381208.937","381208.949","381208.961","381208.972","381208.984","381208.995","381209.006","381209.017","381209.029","381209.04","381209.051","381209.062","381209.073","381209.084","381209.095","381209.106","381209.117","381209.128","381209.139","381209.151","381209.162","381209.173","381209.184","381209.196","381209.207","381209.219","381209.23","381209.242","381209.254","381209.265","381209.277","381209.289","381209.301","381209.314","381209.326","381209.338","381209.351","381209.363","381209.376","381209.388","381209.401","381209.413","381209.426","381209.438","381209.451","381209.464","381209.477","381209.489","381209.502","381209.515","381209.527","381209.54","381209.552","381209.565","381209.577","381209.59","381209.602","381209.615","381209.627","381209.639","381209.651","381209.663","381209.675","381209.687","381209.699","381209.711","381209.723","381209.734","381209.746","381209.757","381209.769","381209.78","381209.791","381209.801","381209.812","381209.823","381209.833","381209.843","381209.854","381209.864","381209.875","381209.885","381209.895","381209.905","381209.915","381209.925","381209.936","381209.946","381209.955","381209.965","381209.975","381209.984","381209.994","381210.003","381210.012","381210.02","381210.028","381210.036","381210.044","381210.051","381210.058","381210.065","381210.071","381210.078","381210.083","381210.089","381210.094","381210.099","381210.104","381210.108","381210.113","381210.117","381210.121","381210.124","381210.128","381210.131","381210.135","381210.138","381210.14","381210.143","381210.145","381210.146","381210.147","381210.148","381210.148","381210.149","381210.149","381210.15","381210.15","381210.151","381210.152","381210.153","381210.154","381210.155","381210.156","381210.157","381210.159","381210.16","381210.162","381210.164","381210.165","381210.167","381210.17","381210.172","381210.174","381210.177","381210.18","381210.183","381210.186","381210.189","381210.192","381210.196","381210.2","381210.204","381210.208","381210.212","381210.216","381210.221","381210.225","381210.23","381210.235","381210.24","381210.246","381210.251","381210.257","381210.263","381210.268","381210.275","381210.281","381210.287","381210.294","381210.3","381210.307","381210.314","381210.321","381210.329","381210.336","381210.344","381210.352","381210.359","381210.367","381210.376","381210.384","381210.393","381210.401","381210.41","381210.419","381210.428","381210.438","381210.447","381210.457","381210.466","381210.476","381210.486","381210.496","381210.506","381210.517","381210.527","381210.538","381210.549","381210.56","381210.571","381210.582","381210.594","381210.605","381210.617","381210.628","381210.64","381210.652","381210.664","381210.677","381210.689","381210.702","381210.714","381210.727","381210.74","381210.753","381210.766","381210.779","381210.793","381210.806","381210.82","381210.833","381210.847","381210.861","381210.875","381210.89","381210.904","381210.918","381210.933","381210.948","381210.962","381210.977","381210.992","381211.008","381211.023","381211.038","381211.054","381211.069","381211.085","381211.101","381211.117","381211.133","381211.149","381211.166","381211.182","381211.199","381211.215","381211.232","381211.249","381211.266","381211.283","381211.3","381211.318","381211.335","381211.353","381211.37","381211.388","381211.406","381211.424","381211.442","381211.46","381211.479","381211.497","381211.516","381211.535","381211.553","381211.572","381211.591","381211.61","381211.63","381211.649","381211.669","381211.688","381211.708","381211.728","381211.747","381211.767","381211.787","381211.808","381211.828","381211.848","381211.869","381211.889","381211.91","381211.931","381211.952","381211.973","381211.994","381212.015","381212.036","381212.057","381212.079","381212.1","381212.122","381212.143","381212.165","381212.187","381212.209","381212.231","381212.253","381212.275","381212.297","381212.32","381212.342","381212.364","381212.387","381212.41","381212.432","381212.455","381212.478","381212.501","381212.524","381212.547","381212.57","381212.593","381212.616","381212.64","381212.663","381212.687","381212.71","381212.734","381212.758","381212.781","381212.805","381212.829","381212.853","381212.877","381212.901","381212.926","381212.95","381212.974","381212.998","381213.023","381213.047","381213.072","381213.096","381213.121","381213.146","381213.171","381213.195","381213.22","381213.245","381213.27","381213.295","381213.32","381213.346","381213.371","381213.396","381213.422","381213.447","381213.472","381213.498","381213.524","381213.549","381213.575","381213.601","381213.626","381213.652","381213.678","381213.704","381213.73","381213.756","381213.782","381213.808","381213.835","381213.861","381213.887","381213.914","381213.94","381213.966","381213.993","381214.02","381214.046","381214.073","381214.1","381214.126","381214.153","381214.18","381214.207","381214.234","381214.261","381214.288","381214.315","381214.343","381214.37","381214.397","381214.424","381214.452","381214.479","381214.507","381214.534","381214.561","381214.589","381214.616","381214.643","381214.671","381214.698","381214.725","381214.752","381214.779","381214.807","381214.834","381214.861","381214.888","381214.915","381214.942","381214.97","381214.997","381215.024","381215.051","381215.078","381215.105","381215.132","381215.16","381215.187","381215.214","381215.241","381215.268","381215.296","381215.323","381215.35","381215.377","381215.404","381215.431","381215.459","381215.486","381215.513","381215.54","381215.567","381215.595","381215.622","381215.649","381215.676","381215.704","381215.731","381215.758","381215.785","381215.812","381215.84","381215.867","381215.894","381215.921","381215.949","381215.976","381216.003","381216.03","381216.058","381216.085","381216.112","381216.139","381216.167","381216.194","381216.221","381216.248","381216.276","381216.303","381216.33","381216.358","381216.385","381216.412","381216.44","381216.467","381216.494","381216.521","381216.548","381216.575","381216.602","381216.629","381216.656","381216.683","381216.71","381216.737","381216.763","381216.79","381216.816","381216.843","381216.869","381216.895","381216.921","381216.947","381216.973","381216.999","381217.024","381217.05","381217.075","381217.1","381217.126","381217.151","381217.176","381217.201","381217.226","381217.25","381217.275","381217.3","381217.324","381217.348","381217.373","381217.397","381217.42","381217.444","381217.468","381217.491","381217.514","381217.537","381217.559","381217.582","381217.604","381217.626","381217.648","381217.669","381217.691","381217.712","381217.733","381217.754","381217.774","381217.794","381217.815","381217.834","381217.854","381217.873","381217.893","381217.911","381217.93","381217.948","381217.966","381217.983","381218","381218.017","381218.034","381218.05","381218.066","381218.081","381218.096","381218.111","381218.126","381218.14","381218.154","381218.168","381218.181","381218.194","381218.207","381218.219","381218.232","381218.244","381218.257","381218.269","381218.281","381218.293","381218.305","381218.317","381218.329","381218.341","381218.353","381218.365","381218.376","381218.388","381218.4","381218.411","381218.423","381218.434","381218.446","381218.457","381218.468","381218.48","381218.491","381218.502","381218.513","381218.525","381218.536","381218.547","381218.559","381218.57","381218.581","381218.592","381218.604","381218.615","381218.626","381218.638","381218.649","381218.66","381218.671","381218.683","381218.694","381218.705","381218.716","381218.728","381218.739","381218.75","381218.762","381218.773","381218.785","381218.796","381218.808","381218.819","381218.831","381218.842","381218.854","381218.866","381218.878","381218.889","381218.901","381218.913","381218.925","381218.937","381218.949","381218.961","381218.973","381218.985","381218.998","381219.01","381219.022","381219.034","381219.046","381219.058","381219.07","381219.082","381219.094","381219.106","381219.119","381219.131","381219.143","381219.155","381219.167","381219.179","381219.191","381219.203","381219.215","381219.227","381219.239","381219.251","381219.264","381219.276","381219.288","381219.3","381219.313","381219.325","381219.338","381219.35","381219.363","381219.375","381219.388","381219.401","381219.414","381219.427","381219.441","381219.454","381219.468","381219.481","381219.495","381219.509","381219.523","381219.537","381219.551","381219.565","381219.58","381219.595","381219.609","381219.624","381219.639","381219.654","381219.67","381219.685","381219.701","381219.716","381219.732","381219.748","381219.765","381219.781","381219.797","381219.814","381219.831","381219.848","381219.865","381219.882","381219.899","381219.916","381219.934","381219.952","381219.97","381219.988","381220.006","381220.024","381220.042","381220.061","381220.079","381220.098","381220.117","381220.135","381220.155","381220.174","381220.193","381220.212","381220.232","381220.251","381220.271","381220.291","381220.31","381220.33","381220.351","381220.371","381220.391","381220.411","381220.432","381220.452","381220.473","381220.494","381220.515","381220.536","381220.557","381220.578","381220.599","381220.62","381220.642","381220.663","381220.685","381220.706","381220.728","381220.75","381220.772","381220.794","381220.816","381220.839","381220.861","381220.883","381220.906","381220.929","381220.951","381220.974","381220.997","381221.02","381221.044","381221.067","381221.09","381221.114","381221.138","381221.162","381221.186","381221.21","381221.234","381221.259","381221.283","381221.308","381221.333","381221.358","381221.383","381221.409","381221.434","381221.46","381221.485","381221.511","381221.538","381221.564","381221.59","381221.617","381221.644","381221.67","381221.698","381221.725","381221.752","381221.779","381221.807","381221.835","381221.862","381221.89","381221.918","381221.946","381221.974","381222.002","381222.03","381222.057","381222.085","381222.113","381222.141","381222.169","381222.196","381222.224","381222.251","381222.279","381222.307","381222.334","381222.362","381222.389","381222.417","381222.444","381222.472","381222.499","381222.527","381222.554","381222.582","381222.609","381222.637","381222.664","381222.692","381222.719","381222.746","381222.774","381222.801","381222.828","381222.856","381222.883","381222.911","381222.938","381222.965","381222.993","381223.02","381223.047","381223.075","381223.102","381223.13","381223.157","381223.184","381223.212","381223.239","381223.266","381223.294","381223.321","381223.348","381223.376","381223.403","381223.431","381223.458","381223.485","381223.513","381223.54","381223.568","381223.595","381223.622","381223.65","381223.677","381223.705","381223.732","381223.76","381223.787","381223.814","381223.842","381223.869","381223.896","381223.924","381223.951","381223.979","381224.006","381224.033","381224.061","381224.088","381224.115","381224.142","381224.17","381224.197","381224.224","381224.252","381224.279","381224.306","381224.333","381224.361","381224.388","381224.416","381224.443","381224.47","381224.498","381224.525","381224.552","381224.579","381224.607","381224.634","381224.661","381224.688","381224.715","381224.742","381224.77","381224.797","381224.824","381224.851","381224.878","381224.906","381224.933","381224.96","381224.988","381225.015","381225.043","381225.07","381225.098","381225.126","381225.153","381225.181","381225.209","381225.237","381225.264","381225.292","381225.319","381225.347","381225.374","381225.402","381225.429","381225.457","381225.484","381225.512","381225.539","381225.566","381225.594","381225.621","381225.648","381225.675","381225.702","381225.729","381225.756","381225.783","381225.81","381225.837","381225.864","381225.891","381225.918","381225.945","381225.972","381225.999","381226.025","381226.052","381226.079","381226.105","381226.132","381226.158","381226.185","381226.212","381226.238","381226.264","381226.291","381226.317","381226.344","381226.371","381226.397","381226.424","381226.451","381226.478","381226.505","381226.532","381226.56","381226.587","381226.614","381226.642","381226.669","381226.697","381226.725","381226.752","381226.78","381226.808","381226.836","381226.863","381226.891","381226.919","381226.947","381226.974","381227.002","381227.03","381227.058","381227.086","381227.114","381227.141","381227.169","381227.197","381227.225","381227.253","381227.281","381227.308","381227.336","381227.364","381227.392","381227.42","381227.447","381227.475","381227.503","381227.531","381227.559","381227.586","381227.614","381227.642","381227.67","381227.697","381227.725","381227.753","381227.78","381227.808","381227.836","381227.863","381227.891","381227.919","381227.946","381227.974","381228.001","381228.029","381228.056","381228.084","381228.111","381228.139","381228.166","381228.194","381228.221","381228.248","381228.276","381228.303","381228.33","381228.358","381228.385","381228.412","381228.44","381228.467","381228.494","381228.522","381228.549","381228.576","381228.603","381228.631","381228.658","381228.685","381228.713","381228.74","381228.767","381228.794","381228.822","381228.849","381228.876","381228.903","381228.931","381228.958","381228.986","381229.013","381229.041","381229.068","381229.096","381229.123","381229.151","381229.179","381229.206","381229.234","381229.261","381229.289","381229.317","381229.344","381229.372","381229.4","381229.427","381229.455","381229.482","381229.51","381229.537","381229.565","381229.592","381229.62","381229.647","381229.674","381229.702","381229.729","381229.756","381229.784","381229.811","381229.838","381229.866","381229.893","381229.92","381229.947","381229.975","381230.002","381230.029","381230.056","381230.084","381230.111","381230.138","381230.165","381230.192","381230.22","381230.247","381230.274","381230.301","381230.329","381230.356","381230.383","381230.41","381230.437","381230.465","381230.492","381230.519","381230.546","381230.574","381230.601","381230.628","381230.655","381230.683","381230.71","381230.737","381230.765","381230.792","381230.82","381230.847","381230.875","381230.902","381230.93","381230.958","381230.985","381231.013","381231.041","381231.068","381231.096","381231.124","381231.151","381231.179","381231.207","381231.234","381231.262","381231.289","381231.317","381231.345","381231.372","381231.4","381231.427","381231.455","381231.482","381231.509","381231.537","381231.564","381231.591","381231.619","381231.646","381231.673","381231.701","381231.728","381231.755","381231.783","381231.81","381231.837","381231.864","381231.892","381231.919","381231.947","381231.974","381232.001","381232.029","381232.056","381232.083","381232.111","381232.138","381232.165","381232.192","381232.219","381232.246","381232.273","381232.3","381232.327","381232.354","381232.38","381232.407","381232.433","381232.459","381232.486","381232.512","381232.538","381232.564","381232.59","381232.616","381232.641","381232.667","381232.693","381232.718","381232.743","381232.769","381232.793","381232.818","381232.843","381232.867","381232.892","381232.916","381232.939","381232.963","381232.985","381233.008","381233.03","381233.052","381233.073","381233.093","381233.113","381233.132","381233.151","381233.169","381233.186","381233.202","381233.218","381233.233","381233.248","381233.262","381233.275","381233.288","381233.301","381233.313","381233.325","381233.337","381233.349","381233.36","381233.372","381233.383","381233.394","381233.405","381233.416","381233.427","381233.438","381233.449","381233.459","381233.47","381233.481","381233.491","381233.502","381233.512","381233.522","381233.533","381233.543","381233.553","381233.563","381233.573","381233.583","381233.593","381233.603","381233.612","381233.622","381233.631","381233.641","381233.65","381233.66","381233.669","381233.678","381233.687","381233.696","381233.705","381233.714","381233.723","381233.732","381233.74","381233.749","381233.758","381233.766","381233.775","381233.783","381233.791","381233.799","381233.807","381233.815","381233.824","381233.832","381233.839","381233.847","381233.855","381233.863","381233.871","381233.879","381233.887","381233.895","381233.903","381233.911","381233.919","381233.927","381233.935","381233.943","381233.951","381233.959","381233.967","381233.975","381233.983","381233.992","381234","381234.008","381234.016","381234.025","381234.033","381234.041","381234.049","381234.058","381234.066","381234.074","381234.082","381234.09","381234.098","381234.106","381234.114","381234.121","381234.129","381234.137","381234.144","381234.152","381234.159","381234.167","381234.174","381234.181","381234.189","381234.196","381234.203","381234.21","381234.216","381234.223","381234.23","381234.237","381234.243","381234.25","381234.256","381234.263","381234.269","381234.276","381234.282","381234.288","381234.294","381234.3","381234.306","381234.312","381234.318","381234.324","381234.33","381234.335","381234.341","381234.346","381234.352","381234.357","381234.363","381234.368","381234.373","381234.378","381234.383","381234.388","381234.393","381234.398","381234.403","381234.408","381234.413","381234.418","381234.422","381234.427","381234.431","381234.436","381234.44","381234.444","381234.449","381234.453","381234.457","381234.461","381234.465","381234.469","381234.473","381234.477","381234.481","381234.484","381234.488","381234.492","381234.495","381234.499","381234.502","381234.505","381234.509","381234.512","381234.515","381234.518","381234.521","381234.524","381234.527","381234.53","381234.533","381234.536","381234.539","381234.541","381234.544","381234.546","381234.549","381234.551","381234.554","381234.556","381234.558","381234.561","381234.563","381234.565","381234.567","381234.569","381234.571","381234.573","381234.574","381234.576","381234.578","381234.579","381234.581","381234.583","381234.585","381234.586","381234.588","381234.59","381234.592","381234.593","381234.595","381234.597","381234.599","381234.602","381234.604","381234.606","381234.609","381234.611","381234.614","381234.617","381234.62","381234.623","381234.626","381234.63","381234.633","381234.637","381234.641","381234.644","381234.648","381234.653","381234.657","381234.661","381234.665","381234.67","381234.675","381234.68","381234.684","381234.69","381234.695","381234.7","381234.705","381234.711","381234.717","381234.722","381234.728","381234.734","381234.741","381234.747","381234.753","381234.759","381234.766","381234.772","381234.779","381234.785","381234.792","381234.799","381234.805","381234.812","381234.818","381234.825","381234.831","381234.838","381234.845","381234.852","381234.858","381234.865","381234.872","381234.879","381234.886","381234.893","381234.9","381234.908","381234.915","381234.922","381234.929","381234.936","381234.944","381234.951","381234.958","381234.965","381234.971","381234.978","381234.985","381234.992","381234.999","381235.005","381235.012","381235.018","381235.025","381235.031","381235.037","381235.043","381235.049","381235.055","381235.061","381235.066","381235.072","381235.077","381235.082","381235.087","381235.092","381235.097","381235.102","381235.107","381235.112","381235.116","381235.121","381235.125","381235.13","381235.134","381235.139","381235.143","381235.148","381235.152","381235.156","381235.161","381235.165","381235.17","381235.174","381235.178","381235.183","381235.187","381235.192","381235.196","381235.201","381235.205","381235.209","381235.214","381235.218","381235.223","381235.227","381235.232","381235.236","381235.241","381235.245","381235.249","381235.254","381235.258","381235.263","381235.267","381235.272","381235.276","381235.281","381235.285","381235.289","381235.294","381235.298","381235.303","381235.307","381235.312","381235.316","381235.321","381235.325","381235.33","381235.334","381235.339","381235.343","381235.348","381235.352","381235.357","381235.361","381235.366","381235.37","381235.374","381235.379","381235.383","381235.388","381235.392","381235.397","381235.401","381235.406","381235.41","381235.415","381235.419","381235.424","381235.428","381235.433","381235.437","381235.442","381235.446","381235.451","381235.455","381235.46","381235.464","381235.469","381235.473","381235.478","381235.482","381235.487","381235.491","381235.496","381235.5","381235.505","381235.509","381235.514","381235.518","381235.523","381235.527","381235.532","381235.536","381235.541","381235.545","381235.55","381235.554","381235.559","381235.563","381235.568","381235.572","381235.577","381235.581","381235.586","381235.59","381235.595","381235.599","381235.604","381235.608","381235.613","381235.617","381235.622","381235.626","381235.631","381235.635","381235.64","381235.644","381235.649","381235.653","381235.658","381235.662","381235.667","381235.671","381235.676","381235.68","381235.685","381235.689","381235.694","381235.699","381235.703","381235.708","381235.712","381235.717","381235.721","381235.726","381235.73","381235.735","381235.739","381235.744","381235.748","381235.753","381235.757","381235.762","381235.766","381235.771","381235.775","381235.78","381235.784","381235.789","381235.793","381235.798","381235.802","381235.807","381235.812","381235.816","381235.821","381235.825","381235.83","381235.834","381235.839","381235.843","381235.848","381235.852","381235.857","381235.861","381235.866","381235.87","381235.875","381235.879","381235.884","381235.888","381235.893","381235.897","381235.902","381235.906","381235.911","381235.915","381235.92","381235.924","381235.929","381235.933","381235.938","381235.942","381235.947","381235.951","381235.956","381235.96","381235.965","381235.969","381235.974","381235.978","381235.983","381235.987","381235.992","381235.996","381236.001","381236.005","381236.009","381236.014","381236.018","381236.022","381236.027","381236.031","381236.035","381236.039","381236.043","381236.047","381236.051","381236.055","381236.059","381236.062","381236.065","381236.068","381236.071","381236.073","381236.075","381236.077","381236.078","381236.079"]
        list3 = ["1","1","1","1","2","2","3","3","3","3","4","4","5","5","5","5","6","6","7","7","8","8","9","9","10","10","11","11","11","12","13","13","13","14","14","15","15","15","15","15","15","15","15","15","15","15","15","14","13","13","12","11","11","9","8","7","7","7","7","6","6","6","6","6","6","6","6","7","7","7","7","7","7","7","7","8","7","7","8","8","8","8","9","9","9","9","9","9","9","9","9","10","10","10","11","11","11","12","12","13","13","13","13","14","14","15","14","15","14","14","14","14","15","14","14","14","14","13","13","13","13","13","13","13","13","13","13","14","13","13","14","14","13","14","14","14","14","14","14","14","15","14","14","14","14","14","14","14","14","14","14","14","14","15","15","15","15","15","15","15","15","15","15","15","14","15","15","14","15","15","14","14","15","15","14","15","15","14","14","14","14","14","14","14","14","14","14","13","14","14","14","14","14","14","14","14","14","14","14","14","14","14","14","14","14","14","13","13","14","13","14","13","13","14","14","14","14","14","14","14","14","14","14","15","15","15","15","15","15","15","15","15","15","15","15","15","15","15","15","15","15","15","15","15","15","15","16","15","15","15","15","15","15","16","15","16","15","16","16","16","16","16","17","17","17","17","17","18","18","19","18","19","19","19","20","20","21","21","22","22","23","23","24","25","25","26","26","27","28","29","29","30","31","31","32","33","33","34","35","35","36","37","37","38","39","39","40","41","41","42","43","43","44","44","45","46","47","47","48","49","49","50","51","51","53","53","54","55","55","56","57","57","58","59","59","60","61","62","63","63","64","65","65","66","67","67","68","69","70","70","71","72","72","73","73","73","73","73","73","73","73","72","71","71","70","69","69","67","65","64","63","61","59","57","55","54","52","50","48","46","44","43","41","40","39","38","37","37","36","36","35","35","35","35","35","35","35","35","35","35","35","35","36","36","37","37","37","37","38","39","39","39","40","40","41","41","42","42","43","43","44","44","45","45","45","46","47","47","47","48","48","49","49","50","50","51","51","52","52","53","53","54","55","55","55","56","57","57","58","59","59","60","60","61","62","62","63","64","65","65","66","67","67","68","69","70","70","71","71","72","73","73","74","75","76","76","77","77","78","79","79","80","81","81","82","82","83","84","84","85","86","86","87","87","88","88","89","89","90","91","91","91","92","93","93","93","94","95","95","96","96","96","97","97","98","98","99","99","100","100","100","100","100","100","100","100","100","100","100","99","99","99","99","99","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","99","98","98","98","99","98","98","99","98","99","99","98","99","99","99","99","99","98","99","99","99","99","99","99","99","99","99","99","99","99","99","99","100","99","99","100","99","100","100","100","100","100","100","100","99","100","100","99","99","99","99","99","98","98","98","98","98","98","98","98","98","98","99","98","98","98","99","99","99","99","99","98","99","99","99","99","99","99","98","99","99","98","99","99","99","99","99","98","99","99","99","98","99","99","98","99","98","99","98","99","99","98","98","99","98","99","98","98","99","98","98","99","98","99","98","99","98","99","98","99","98","99","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","97","98","97","98","98","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","98","97","97","97","97","97","98","97","97","97","97","98","97","97","97","97","97","98","97","97","97","97","97","97","98","97","97","97","98","97","97","97","97","97","97","97","98","97","98","97","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","99","99","98","99","99","99","99","99","99","99","99","99","98","98","98","98","98","98","98","98","98","98","98","97","97","97","97","97","97","97","96","96","96","97","96","96","96","96","96","96","96","96","96","97","96","97","97","97","97","97","97","97","97","98","97","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","97","98","98","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","96","97","96","97","97","96","96","96","97","96","97","96","97","96","96","96","96","96","97","96","96","96","97","96","97","96","96","97","97","97","96","97","97","97","97","97","97","97","97","97","97","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","97","97","97","97","97","97","96","96","96","96","96","96","96","96","96","96","96","96","96","96","96","96","97","96","97","97","97","97","98","97","98","98","98","98","98","98","98","98","98","98","98","99","98","99","98","98","98","99","98","98","98","98","98","98","98","97","97","97","97","96","96","97","96","96","96","96","95","95","95","95","95","95","95","95","94","94","94","94","93","94","93","93","93","93","93","93","93","92","92","92","92","92","91","91","91","91","91","91","91","91","91","91","91","91","91","92","92","92","93","93","93","94","95","95","95","96","96","96","97","97","98","98","98","98","99","98","99","99","99","99","99","99","99","99","99","99","99","98","99","98","99","98","98","98","98","98","98","98","98","98","98","98","98","97","97","97","97","97","96","96","96","96","96","96","96","96","96","95","95","95","95","95","95","94","94","94","94","93","93","93","93","93","93","93","93","93","93","93","94","94","95","95","95","95","96","96","97","97","98","98","98","99","100","100","100","100","100","100","100","100","100","100","100","100","100","100","99","100","99","99","99","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","97","98","98","97","98","98","97","97","98","97","97","97","97","97","97","97","97","97","97","98","97","97","97","98","97","98","97","98","97","98","97","98","97","97","98","97","97","97","98","97","97","98","98","97","98","98","98","98","98","98","98","98","98","98","98","98","98","99","99","98","99","99","99","99","99","99","99","99","99","99","98","99","99","99","98","98","99","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","97","98","97","97","98","97","97","98","97","97","98","97","98","97","98","97","98","97","98","98","98","98","97","98","98","98","98","97","98","98","97","98","97","97","97","98","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","98","97","97","97","98","98","98","98","98","98","98","98","98","98","98","98","98","99","99","99","99","99","99","99","99","99","99","99","99","98","99","99","99","98","99","99","98","99","99","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","97","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","97","98","98","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","97","98","97","97","98","98","98","98","98","98","98","98","97","98","98","98","97","98","97","97","98","98","97","97","97","98","97","97","98","97","97","98","97","98","97","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","99","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","99","98","98","99","98","98","98","98","99","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","97","97","97","97","96","96","96","96","95","95","95","94","93","93","93","92","92","92","91","91","90","90","89","89","88","88","87","87","87","85","85","84","83","82","81","80","79","78","77","76","75","75","74","73","73","73","73","73","73","73","73","73","73","73","73","73","74","73","73","73","74","73","74","74","73","74","74","74","74","74","74","74","74","74","74","74","74","74","74","74","75","74","74","74","75","74","74","74","75","74","74","75","75","75","75","75","75","75","75","75","76","76","76","77","77","77","78","78","78","78","79","79","80","80","81","81","82","82","83","83","83","84","84","85","85","85","86","87","87","87","88","89","89","89","90","90","91","91","92","92","93","93","93","94","95","95","96","96","96","96","97","96","97","97","96","97","97","97","97","97","97","97","97","97","97","97","98","98","98","98","98","98","98","99","99","99","100","99","99","99","99","100","99","100","99","100","100","100","100","99","100","100","99","100","100","99","100","99","100","100","99","99","100","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","98","99","99","98","99","99","99","99","99","98","99","99","98","99","99","99","99","99","98","98","99","99","98","99","99","98","99","98","98","99","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","97","98","98","98","98","98","97","98","98","98","98","98","98","98","98","98","98","98","99","98","99","99","98","99","99","99","98","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","98","99","99","99","99","99","98","99","99","98","99","99","98","98","99","98","99","98","99","98","99","98","99","98","99","98","99","98","98","99","99","98","98","99","98","98","99","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","97","98","97","98","98","98","97","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","99","98","98","99","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","99","98","98","99","98","98","99","98","99","99","99","99","99","99","99","99","99","99","99","98","99","99","99","99","99","99","98","99","99","99","99","99","99","99","99","98","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","100","99","99","100","99","100","100","100","100","100","100","100","100","100","100","100","100","100","100","100","100","100","100","100","100","100","100","100","100","100","100","101","100","100","100","100","100","99","98","98","98","98","97","97","96","97","96","97","96","96","97","97","96","97","97","97","97","97","98","98","97","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","97","98","97","96","97","96","96","96","96","96","96","96","96","96","95","96","95","95","95","95","95","95","94","95","95","95","95","95","94","95","95","95","95","95","95","95","95","95","95","96","96","96","96","96","96","96","96","96","96","96","96","96","97","96","97","97","97","97","97","97","98","98","98","98","98","98","98","99","99","98","99","99","99","99","99","99","100","99","99","99","99","99","99","99","99","99","99","99","98","99","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","97","98","98","98","97","98","98","98","98","98","98","98","98","98","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","98","99","98","98","98","98","98","98","98","98","98","98","98","98","98","98","99","98","99","99","99","99","100","99","99","99","99","100","99","99","100","99","99","99","99","99","99","100","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","98","99","99","99","99","98","99","99","98","99","98","98","99","98","98","98","98","99","98","98","99","98","99","98","98","98","98","99","98","99","98","99","99","98","99","98","99","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","97","98","98","98","98","97","98","98","98","98","97","98","98","98","98","97","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","99","98","99","99","99","99","99","98","99","99","98","99","99","99","99","98","99","98","99","99","98","98","99","99","98","99","98","99","99","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","97","98","98","98","98","97","98","98","98","98","97","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","99","99","99","98","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","98","99","99","99","98","99","98","99","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","99","98","98","98","98","98","98","98","99","98","99","98","98","98","98","98","98","98","98","98","99","98","99","98","98","99","98","99","98","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","98","98","98","98","98","97","97","96","96","96","96","95","95","95","95","94","93","93","93","92","92","91","91","91","90","90","89","89","88","88","87","87","86","85","85","85","84","84","84","83","82","82","81","81","80","80","79","79","78","77","77","77","76","75","75","75","74","73","73","72","71","71","70","70","69","69","68","68","67","66","66","65","65","64","64","63","63","62","62","61","61","60","59","58","57","57","55","54","53","51","50","48","46","45","43","42","42","41","40","40","40","39","40","39","39","40","40","40","40","40","39","40","40","39","40","39","40","40","39","40","39","39","39","40","39","39","39","39","39","39","39","39","39","39","39","38","39","39","38","38","38","38","38","37","37","37","37","36","35","35","35","34","33","33","32","31","31","30","30","29","29","28","27","27","26","25","25","25","24","23","23","22","22","21","21","20","19","19","18","18","17","17","16","16","15","14","14","14","14","14","13","14","13","13","13","13","14","13","13","14","13","14","13","14","14","14","14","14","14","14","14","14","13","14","14","14","14","14","14","14","14","14","14","14","14","13","14","14","14","14","14","14","14","14","14","14","13","14","14","14","14","13","14","14","13","14","14","13","14","14","14","14","13","14","14","13","14","13","14","14","13","14","13","14","13","14","13","14","13","13","14","13","14","13","13","13","14","13","13","14","13","13","13","13","14","13","13","13","13","13","13","13","13","13","13","13","13","13","13","13","13","13","13","13","13","13","13","13","13","13","13","13","13","13","13","13","13","13","12","13","12","12","11","11","10","10","10","9","9","9","8","9","8","8","8","8","8","9","9","9","9","9","9","10","10","11","11","12","12","13","13","13","13","13","13","13","13","14","13","13","13","13","13","13","13","13","13","13","13","13","12","13","12","12","12","12","11","11","11","11","11","10","10","10","10","10","9","9","9","9","9","9","9","9","9","8","8","8","8","7","7","6","6","5","5","4","4","4","4","3","3","2","2","1","1","0","1","1","2","2","2","2","3","3","4","4","4","4","5","5","5","5","6","6","7","7","7","7","8","8","8","9","9","9","10","10","10","11","11","11","11","12","12","12","13","13","13","13","13","13","14","13","14","14","14","14","14","14","14","14","14","14","14","14","14","14","14","14","14","14","14","14","14","14","14","14","14","14","13","14","14","14","14","14","13","14","14","14","13","14","14","14","13","14","13","14","13","14","13","14","13","14","13","14","13","14","13","14","13","14","13","13","14","13","14","13","13","14","13","13","14","13","13","13","14","13","13","14","13","14","13","14","13","13","14","13","14","13","13","13","14","13","13","14","13","13","14","13","14","13","13","14","13","14","13","13","14","13","14","13","13","14","13","14","13","13","14","13","14","13","14","13","14","13","14","13","13","14","13","14","13","13","14","13","14","13","13","14","13","14","13","13","14","13","14","13","14","13","14","13","14","13","14","13","14","13","14","14","13","13","14","13","14","13","14","13","14","13","14","13","13","14","13","14","13","13","14","13","14","13","13","14","13","13","14","13","13","13","14","13","13","14","13","13","13","14","13","13","13","13","14","13","13","13","13","14","13","13","13","13","13","13","13","13","13","13","14","13","13","13","13","13","14","13","13","13","14","13","13","13","14","13","13","13","13","14","13","13","14","13","13","13","14","13","13","13","14","13","13","13","14","13","13","14","13","14","14","14","15","15","16","16","16","17","17","18","18","19","19","20","20","21","21","21","22","23","23","23","24","24","25","25","26","26","27","27","27","27","28","29","29","29","30","30","31","31","31","32","32","33","33","34","34","34","35","35","35","36","36","37","37","37","38","38","39","39","39","40","40","41","40","41","41","41","41","41","41","40","40","40","39","40","39","39","39","39","38","38","38","37","37","37","37","36","36","35","35","35","35","34","34","33","33","33","33","32","31","31","31","31","31","30","30","30","29","29","29","29","29","28","28","28","28","27","27","27","27","27","26","26","26","26","26","25","25","25","25","25","25","25","25","25","25","25","24","25","24","24","24","24","24","24","24","24","24","24","24","25","25","25","25","25","25","25","26","26","26","26","26","27","27","27","27","27","27","28","27","28","28","28","29","28","29","29","29","29","29","29","29","29","29","29","29","29","28","29","29","29","28","29","28","28","29","28","29","29","29","29","29","29","29","29","29","30","30","30","30","30","31","31","31","31","31","32","32","32","33","33","33","34","34","35","35","35","36","36","37","37","37","37","38","38","39","39","39","39","40","40","41","41","41","42","42","43","43","43","43","43","43","44","43","44","44","44","43","44","43","44","43","43","43","43","43","43","43","43","43","43","43","43","43","43","43","43","43","44","43","43","43","44","44","44","44","44","44","45","44","44","44","45","45","45","45","45","45","45","45","45","45","45","45","45","44","45","44","44","44","43","43","43","43","43","42","42","42","41","41","41","41","40","40","41","40","40","40","40","40","40","40","40","40","40","40","40","41","41","41","41","41","42","42","42","43","43","44","44","45","45","45","46","46","47","47","48","48","49","49","50","50","51","51","52","53","53","53","54","55","55","56","56","57","57","58","58","59","59","60","60","61","61","62","62","63","63","63","64","64","65","65","66","66","66","67","67","68","68","69","69","69","70","70","70","71","71","71","72","73","72","73","74","73","74","74","75","75","75","76","76","76","77","77","77","78","78","79","79","79","80","80","80","81","81","82","82","83","83","83","84","85","85","85","85","86","86","87","87","88","88","89","89","89","90","91","91","91","91","92","92","93","93","94","94","95","95","95","95","96","96","97","96","97","97","97","97","98","97","98","97","98","97","97","98","97","98","97","98","98","97","98","98","97","98","97","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","99","98","98","99","98","98","99","98","99","98","99","99","99","98","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","100","99","99","100","99","100","100","100","100","100","100","100","100","100","100","100","100","100","100","99","100","100","100","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","98","99","98","99","99","99","99","99","99","99","99","99","99","99","100","99","100","99","100","99","100","100","99","100","99","100","100","99","100","99","99","100","100","99","99","100","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","98","99","98","99","98","98","98","98","98","98","98","97","97","97","96","96","96","96","96","95","95","95","95","94","94","94","93","93","93","93","92","93","92","92","92","92","91","91","90","90","89","88","88","87","86","86","85","84","84","83","83","82","81","81","80","79","79","78","77","77","76","75","75","74","73","72","72","71","70","69","69","68","68","67","67","65","65","64","63","63","62","61","61","60","60","59","58","57","57","56","56","55","55","53","53","53","52","52","52","51","51","51","51","51","51","51","50","50","50","50","50","49","49","49","49","48","48","48","48","47","47","47","47","47","46","47","46","46","46","46","46","45","46","46","45","45","45","45","45","45","45","45","45","45","45","45","45","45","45","45","45","45","44","44","44","43","43","43","43","43","43","43","42","42","42","41","41","41","41","41","41","40","40","40","39","39","39","39","39","38","38","37","37","37","36","36","35","35","35","34","34","33","32","32","31","30","30","29","29","28","27","27","26","25","25","24","24","23","23","23","23","23","22","22","22","22","22","22","22","21","22","21","21","21","20","19","19","18","17","17","17","15","15","15","14","14","14","13","13","12","11","9","9","7","7","4","4","1","1","0","1","1","2","2","3","3","4","4","5","5","6","6","7","7","8","9","10","10","11","11","12","13","13","13","14","15","15","16","17","17","18","18","19","20","21","21","22","23","23","23","24","25","25","26","27","27","28","29","29","30","30","31","32","32","33","33","34","35","35","36","37","37","38","39","39","40","41","41","41","43","43","44","44","45","45","46","46","47","47","48","49","50","50","51","51","52","53","53","54","55","55","56","56","57","57","58","58","59","59","60","61","61","61","62","62","63","63","64","64","65","65","66","66","66","67","67","67","68","68","69","69","69","70","70","71","71","71","71","72","72","73","73","73","74","74","75","75","75","75","76","76","77","77","77","77","77","78","78","79","79","79","79","80","80","80","81","81","81","81","82","82","83","82","83","83","83","83","84","84","84","85","85","85","85","85","86","86","86","86","87","87","87","87","87","88","88","88","89","89","89","89","89","90","90","90","91","90","91","91","91","91","92","91","92","92","92","93","93","93","93","93","93","94","94","94","95","94","95","95","95","95","95","96","96","96","96","96","96","96","96","96","97","96","96","96","96","96","96","96","96","96","96","96","96","96","97","97","97","97","97","97","98","98","98","98","98","98","98","98","99","98","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","98","99","99","98","99","99","99","98","99","99","99","98","99","99","99","99","99","98","99","99","99","99","98","99","99","99","98","98","99","99","98","98","99","98","99","98","98","99","98","98","98","99","98","98","98","98","99","98","98","98","98","98","98","98","98","99","98","99","98","98","98","98","98","98","99","98","99","98","98","99","98","99","98","99","99","98","99","99","98","99","99","98","98","99","98","98","98","99","98","98","98","99","98","99","98","98","98","99","98","99","98","98","99","98","98","99","98","98","98","99","98","99","98","98","98","98","98","98","98","99","98","98","98","99","98","98","99","98","99","99","98","99","98","99","99","99","99","98","99","98","99","98","99","98","98","99","98","98","99","98","98","98","99","98","98","98","99","98","98","98","99","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","99","98","98","98","98","99","98","98","98","98","99","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","99","98","99","98","99","98","98","98","99","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","99","99","98","99","99","99","99","99","99","99","99","99","99","99","99","98","99","99","98","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","98","99","99","99","99","99","98","98","99","98","98","98","98","97","97","97","97","96","96","96","96","95","95","95","95","94","94","94","93","93","93","93","91","91","91","90","89","89","87","87","85","84","83","82","80","78","76","74","72","69","67","65","62","59","57","55","53","51","49","48","47","46","45","44","43","43","43","42","42","42","41","41","41","41","41","41","40","40","39","39","39","39","39","38","39","38","38","37","37","37","37","37","37","36","36","36","36","36","35","35","35","35","35","35","35","35","35","34","34","34","34","33","33","33","33","33","33","33","33","33","33","33","33","33","33","33","33","33","33","33","33","33","33","33","33","33","33","33","33","33","33","33","33","33","33","33","33","33","33","32","32","32","32","32","32","31","31","31","31","31","31","31","31","30","30","30","29","30","29","29","29","29","29","28","29","28","28","28","27","27","27","27","27","26","26","26","25","24","24","23","23","23","22","22","21","21","20","20","20","19","19","19","19","18","18","18","17","17","17","17","17","16","16","16","15","15","15","15","15","15","16","15","15","16","17","17","17","17","18","18","19","19","19","20","20","21","21","21","22","23","23","24","24","25","25","26","27","28","28","29","29","30","31","32","32","33","33","34","35","35","36","36","36","37","37","38","38","38","38","38","38","38","38","38","39","38","38","38","38","38","38","39","38","39","38","39","38","39","38","39","38","39","39","38","38","38","38","38","37","37","37","37","37","36","36","35","35","35","35","34","34","34","34","34","33","33","33","33","34","33","33","33","33","33","33","33","33","33","33","33","33","32","32","31","31","31","29","29","29","28","27","27","26","25","25","24","24","23","23","23","23","23","22","22","22","21","21","21","21","20","19","19","19","18","17","16","15","14","13","11","9","8","5","5","2","2","0","1","1","1","1","2","2","4","4","5","5","6","6","7","7","9","9","9","11","11","12","12","13","14","15","15","16","17","17","19","19","20","21","21","22","23","23","24","25","26","26","27","28","28","29","30","31","31","32","33","34","34","35","36","36","37","38","39","39","40","41","42","43","43","44","45","45","47","47","48","48","49","50","51","51","52","53","54","54","55","56","56","57","58","58","59","59","60","61","61","62","62","63","63","64","64","65","65","66","66","67","67","67","68","69","69","69","70","70","71","71","71","72","72","73","73","73","74","75","75","75","75","75","75","74","75","74","74","74","74","74","74","74","73","73","73","73","73","73","73","73","73","73","73","73","72","73","72","72","72","72","72","72","71","72","72","71","72","72","72","72","73","73","74","74","75","75","75","76","77","77","77","78","79","79","79","80","81","81","81","82","82","83","83","83","84","84","85","85","85","86","86","87","87","88","88","88","89","89","89","89","90","90","91","91","91","91","91","92","92","92","92","93","93","93","93","93","94","94","94","94","95","95","95","95","95","95","96","95","96","96","96","96","96","97","97","97","97","98","97","98","98","98","98","98","98","99","99","99","99","99","100","100","100","100","100","100","100","100","100","101","100","100","101","101","100","100","100","99","100","99","99","99","99","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","99","98","98","99","98","99","98","99","99","98","99","99","99","99","99","99","99","99","99","99","99","98","99","99","98","99","98","99","98","99","99","98","98","99","99","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","99","98","98","98","99","98","99","98","98","99","98","98","98","98","98","98","98","98","98","99","98","98","99","98","99","98","99","99","98","98","99","98","99","98","98","99","98","98","98","98","98","98","99","98","98","99","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","99","99","98","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","100","99","100","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","98","99","98","99","98","99","99","98","98","99","98","98","99","99","98","98","99","98","98","99","98","99","98","99","99","98","99","99","99","98","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","98","99","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","99","98","98","99","99","99","99","99","99","99","99","99","99","99","98","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","100","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","98","99","98","99","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","99","98","98","98","99","99","98","99","99","99","99","99","99","99","99","99","100","99","100","100","100","100","100","100","100","100","100","99","100","99","100","99","99","99","99","99","99","99","99","99","99","98","99","99","99","98","98","99","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","99","98","98","99","98","98","99","98","98","99","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","99","98","98","99","99","98","99","99","98","99","99","99","99","99","98","99","99","99","99","99","99","99","98","99","99","98","99","98","99","99","98","99","99","99","98","99","99","99","98","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","98","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","98","99","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","97","97","96","97","96","95","95","95","94","93","93","91","91","90","89","89","87","87","86","85","85","83","83","81","80","79","78","77","75","73","71","69","67","65","63","61","58","56","54","52","50","48","46","45","44","43","43","42","41","41","41","40","40","40","40","40","39","39","39","39","39","39","39","39","40","39","40","40","40","41","41","41","41","41","42","42","42","43","43","43","43","44","44","45","45","45","45","45","45","45","45","45","45","45","45","45","45","45","45","45","44","45","44","44","43","44","43","43","43","43","42","42","42","41","41","40","40","39","39","38","38","37","37","37","37","37","37","37","36","36","36","36","36","36","35","35","34","33","33","32","31","30","29","28","27","26","25","23","23","21","20","19","18","17","16","16","15","14","13","13","13","12","11","10","9","8","7","7","3","3","0","1","1","1","1","1","1","3","3","3","3","4","4","5","5","6","6","7","7","8","8","9","9","10","11","11","12","12","13","13","14","15","15","16","16","17","17","18","19","19","20","20","21","21","22","23","23","23","24","25","25","26","26","27","27","28","29","29","30","30","31","31","32","32","33","33","34","35","35","35","36","36","37","37","38","39","39","39","40","40","41","41","42","42","43","43","44","44","45","45","45","46","46","47","47","48","48","49","49","49","50","50","51","51","51","52","52","53","53","53","54","55","55","55","56","56","57","57","57","58","58","59","59","59","60","61","61","61","62","62","62","63","63","63","64","64","65","65","65","66","67","67","67","68","68","69","69","69","70","70","70","71","71","71","72","72","73","73","73","73","74","75","75","75","75","75","76","77","77","77","77","77","78","78","78","79","79","79","79","80","80","80","81","81","81","81","82","82","83","82","83","83","83","84","84","84","85","85","85","85","85","85","86","86","86","87","87","87","87","87","87","88","88","88","89","89","89","89","90","89","90","90","90","90","91","91","91","91","91","92","92","92","92","92","93","93","93","93","93","94","94","94","94","94","95","95","95","95","95","95","96","96","96","96","96","96","96","97","97","97","97","98","98","98","98","98","98","98","98","98","98","98","98","97","98","97","97","98","97","97","97","97","97","97","97","97","97","98","97","97","97","97","97","97","97","97","97","97","97","97","97","97","98","97","98","97","98","97","98","97","98","97","98","97","97","98","97","98","97","97","98","97","98","98","98","98","97","98","98","98","98","98","98","98","98","98","98","98","98","97","97","97","96","97","96","96","96","96","96","95","95","95","95","94","94","93","93","93","93","92","91","91","91","91","90","90","89","89","89","89","88","87","87","87","86","85","85","84","83","82","82","81","81","79","79","78","77","77","75","75","75","73","73","72","71","70","69","69","67","66","65","63","63","61","60","59","58","56","55","54","53","52","51","49","49","47","47","46","45","45","44","44","44","43","43","43","43","42","42","42","42","41","41","41","41","41","41","41","40","40","40","40","40","40","40","40","40","40","40","40","40","40","40","40","40","40","40","40","40","40","40","40","40","40","41","41","41","41","41","41","41","41","41","41","42","42","42","42","42","43","43","43","43","43","43","43","43","43","43","43","43","43","43","43","43","43","43","43","43","43","43","43","43","43","43","43","43","43","43","44","43","44","44","45","44","45","45","46","46","47","47","47","48","48","48","49","49","49","50","51","51","51","52","53","53","53","54","55","55","56","56","57","57","58","58","59","59","60","61","61","61","62","62","63","63","64","64","65","65","65","66","66","67","67","67","68","68","69","69","69","70","71","71","71","71","72","72","72","73","73","73","74","74","75","75","75","76","76","76","77","77","77","77","78","78","79","79","79","80","80","81","81","81","82","82","82","83","83","83","84","85","85","85","86","87","87","88","88","89","89","90","90","91","91","92","93","93","94","94","95","95","96","96","97","97","98","98","98","99","99","100","100","100","100","100","100","100","100","99","100","99","99","99","99","98","99","98","99","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","97","98","97","98","97","97","97","97","97","97","97","98","97","98","98","98","98","98","98","99","99","99","99","99","99","99","99","99","99","99","98","98","98","98","98","98","98","98","98","97","98","97","97","97","97","97","96","96","96","96","96","96","96","96","96","96","95","96","95","95","95","95","95","95","95","95","95","95","96","96","96","96","97","97","98","98","98","98","99","99","99","99","99","99","99","99","99","99","100","99","100","99","100","100","100","99","100","100","100","100","100","99","100","99","100","100","99","100","99","100","100","99","100","99","99","99","100","99","99","99","99","99","99","99","98","99","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","97","98","98","98","98","97","98","98","98","98","98","98","98","98","98","98","98","99","98","99","99","99","99","99","99","99","99","99","99","98","99","99","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","97","98","98","97","98","97","97","98","97","97","98","97","97","98","97","97","97","98","97","98","98","97","98","98","98","98","98","98","98","98","98","98","99","98","99","99","99","99","99","99","99","99","99","99","99","99","99","99","99","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","98","97","97","97","97","97","97","96","96","95","95","95","95","94","94","94","93","93","93","93","92","92","92","91","91","90","89","89","89","87","87","86","85","83","82","81","79","77","75","73","71","69","66","64","61","59","57","54","52","50","48","46","45","44","43","42","41","41","40","40","39","39","39","39","39","38","38","38","37","37","37","37","37","37","36","36","36","35","35","35","35","34","34","34","34","33","33","33","33","32","32","32","31","31","31","31","30","31","30","30","29","29","29","29","29","29","28","28","28","28","28","28","28","28","28","28","28","28","29","28","28","29","29","29","29","29","29","29","29","29","29","29","29","29","29","29","29","29","28","28","28","28","27","27","27","27","27","26","26","26","25","25","25","25","25","24","24","24","23","23","23","23","23","23","22","22","21","21","21","21","21","20","21","20","20","19","19","19","19","19","19","18","18","18","17","17","17","17","17","17","16","16","16","15","15","15","15","15","15","14","14","14","13","14","13","13","13","13","13","12","12","12","12","11","11","11","11","11","10","10","10","10","9","9","9","9","9","9","8","8","8","7","7","7","7","7","7","6","6","6","6","5","5","5","5","5","5","6","6","6","6","7","7","7","7","8","8","8","9","9","10","10","11","11","12","12","12","13","13","14","14","15","15","15","16","16","17","17","18","18","18","19","19","20","21","21","21","22","22","22","22","23","23","23","23","23","23","23","23","23","23","23","23","24","24","24","24","24","25","25","25","25","26","25","26","25","25","25","25","25","25","24","24","24","24","24","23","23","23","23","22","22","21","21","20","20","19","19","18","18","18","17","17","17","17","17","16","16","15","16","15","15","15","15","15","15","15","15","15","15","15","15","15","15","15","15","15","15","15","15","15","15","15","15","16","15","15","16","15","15","15","15","15","16","15","15","16","15","15","15","15","16","15","15","16","15","15","16","15","16","15","16","15","15","16","15","15","15","16","15","15","16","15","15","16","15","16","15","16","15","16","16","15","16","15","16","15","16","16","15","16","16","16","15","16","16","15","16","15","16","16","15","16","15","16","15","16","16","15","16","15","16","16","15","16","16","15","16","16","15","16","15","16","16","15","16","16","15","16","16","15","16","16","15","16","16","16","16","15","16","16","15","16","16","16","15","16","16","16","16","15","16","16","16","16","15","16","16","15","16","16","16","15","16","16","15","16","16","15","16","15","16","16","15","16","15","16","16","15","16","15","16","16","15","16","15","16","15","16","15","16","16","15","16","15","16","16","15","16","15","16","15","16","15","15","15","15","15","15","15","15","15","15","14","13","13","13","12","11","11","10","9","8","7","7","5","5"]
        """
        if route == "JTJED":
            """
            #signalkm = ["0","0.48","2.58","3.6","4.68","5.88","6.58","7.18","7.68","7.98","10.98","11.38","12.38","13.58","15.18","15.78","16.98","17.68","18.18","18.88","19.18","22.68","23.08","24.18","24.18","27.98","28.58","29.58","30.18","31.18","31.48","32.18","35.48","36.68","37.98","39.18","39.18","39.78","44.38","45.48","46.48","47.18","47.68","47.98","52.08","52.48","53.48","54.18","54.78","55.18","59.28","59.78","60.78","60.18","64.58","64.98","66.08","66.18","67.48","67.68","68.88","69.28","70.28","71.58","72.38","74.78","75.78","76.18","76.98","77.38","84.88","85.28","86.28","87.02","87.38","87.78","92.98","94.58","95.58","96.33","96.88","97.28","99.88","101.28","102.48","103.32","103.78","104.18","107.58","107.98","109.08","110.98","111.38","112.48","113.33","114.38","115.28","116.7","116.38","116.88","117.08","117.68","118.68","120.08","120.18","120.68","122.58","123.18","124.18","125.58","126.48","126.88","127.88","128.98","129.98","130.18","130.58","132.58","132.98","133.98","137.38","138.68","139.68","140.68","140.78","141.28","143.18","143.88","144.78","146.58","147.08","148.38","149.48","151.58","151.28","152.28","153.18","153.48","153.78","155.68","156.18","157.38","158.18","158.68","159.38","160.78","161.38","162.38","162.68","164.18","165.18","165.98","166.38","166.68","169.28","169.98","170.98","171.88","172.48","173.58","174.38","174.68","175.08","176.58","176.98","177.08","178.08","179.28","179.28"]
            signalkm = ["213.02","213.5","215.6","216.62","217.7","218.9","219.6","220.2","220.7","221","224","224.4","225.4","226.6","228.2","228.8","230","230.7","231.2","231.9","232.2","235.7","236.1","237.2","237.2","241","241.6","242.6","243.2","244.2","244.5","245.2","248.5","249.7","251","252.2","252.2","252.8","257.4","258.5","259.5","260.2","260.7","261","265.1","265.5","266.5","267.2","267.8","268.2","272.3","272.8","273.8","273.2","277.6","278","279.1","279.2","280.5","280.7","281.9","282.3","283.3","284.6","285.4","287.8","288.8","289.2","290","290.4","297.9","298.3","299.3","300.04","300.4","300.8","306","307.6","308.6","309.35","309.9","310.3","312.9","314.3","315.5","316.34","316.8","317.2","320.6","321","322.1","324","324.4","325.5","326.35","327.4","328.3","329.72","329.4","329.9","330.1","330.7","331.7","333.1","333.2","333.7","335.6","336.2","337.2","338.6","339.5","339.9","340.9","342","343","343.2","343.6","345.6","346","347","350.4","351.7","352.7","353.7","353.8","354.3","356.2","356.9","357.8","359.6","360.1","361.4","362.5","364.6","364.3","365.3","366.2","366.5","366.8","368.7","369.2","370.4","371.2","371.7","372.4","373.8","374.4","375.4","375.7","377.2","378.2","379","379.4","379.7","382.3","383","384","384.9","385.5","386.6","387.4","387.7","388.1","389.6","390","390.1","391.1","392.3","392.3","392.8","393.7"]
            #signalname = ["JTJ","STARTER","INNER STARTER I","INNER STARTER II","LSS","GSS","HOME","TPT","STARTER","LSS","IB GWB","IB DISTANT","IB HOME/GD","GSS","GWB","GD","GSS","HOME","KEY","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","KNNT","GWB","DISTANT","HOME","SLY","STARTER","LSS/GD","GSS","GWB","DISTANT","HOME","DST","STARTER","LSS","GWB","DISTANT","HOME","DPI","STARTER","LSS","GWB","DISTANT","HOME","MAP","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","R","GWB","DISTANT","HOME","BDY","STARTER","LSS","GWB","GD","G/IB D","IBH","GWB","DISTANT","HOME","BQI","STARTER","LSS","GWB","DISTANT","HOME","LCR","STARTER","LSS","GWB","DISTANT","HOME","DSPT","STARTER","LSS","GWB","DISTANT","HOME","TNT","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","KPPR","LSS/DISTANT","HOME","MGSJ","STARTER","LSS","GWB","DISTANT","HOME","SA","STARTER","LSS","GWB","GD","LC 115 GSS/IBD","IBS","GWB","GD","LC 115 A GSS/D","HOME","VRPD","STARTER","LSS","GWB","IBD","IBS","GWB","DISTANT","HOME","DC","STARTER","LSS","GWB","GD","LC 116 C GSS","GWB","IBD","IBS/GD","LC 116 E GSS","GWB","DISTANT","HOME","MVPM","STARTER","LSS","GWB","DISTANT","HOME","SGE","STARTER","LSS","GWB","GD","LC 118 GSS","GWB","DISTANT","HOME","ANU","STARTER","LSS","GWB","GD","LC 121 A GSS","GWB","DISTANT","HOME","CV","STARTER","LSS","GWB","GD","LC 121 C GSS/D","HOME","ED","ED"]
            signalname = ["JTJ","STARTER","INNER STARTER I","INNER STARTER II","LSS","GSS","HOME","TPT","STARTER","LSS","IB GWB","IB DISTANT","IB HOME/GD","GSS","GWB","GD","GSS","HOME","KEY","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","KNNT","GWB","DISTANT","HOME","SLY","STARTER","LSS/GD","GSS","GWB","DISTANT","HOME","DST","STARTER","LSS","GWB","DISTANT","HOME","DPI","STARTER","LSS","GWB","DISTANT","HOME","MAP","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","R","GWB","DISTANT","HOME","BDY","STARTER","LSS","GWB","GD","G/IB D","IBH","GWB","DISTANT","HOME","BQI","STARTER","LSS","GWB","DISTANT","HOME","LCR","STARTER","LSS","GWB","DISTANT","HOME","DSPT","STARTER","LSS","GWB","DISTANT","HOME","TNT","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","KPPR","LSS/DISTANT","HOME","MGSJ","STARTER","LSS","GWB","DISTANT","HOME","SA","STARTER","LSS","GWB","GD","LC 115 GSS/IBD","IBS","GWB","GD","LC 115 A GSS/D","HOME","VRPD","STARTER","LSS","GWB","IBD","IBS","GWB","DISTANT","HOME","DC","STARTER","LSS","GWB","GD","LC 116 C GSS","GWB","IBD","IBS/GD","LC 116 E GSS","GWB","DISTANT","HOME","MVPM","STARTER","LSS","GWB","DISTANT","HOME","SGE","STARTER","LSS","GWB","GD","LC 118 GSS","GWB","DISTANT","HOME","ANU","STARTER","LSS","GWB","GD","LC 121 A GSS","GWB","DISTANT","HOME","CV","STARTER","LSS","GWB","GD","LC 121 C GSS/D","HOME","ED","ED","STARTER","LSS"]
            nstn = ["JTJ","TPT","KEY","SLY","DST","DPI","MAP","BDY","BQI","LCR","DSPT","TNT","KPPR","MGSJ","SA","VRPD","DC","MVPM","SGE","ANU","CV","ED"]
            #input data for annotations
            annot = ["JTJ","TPT","KEY","SLY","DST","DPI","MAP","BDY","BQI","LCR","DSPT","TNT","KPPR","MGSJ","SA","VRPD","DC","MVPM","SGE","ANU","CV","ED"]
            #dstn = [0.00,7.18,18.18,30.18,39.18,46.48,54.18,66.18,76.18,96.33,103.32,113.33,116.7,120.08,129.98,140.68,153.18,158.18,165.98,174.38,179.28]
            dstn = [0.00,7.18,18.18,30.18,39.18,47.18,54.18,66.18,76.18,87.02,96.33,103.32,113.33,116.7,120.08,129.98,140.68,153.18,158.18,165.98,174.38,179.28]
            dstnannot = ["0","7.18","18.18","30.18","39.18","47.18","54.18","66.18","76.18","87.02","96.33","103.32","113.33","116.7","120.08","129.98","140.68","153.18","158.18","165.98","174.38","179.28"]
            #dstnannot = ["0","4.91","13.31","21.11","26.11","38.61","49.31","59.21","62.79","65.96","75.97","82.96","92.27","102.44","112.48","125.03","132.02","140.3","148.65","160.53","171.91","179.29"]
            
            signalkm = ["0","0.48","2.58","3.6","4.68","5.88","6.58","7.18","7.68","7.98","10.98","11.38","12.38","13.58","15.18","15.78","16.98","17.68","18.18","18.88","19.18","22.68","23.08","24.18","24.18","27.98","28.58","29.58","30.18","31.18","31.48","32.18","35.48","36.68","37.98","39.18","39.18","39.78","44.38","45.48","46.48","47.18","47.68","47.98","52.08","52.48","53.48","54.18","54.78","55.18","59.28","59.78","60.78","60.88","65.28","65.68","66.78","66.88","68.18","68.38","69.58","69.98","70.98","72.28","73.08","75.48","76.48","76.88","77.68","78.08","85.58","86.38","87.38","88.12","88.48","88.88","94.08","95.68","96.68","97.43","97.98","98.38","100.98","102.38","103.58","104.42","104.88","105.28","108.68","109.08","110.18","112.08","112.48","113.58","114.43","115.48","116.38","117.8","118.12","118.62","118.82","119.62","120.62","122.02","122.12","122.62","124.52","125.32","126.32","127.72","128.62","129.42","130.42","131.52","132.52","132.72","133.12","135.12","135.92","136.92","140.32","141.62","142.62","143.62","143.72","144.22","146.12","147.12","148.12","149.92","150.72","152.02","153.12","155.22","156.02","157.02","157.92","158.22","158.52","160.42","161.42","162.62","163.42","163.92","164.82","166.22","167.12","168.12","168.42","169.92","170.92","171.72","172.12","172.42","175.02","175.72","176.72","177.62","178.22","179.32","180.12","180.42","180.82","182.32","182.72","182.82","183.82","185.02","185.02","185.52","186.42"]
            signalname = ["JTJ","STARTER","INNER STARTER I","INNER STARTER II","LSS","GSS","HOME","TPT","STARTER","LSS","IB GWB","IB DISTANT","IB HOME/GD","GSS","GWB","GD","GSS","HOME","KEY","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","KNNT","GWB","DISTANT","HOME","SLY","STARTER","LSS/GD","GSS","GWB","DISTANT","HOME","DST","STARTER","LSS","GWB","DISTANT","HOME","DPI","STARTER","LSS","GWB","DISTANT","HOME","MAP","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","THONGNUR NBS","GWB","DISTANT","HOME","BDY","STARTER","LSS","GWB","GD","G/IB D","IBH","GWB","DISTANT","HOME","BQI","STARTER","LSS","GWB","DISTANT","HOME","LCR","STARTER","LSS","GWB","DISTANT","HOME","DSPT","STARTER","LSS","GWB","DISTANT","HOME","TNT","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","KPPR","LSS/DISTANT","HOME","MGSJ","STARTER","LSS","GWB","DISTANT","HOME","SA","STARTER","LSS","GWB","GD","LC 115 GSS/IBD","IBS","GWB","GD","LC 115 A GSS/D","HOME","VRPD","STARTER","LSS","GWB","IBD","IBS","GWB","DISTANT","HOME","DC","STARTER","LSS","GWB","GD","LC 116 C GSS","GWB","IBD","IBS/GD","LC 116 E GSS","GWB","DISTANT","HOME","MVPM","STARTER","LSS","GWB","DISTANT","HOME","SGE","STARTER","LSS","GWB","GD","LC 118 GSS","GWB","DISTANT","HOME","ANU","STARTER","LSS","GWB","GD","LC 121 A GSS","GWB","DISTANT","HOME","CV","STARTER","LSS","GWB","GD","LC 121 C GSS/D","HOME","ED","ED","STARTER","LSS"]
            nstn = ["JTJ","TPT","KEY","SLY","DST","DPI","MAP","BDY","BQI","LCR","DSPT","TNT","KPPR","MGSJ","SA","VRPD","DC","MVPM","SGE","ANU","CV","ED"]
            annot = ["JTJ","TPT","KEY","SLY","DST","DPI","MAP","BDY","BQI","LCR","DSPT","TNT","KPPR","MGSJ","SA","VRPD","DC","MVPM","SGE","ANU","CV","ED"]
            dstn = [0,7.18,18.18,30.18,39.18,47.18,54.18,66.88,76.88,88.12,97.43,104.42,114.43,117.8,122.02,132.52,143.62,157.92,163.42,171.72,180.12,185.02]
            dstnannot = ["0","7.18","18.18","30.18","39.18","47.18","54.18","66.88","76.88","88.12","97.43","104.42","114.43","117.8","122.02","132.52","143.62","157.92","163.42","171.72","180.12","185.02"]
            """


            signalkm = ["0","0.48","2.58","3.6","4.68","5.88","6.58","7.18","7.68","7.98","10.98","11.38","12.38","13.58","15.18","15.78","16.98","17.68","18.18","18.88","19.18","22.68","23.08","24.18","24.18","27.98","28.58","29.58","30.18","31.18","31.48","32.18","35.48","36.68","37.98","39.18","39.18","39.78","44.38","45.48","46.48","47.18","47.68","47.98","52.08","52.78","53.78","54.48","55.08","55.48","59.58","60.38","61.38","61.48","65.88","66.68","67.78","67.88","69.18","69.38","70.58","71.28","72.28","73.58","74.38","76.78","77.78","78.18","78.98","79.38","86.88","87.68","88.68","89.42","89.78","90.18","95.38","96.98","97.98","98.73","99.28","99.68","102.48","103.88","105.08","105.92","106.38","106.78","110.18","110.58","111.68","113.58","114.38","115.48","116.33","117.38","118.28","119.7","120.02","120.52","120.92","121.72","122.72","124.12","124.22","124.72","126.62","127.42","128.42","129.82","130.72","131.62","132.62","133.72","134.72","134.92","135.32","137.32","138.12","139.12","142.32","143.62","144.62","145.62","145.72","146.22","148.12","149.12","150.12","151.92","152.92","154.22","155.42","157.42","158.32","159.32","160.12","160.42","160.72","162.62","163.62","164.92","165.72","166.22","167.12","168.52","169.32","170.32","170.62","172.12","173.12","173.92","174.32","174.62","176.92","177.62","178.62","179.42","180.02","181.12","181.92","182.22","182.62","184.12","184.52","184.62","185.62","186.82","186.82","187.32"]
            signalname = ["JTJ","STARTER","INNER STARTER I","INNER STARTER II","LSS","GSS","HOME","TPT","STARTER","LSS","IB GWB","IB DISTANT","IB HOME/GD","GSS","GWB","GD","GSS","HOME","KEY","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","KNNT","GWB","DISTANT","HOME","SLY","STARTER","LSS/GD","GSS","GWB","DISTANT","HOME","DST","STARTER","LSS","GWB","DISTANT","HOME","DPI","STARTER","LSS","GWB","DISTANT","HOME","MAP","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","THONGNUR NBS","GWB","DISTANT","HOME","BDY","STARTER","LSS","GWB","GD","G/IB D","IBH","GWB","DISTANT","HOME","BQI","STARTER","LSS","GWB","DISTANT","HOME","LCR","STARTER","LSS","GWB","DISTANT","HOME","DSPT","STARTER","LSS","GWB","DISTANT","HOME","TNT","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","KPPR","LSS/DISTANT","HOME","MGSJ","STARTER","LSS","GWB","DISTANT","HOME","SA","STARTER","LSS","GWB","GD","LC 115 GSS/IBD","IBS","GWB","GD","LC 115 A GSS/D","HOME","VRPD","STARTER","LSS","GWB","IBD","IBS","GWB","DISTANT","HOME","DC","STARTER","LSS","GWB","GD","LC 116 C GSS","GWB","IBD","IBS/GD","LC 116 E GSS","GWB","DISTANT","HOME","MVPM","STARTER","LSS","GWB","DISTANT","HOME","SGE","STARTER","LSS","GWB","GD","LC 118 GSS","GWB","DISTANT","HOME","ANU","STARTER","LSS","GWB","GD","LC 121 A GSS","GWB","DISTANT","HOME","CV","STARTER","LSS","GWB","GD","LC 121 C GSS/D","HOME","ED","ED","STARTER"]
            nstn = ["JTJ","TPT","KEY","SLY","DST","DPI","MAP","BDY","BQI","LCR","DSPT","TNT","KPPR","MGSJ","SA","VRPD","DC","MVPM","SGE","ANU","CV","ED"]
            annot = ["JTJ","TPT","KEY","SLY","DST","DPI","MAP","BDY","BQI","LCR","DSPT","TNT","KPPR","MGSJ","SA","VRPD","DC","MVPM","SGE","ANU","CV","ED"]
            dstn = [0,7.18,18.18,30.18,39.18,47.18,54.48,67.88,78.18,89.42,98.73,105.92,116.33,119.7,124.12,134.72,145.62,160.12,165.72,173.92,181.92,186.82]
            dstnannot = ["0","7.18","18.18","30.18","39.18","47.18","54.48","67.88","78.18","89.42","98.73","105.92","116.33","119.7","124.12","134.72","145.62","160.12","165.72","173.92","181.92","186.82"]


        if route == "EDJTJ":
            """
            signalkm = ["0","0.11","0.61","1.71","2.21","3.41","4.91","5.41","5.61","5.41","5.81","7.31","8.11","11.11","11.41","12.51","13.31","14.21","13.51","13.51","13.91","15.01","16.11","18.51","18.91","19.91","21.11","21.11","21.81","23.71","24.31","25.51","26.11","26.51","26.81","29.71","30.11","31.11","32.11","33.11","35.51","36.81","38.01","38.61","39.11","39.51","43.11","43.41","44.51","47.01","47.61","48.71","49.31","49.81","50.21","50.91","52.11","52.51","53.61","54.61","56.61","57.11","58.41","59.21","59.81","60.51","60.71","61.11","62.11","62.79","63.51","63.91","64.81","65.96","66.11","66.71","68.61","69.21","70.21","73.21","73.61","74.71","75.97","76.31","76.71","80.21","80.81","81.81","82.96","83.01","83.41","88.61","90.21","91.21","92.27","92.51","92.91","95.91","96.41","97.41","98.21","98.61","99.71","100.21","100.61","101.71","102.44","103.01","103.41","105.31","105.81","106.81","107.11","107.51","108.51","110.11","110.51","111.61","112.48","112.61","112.91","116.51","116.91","117.91","118.91","121.31","122.61","123.61","125.03","125.11","125.41","129.21","130.11","131.11","132.02","132.31","132.71","137.71","138.31","139.31","140.3","140.61","141.11","144.81","145.51","146.51","147.51","148.65","148.81","149.21","152.71","153.11","154.11","155.11","157.71","158.11","159.11","160.11","160.53","161.01","161.41","163.81","164.21","165.21","165.81","169.41","169.81","171.21","171.91","172.41","172.91","174.61","176.11","177.91","179.29","179.91"]
            signalname = ["ED","STARTER","LSS","GWB","DISTANT","HOME","CV","STARTER","LSS/GD","LC 120 B GSS","GD","LC 120 A GSS/GD","LCC 119 GSS","GWB","DISTANT","HOME","ANU","STARTER","LSS","GWB","GD","LC 118 C GSS/GD","LC 118 B GSS","GWB","DISTANT","HOME","SGE","STARTER","LSS","GWB","DISTANT","HOME","MVPM","STARTER","LSS","GWB","GD","LC 117 A GSS/IBD","IBS/GD","LC 117  GSS","GWB","DISTANT","HOME","DC","STARTER","LSS","GWB","IBD","IBS","GWB","DISTANT","HOME","VRPD","STARTER","LSS/GD","LC 115 A GSS","IB GWB","IBD","IBS/GD","LC 115 GSS","GWB","DISTANT","HOME","SA","STARTER","LSS","GWB","DISTANT","HOME","MGSJ","STARTER","LSS/DISTANT","HOME","KPPR","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","TNT","STARTER","LSS","GWB","DISTANT","HOME","DSPT","STARTER","LSS","GWB","DISTANT","HOME","LCR","STARTER","LSS","GWB","GD","GSS      ","GWB    ","GD    ","GSS","GWB","DISTANT","HOME","BQI","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","GD","GSS","GWB","DISTANT","HOME","BDY","STARTER","LSS","GWB","GD","GSS/ IB DIS","IB HOME","GWB","DISTANT","HOME","MAP","STARTER","LSS","GWB","DISTANT","HOME","DPI","STARTER","LSS","GWB","DISTANT","HOME","DST","STARTER","LSS","GWB","GD","GSS/DISTANT","HOME","SLY","STARTER","LSS","GWB","GD","GSS/ IB D","IB HOME","GWB","DISTANT","INNER HOME","ROUTING HOM","KEY","STARTER","LSS","GWB","GD","GSS/ IB D","IB HOME","GWB","DISATANT","HOME","TPT","STARTER","LSS/D","HOME","HOME","HOME","JTJ","STARTER"]
            nstn = ["ED","CV","ANU","SGE","MVPM","DC","VRPD","SA","MGSJ","KPPR","TNT","DSPT","LCR","BQI","BDY","MAP","DPI","DST","SLY","KEY","TPT","JTJ"]
            annot = ["ED","CV","ANU","SGE","MVPM","DC","VRPD","SA","MGSJ","KPPR","TNT","DSPT","LCR","BQI","BDY","MAP","DPI","DST","SLY","KEY","TPT","JTJ"]
            dstn = [0,4.91,13.31,21.11,26.11,38.61,49.31,59.21,62.79,65.96,75.97,82.96,92.27,102.44,112.48,125.03,132.02,140.3,148.65,160.53,171.91,179.29]
            dstnannot = ["0","4.91","13.31","21.11","26.11","38.61","49.31","59.21","62.79","65.96","75.97","82.96","92.27","102.44","112.48","125.03","132.02","140.3","148.65","160.53","171.91","179.29"]
            
            signalkm = ["0","0.21","0.61","1.71","2.21","3.41","4.91","5.41","5.61","5.41","5.81","7.31","8.11","11.11","11.41","12.51","13.31","14.21","13.51","13.51","13.91","15.01","16.11","18.51","18.91","19.91","21.11","21.11","21.81","23.71","24.31","25.51","26.11","26.51","26.81","29.61","30.41","31.41","32.41","33.61","36.11","37.51","38.81","39.41","39.91","40.31","44.21","45.01","46.11","49.01","49.81","50.81","51.41","51.91","52.31","53.01","54.31","55.11","56.21","57.21","59.21","60.01","61.31","62.11","62.71","63.41","63.61","64.11","65.11","65.79","66.51","66.91","67.81","68.96","69.11","69.71","71.61","72.41","73.41","76.51","77.21","78.31","79.57","79.91","80.31","83.81","84.61","85.61","86.76","86.81","87.21","92.51","94.11","95.11","96.17","96.41","96.81","100.01","100.81","101.81","102.61","103.41","104.51","105.01","105.41","106.51","107.24","107.81","108.21","110.11","110.61","111.61","111.91","112.31","113.31","114.91","115.71","116.81","117.68","117.81","118.11","121.71","122.41","123.41","124.41","126.81","128.11","129.11","130.53","130.61","130.91","134.71","135.61","136.61","137.52","137.81","138.21","143.21","144.01","145.01","146","146.31","146.81","150.51","151.21","152.21","153.21","154.35","154.51","154.91","158.41","159.11","160.11","161.11","163.71","164.11","165.11","166.11","166.53","167.01","167.41","169.81","170.21","171.21","171.81","175.41","175.81","177.31","177.91","178.41","178.91","180.61","182.11","184.01","185.3","186.01"]
            signalname = ["ED","STARTER","LSS","GWB","DISTANT","HOME","CV","STARTER","LSS/GD","LC 120 B GSS","GD","LC 120 A GSS/GD","LCC 119 GSS","GWB","DISTANT","HOME","ANU","STARTER","LSS","GWB","GD","LC 118 C GSS/GD","LC 118 B GSS","GWB","DISTANT","HOME","SGE","STARTER","LSS","GWB","DISTANT","HOME","MVPM","STARTER","LSS","GWB","GD","LC 117 A GSS/IBD","IBS/GD","LC 117  GSS","GWB","DISTANT","HOME","DC","STARTER","LSS","GWB","IBD","IBS","GWB","DISTANT","HOME","VRPD","STARTER","LSS/GD","LC 115 A GSS","IB GWB","IBD","IBS/GD","LC 115 GSS","GWB","DISTANT","HOME","SA","STARTER","LSS","GWB","DISTANT","HOME","MGSJ","STARTER","LSS/DISTANT","HOME","KPPR","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","TNT","STARTER","LSS","GWB","DISTANT","HOME","DSPT","STARTER","LSS","GWB","DISTANT","HOME","LCR","STARTER","LSS","GWB","GD","GSS      ","GWB    ","GD    ","GSS","GWB","DISTANT","HOME","BQI","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","GD","GSS","GWB","DISTANT","HOME","BDY","STARTER","LSS","GWB","GD","GSS/ IB DIS","IB HOME","GWB","DISTANT","HOME","MAP","STARTER","LSS","GWB","DISTANT","HOME","DPI","STARTER","LSS","GWB","DISTANT","HOME","DST","STARTER","LSS","GWB","GD","GSS/DISTANT","HOME","SLY","STARTER","LSS","GWB","GD","GSS/ IB D","IB HOME","GWB","DISTANT","INNER HOME","ROUTING HOME","KEY","STARTER","LSS","GWB","GD","GSS/ IB D","IB HOME","GWB","DISATANT","HOME","TPT","STARTER","LSS/D","HOME","HOME","HOME","JTJ","STARTER"]
            nstn = ["ED","CV","ANU","SGE","MVPM","DC","VRPD","SA","MGSJ","KPPR","TNT","DSPT","LCR","BQI","BDY","MAP","DPI","DST","SLY","KEY","TPT","JTJ"]
            annot = ["ED","CV","ANU","SGE","MVPM","DC","VRPD","SA","MGSJ","KPPR","TNT","DSPT","LCR","BQI","BDY","MAP","DPI","DST","SLY","KEY","TPT","JTJ"]
            dstn = [0,4.91,13.31,21.11,26.11,39.41,51.41,62.11,65.79,68.96,79.57,86.76,96.17,107.24,117.68,130.53,137.52,146,154.35,166.53,177.91,185.3]
            dstnannot = ["0","4.91","13.31","21.11","26.11","39.41","51.41","62.11","65.79","68.96","79.57","86.76","96.17","107.24","117.68","130.53","137.52","146","154.35","166.53","177.91","185.3"]

            """

            signalkm = ["0","0.21","0.61","1.71","2.21","3.41","4.91","5.41","5.61","5.81","6.21","7.41","8.21","11.21","11.51","12.61","13.41","13.61","14.01","14.41","14.81","15.91","17.01","18.81","19.21","20.21","21.41","21.51","22.21","23.91","24.71","25.81","26.41","26.81","27.11","29.81","30.81","31.81","32.81","34.01","36.71","37.91","39.11","39.71","40.21","40.61","44.11","44.91","46.11","48.41","49.11","50.11","50.71","51.21","51.61","52.31","53.61","54.41","55.51","56.51","58.51","59.31","60.61","61.41","62.01","62.71","62.91","63.41","64.41","65.09","65.81","66.21","66.91","68.06","68.26","68.86","70.36","71.16","72.16","74.96","75.66","76.76","78.02","78.36","78.76","82.26","83.06","84.06","85.21","85.41","85.81","91.11","92.71","93.71","94.77","95.01","95.41","98.61","99.41","100.41","101.21","102.01","103.11","103.61","104.01","105.11","105.84","106.41","106.81","108.71","109.21","110.21","110.51","111.01","112.01","113.61","114.41","115.51","116.38","116.51","116.81","120.41","121.11","122.11","123.11","125.51","126.81","127.81","129.23","129.33","129.63","133.43","134.33","135.33","136.24","136.54","136.94","141.74","142.54","143.54","144.53","144.84","145.34","148.64","149.34","150.34","151.34","152.48","152.64","153.04","155.84","156.54","157.54","158.54","160.84","161.34","162.34","163.34","163.76","164.24","164.64","166.24","166.64","167.64","168.44","171.54","172.04","173.54","174.14","174.64","175.14","176.84","178.34","180.24","181.53","182.24"]
            signalname = ["ED","STARTER","LSS","GWB","DISTANT","HOME","CV","STARTER","LSS/GD","LC 120 B GSS","GD","LC 120 A GSS/GD","LCC 119 GSS","GWB","DISTANT","HOME","ANU","STARTER","LSS","GWB","GD","LC 118 C GSS/GD","LC 118 B GSS","GWB","DISTANT","HOME","SGE","STARTER","LSS","GWB","DISTANT","HOME","MVPM","STARTER","LSS","GWB","GD","LC 117 A GSS/IBD","IBS/GD","LC 117  GSS","GWB","DISTANT","HOME","DC","STARTER","LSS","GWB","IBD","IBS","GWB","DISTANT","HOME","VRPD","STARTER","LSS/GD","LC 115 A GSS","IB GWB","IBD","IBS/GD","LC 115 GSS","GWB","DISTANT","HOME","SA","STARTER","LSS","GWB","DISTANT","HOME","MGSJ","STARTER","LSS/DISTANT","HOME","KPPR","STARTER","LSS"," GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","TNT","STARTER","LSS","GWB","DISTANT","HOME","DSPT","STARTER","LSS","GWB","DISTANT","HOME","LCR","STARTER","LSS","GWB","GD","GSS      ","GWB    ","GD    ","GSS","GWB","DISTANT","HOME","BQI","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","GD","GSS","GWB","DISTANT","HOME","BDY","STARTER","LSS","GWB","GD","GSS/ IB DIS","IB HOME","GWB","DISTANT","HOME","MAP","STARTER","LSS","GWB","DISTANT","HOME","DPI","STARTER","LSS","GWB","DISTANT","HOME","DST","STARTER","LSS","GWB","GD","GSS/DISTANT","HOME","SLY","STARTER","LSS","GWB","GD","GSS/ IB D","IB HOME","GWB","DISTANT","INNER HOME","ROUTING HOME","KEY","STARTER","LSS","GWB","GD","GSS/ IB D","IB HOME","GWB","DISATANT","HOME","TPT","STARTER","LSS/D","HOME","HOME","HOME","JTJ","STARTER"]
            nstn = ["ED","CV","ANU","SGE","MVPM","DC","VRPD","SA","MGSJ","KPPR","TNT","DSPT","LCR","BQI","BDY","MAP","DPI","DST","SLY","KEY","TPT","JTJ"]
            annot = ["ED","CV","ANU","SGE","MVPM","DC","VRPD","SA","MGSJ","KPPR","TNT","DSPT","LCR","BQI","BDY","MAP","DPI","DST","SLY","KEY","TPT","JTJ"]
            dstn = [0,4.91,13.41,21.41,26.41,39.71,50.71,61.41,65.09,68.06,78.02,85.21,94.77,105.84,116.38,129.23,136.24,144.53,152.48,163.76,174.14,181.53]
            dstnannot = ["0","4.91","13.41","21.41","26.41","39.71","50.71","61.41","65.09","68.06","78.02","85.21","94.77","105.84","116.38","129.23","136.24","144.53","152.48","163.76","174.14","181.53"]

        if route == "SATPJ":
            signalkm = ["0","0","0.09","10.05","11.01","12.02","13.3","13.4","13.8","14.01","23.06","24.02","25.02","25.87","26.4","26.8","32.02","33.5","36.07","37.05","38.05","39.6","39.8","40.01","48.08","49.04","50.04","51.4","51.8","52.01","56.01","57.8","67","67.06","68.06","69.5","69.9","70.02","72.05","78.5","81.06","82.02","83.02","83.08","85","85.2","85.6","87.6","88.9","90.4","90.8","91.8","92","92.4","93.4","94.52","94.6","94.9","95.2","95.6","96.6","98.9","99.4","100.4","101","101.5","102.5","103.48","103.7","104","107.6","107.64","107.8","108.5","109.5","110.41","110.7","111","112.5","112.8","113.9","114.2","114.8","115.2","116.3","116.3","116.7","117.8","118.49","118.7","119.2","120.2","120.6","122.1","122.2","123","123.5","124.3","124.7","125","125.7","126.1","126.5","127.6","128.3","128.7","129.8","130.4","130.9","131.9","133.19","133.5","133.8","134.6","136","136.4","137.4","138.4","139.17","140.37","140.57","141.67","143.57","144.09","144.67","144.87","145.77","146.07","147.17","149.06","150.27","150.77","151.77","153.47","154.82","154.97","155.42","156.02","156.62","157.72","158.72","159.02","159.72","160.72","162.02","162.58","163.68","163.98","164.98","164.98","165.38","167.08"]
            signalname = ["SA","STARTER","LSS","GWB","DISTANT","HOME","MALR","STARTER","LSS/GD","GSS","GWB","DISTANT","HOME","RASP","STARTER","LSS","SWB","PCTM - H","GWB","DISTANT","HOME","KLGN","STARTER","LSS","GWB","DISTANT","HOME","NMKL","STARTER","LSS","SWB","LDVD - H","GWB","DISTANT","HOME","MONR","STARTER","LSS","SWB","VNGL - H","GWB","GD","G/D","HOME","KRR","STARTER","LSS/GD","LC39 GSS/GD","LC40GSS","GWB","GD","LC41/GSS","GWB","DISTANT","HOME","VRQ","STARTER","LSS","GWB","GD","LC43/GSS","GWB","GD","LC44/GSS","GWB","DISTANT","HOME","MYU","STARTER","LSS","SITHALAVAI","SEV","GWB","DISTANT","HOME","MMH","STARTER","LSS","GWB","DISTANT","HOME","LP","STARTER","LSS/GD","GWB","GSS LC49","GD","LC50 GSS","TIC","GWB","GD","GSS LC49","GWB","D/GD","LC53 GSS/GD","LC54 GSS","HOME","KLT","STARTER","LSS/GD","LC55 GSS","GWB","G/D","LC NO 57 GSS","GWB","GD","LC59 GSS","GWB","DISTANT","HOME","PLI","STARTER","LSS/GD","LC63 GSS","GWB","GD","LC64 GSS/D","HOME","PGN","STARTER","LSS/GD","LC67/GSS/D","HOME","EL","STARTER","LSS","GWB","GD","LC73 GSS","JPM","GWB","GD/D","LC75 GSS","HOME","MTNL","STARTER","LSS/GD","LC 78 GSS","GWB","GD","LCNO 80 GSS","GWB","GD","LC 83 GSS/D","HOME","TP","STARTER","LSS/D","PALAKARAI","TPE","HOME","TPJ"]
            nstn = ["SA","MALR","RASP","KLGN","NMKL","MONR","KRR","VRQ","MYU","MMH","LP","KLT","PLI","PGN","EL","MTNL","TP","TPJ"]
            annot = ["SA","MALR","RASP","KLGN","NMKL","MONR","KRR","VRQ","MYU","MMH","LP","KLT","PLI","PGN","EL","MTNL","TP","TPJ"]
            dstn = [0,13.3,25.87,39.6,51.4,69.5,85,94.52,103.48,110.41,114.2,124.3,133.19,139.17,144.09,154.82,162.58,167.08]
            dstnannot = ["0","13.3","25.87","39.6","51.4","69.5","85","94.52","103.48","110.41","114.2","124.3","133.19","139.17","144.09","154.82","162.58","167.08"]



        if route == "TPJSA":
            signalkm = ["0","0.4","0.6","1.9","2.5","3.2","4.34","4.4","4.6","5.7","5.9","6.4","7.5","8","8.6","9.6","10.3","11.05","11.5","11.8","12.6","14","14.6","15.6","15.81","16.7","17.1","18.1","19.48","19.5","19.8","20.3","21.6","22.53","22.8","23","23.8","25.2","25.6","26.6","27.6","28.51","28.8","29","30","30.4","31.4","32.3","32.6","33.6","34","34.4","35.4","36.4","37.4","37.6","37.9","38.3","39","39.6","40","41","42","42.5","43.4","43.9","46.2","46.8","47.3","47.5","49","49.4","50.4","51.19","51.6","51.9","54.06","55.8","56.3","57.3","58.22","58.7","58.9","59.2","59.7","60.8","63.2","63.6","64.6","64.8","65.5","66.5","67.18","67.7","68","69.4","70.8","71.2","71.9","72.2","73.6","75.6","76.7","77.3","78.2","79.15","87.21","89.21","90.17","92.17","92.72","93.19","94.13","103.22","104.42","108.16","108.21","110.18","110.82","111.21","112.14","120.18","121.14","122.18","122.62","123.21","124.14","128.14","128.72","134.15","134.21","135.21","136.35","137.14","137.17","146.19","147.15","148.16","148.21","148.92","150.14","150.17","160.15","160.19","161.2","162.22"]
            signalname = ["TPJ","STARTER","LSS/D","PALAKARAI","TPE","HOME","TP","STARTER","LSS/GD","LC 83 GSS","GWB","GD","LC 82 GSS","GWB","D/GD","LC 78 GSS","HOME","MTNL","STARTER","LSS/GD","LC75 GSS","GWB","GD","LC73,72,71 GSS","JPM HALT","GWB","DISTANT","HOME","EL","STARTER","LSS/GD","LC 67 GSS/D","HOME","PGN","STARTER","LSS/GD","LC64 GSS","GWB","GD","LC 63 GSS/D","HOME","PLI","STARTER","LSS","GWB","GD","LC 59 GSS","GWB","GD","LC57 GSS","GWB","GD","LC 55 GSS/D","HOME","KLT","STARTER","LSS/GD","LC54 GSS/GD","LC53 GSS","GWB","GD","GSS LC52","GWB","GD","GSS/GWB","GD","HOME","LP","STARTER","LSS","GWB","DISTANT","HOME","MMH","STARTER","LSS","SEV HALT","GWB","DISTANT","HOME","MYU","STARTER","LSS","GWB","GD","LC44 GSS","GWB","GD","LC43 GSS","GWB","DISTANT","HOME","VRQ","STARTER","LSS/GD","LC41 GSS","GWB","GD","49 GSS/D","LC40 GSS/GD","LC39 GSS/D","HOME","KRR","STARTER","LSS/GD","GSS","SWB","GWB","DISTANT","HOME","MONR","STARTER","LSS","SWB","LDVD - H","GWB","DISTANT","HOME","NMKL","STARTER","LSS","GWB","DISTANT","HOME","KLGN","STARTER","LSS","SWB","PCTM","GWB","DISTANT","HOME","RASP","STARTER","LSS","GWB","DISTANT","GSS/DISTANT","HOME","MALR","STARTER","LSS","GWB","DISTANT","HOME","SA"]
            nstn = ["TPJ","TP","MTNL","EL","PGN","PLI","KLT","LP","MMH","MYU","VRQ","KRR","MONR","NMKL","KLGN","RASP","MALR","SA"]
            annot = ["TPJ","TP","MTNL","EL","PGN","PLI","KLT","LP","MMH","MYU","VRQ","KRR","MONR","NMKL","KLGN","RASP","MALR","SA"]
            dstn = [0,4.34,11.05,19.48,22.53,28.51,37.4,46.8,51.19,58.22,67.18,76.7,92.72,110.82,122.62,136.35,148.92,162.22]
            dstnannot = ["0","4.34","11.05","19.48","22.53","28.51","37.4","46.8","51.19","58.22","67.18","76.7","92.72","110.82","122.62","136.35","148.92","162.22"]



        if route == "EDTPJ":
            signalkm = ["0.00","0.00","0.70","1.13","1.18","2.05","3.06","4.06","5.06","6.07","7.19","8.15","8.21","9.08","10.09","11.00","11.14","11.22","13.19","14.06","15.06","15.11","15.19","16.20","18.07","19.00","19.11","19.20","20.08","21.08","22.12","22.19","23.07","24.07","24.20","26.01","27.09","27.18","28.17","29.18","30.15","31.05","32.00","32.08","32.14","33.05","33.14","34.15","35.15","36.02","37.03","38.00","38.15","38.24","39.05","39.13","40.12","41.15","42.05","43.06","48.04","48.13","49.14","50.00","50.19","51.01","52.19","53.07","54.00","54.09","55.11","55.17","56.03","57.07","61.16","62.08","63.10","64.80","65.00","65.40","67.40","68.70","70.20","70.60","71.60","71.80","72.20","73.20","74.32","74.40","74.70","75.00","75.40","76.40","78.70","79.20","80.20","80.80","81.30","82.30","83.28","83.50","83.80","87.4","87.44","87.60","88.30","89.30","90.21","90.50","90.80","92.30","92.60","93.70","94.60","94.76","95.00","96.10","96.10","96.50","97.60","98.29","98.50","99.00","100.00","100.40","101.90","102.00","102.80","103.30","104.10","104.50","104.80","105.50","105.90","106.30","107.40","108.10","108.50","109.60","110.20","110.70","111.70","112.99","113.30","113.60","114.40","115.80","116.20","117.20","118.20","118.97","119.40","119.60","120.70","121.50","122.02","122.02","122.60","122.80","123.50","123.80","124.90","125.69","126.90","127.40","128.40","129.10","130.45","130.60","130.90","131.50","132.10","132.60","133.60","133.90","134.30","135.30","136.60","137.16","137.70","138.00","139.00","139.00","139.40","141.10"]
            signalname = ["ED","STARTER","LSS/GD","LC 12D GSS","GWB","GD","LC 3 GSS/GD","LC 4 GSS/GD","LC 5 GSS/GD","LC 6 GSS/GD","LC 8 GSS/GD","LC 9 GSS","GWB","DISTANT","HOME","CVD","STARTER","LSS","GWB","GD","LC13 GSS","GWB","GD","LC14 GSS/D","HOME","PAS","STARTER","LC18 LSS/GD","LC18 GSS/GD","LC19 GSS/GD","LC20 GSS","GWB","GD","LC22 GSS/GD","LC23 GSS/GD","LC24 GSS","GWB","GD","LC25 GSS/GD","LC26 GSS/GD/D","LC27 GSS","HOME","URL","STARTER","LSS","GWB","GD","LC28A GSS","GWB","DISTANT","HOME","KMD","STARTER","LSS","GWB","GD","LC31A GSS ","GWB","GD","LC32 GSS ","GWB","DISTANT","HOME","PGR","STARTER","LSS","GWB","DISTANT","HOME","MPLM","STARTER","LSS/GD","LC34 GSS/GD","LC35 GSS","GWB","DISTANT","HOME","KRR","STARTER","LSS/GD","LC39 GSS/GD","LC40GSS","GWB","GD","LC41/GSS","GWB","DISTANT","HOME","VRQ","STARTER","LSS","GWB","GD","LC43/GSS","GWB","GD","LC44/GSS","GWB","DISTANT","HOME","MYU","STARTER","LSS","SITHALAVAI","SEV","GWB","DISTANT","HOME","MMH","STARTER","LSS","GWB","DISTANT","HOME","LP","STARTER","LSS/GD","GWB","GSS LC49","GD","LC50 GSS","TIC","GWB","GD","GSS LC49","GWB","D/GD","LC53 GSS/GD","LC54 GSS","HOME","KLT","STARTER","LSS/GD","LC55 GSS","GWB","G/D","LC NO 57 GSS","GWB","GD","LC59 GSS","GWB","DISTANT","HOME","PLI","STARTER","LSS/GD","LC63 GSS","GWB","GD","LC64 GSS/D","HOME","PGN","STARTER","LSS/GD","LC67/GSS/D","HOME","EL","ELAMANUR","STARTER","LSS","GWB","GD","LC73 GSS","JPM","GWB","GD/D","LC75 GSS","HOME","MTNL","STARTER","LSS/GD","LC 78 GSS","GWB","GD","LCNO 80 GSS","GWB","GD","LC 83 GSS/D","HOME","TP","STARTER","LSS/D","PALAKARAI","TPE","HOME","TPJ"]
            nstn = ["ED","CVD","PAS","URL","KMD","PGR","MPLM","KRR","VRQ","MYU","MMH","LP","KLT","PLI","PGN","EL","MTNL","TP","TPJ"]
            annot = ["ED","CVD","PAS","URL","KMD","PGR","MPLM","KRR","VRQ","MYU","MMH","LP","KLT","PLI","PGN","EL","MTNL","TP","TPJ"]
            dstn = [0.00,11.00,19.00,32.00,38.00,50.00,54.09,64.80,74.32,83.28,90.21,94.60,104.10,112.99,118.97,122.02,130.45,137.16,141.10]
            dstnannot = ["0.00","11.00","19.00","32.00","38.00","50.00","54.09","64.80","74.32","83.28","90.21","94.60","104.10","112.99","118.97","122.02","130.45","137.16","141.10"]

        if route == "TPJED":
            signalkm = ["0","0.4","0.6","1.9","2.5","3.2","4.34","4.4","4.6","5.7","5.9","6.4","7.5","8","8.6","9.6","10.3","11.05","11.5","11.8","12.6","14","14.6","15.6","15.81","16.7","17.1","18.1","19.48","19.5","19.8","20.3","21.6","22.53","22.8","23","23.8","25.2","25.6","26.6","27.6","28.51","28.8","29","30","30.4","31.4","32.3","32.6","33.6","34","34.4","35.4","36.4","37.4","37.6","37.9","38.3","39","39.6","40","41","42","42.5","43.4","43.9","46.2","46.8","47.3","47.5","49","49.4","50.4","51.19","51.6","51.9","54.06","55.8","56.3","57.3","58.22","58.7","58.9","59.2","59.7","60.8","63.2","63.6","64.6","64.8","65.5","66.5","67.18","67.7","68","69.4","70.8","71.2","71.9","72.2","73.6","75.6","76.7","77.4","78.4","82.5","83.4","84.4","85.4","86.3","86.3","87.3","87.4","89.4","89.5","90.5","91.06","91.5","92.4","96.5","97.4","98.4","99.5","100.3","101.3","101.4","101.5","102.5","103.3","104.3","104.4","105.4","105.5","106.5","107.5","108.4","109.22","109.4","110.4","110.4","110.5","111.5","112.5","114.3","114.4","115.4","116.5","117.4","117.5","117.5","119.3","120.4","120.5","121.5","122.58","123.4","123.4","124.5","124.5","125.4","126.4","127.5","128.4","129.5","130.17","131.3","131.4","131.4","131.5","132.5","133.5","135.4","136.4","137.4","138.4","138.5","139.5","140.5","141.5"]
            signalname = ["TPJ","STARTER","LSS/D","PALAKARAI","TPE","HOME","TP","STARTER","LSS/GD","LC 83 GSS","GWB","GD","LC 82 GSS","GWB","D/GD","LC 78 GSS","HOME","MTNL","STARTER","LSS/GD","LC75 GSS","GWB","GD","LC73,72,71 GSS","JPM HALT","GWB","DISTANT","HOME","EL","STARTER","LSS/GD","LC 67 GSS/D","HOME","PGN","STARTER","LSS/GD","LC64 GSS","GWB","GD","LC 63 GSS/D","HOME","PLI","STARTER","LSS","GWB","GD","LC 59 GSS","GWB","GD","LC57 GSS","GWB","GD","LC 55 GSS/D","HOME","KLT","STARTER","LSS/GD","LC54 GSS/GD","LC53 GSS","GWB","GD","GSS LC52","GWB","GD","GSS/GWB","GD","HOME","LP","STARTER","LSS","GWB","DISTANT","HOME","MMH","STARTER","LSS","SEV HALT","GWB","DISTANT","HOME","MYU","STARTER","LSS","GWB","GD","LC44 GSS","GWB","GD","LC43 GSS","GWB","DISTANT","HOME","VRQ","STARTER","LSS/GD","LC41 GSS","GWB","GD","49 GSS/D","LC40 GSS/GD","LC39 GSS/D","HOME","KRR","STARTER","LSS","GWB","GD","LC35 GSS/GD","LC34 GSS/D","HOME","MPLM","STARTER","LSS","GWB","DISTANT","HOME","PGR","STARTER","LSS","GWB","GD","LC32 GSS","GWB","GD","LC31A GSS","GWB","DISTANT","HOME","KMD","STARTER","LSS","GWB","GD","LC28A GSS","GWB","DISTANT","URL","HOME","STARTER","LSS/GD","LC27 GSS/GD","LC26 GSS/GD","LC25 GSS","GWB","GD","LC24 GSS/GD","LC23 GSS/GD","LC22 GSS","GWB","GD","LC20 GSS/GD","LC19 GSS/GD","LC18 GSS/D","HOME","PAS","STARTER","LSS/GD","LC14 GSS","GWB","GD","LC13 GSS","GWB","GD","HOME","CVD","STARTER","LSS","GWB","GD","LC9 GSS/GD","LC8 GSS/GD","LC6 GSS/GD","LC5 GSS/GD","LC4 GSS/GD","GWB","GD","LC121D GSS/D","HOME","ED"]
            nstn = ["TPJ","TP","MTNL","EL","PGN","PLI","KLT","LP","MMH","MYU","VRQ","KRR","MPLM","PGR","KMD","URL","PAS","CVD","ED"]
            annot = ["TPJ","TP","MTNL","EL","PGN","PLI","KLT","LP","MMH","MYU","VRQ","KRR","MPLM","PGR","KMD","URL","PAS","CVD","ED"]
            dstn = [0,4.34,11.05,19.48,22.53,28.51,37.4,46.8,51.19,58.22,67.18,76.7,86.3,91.06,103.3,109.22,122.58,130.17,141.5]
            dstnannot = ["0","4.34","11.05","19.48","22.53","28.51","37.4","46.8","51.19","58.22","67.18","76.7","86.3","91.06","103.3","109.22","122.58","130.17","141.5"]


        if route == "EDIGU":
            signalkm = ["0","0.08","1.1","3.9","5.1","6.1","6.39","7.1","10.85","10.94","11.92","13.59","14.3","14.5","16.8","16.92","17.92","19.05","19.8","19.9","23.92","25.3","26.3","26.7","26.82","26.9","28.94","30.5","30.8","33.86","33.96","35.02","36.65","36.9","37","40.86","42.1","43.3","43.8","43.98","44.98","45.98","46.84","47.9","48.94","50.16","50.6","51","55.9","56.3","57.3","58.2","58.6","58.9","59.4","59.8","60.9","62.2","63.2","63.8","64.8","65.5","65.9","67.1","67.8","67.8","68.7","68.8","69.8","70.8","72.7","72.8","73.8","73.8","75.9","76.3","77","77.3","79.8","79.9","80.9","82.8"]
            signalname = ["ED","STARTER","LSS","GWB","DISTANT","HOME","TPM","STARTER","GWB","DISTANT","HOME","PY","STARTER","LSS","GWB","DISTANT","HOME","IGR","STARTER","LSS","GWB","DISTANT","HOME","VZ","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","UKL","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","GD","GSS","GWB","DISTANT","HOME","HOME 2","TUP","STARTER","LSS","GWB","DISTANT","HOME","VNJ","STARTER","LSS","GWB","GD","GSS/DISTANT","GSS","GWB","GD","GSS","GWB","DISTANT","HOME","SNO","STARTER","LSS/GD","GSS","GWB","GD","GSS","GWB","GD","G/D","HOME","SUU","STARTER","LSS","GWB","DISTANT","HOME","IGU"]
            nstn = ["ED","TPM","PY","IGR","VZ","UKL","TUP","VNJ","SNO","SUU","IGU"]
            annot = ["ED","TPM","PY","IGR","VZ","UKL","TUP","VNJ","SNO","SUU","IGU"]
            dstn = [0,6.39,13.59,19.05,26.7,36.65,50.16,58.2,67.8,76.3,82.8]
            dstnannot = ["0","6.39","13.59","19.05","26.7","36.65","50.16","58.2","67.8","76.3","82.8"]

        if route == "IGUED":
            signalkm = ["0","0.5","1.3","3.8","4.3","5.3","6.6","6.5","6.9","8.3","9.5","10","11.1","11.5","12","13","14","15.1","15.2","15.6","15.9","16.5","17.6","18.7","19.2","20.2","21.6","22.2","22.8","23.8","24.7","24.9","25.4","30.3","30.7","31.7","32.8","32.9","33.4","34.2","35.2","35.8","36.8","37.8","38.4","39.6","43","43.5","45.4","46.3","46.8","47.2","54","54.3","55.4","55.9","56.6","57","60.6","61.8","62.9","63.9","64.1","64.4","67.1","67.5","68.4","69.8","69.7","70.1","74.1","74.7","75.7","76.9","76.9","77.1","78.6","79.2","80.9","81.3","82.2","82.9"]
            signalname = ["IGU","STARTER","LSS","GWB","DISTANT","HOME","SUU","STARTER","LSS/GD","LC142 GSS","GWB","G D","LC141 GSS","GWB","GD","LC140 GSS/DIST","HOME","SNO","STARTER","LSS","GWB","G D","LC137 GSS","GWB","GD","LC136 GSS/GD","LC 135 GSS","GWB","DISTANT","HOME","VNJ","STARTER","LSS","GWB","DISTANT","HOME","TUP","STARTER","LSS/G D","GSS","GWB","GD","LC 131 E GSS","IBS GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","UKL","STARTER","LSS ","GWB","DISTANT","HOME","VZ","STARTER","LSS ","GWB","DISTANT","HOME","IGR","STARTER","LSS","GWB","DISTANT","HOME","PY","STARTER","LSS","GWB","DISTANT","HOME","TPM","STARTER","LSS","GWB","GD","GSS/DISTANT","HOME","RT HOME","ED"]
            nstn = ["IGU","SUU","SNO","VNJ","TUP","UKL","VZ","IGR","PY","TPM","ED"]
            annot = ["IGU","SUU","SNO","VNJ","TUP","UKL","VZ","IGR","PY","TPM","ED"]
            dstn = [0,6.6,15.1,24.7,32.8,46.3,55.9,63.9,69.8,76.9,82.9]
            dstnannot = ["0","6.6","15.1","24.7","32.8","46.3","55.9","63.9","69.8","76.9","82.9"]

        if route == "KRRDG":
            signalkm = ["0","0.1","1","5.8","6.3","7.3","13","13.5","14.5","15.3","15.7","16","26.8","27.3","28.3","29.1","29.5","29.9","32.6","33","34","42.3","42.7","43.7","44.7","45.3","46.3","49.9","50.5","51.5","52.5","53.3","53.7","54","69.9","70.3","71.3","72.2","73.6","73.6"]
            signalname = ["KRR","Starter","LSS/GD","GSS","LC 04 GWB","GD","GSS","GWB","DISTANT","HOME","VEI","STARTER","LSS","GWB","DISTANT","HOME","PALM","STARTER","LSS","GWB","GD","LC 12 GSS","GWB","GD","LC 18 GSS","GWB","GD","LC 20 GSS","GWB","GD","LC 22 GSS/D","HOME","EDU","STARTER","LSS","GWB","GD/D","GSS","HOME","DG"]
            nstn = ["KRR","VEI","PALM","EDU","DG"]
            annot = ["KRR","VEI","PALM","EDU","DG"]
            dstn = [0,15.7,29.5,53.7,73.6]
            dstnannot = ["0","15.7","29.5","53.7","73.6"]

        if route == "DGKRR":
            signalkm = ["0","1","1.2","17.8","18.4","19.4","20.3","20.5","20.8","21.5","25.3","25.8","26.8","28.1","28.4","29.4","37.7","38.1","39.1","41.9","42.4","43.5","44.5","44.7","45","55.9","56.4","57.4","58.3","58.5","58.8","64.4","64.8","65.8","70.7","71.1","72.1","72.7","73.9"]
            signalname = ["DG","STARTER","LSS","GWB","DISTANT","HOME","EDU","STARTER","LSS/GD","LC 22  GSS","GWB","GD","LC 20 GSS","GWB","GD","LC 18 GSS","GWB","GD","LC 12 GSS","GWB","DISTANT","HOME","PALM","STARTER","LSS","GWB","DISTANT","HOME","VEI","STARTER","LSS","GWB","GD","LC 4 GSS","GWB","GD/D","GSS","HOME","KRR"]
            nstn = ["DG","EDU","PALM","VEI","KRR"]
            annot = ["DG","EDU","PALM","VEI","KRR"]
            dstn = [0,20.3,44.5,58.3,73.9]
            dstnannot = ["0","20.3","44.5","58.3","73.9"]


        if route == "EDPGTA":
            signalkm = ["0","0.08","1.1","3.9","5.1","6.1","6.39","7.1","10.85","10.94","11.92","13.59","14.3","14.5","16.8","16.92","17.92","19.05","19.8","19.9","23.92","25.3","26.3","26.7","26.82","26.9","28.94","30.5","30.8","33.86","33.96","35.02","36.65","36.9","37","40.86","42.1","43.3","43.8","43.98","44.98","45.98","46.84","47.9","48.94","50.16","50.6","51","55.9","56.3","57.3","58.2","58.6","58.9","59.4","59.8","60.9","62.2","63.2","63.8","64.8","65.5","65.9","67.1","67.8","67.8","68.7","68.8","69.8","70.8","72.7","72.8","73.8","73.8","75.9","76.3","77","77.3","79.8","79.9","80.9","82.8","83.1","83.6","88.8","88.9","89.9","90","90.8","91.8","93.2","93.6","93.9","100.5","100.8","102","102.8","103.7","103.9","105.6","105.8","107.7","107.7","114.6","114.8","115.8","116.6","117.1","117.8","122.6","121.8","122.7","122.7","122.9","123.9","125.6","125.9","127.7","128.5","128.8","129.72","130.86","138.6","138.9","139.9","140.9","141.7"]
            signalname = ["ED","STARTER","LSS","GWB","DISTANT","HOME","TPM","STARTER","GWB","DISTANT","HOME","PY","STARTER","LSS","GWB","DISTANT","HOME","IGR","STARTER","LSS","GWB","DISTANT","HOME","VZ","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","UKL","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","GD","GSS","GWB","DISTANT","HOME","HOME 2","TUP","STARTER","LSS","GWB","DISTANT","HOME","VNJ","STARTER","LSS","GWB","GD","GSS/DISTANT","GSS","GWB","GD","GSS","GWB","DISTANT","HOME","SNO","STARTER","LSS/GD","GSS","GWB","GD","GSS","GWB","GD","G/D","HOME","SUU","STARTER","LSS","GWB","DISTANT","HOME","IGU","STARTER","LSS","GWB","GD","GSS","GWB","DISTANT","HOME","PTJ","STARTER","LSS","GWB","DISTANT","HOME","MDKI","STARTER","LSS","GWB","DISTANT","HOME","ETTIMADAI","GWB","DISTANT","HOME","WAL","STARTER","LSS","GWB","DISTANT","HOME","CHULLIMADAI","GD","GSS","GWB","DISTANT","HOME","KJKD","STARTER","LSS/GD","GSS","GWB","DISTANT","INNER HOME","HOME","PGT"]
            nstn = ["ED","TPM","PY","IGR","VZ","UKL","TUP","VNJ","SNO","SUU","IGU","PTJ","MDKI","WAL","KJKD","PGT"]
            annot = ["ED","TPM","PY","IGR","VZ","UKL","TUP","VNJ","SNO","SUU","IGU","PTJ","MDKI","WAL","KJKD","PGT"]
            dstn = [0,6.39,13.59,19.05,26.7,36.65,50.16,58.2,67.8,76.3,82.8,93.2,102.8,116.6,128.5,141.7]
            dstnannot = ["0","6.39","13.59","19.05","26.7","36.65","50.16","58.2","67.8","76.3","82.8","93.2","102.8","116.6","128.5","141.7"] 

        if route == "PGTEDA":
            signalkm = ["0","0.55","0.94","4.2","4.9","5.9","5.9","8.1","9.7","10.68","11.9","12.98","13.56","13.84","17.2","17.98","19.92","19.92","22.1","22.82","23.8","23.98","24.92","25.7","31.1","31.88","32.92","33.98","35.1","36.66","37.64","38.24","38.78","38.98","45.1","45.94","46.86","48.5","48.8","49.1","49.3","49.7","50.7","52.5","52.9","53.9","56.4","56.8","57.8","58.8","59.3","60.1","62.6","63.1","64.1","65.1","65.3","65.7","67.1","68.3","68.8","69.9","70.3","70.8","71.8","72.8","73.9","74","74.4","74.7","75.3","76.4","77.5","78","79","80.4","81","81.6","82.6","83.5","83.7","84.2","89.1","89.5","90.5","91.6","91.7","92.2","93","94","94.6","95.6","96.6","97.2","98.4","101.8","102.3","104.2","105.1","105.6","106","112.8","113.1","114.2","114.7","115.4","115.8","119.4","120.6","121.7","122.7","122.9","123.2","125.9","126.3","127.2","128.1","128.5","128.9","132.9","133.5","134.5","135.7","135.7","135.9","137.4","139.7","140.1","141","141.7"]
            signalname = ["PGT","STARTER","LSS","GWB","DISTANT","HOME","KOTTEKAD","GWB","GD","GSS","HOME","KJKD","STARTER","LSS","GWB","DISTANT","HOME","CHULLIMADA","GWB","DISTANT","HOME","WAL","STARTER","LSS","GWB","DISTANT","HOME","ETTIMADAI","GWB","DISTANT","HOME","MDKI","STARTER","LSS","GWB","DISTANT","HOME","PTJ","STARTER","LSS","GWB","G D","LC 147 GSS","GWB","G D","LC 146 GSS","GWB","DISTANT","HOME","IGU","STARTER","LSS","GWB","DISTANT","HOME","SUU","STARTER","LSS/GD","LC142 GSS","GWB","G D","LC141 GSS","GWB","GD","LC140 GSS/DIST","HOME","SNO","STARTER","LSS","GWB","G D","LC137 GSS","GWB","GD","LC136 GSS/GD","LC 135 GSS","GWB","DISTANT","HOME","VNJ","STARTER","LSS","GWB","DISTANT","HOME","TUP","STARTER","LSS/G D","GSS","GWB","GD","LC 131 E GSS","IBS GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","UKL","STARTER","LSS ","GWB","DISTANT","HOME","VZ","STARTER","LSS ","GWB","DISTANT","HOME","IGR","STARTER","LSS","GWB","DISTANT","HOME","PY","STARTER","LSS","GWB","DISTANT","HOME","TPM","STARTER","LSS","GWB","DISTANT","HOME","RT HOME","ED"]
            nstn = ["PGT","KJKD","WAL","MDKI","PTJ","IGU","SUU","SNO","VNJ","TUP","UKL","VZ","IGR","PY","TPM","ED"]
            annot = ["PGT","KJKD","WAL","MDKI","PTJ","IGU","SUU","SNO","VNJ","TUP","UKL","VZ","IGR","PY","TPM","ED"]
            dstn = [0,12.98,23.98,38.24,48.5,58.8,65.1,73.9,83.5,91.6,105.1,114.7,122.7,128.1,135.7,141.7]
            dstnannot = ["0","12.98","23.98","38.24","48.5","58.8","65.1","73.9","83.5","91.6","105.1","114.7","122.7","128.1","135.7","141.7"]

        if route == "EDPGTB":
            signalkm = ["0","0.08","1.1","3.9","5.1","6.1","6.39","7.1","10.85","10.94","11.92","13.59","14.3","14.5","16.8","16.92","17.92","19.05","19.8","19.9","23.92","25.3","26.3","26.7","26.82","26.9","28.94","30.5","30.8","33.86","33.96","35.02","36.65","36.9","37","40.86","42.1","43.3","43.8","43.98","44.98","45.98","46.84","47.9","48.94","50.16","50.6","51","55.9","56.3","57.3","58.2","58.6","58.9","59.4","59.8","60.9","62.2","63.2","63.8","64.8","65.5","65.9","67.1","67.8","67.8","68.7","68.8","69.8","70.8","72.7","72.8","73.8","73.8","75.9","76.3","77","77.3","79.8","79.9","80.9","82.8","83.1","83.6","88.8","88.9","89.9","90","90.8","91.8","93.2","93.6","93.95","100.83","101.99","102.8","103.72","103.93","105.85","107.73","107.74","114.71","115.89","116.6","117.15","117.75","123","124.1","124.8","125.93","127.09","128.5","128.86","129.71","130.85","133.87","134.91","136.228","138.97","139.95","140.87","141.7"]
            signalname = ["ED","STARTER","LSS","GWB","DISTANT","HOME","TPM","STARTER","GWB","DISTANT","HOME","PY","STARTER","LSS","GWB","DISTANT","HOME","IGR","STARTER","LSS","GWB","DISTANT","HOME","VZ","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","UKL","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","GD","GSS","GWB","DISTANT","HOME","HOME 2","TUP","STARTER","LSS","GWB","DISTANT","HOME","VNJ","STARTER","LSS","GWB","GD","GSS/DISTANT","GSS","GWB","GD","GSS","GWB","DISTANT","HOME","SNO","STARTER","LSS/GD","GSS","GWB","GD","GSS","GWB","GD","G/D","HOME","SUU","STARTER","LSS","GWB","DISTANT","HOME","IGU","STARTER","LSS","GWB","GD","GSS","GWB","DISTANT","HOME","PTJ","STARTER","LSS","DISTANT","HOME","MDKI","STARTER","LSS","DISTANT","HOME","ETTIMADAI","DISTANT","HOME","WAL","STARTER","LSS","GWB","IBD","IBSS","DISTANT","HOME","KJKD","STARTER","LSS/GD","GSS","DISTANT","HOME","KOTTAKADU","DISTANT","HOME","INNER HOME","PGT"]
            nstn = ["ED","TPM","PY","IGR","VZ","UKL","TUP","VNJ","SNO","SUU","IGU","PTJ","MDKI","WAL","KJKD","PGT"]
            annot = ["ED","TPM","PY","IGR","VZ","UKL","TUP","VNJ","SNO","SUU","IGU","PTJ","MDKI","WAL","KJKD","PGT"]
            dstn = [0,6.39,13.59,19.05,26.7,36.65,50.16,58.2,67.8,76.3,82.8,93.2,102.8,116.6,128.5,141.7]
            dstnannot = ["0","6.39","13.59","19.05","26.7","36.65","50.16","58.2","67.8","76.3","82.8","93.2","102.8","116.6","128.5","141.7"]

        if route == "PGTEDB":
            signalkm = ["0","0.55","0.95","4.2","4.91","5.91","5.91","9.1","9.71","10.69","11.91","13.48","13.57","13.87","16.9","17.6","18.7","22.1","22.75","23.73","24.491","24.87","25.65","31.1","31.89","32.93","33.5","35.1","36.67","37.65","37.98","38.79","38.99","45.2","45.95","46.87","48.5","48.8","49.1","49.3","49.7","50.7","56.4","56.8","57.8","58.8","59.3","60.1","62.6","63.1","64.1","65.4","65.6","65.7","67.1","68.3","68.8","69.9","70.3","70.8","71.8","72.8","73.9","74","74.4","74.7","75.3","76.4","77.5","78","79","80.4","81","81.6","82.6","83.5","83.7","84.2","89.1","89.5","90.5","91.6","91.7","92.2","93","94","94.6","95.6","96.6","97.2","98.4","101.8","102.3","104.2","105.1","105.6","106","112.8","113.1","114.2","114.7","115.4","115.8","119.4","120.6","121.7","122.7","122.9","123.2","125.9","126.3","127.2","128.1","128.5","128.9","132.9","133.5","134.5","135.7","135.7","135.9","137.4","139.7","140.1","141","141.7"]
            signalname = ["PGT","STARTER","LSS","GWB","DISTANT","HOME","KOTTAIKADU","GWB","GD","GSS","HOME","KJKD","STARTER","LSS","GWB","IBD","IBSS","GWB","DISTANT","HOME","WAL","STARTER","LSS","GWB","DISTANT","HOME","ETTIMADAI","GWB","DISTANT","HOME","MDKI","STARTER","LSS","GWB","DISTANT","HOME","PTJ","STARTER","LSS","GWB","G D","LC 147 GSS","GWB","DISTANT","HOME","IGU","STARTER","LSS","GWB","DISTANT","HOME","SUU","STARTER","LSS/GD","LC142 GSS","GWB","G D","LC141 GSS","GWB","GD","LC140 GSS/DIST","HOME","SNO","STARTER","LSS","GWB","G D","LC137 GSS","GWB","GD","LC136 GSS/GD","LC 135 GSS","GWB","DISTANT","HOME","VNJ","STARTER","LSS","GWB","DISTANT","HOME","TUP","STARTER","LSS/G D","GSS","GWB","GD","LC 131 E GSS","IBS GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","UKL","STARTER","LSS ","GWB","DISTANT","HOME","VZ","STARTER","LSS ","GWB","DISTANT","HOME","IGR","STARTER","LSS","GWB","DISTANT","HOME","PY","STARTER","LSS","GWB","DISTANT","HOME","TPM","STARTER","LSS","GWB","DISTANT","HOME","RT HOME","ED"]
            nstn = ["PGT","KJKD","WAL","MDKI","PTJ","IGU","SUU","SNO","VNJ","TUP","UKL","VZ","IGR","PY","TPM","ED"]
            annot = ["PGT","KJKD","WAL","MDKI","PTJ","IGU","SUU","SNO","VNJ","TUP","UKL","VZ","IGR","PY","TPM","ED"]
            dstn = [0,13.48,24.491,37.98,48.5,58.8,65.4,73.9,83.5,91.6,105.1,114.7,122.7,128.1,135.7,141.7]
            dstnannot = ["0","13.48","24.491","37.98","48.5","58.8","65.4","73.9","83.5","91.6","105.1","114.7","122.7","128.1","135.7","141.7"]


        """
        if route == "EDJTJ":
            signalkm = ["213.02","213.5","215.6","216.62","217.7","218.9","219.6","220.2","220.7","221","224","224.4","225.4","226.6","228.2","228.8","230","230.7","231.2","231.9","232.2","235.7","236.1","237.2","237.2","241","241.6","242.6","243.2","244.2","244.5","245.2","248.5","249.7","251","252.2","252.2","252.8","257.4","258.5","259.5","260.2","260.7","261","265.1","265.5","266.5","267.2","267.8","268.2","272.3","272.8","273.8","273.2","277.6","278","279.1","279.2","280.5","280.7","281.9","282.3","283.3","284.6","285.4","287.8","288.8","289.2","290","290.4","297.9","298.3","299.3","300.04","300.4","300.8","306","307.6","308.6","309.35","309.9","310.3","312.9","314.3","315.5","316.34","316.8","317.2","320.6","321","322.1","324","324.4","325.5","326.35","327.4","328.3","329.72","329.4","329.9","330.1","330.7","331.7","333.1","333.2","333.7","335.6","336.2","337.2","338.6","339.5","339.9","340.9","342","343","343.2","343.6","345.6","346","347","350.4","351.7","352.7","353.7","353.8","354.3","356.2","356.9","357.8","359.6","360.1","361.4","362.5","364.6","364.3","365.3","366.2","366.5","366.8","368.7","369.2","370.4","371.2","371.7","372.4","373.8","374.4","375.4","375.7","377.2","378.2","379","379.4","379.7","382.3","383","384","384.9","385.5","386.6","387.4","387.7","388.1","389.6","390","390.1","391.1","392.3","392.3","392.8","393.7"]
            signalname = ["JTJ","STARTER","INNER STARTER I","INNER STARTER II","LSS","GSS","HOME","TPT","STARTER","LSS","IB GWB","IB DISTANT","IB HOME/GD","GSS","GWB","GD","GSS","HOME","KEY","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","KNNT","GWB","DISTANT","HOME","SLY","STARTER","LSS/GD","GSS","GWB","DISTANT","HOME","DST","STARTER","LSS","GWB","DISTANT","HOME","DPI","STARTER","LSS","GWB","DISTANT","HOME","MAP","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","R","GWB","DISTANT","HOME","BDY","STARTER","LSS","GWB","GD","G/IB D","IBH","GWB","DISTANT","HOME","BQI","STARTER","LSS","GWB","DISTANT","HOME","LCR","STARTER","LSS","GWB","DISTANT","HOME","DSPT","STARTER","LSS","GWB","DISTANT","HOME","TNT","STARTER","LSS","IB GWB","IB DISTANT","IB HOME","GWB","DISTANT","HOME","KPPR","LSS/DISTANT","HOME","MGSJ","STARTER","LSS","GWB","DISTANT","HOME","SA","STARTER","LSS","GWB","GD","LC 115 GSS/IBD","IBS","GWB","GD","LC 115 A GSS/D","HOME","VRPD","STARTER","LSS","GWB","IBD","IBS","GWB","DISTANT","HOME","DC","STARTER","LSS","GWB","GD","LC 116 C GSS","GWB","IBD","IBS/GD","LC 116 E GSS","GWB","DISTANT","HOME","MVPM","STARTER","LSS","GWB","DISTANT","HOME","SGE","STARTER","LSS","GWB","GD","LC 118 GSS","GWB","DISTANT","HOME","ANU","STARTER","LSS","GWB","GD","LC 121 A GSS","GWB","DISTANT","HOME","CV","STARTER","LSS","GWB","GD","LC 121 C GSS/D","HOME","ED","ED","STARTER","LSS"]
            nstn = ["JTJ","TPT","KEY","SLY","DST","DPI","MAP","BDY","BQI","LCR","DSPT","TNT","KPPR","MGSJ","SA","VRPD","DC","MVPM","SGE","ANU","CV","ED"]
            #input data for annotations
            annot = ["JTJ","TPT","KEY","SLY","DST","DPI","MAP","BDY","BQI","LCR","DSPT","TNT","KPPR","MGSJ","SA","VRPD","DC","MVPM","SGE","ANU","CV","ED"]
        """
        list1 = distancelist
        list3 = speedlist
        sstn = data.get("ststn")
        #print(count)
        #nstn = annot
        #nstn = ["SA","LCR","BDY","KEY","TPT"]
        #dstn = [0,30,80,100,110]
        if nstn[0]  == sstn:
            start = 1
        else:
            start = 0
        count = len(nstn)
        signalkmfinal = list(map(float,signalkm))
        minvalue = signalkmfinal[0]
        for index in range(len(signalkmfinal)):
            signalkmfinal[index] = signalkmfinal[index]-minvalue
            signalkmfinal[index] = round(signalkmfinal[index],2)
        #print(signalkmfinal)
        #print(signalname)
        list4 = list(map(float,list3))
        minvalue = list4[0]

        #print (len(list3))
        for index in range(len(list4)):
            #list4[index] = list4[index]-minvalue
            list4[index] = round(list4[index],2)
        list2 = list(map(float,list1))
        minvalue = list2[0]
        #print (len(list1))
        for index in range(len(list2)):
            list2[index] = list2[index]-minvalue
            list2[index] = round(list2[index],2)
        for index in range(len(nstn)):
            if sstn == nstn[index]:
                for x in range(len(nstn)):
                    nstn[x] = list2[0]+dstn[x]-dstn[index]
        #print(nstn)
        stn = [0]*count
        if start == 0:
            for neg in range(len(nstn)):
                if nstn[neg] > 0:
                    stn[neg] = nstn[neg]
            #print(stn)
            stn = list(dict.fromkeys(stn))
        if start == 1:
            for neg in range(len(nstn)):
                stn[neg] = nstn[neg]
            #print(stn)
            stn = list(stn)
        #print(stn)
        #print(sstn)
        #print(nstn[0])
        spd = [0]*(len(stn))
        #print(round(list2[800],0))
        highlightxtime = [0] * len(stn)
        for x in range(len(stn)):
            for index in range(len(list2)):
                if round(stn[x],1) == round(list2[index],1):
                    spd[x] = list4[index]
                    highlightxtime[x] = index
                    break
        #for sp in range(len(list2):
        print(stn)
        print(spd)
        x=list2
        y=list4
        highlightx = stn
        highlighty = spd
        #print(highlightx)
        if len(stn) != len(annot):
            print(annot)
            for i in range(len(annot)):
                annotlen = len(annot)
                stnlen = len(stn)
                if annotlen-stnlen !=0:
                    del annot[0]
                    #print(annot)
                    #print(stn)
        #highlight = list(highlight)
        #print (highlightx)
        #print(annot)
        
        #graphsizef = (len(list2))/500
        
        graphsizef = 50

        mps = data.get("mps")
        #print(type(mps))
        mpscount = 0
        for i in range(len(y)):
            if y[i] > float(mps):
                mpscount = 1
                mpsvalue = "MPS reached and violated"
        if mpscount == 0:
            for i in range(len(y)):
                if y[i] == float(mps):
                    mpscount = 2
                    mpsvalue = "MPS reached"
        if mpscount == 0:
            mpsvalue = "MPS not reached"

        mpsmaxvalue = str(max(y)) + " KM/hr"


        mpsrange = float(mps) - 5
        mpsstart = []
        mpsend = []
        mpsstartvalue = 0

        for i in range(len(y)):
            if mpsstartvalue == 0:
                if y[i] == mpsrange:
                    mpsstart.append(x[i])
                    #mpsstart.append(y[i])
                    mpsstartvalue = 1
            if mpsstartvalue == 1:
                if y[i] < mpsrange:
                    mpsend.append(x[i])
                    #mpsend.append(y[i])
                    mpsstartvalue = 0
        #print(len(mpsstart))
        #print(len(mpsend))

        #print(mpsstart)
        #print(mpsend)

        mpsdistancetravelled = 0
        for i in range(len(mpsstart)):
            mpsdistancetravelled = mpsdistancetravelled + (mpsend[i] - mpsstart[i])

        mpsviolatedstart = []
        mpsviolatedend = []
        for i in range(len(y)):
            if mpsstartvalue == 0:
                if y[i] > float(mps):
                    mpsviolatedstart.append(x[i])
                    mpsstartvalue = 1
            if mpsstartvalue == 1:
                if y[i] <= float(mps):
                    mpsviolatedend.append(x[i])
                    mpsstartvalue = 0
        #print(mpsviolatedstart)
        #print(mpsviolatedend)
        mpsviolateddistancetravelled = 0
        for i in range(len(mpsviolatedstart)):
            mpsviolateddistancetravelled = mpsviolateddistancetravelled + (mpsviolatedend[i] - mpsviolatedstart[i])
        mpsdistancetravelled = round(mpsdistancetravelled,2)
        mpsviolateddistancetravelled = round(mpsviolateddistancetravelled,2)

        totaldistance = round((x[len(x)-1] - x[0]),2)

        print(str(mpsviolateddistancetravelled) + ' Kms travelled by violating Maximum Permissible Speed and travelled at a speed above ' + str(mps) + ' KM/hr')
        print(str(mpsdistancetravelled) + ' Kms travelled at Maximum Permissible Speed of range ' + str(mpsrange) + 'KM/hr and ' + str(mpsmaxvalue))
        mpsdistance = 'Out of ' + str(totaldistance) + ' KMS - ' + str(round(((mpsdistancetravelled/totaldistance) * 100),2))+ '%  of distance - ' + str(mpsdistancetravelled) + ' Kms travelled at Maximum Permissible Speed of range ' + str(mpsrange) + 'KM/hr and ' + str(mpsmaxvalue)
        mpsviolateddistance ='Out of ' + str(totaldistance) + ' KMS - ' + str(round(((mpsviolateddistancetravelled/totaldistance) * 100),2))+ '%  of distance - ' +  str(mpsviolateddistancetravelled) + ' Kms travelled by violating Maximum Permissible Speed and travelled at a speed above ' + str(mps) + ' KM/hr'



        #no of stopping stations
        zerovaluelist = []
        for i in range((len(y))):
            if y[i] == 0 and i<(len(y)-11):
                zerovaluelist.append(x[i])
        #print(zerovaluelist)
        #zerovalue = zerovaluelist.copy()
        #print((len(zerovaluelist)-1))
        for i in reversed(range(len(zerovaluelist))):
            if i != 0:
                if zerovaluelist[i] - zerovaluelist[i-1] <= 0.1:
                    #print(zerovaluelist[i] - zerovaluelist[i-1])
                    zerovaluelist.pop(i)
        #print(zerovaluelist)
        
        zerocount = 0
        zeroindex = []
        #print(signalkmfinal)
        for a in range(len(zerovaluelist)):
            for i in range(len(signalkmfinal)):
                if zerovaluelist[a]<signalkmfinal[i] and zerocount == 0:
                    zeroindex.append(i)
                    print(signalkmfinal[i])
                    zerocount = 1
            zerocount = 0
        zeroindex = list(dict.fromkeys(zeroindex))


        zeroindexcopy = zeroindex.copy()
        for i in range(len(zeroindex)):
            if signalname[zeroindex[i]] == "STARTER":
                zeroindexcopy[i] = zeroindex[i]-1
            if signalname[zeroindex[i]] == "HOME":
                zeroindexcopy[i] = zeroindex[i] +1 
            #print(signalname[zeroindex[i]])  
        #print(zeroindex)
        #print(zeroindexcopy)
        
        zeroindexfound = [0] * len(zeroindexcopy)


        for a in range(len(zeroindexcopy)):
            for i in range(len(annot)):
                if signalname[zeroindexcopy[a]] == annot[i]:
                    zeroindexfound[a] = 1

        zeroindexvalue = 0

        

        for i in range(len(zeroindexfound)):
            if zeroindexfound[i] == 0:
                #zeroindexvalue = zeroindexcopy[i]
                for p in range(10):
                    for a in range(len(annot)):
                        if (zeroindexfound[i]+p)<(len(signalname)-1):
                            if signalname[zeroindexfound[i]+p] == annot[a] and zeroindexfound[i] == 0:
                                zeroindexfound[i] = 1
                                zeroindexcopy[i] = zeroindexfound[i]+p
                    

        #print(zeroindexfound)
        #print(zeroindexcopy)
        #zerofinalreached = 0
        """
        for a in range(len(signalname)):
            if annot[len(annot)-1] == signalname[a]:
                if (signalkmfinal[a]-(x[len(x)-1]))<0:
                    zerofinalreached = 0
                else:
                    zerofinalreached = 1
        print(zerofinalreached)
        """

        stoppingstations = []
        for i in range(len(zeroindexcopy)):
            if signalname[zeroindexcopy[i]] == signalname[zeroindex[i]]:
                stoppingstations.append(str(signalname[zeroindexcopy[i]]) + " -- station")
                #print(str(signalname[zeroindexcopy[i]]) + " -- station")
            elif signalname[zeroindex[i]] == "GWB":
                if "home" in signalname[zeroindex[i]+2].lower():
                    stoppingstations.append(str(signalname[zeroindex[i]+3]) + " -- " + str(signalname[zeroindex[i]]))    
                else:
                    stoppingstations.append(str(signalname[zeroindex[i]+2]) + " -- " + str(signalname[zeroindex[i]]))
            elif "distant" in signalname[zeroindex[i]].lower():
                stoppingstations.append(str(signalname[zeroindex[i]+2]) + " -- " + str(signalname[zeroindex[i]]))
            else:
                stoppingstations.append(str(signalname[zeroindexcopy[i]]) + " -- " + str(signalname[zeroindex[i]]))
            """
            if i == (len(zeroindexcopy)-1) and zerofinalreached == 1:
                stoppingstations.append(str(annot[len(annot)-1]) + " -- station")
            """
                #print(str(signalname[zeroindexcopy[i]]) + " -- " + str(signalname[zeroindex[i]]))
        #print(stoppingstations)
        if (signalname[zeroindex[len(zeroindexcopy)-1]] != annot[len(annot)-1]):
            stoppingstations.append(str(annot[len(annot)-1]) + " -- station")

        #late signals
        latecount = 0
        latestart = []
        lateend = []
        for i in range(len(x)-250):
            for k in range(250):
                if y[i] >= 45:
                    if y[i+k] <= ((y[i] * 0.6)) and y[i+k] < (33):
                        if ((x[i+k]-x[i]) * 1000) < 1000 and i == latecount:
                            latestart.append(i)
                            lateend.append(i+k)
                            latecount = i+k
            if latecount == i:
                latecount = i+1

                """
                for k in range(100):
                    if y[i+k] <= ((y[i] * 0.6)) and i == latecount:
                        latestart.append(i)
                        lateend.append(i+k)
                        latecount = i+k
                        print(latecount)
                        #print("yes")
                """
           
        print(latestart)
        print(lateend)
        lateendspeed = []
        for i in range(len(lateend)):
            lateendspeed.append(y[lateend[i]])
        print(lateendspeed)
        print("---")

        lateendappend = []
        lateendappenddistance = []
        lateendappendspeed = []
        lateendfinal = []
        lateendminvalue = 0
        lateendminvalueindex = 0


        for i in range(len(lateend)):
            if (lateend[i]+250)< len(x):
                for k in range(250):
                    lateendappend.append(lateend[i]+k)
                    lateendappenddistance.append(x[lateend[i]+k])
                    lateendappendspeed.append(y[lateend[i]+k])
                lateendminvalue = lateendappendspeed[0]
                for a in range(len(lateendappend)):
                    if lateendappendspeed[a] <= lateendminvalue:
                        lateendminvalue = lateendappendspeed[a]
                        lateendminvalueindex = lateendappend[a]
                lateendfinal.append(lateendminvalueindex)
                lateendappend = []
                lateendappenddistance = []
                lateendappendspeed = []
                lateendminvalue = 0
                lateendminvalueindex = 0
        print(lateendfinal)

        lateendspeed = []
        lateenddistance = []
        for i in range(len(lateendfinal)):
            lateendspeed.append(y[lateendfinal[i]])
            lateenddistance.append(x[lateendfinal[i]])
        print(lateenddistance)
        print(lateendspeed)

        lateendfound = [0] * len(lateenddistance)
        lateendfoundindex = [0] * len(lateenddistance)
        latebefaft = [0] * len(lateenddistance)
        for i in range(len(lateenddistance)):
            for k in range(len(signalkmfinal)):
                for a in range(250):
                    if round((lateenddistance[i]+(0.01*a)),2) == signalkmfinal[k] and lateendfound[i] == 0:
                        lateendfound[i] = signalname[k]
                        lateendfoundindex[i] = k
        #print(lateenddistance)
        #print(lateendfound)
        for i in range(len(lateenddistance)):
            for k in reversed(range(len(signalkmfinal))):
                if lateendfound[i] == 0:
                    for p in range(50):
                        #print(round((lateenddistance[i]-(0.01*a)),2))
                        #print(signalkmfinal[k])
                        if round((lateenddistance[i]-(0.01*p)),2) == signalkmfinal[k]:
                            lateendfound[i] = signalname[k]
                            lateendfoundindex[i] = k
                            latebefaft[i] = 1

        print(lateendfound)
        print(lateendfoundindex)
        #print("kkk")
        print(latebefaft)

        lateendfoundindexcopy = lateendfoundindex.copy()
        for i in range(len(lateendfoundindex)):    
            if "STARTER" in signalname[lateendfoundindex[i]]:
                lateendfoundindexcopy[i] = lateendfoundindex[i]-1
            for a in range(3):
                if (lateendfoundindex[i]+a)<(len(signalname)):
                    if "HOME" in signalname[lateendfoundindex[i]+a]:
                        lateendfoundindexcopy[i] = lateendfoundindex[i] +(a+1)
            print(signalname[lateendfoundindexcopy[i]])
        print(lateendfoundindexcopy)



        lateindexfound = [0] * len(lateendfoundindexcopy)


        for a in range(len(lateendfoundindexcopy)):
            for i in range(len(annot)):
                if signalname[lateendfoundindexcopy[a]] == annot[i]:
                    lateindexfound[a] = 1

        lateindexvalue = 0

        lateindexfoundannot = lateindexfound.copy()
        print(lateindexfound)
        for i in range(len(lateendfoundindexcopy)):
            if lateindexfound[i] == 0:
                for a in range(20):
                    for k in range(len(annot)):
                        if (lateendfoundindexcopy[i]+a)<(len(signalname)-1):
                            if (signalname[lateendfoundindexcopy[i]+a]) == annot[k] and lateindexfoundannot[i] == 0:
                                lateendfoundindexcopy[i] = (lateendfoundindexcopy[i]+a)
                                lateindexfoundannot[i] = 1
                                #print(signalname[lateendfoundindexcopy[i]+a])
                                #print("---")
                                #print(annot[k])
        print(lateindexfoundannot)
        print(lateendfoundindexcopy)
        print(lateindexfound)
        
        cautiondetected = []
        for a in range(len(lateendfoundindexcopy)):
            if lateendfoundindexcopy[a] == 0:
                cautiondetected.append(a)
        print(cautiondetected)


        latesignals = []
        for i in range(len(lateendfoundindexcopy)):
            if lateendfoundindexcopy[i] != 0:
                if latebefaft[i] == 0:
                    if signalname[lateendfoundindexcopy[i]] == signalname[lateendfoundindex[i]]:
                        latesignals.append(str(signalname[lateendfoundindexcopy[i]]) + " -- station")
                    elif signalname[lateendfoundindexcopy[i]] != signalname[lateendfoundindex[i]] and lateindexfound[i] == 0:
                        for a in range(len(annot)):
                            if signalname[lateendfoundindexcopy[i]] == annot[a]:
                                lateindexfound[i] = annot[a-1]
                        if signalname[lateendfoundindex[i]] == "GWB":
                            latesignals.append(str(signalname[lateendfoundindex[i]]) + " before " + str(signalname[lateendfoundindex[i]+1]) + " between " +str(lateindexfound[i])+ " -- " + str(signalname[lateendfoundindexcopy[i]]))
                        else:
                            latesignals.append(str(signalname[lateendfoundindex[i]]) + " between " +str(lateindexfound[i])+ " -- " + str(signalname[lateendfoundindexcopy[i]]))
                    elif signalname[lateendfoundindexcopy[i]] != signalname[lateendfoundindex[i]] and lateindexfound[i] !=0:
                        latesignals.append(str(signalname[lateendfoundindexcopy[i]]) + " -- " + str(signalname[lateendfoundindex[i]]))
                if latebefaft[i] == 1:
                    for a in range(len(annot)):
                        if signalname[lateendfoundindexcopy[i]] == annot[a]:
                            lateindexfound[i] = annot[a-1]
                    if "starter" in signalname[lateendfoundindex[i]].lower() or "lss" in signalname[lateendfoundindex[i]].lower():
                        latesignals.append("After passing " + str(signalname[lateendfoundindex[i]]) + " of " + str(signalname[lateendfoundindexcopy[i]]))
                    else:
                        latesignals.append("After passing " + str(signalname[lateendfoundindex[i]]) + " between " +str(lateindexfound[i])+ " -- " + str(signalname[lateendfoundindexcopy[i]]))







        datetimesplit = []
        """
        #print(datelist)
        timestart = []
        timeend = []
        datestart = []
        dateend = []
        for i in range(len(x)):
            if i < (len(x)-1):
                if y[i] == 0:
                    if "time" in timevalue.lower() and "date" in timevalue.lower():
                        timestart.append(timelist[i])
                        timeend.append(timelist[i+1])
                    else:
                        timestart.append(timelist[i])
                        timeend.append(timelist[i+1])
                        datestart.append(datelist[i])
                        dateend.append(datelist[i+1])
        """


        #avg speed with detention
        print(timelist[10])
        print(datelist[10])

        if "time" in timevalue.lower() and "date" in timevalue.lower():
            for i in range(len(timelist)):
                #timesplit1.append(timelist[i].split(" "))
                datetimesplit.append(datetime.strptime(timelist[i], ' %d/%m/%y %H:%M:%S'))
            totalsecondswd = (datetimesplit[len(datetimesplit)-1]-datetimesplit[0]).total_seconds()
            #print(totalsecondswd)
        else:
            for i in range(len(timelist)):
                #timesplit1.append(timelist[i])
                #datesplit1.append(datelist[i])
                if (type(datelist[i]).__name__) == "str":
                    datesplit = datelist[i].split('/')
                    #print(datesplit)
                    #print("jjjjj")
                    if len(datesplit[2]) == 4:
                        datelist[i] = datetime.strptime(datelist[i], '%d/%m/%Y')
                    if len(datesplit[2]) == 2:
                        datelist[i] = datetime.strptime(datelist[i], '%d/%m/%y')
                    datelist[i] = datelist[i].date()
                if (type(timelist[i]).__name__) == "str":
                    timelist[i] = datetime.strptime(timelist[i], '%H:%M:%S')
                    timelist[i] = timelist[i].time()
                datetimesplit.append(datetime.combine(datelist[i],timelist[i]))
            totalsecondswd = (datetimesplit[len(datetimesplit)-1]-datetimesplit[0]).total_seconds()
        #print(totalsecondswd)
        avgspdwd = ((x[len(x)-1])*1000)/(totalsecondswd)
        #print(avgspdwd)
        avgspdwd = round((avgspdwd * (18/5)),2)
        #print(avgspdwd)
        avgspdwd = "Average Speed of the train with detention is " + str(avgspdwd) + " KM/Hr"
        #print(len(timelist))
        #print(len(x))


        datetimesplit1 = []
       

        if "time" in timevalue.lower() and "date" in timevalue.lower():
            for i in range(len(x)):
                datetimesplit1.append(datetime.strptime(timelist[i], ' %d/%m/%y %H:%M:%S'))
        else:
            for i in range(len(x)):
                if (type(datelist[i]).__name__) == "str":
                    datesplit = datelist[i].split('/')
                    if len(datesplit[2]) == 4:
                        datelist[i] = datetime.strptime(datelist[i], '%d/%m/%Y')
                    if len(datesplit[2]) == 2:
                        datelist[i] = datetime.strptime(datelist[i], '%d/%m/%y')
                if (type(timelist[i]).__name__) == "str":
                    timelist[i] = datetime.strptime(timelist[i], '%H:%M:%S')
                    timelist[i] = timelist[i].time()
                datetimesplit1.append(datetime.combine(datelist[i],timelist[i]))

        totalsecondswod = 0
        


        for i in range(len(datetimesplit1)):
            if i <(len(datetimesplit1)-1):
                if y[i] == 0:
                    totalsecondswod = totalsecondswod
                else:
                    totalsecondswod = totalsecondswod + ((datetimesplit1[i+1]-datetimesplit1[i]).total_seconds())
        #print(totalsecondswod)
        #print(totalsecondswd)



        avgspdwod = ((x[len(x)-1])*1000)/(totalsecondswod)
        #print(avgspdwod)
        avgspdwod = round((avgspdwod * (18/5)),2)
        #print(avgspdwod)
        avgspdwod = "Average Speed of the train without detention is " + str(avgspdwod) + " KM/Hr"

        """
        #avg speed without detention
        datetimesplitstart = []
        datetimesplitend = []
        totalsecondsarrwod = []
        if len(timestart) == len(datestart):
            for i in range(len(timestart)):
                if (type(type(datestart[i]).__name__)) == "str":
                    datestart[i] = datetime.strptime(datestart[i], '%d/%m/%y')
                    dateend[i] = datetime.strptime(dateend[i], '%d/%m/%y')
                datetimesplitstart.append(datetime.combine(datestart[i],timestart[i]))
                datetimesplitend.append(datetime.combine(dateend[i],timeend[i]))
        else:
            for i in range(len(timestart)):
                datetimesplitstart.append(datetime.strptime(timestart[i], ' %d/%m/%y %H:%M:%S'))
                datetimesplitend.append(datetime.strptime(timelist[i], ' %d/%m/%y %H:%M:%S'))

        print(datetimesplitstart)
        print(datetimesplitend)

        for i in range(len(datetimesplitstart)):
            totalsecondsarrwod.append((datetimesplitend[i]-datetimesplitstart[i]).total_seconds())
        for i in range(len(totalsecondsarrwod)):
            if totalsecondsarrwod[i]<0:
                totalsecondsarrwod[i] = totalsecondsarrwod[i]*(-1)
        totalsecondswod = sum(totalsecondsarrwod)
        print(totalsecondsarrwod)
        print(totalsecondswod)
        print(totalsecondswd)
        totalsecondswod = totalsecondswd-totalsecondswod

        avgspdwod = (((x[len(x)-1])-(0.01*(len(totalsecondsarrwod)+1)))*1000)/(totalsecondswod)
        print(avgspdwod)
        avgspdwod = round((avgspdwod * (18/5)),2)
        print(avgspdwod)

        """


        
        #print(mpsvalue)
        #print(max(y))

        """
        nxtstnindexvalue = 0
        bftindexstart = 0
        bftindexend = 0
        for i in range(len(x)):
            if x[i] == dstn[1]:
                nxtstnindexvalue = i
                break
        print(nxtstnindexvalue)

        for i in range(nxtstnindexvalue):
            if y[i] == 15:
                bftindexstart = i
                bftvalue= 15
                break
        print(bftindexstart)

        for i in range(bftindexstart, nxtstnindexvalue):
            if (y[i] < y[i+1]) and (y[i] != 14 or y[i] != 15):
                bftindexend = i
                break



        print(bftindexend)
        #print("xxxxxxxxx")

        if bftindexend != 0:
            bftdistance = []
            bftspeed = []
            bftdistancetravelled = 0
            for i in range(bftindexstart,(bftindexend+1)):
                bftdistance.append(x[i])
                bftspeed.append(y[i])
            bftdistancetravelled = (x[bftindexend+1] - x[bftindexstart]) * 1000
            print(bftdistance)
            print(bftspeed)
            print(str(bftdistancetravelled) + "mts")

            finalbftvalue = 0
            reqbftvalue = 0

            finalbftvalue = (bftspeed[len(bftspeed) - 1])
            reqbftvalue = 15 * 0.6
            if finalbftvalue <= reqbftvalue:
                print("bft done correctly")
            else:
                bftdistance = []
                bftspeed = []
                bftdistancetravelled = 0
                print("bft not done correctly")
    
            for i in range(len(bftspeed)):
                if bftspeed[i] >15:
                    print("bft poor")
                    break
        
       

        """

        blksecmaxspeed = []
        blksecminspeed = []
        blksecmaxspeedcount = []
        stnindexvalue = [0] * 6
        for a in range(6):
            for i in range(len(x)):
                if dstn[a] == x[i]:
                    stnindexvalue[a] = i
                    #print(x[i])
                    #print(dstn[a])

        for a in range(len(stnindexvalue)):
            if stnindexvalue[a] == 0:
                for i in range(len(x)):
                    if x[i] == round(dstn[a],1):
                        stnindexvalue[a] = i
                        break
        reqvalue = 0
        for a in range(len(stnindexvalue)):
            if stnindexvalue[a] == 0:
                for i in range(len(x)):
                    for p in range(10):
                        reqvalue = round(((dstn[a]) + (p/100)),2)
                        if x[i] == reqvalue:
                            stnindexvalue[a] = i

        print(stnindexvalue)

        startindex = 0
        endindex = 0
        for a in range(len(stnindexvalue)):
            if startindex == 0:
                if a < (len(stnindexvalue)-1):
                    for i in range(stnindexvalue[a], (stnindexvalue[a+1])):
                        for k in range(100):
                            if y[i] >= 45:
                                if y[i+k] <= ((y[i] * 0.6)):
                                    if ((x[i+k]-x[i]) * 1000) < 700:
                                        startindex = i
                                        endindex = (i+k)
                                        break
        #print(startindex)
        #print(endindex)
        if startindex != 0:
            finalmaxspeed = []
            finalmaxspeedindex = [] 
            for i in range(startindex,(startindex-100),-1):
                finalmaxspeed.append(y[i])
                finalmaxspeedindex.append(i)
            finalmaxvalue = max(finalmaxspeed)

            #print(finalmaxspeed)
            #print(finalmaxspeedindex)
            
            for i in reversed(range((len(finalmaxspeed)-1),-1,-1)):
                if finalmaxvalue == finalmaxspeed[i]:
                    startindex = finalmaxspeedindex[i]

            #print(startindex)
            maxfinalspeed = y[startindex]
            #print(maxfinalspeed)
            for a in range(endindex,(startindex-1),-1):
                if y[a] == maxfinalspeed or y[a] == (maxfinalspeed-1):
                    startindex = a
                    break

            bptdistancevalues = []
            bptspeedvalues = []
            bptdistancevaluestime = []

            for i in range(startindex,endindex+1):
                bptdistancevalues.append(x[i])
                bptspeedvalues.append(y[i])
                bptdistancevaluestime.append(datetimesplit[i].time())
            #print(bptdistancevalues)
            #print(bptspeedvalues)

            bptdistancetravelled = 0
            bptdistancetravelled = round((bptdistancevalues[(len(bptdistancevalues)-1)] - bptdistancevalues[0]) * 1000)

            print(str(bptdistancetravelled) + " mts travelled from speed " +str(y[startindex]) + " to " + str(y[endindex]))
            bptvalue = str(bptdistancetravelled) + " mts travelled from speed " +str(y[startindex]) + " to " + str(y[endindex])
        if startindex == 0:
            print("Brake power test not done properly")
            bptvalue = "Brake power test not done properly"
            bptdistancevalues = []
            bptspeedvalues = []




        bftstartindex = 0
        bftendindex = 0
        for i in range(stnindexvalue[0],(stnindexvalue[0]+500)):
            if y[i] >= 10 and y[i] <= 16:
                if bftstartindex == 0:
                    for k in range(50):
                        if y[i+k] <= ((y[i] * 0.65)):
                            bftstartindex = i
                            bftendindex = i+k
                            break
        print(bftstartindex)
        print(bftendindex)
        bftvalue = ""

        if bftstartindex !=0: 
            bftfinalmaxspeed = []
            bftfinalmaxspeedindex = []
            for i in range(bftstartindex, (bftstartindex-10),-1):
                bftfinalmaxspeed.append(y[i])
                bftfinalmaxspeedindex.append(i)
            bftfinalmaxvalue = max(bftfinalmaxspeed)

            for i in reversed(range((len(bftfinalmaxspeed)-1),-1,-1)):
                if bftfinalmaxvalue == bftfinalmaxspeed[i]:
                    bftstartindex = bftfinalmaxspeedindex[i]
            print(bftstartindex)
            bftmaxfinalspeed = y[bftstartindex]
            print(bftmaxfinalspeed)
            for a in range(bftendindex,(bftstartindex-1),-1):
                if y[a] == bftmaxfinalspeed:
                    startindex = a
                    break
            bftdistancevalues = []
            bftspeedvalues = []
            bftdistancevaluestime= []
            for i in range(bftstartindex, bftendindex+1):
                bftdistancevalues.append(x[i])
                bftspeedvalues.append(y[i])
                bftdistancevaluestime.append(datetimesplit[i].time())
            #print(bftdistancevalues)
            #print(bftspeedvalues)

            bftdistancetravelled = 0
            bftdistancetravelled = round((bftdistancevalues[(len(bftdistancevalues)-1)] - bftdistancevalues[0]) * 1000)
            print(str(bftdistancetravelled) + " mts travelled from speed " +str(y[bftstartindex]) + " to " + str(y[bftendindex])+ str(bftstartindex) +"dsjflk"+ str(bftendindex))
            bftvalue = str(bftdistancetravelled) + " mts travelled from speed " +str(y[bftstartindex]) + " to " + str(y[bftendindex])
        if bftstartindex == 0:
            print("Brake feel test not done properly")
            bftvalue = "Brake feel test not done properly"
            bftdistancevalues = []
            bftspeedvalues = []

        """


        for i in range(len(stnindexvalue)):
            if i < (len(stnindexvalue)-1):
                blksecmaxspeedvalues = []
                for a in range(stnindexvalue[i],stnindexvalue[i+1]):
                    blksecmaxspeedvalues.append(y[a])
                #print(blksecmaxspeedvalues.count(blksecmaxspeed[i]))
                #blksecmaxspeed.append(max(blksecmaxspeedvalues))
                blksecminspeed.append(min(blksecmaxspeedvalues))
                #blksecmaxspeedcount.append(blksecmaxspeedvalues.count(blksecmaxspeed[i]))
                   


        #print(blksecmaxspeed)
        #print(blksecmaxspeedcount)
        #print(blksecmaxspeedreduced)

        #print(blksecmaxspeed)
        #print(blksecmaxspeedreduced)
        blksecminspeedindex = [1] * (len(blksecminspeed))
        for i in range(len(stnindexvalue)):
            if i <((len(stnindexvalue)) - 1):
                for a in range(stnindexvalue[i], stnindexvalue[i+1]):
                    if y[a] == blksecminspeed[i]:
                        blksecminspeedindex[i] = a
        print(blksecminspeedindex)

        for i in range(len(stnindexvalue)):
            if i < (len(stnindexvalue)-1):
                blksecmaxspeedvalues = []
                for a in range(stnindexvalue[i], blksecminspeedindex[i]):
                    blksecmaxspeedvalues.append(y[a])
                print(blksecmaxspeedvalues)
                blksecmaxspeed.append(max(blksecmaxspeedvalues, default=0))

        blksecmaxspeedreduceddup = []

        blksecmaxspeedreduced = []
        for i in range(len(blksecmaxspeed)):
            blksecmaxspeedreduced.append((round((blksecmaxspeed[i] * 0.6)-2)))

        blksecmaxspeedindex = [1] * len(blksecmaxspeed)


        blksecmaxspeedreduceddup = blksecmaxspeedreduced.copy()
        
        
        blksecmaxspeeddup = []
        blksecmaxspeeddup = blksecmaxspeed.copy()

        for i in range(len(stnindexvalue)):
            if i < (len(stnindexvalue)-1):
                for a in range(stnindexvalue[i], blksecminspeedindex[i]):
                    if y[a] == blksecmaxspeeddup[i]:
                        blksecmaxspeedindex[i] = a
                        blksecmaxspeeddup[i] = "found"
        for i in range(len(blksecmaxspeedindex)):
            if blksecmaxspeedindex[i] == 1:
                blksecminspeedindex[i] = 1
                blksecmaxspeed[i] = "ignore"
                blksecmaxspeedreduced[i] = "ignore"



        blksecmaxspeedreduceddup = blksecmaxspeedreduced.copy()
        blksecmaxspeedreducedindex = [1] * len(blksecmaxspeedreduced)
        #print("0000000")
        #print(blksecmaxspeed)
        #print(blksecmaxspeeddup)
        #print(blksecmaxspeedreduced)
        #print("0000000")
  

        for i in range(len(blksecmaxspeedindex)):
            for a in range(blksecmaxspeedindex[i], blksecminspeedindex[i]):
                if y[a] == blksecmaxspeedreduceddup[i]:
                    blksecmaxspeedreducedindex[i] = a
                    blksecmaxspeedreduceddup[i] = "found"

    
        #blksecfinalspeed = []
        #blksecfinalspeedreduced = []
        for i in range(len(blksecmaxspeeddup)):
            if blksecmaxspeeddup[i] != "found" or blksecmaxspeedreduceddup[i] != "found":
                #blksecmaxspeed[i] = "ignore"
                #blksecfinalspeed.append(blksecmaxspeed[i])
                #blksecfinalspeedreduced.append(blksecmaxspeedreduced[i])
                #print("found") 
                blksecmaxspeed[i] = "ignore"
                blksecmaxspeedreduced[i] = "ignore"
                blksecminspeed[i] = "ignore"
        #print(blksecmaxspeed)
        #print(blksecmaxspeedreduced)
        print(blksecmaxspeed)
        print(blksecmaxspeedindex)
        print("xxxxxxx")
        print(blksecmaxspeedreduced)
        print(blksecmaxspeedreducedindex)
        print("Xxxxxxxxx")
        print(blksecminspeed)
        print(blksecminspeedindex)
        print("Xxxxxxxxx")
        print(blksecmaxspeeddup)
        print(blksecmaxspeedreduceddup)


        for i in range(len(blksecmaxspeeddup)):
            if blksecmaxspeed[i] != "ignore" or blksecmaxspeedreduced[i] != "ignore":
                if blksecmaxspeed[i] < 40:
                    blksecmaxspeed[i] = "ignore"
                    blksecmaxspeedindex[i] = 1
                    blksecmaxspeedreduced[i] = "ignore"
                    blksecmaxspeedreducedindex[i] = 1
                    blksecminspeed[i] = "ignore"
                    blksecminspeedindex[i] = 1


        print(blksecmaxspeed)
        print(blksecmaxspeedindex)
        print("xxxxxxx")
        print(blksecmaxspeedreduced)
        print(blksecmaxspeedreducedindex)
        print("Xxxxxxxxx")
        print(blksecminspeed)
        print(blksecminspeedindex)
        print("Xxxxxxxxx")
        print(blksecmaxspeeddup)
        print(blksecmaxspeedreduceddup)


        blksecmaxfirstindex = [1] * len(stnindexvalue)
        blksecreducedindex = [1] * len(stnindexvalue)
        #print(stnindexvalue)
        blksecmaxspeeddup = blksecmaxspeed.copy()
        blksecmaxspeedreduceddup = blksecmaxspeedreduced.copy()

        blksecindexstart = []
        blksecreducedindex = []

        for i in range(len(blksecmaxspeedindex)):
            if blksecmaxspeedindex[i] != 1:
                blksecindexstart.append(blksecmaxspeedindex[i])
                blksecreducedindex.append(blksecmaxspeedreducedindex[i])



        #print(blksecindexstart)
        
        #blksecindexstart = list(set(blksecindexstart))
        #blksecreducedindex = list(set(blksecreducedindex))


        print(blksecindexstart)
        print(blksecreducedindex)

        print("sdlijf")

        bptdistancevalues = []
        bptspeedvalues = []


        for i in range(len(blksecindexstart)):
            if blksecreducedindex[i] == 1:
                blksecindexstart[i] = 1
                blksecminspeedindex[i] = 1
        
        print(blksecindexstart)
        print(blksecreducedindex)
        print(blksecminspeedindex)

        bptindexvalue = 0
        for i in range(len(blksecindexstart)):
            if (blksecreducedindex[i] - blksecindexstart[i] >=1):
                bptindexvalue = i
                break

        finalindexvalue = blksecindexstart[bptindexvalue]
        finalindex = []
        finalindexspeed =  []
        finalindexcount = 0
        for i in range((blksecreducedindex[bptindexvalue]),(blksecindexstart[bptindexvalue]-1), -1):
            if y[i]-y[i-1] > 0:
                #finalindexvalue = i
                finalindexcount = finalindexcount + 1
                finalindex.append(i)
                finalindexspeed.append(y[i])
                if finalindexcount > 10:
                    finalindexspeedmax = 0
                    finalindexspeedmax = max(finalindexspeed[a])
                    for a in reversed(range((len(finalindex)-1),-1,-1)):
                        if finalindexspeed[a] == finalindexspeedmax:
                            finalindexvalue = finalindex[a]
                            break
        blksecindexstart[bptindexvalue] = finalindexvalue


        print(blksecindexstart)
        print(blksecreducedindex)

        for i in range(len(blksecindexstart)):
            if blksecreducedindex[i] - blksecindexstart[i] > 150:
                blksecindexstart[i] = 1
                blksecreducedindex[i] = 1

        for i in range(len(blksecindexstart)):
            if blksecindexstart[i] != 1 or blksecreducedindex[i] != 1:
                bptindexvalue = i
                break
        #maxfinalindex = 0
        maxfinalspeed = 0
        for i in range(len(blksecindexstart)):
            #maxfinalindex = blksecreducedindex[i]
            maxfinalspeed = y[blksecindexstart[i]]
            for a in range(blksecreducedindex[i],(blksecindexstart[i]-1),-1):
                if y[a] == maxfinalspeed or y[a] == (maxfinalspeed-1):
                    blksecindexstart[i] = a
                    break
        print(blksecindexstart)
        print(blksecreducedindex)

        for i in range(blksecindexstart[bptindexvalue], (blksecreducedindex[bptindexvalue] +1)):
            bptdistancevalues.append(x[i])
            bptspeedvalues.append(y[i])


        bptdistancetravelled = 0
        bptdistancetravelled = round((bptdistancevalues[(len(bptdistancevalues)-1)] - bptdistancevalues[0]) * 1000)
        
        print(bptdistancevalues)
        print(bptspeedvalues)
        print(str(bptdistancetravelled) + "mts travelled")
        
        """








        cautionorderlength = int(data.get("cautionorderlength"))
        cautionorders  =  data.get("Cautionorders")
        print(cautionorders)
        req = json.loads(cautionorders)
        #print(type(cautionorders))
        #print(req[0]['startingkm'])
        print(cautionorderlength)
        skm = []
        ekm = []
        speed = []
        cautioncheckboxvalue = []
        for i in range(cautionorderlength):
            #print(req[i]['Speed'])
            skm.append(req[i]['startingkm'])
            ekm.append(req[i]['endingkm'])
            speed.append(req[i]['Speed'])
            if route == "TPJSA" or route == "SATPJ" or route == "CBEJTJ" or route == "JTJCBE" or route == "CBESA" or route == "SACBE":
                cautioncheckboxvalue.append(req[i]['cautioncheckboxvalue'])
                #route = "JTJED"

        #print(skm)
        #print(ekm)
        #print(speed)
        #print(cautioncheckboxvalue)
        #print("----------------")
        xvalue = []
        for i in range(cautionorderlength):
            xvalue.append(skm[i].split("/"))
        #print(xvalue)
        #print(len(xvalue))
        yvalue = []
        for i in range(cautionorderlength):
            yvalue.append(ekm[i].split("/"))
        sa = []
        sb = []
        for i in range(len(xvalue)):
            sa.append(xvalue[i][0])
            sb.append(xvalue[i][1])
            #print(xvalue[i][0])
            #print(xvalue[i][1])
        ea = []
        eb = []
        for i in range(len(yvalue)):
            ea.append(yvalue[i][0])
            eb.append(yvalue[i][1])
        #print(sa)
        #print(sb)
        #print(ea)
        #print(eb)
        #print(speed)
        cautionmin = 0
        cautiondata = 0
        if route == "JTJED":
            cautiondata = 1
            cautionmin = "JTJED"
        if route == "EDJTJ":
            cautiondata = 1
            cautionmin = "EDJTJ"
        if route == "TPTED":
            cautiondata = 1
            cautionmin = "TPTED"
        if route == "EDTPT":
            cautiondata = 1
            cautionmin = "EDTPT"
        if route == "EDTPJ":
            cautiondata = 1
            cautionmin = "EDTPJ"
        if route == "TPJED":
            cautiondata = 1
            cautionmin = "TPJED"
        if route == "SATPJ":
            cautiondata = 1
            cautionmin = "SATPJ"
        if route == "TPJSA":
            cautiondata = 1
            cautionmin = "TPJSA"
        if route == "CBEJTJ":
            cautiondata = 1
            cautionmin = "CBEJTJ"
        if route == "JTJCBE":
            cautiondata = 1
            cautionmin = "JTJCBE"
        if route == "CBESA":
            cautiondata = 1
            cautionmin = "CBESA"
        if route == "SACBE":
            cautiondata = 1
            cautionmin = "SACBE"
        if route == "EDIGU":
            cautiondata = 1
            cautionmin = "EDIGU"
        if route == "IGUED":
            cautiondata = 1
            cautionmin = "IGUED"
        if route == "KRRDG":
            cautiondata = 1
            cautionmin = "KRRDG"
        if route == "DGKRR":
            cautiondata = 1
            cautionmin = "DGKRR"
        if route == "EDPGTA":
            cautiondata = 1
            cautionmin = "EDPGTA"
        if route == "PGTEDA":
            cautiondata = 1
            cautionmin = "PGTEDA"

        if route == "EDPGTB":
            cautiondata = 1
            cautionmin = "EDPGTB"
        if route == "PGTEDB":
            cautiondata = 1
            cautionmin = "PGTEDB"

        start = []
        end = []
        for i in range(len(sa)):
            if len(sb[i])>=3:
                sb[i] = int(sb[i])
                sa[i] = int(sa[i])
                sb[i] = sb[i]/1000
                #print(sb[i])
                sa[i] = sa[i] + sb[i]
                sa[i] = round(sa[i],2)
                sa[i] = str(sa[i])
                start.append(sa[i])
                #print("yes")
            else:
                sb[i] = int(sb[i])
                sa[i] = int(sa[i])
                #print(sa[i])
                #print(sb[i])
                if route == "JTJED" or route == "EDJTJ" or route == "TPTED" or route == "EDTPT" or route == "CBEJTJ" or route == "JTJCBE" or route == "CBESA" or route == "SACBE" or route == "EDTPJ" or route == "TPJED" or route == "EDIGU" or route == "IGUED" or route == "KRRDG" or route == "DGKRR" or route == "EDPGTA" or route == "PGTEDA" or route == "EDPGTB" or route == "PGTEDB":
                    sb[i] = ((sb[i]/2)*72)/1000
                if route == "SATPJ" or route == "TPJSA":
                    sb[i] = ((sb[i])*72)/1000
                #print(type(sb[i]))
                #print(sb[i])
                sa[i] = sa[i] + sb[i]
                sa[i] = round(sa[i],2)
                #print(sa[i])
                sa[i] = str(sa[i])
                start.append(sa[i])
        for i in range(len(ea)):
            if len(eb[i])>=3:
                eb[i] = int(eb[i])
                ea[i] = int(ea[i])
                eb[i] = eb[i]/1000
                #print(eb[i])
                if route == "JTJED" or route == "TPTED" or route == "EDTPJ" or route == "SATPJ" or route == "JTJCBE" or route == "SACBE" or route == "EDIGU" or route == "KRRDG" or route == "EDPGTA" or route == "EDPGTB":
                    ea[i] = ea[i] + eb[i] +0.7
                if route == "EDJTJ" or route == "EDTPT" or route == "TPJED" or route == "TPJSA" or route == "CBEJTJ" or route == "CBESA" or route == "IGUED" or route == "DGKRR" or route == "PGTEDA" or route == "PGTEDB":
                    ea[i] = ea[i] + eb[i] - 0.7
                ea[i] = round(ea[i],2)
                ea[i] = str(ea[i])
                end.append(ea[i])
                #print("yes")
            else:
                eb[i] = int(eb[i])
                ea[i] = int(ea[i])
                #print(ea[i])
                #print(eb[i])
                if route == "JTJED" or route == "EDJTJ" or route == "TPTED" or route == "EDTPT" or route == "CBEJTJ" or route == "JTJCBE" or route == "CBESA" or route == "SACBE" or route == "EDTPJ" or route == "TPJED" or route == "EDIGU" or route == "IGUED" or route == "KRRDG" or route == "DGKRR" or route == "EDPGTA" or route == "PGTEDA" or route == "EDPGTB" or route == "PGTEDB":
                    eb[i] = ((eb[i]/2)*72)/1000
                if route == "SATPJ" or route == "TPJSA":
                    eb[i] = ((eb[i])*72)/1000
                #print(type(eb[i]))
                #print(eb[i])
                if route == "JTJED" or route == "TPTED" or route == "EDTPJ" or route == "SATPJ" or route == "JTJCBE" or route == "SACBE" or route == "EDIGU" or route == "KRRDG" or route == "EDPGTA" or route == "EDPGTB":
                    ea[i] = ea[i] + eb[i] +0.7
                if route == "EDJTJ" or route == "EDTPT" or route == "TPJED" or route == "TPJSA" or route == "CBEJTJ" or route == "CBESA" or  route == "IGUED" or route == "DGKRR" or route == "PGTEDA" or route == "PGTEDB":
                    ea[i] = ea[i] + eb[i] - 0.7
                ea[i] = round(ea[i],2)
                #print(ea[i])
                ea[i] = str(ea[i])
                end.append(ea[i])
        print(start)
        print(end)
        print(speed)
        
    
        found = [0] * len(start)
        founde = [0] * len(end)
        cautionindexstart = [1] * len(start)
        cautionindexend = [1] * len(end)
        if cautiondata == 1:
            if cautionmin == "JTJED":
                cautionminvalue = 213.02
            if cautionmin == "EDJTJ":
                cautionminvalue = 392.31
            if cautionmin == "TPTED":
                cautionminvalue = 220.2
            if cautionmin == "EDTPT":
                cautionminvalue = 392.31
            if cautionmin == "EDTPJ":
                cautionminvalue = 0
            if cautionmin == "TPJED":
                cautionminvalue = 141.5
            if cautionmin == "EDIGU":
                cautionminvalue = 392.3
            if cautionmin == "IGUED":
                cautionminvalue = 475.2
            if cautionmin == "KRRDG":
                cautionminvalue = 0
            if cautionmin == "DGKRR":
                cautionminvalue = 73.9
            if cautionmin == "EDPGTA":
                cautionminvalue = 392.3
            if cautionmin == "PGTEDA":
                cautionminvalue = 534.00
            if cautionmin == "EDPGTB":
                cautionminvalue = 392.3
            if cautionmin == "PGTEDB":
                cautionminvalue = 534.00


            

            for i in range(len(start)):
                start[i] = float(start[i])
                if cautionmin == "JTJED" or cautionmin == "EDJTJ" or cautionmin == "TPTED" or cautionmin == "EDTPT" or cautionmin == "EDTPJ" or cautionmin == "TPJED" or cautionmin =="EDIGU" or cautionmin == "IGUED" or cautionmin =="KRRDG" or cautionmin == "DGKRR" or cautionmin == "EDPGTA" or cautionmin == "PGTEDA" or cautionmin == "EDPGTB" or cautionmin == "PGTEDB":
                    start[i] = start[i] - cautionminvalue
                if cautionmin == "SATPJ":
                    if cautioncheckboxvalue[i] == "SA - KRR":
                        cautionminvalue = 0
                    if cautioncheckboxvalue[i] == "KRR - TPJ":
                        cautionminvalue = -20
                    start[i] = start[i] - cautionminvalue
                if cautionmin == "TPJSA":
                    if cautioncheckboxvalue[i] == "KRR - TPJ":
                        cautionminvalue = 141.5
                    if cautioncheckboxvalue[i] == "SA - KRR":
                        cautionminvalue = 162.22
                    start[i] = start[i] - cautionminvalue
                if cautionmin == "JTJCBE":
                    if cautioncheckboxvalue[i] == "JTJ - IGU":
                        cautionminvalue = 213.02
                    if cautioncheckboxvalue[i] == "IGU - CBE":
                        cautionminvalue = -267.92
                    start[i] = start[i] - cautionminvalue
                if cautionmin == "CBEJTJ":
                    if cautioncheckboxvalue[i] == "IGU - CBE":
                        cautionminvalue = 17.80
                    if cautioncheckboxvalue[i] == "JTJ - IGU":
                        cautionminvalue = 493.1
                if cautionmin == "SACBE":
                    if cautioncheckboxvalue[i] == "SA - IGU":
                        cautionminvalue = 333.1
                    if cautioncheckboxvalue[i] == "IGU - CBE":
                        cautionminvalue = -142
                    start[i] = start[i] - cautionminvalue
                if cautionmin == "CBESA":
                    if cautioncheckboxvalue[i] == "IGU - CBE":
                        cautionminvalue = 17.80
                    if cautioncheckboxvalue[i] == "SA - IGU":
                        cautionminvalue = 493.1
                    start[i] = start[i] - cautionminvalue
                start[i] = round(start[i],2)
                #EDJTJ removing negative values
                if start[i] < 0:
                    start[i] = start[i] * (-1)
                start[i] = str(start[i])
            #print(start)
            start = list(map(float,start))
            for s in range(len(start)):
                for i in range(len(x)):
                    if x[i] == start[s]:
                        #print("found")
                        #print(start[s])
                        #print(x[i])
                        found[s] = 1
                        cautionindexstart[s] = i
            #print(found)
            for i in range(len(found)):
                if found[i] == 0:
                    start[i] = round(start[i],1)
                    for v in range(len(x)):
                        if x[v] == start[i]:
                            #print("found again")
                            found[i] = 1
                            cautionindexstart[i] = v
            #print(found)
            #print(start)
            for a in range(10):
                for i in range(len(found)):
                    if found[i] == 0:
                        start[i] = start[i] + (a/100)
                        for v in range(len(x)):
                            if x[v] == start[i]:
                                found[i] = 1
                                cautionindexstart[i] = v
            #print(found)
            #print(start)
            for i in range(len(end)):
                end[i] = float(end[i])
                if cautionmin == "JTJED" or cautionmin == "EDJTJ" or cautionmin == "TPTED" or cautionmin == "EDTPT" or cautionmin == "EDTPJ" or cautionmin == "TPJED" or cautionmin == "EDIGU" or cautionmin == "IGUED" or cautionmin =="KRRDG" or cautionmin == "DGKRR" or cautionmin == "EDPGTA" or cautionmin == "PGTEDA" or cautionmin == "EDPGTB" or cautionmin == "PGTEDB":
                    end[i] = end[i] - cautionminvalue
                if cautionmin == "SATPJ":
                    if cautioncheckboxvalue[i] == "SA - KRR":
                        cautionminvalue = 0
                    if cautioncheckboxvalue[i] == "KRR - TPJ":
                        cautionminvalue = -20
                    end[i] = end[i] - cautionminvalue
                if cautionmin == "TPJSA":
                    if cautioncheckboxvalue[i] == "KRR - TPJ":
                        cautionminvalue = 141.5
                    if cautioncheckboxvalue[i] == "SA - KRR":
                        cautionminvalue = 162.22
                    end[i] = end[i] - cautionminvalue
                if cautionmin == "JTJCBE":
                    if cautioncheckboxvalue[i] == "JTJ - IGU":
                        cautionminvalue = 213.02
                    if cautioncheckboxvalue[i] == "IGU - CBE":
                        cautionminvalue = -267.92
                    end[i] = end[i] - cautionminvalue
                if cautionmin == "CBEJTJ":
                    if cautioncheckboxvalue[i] == "IGU - CBE":
                        cautionminvalue = 17.80
                    if cautioncheckboxvalue[i] == "JTJ - IGU":
                        cautionminvalue = 493.1
                    end[i] = end[i] - cautionminvalue
                if cautionmin == "SACBE":
                    if cautioncheckboxvalue[i] == "SA - IGU":
                        cautionminvalue = 333.1
                    if cautioncheckboxvalue[i] == "IGU - CBE":
                        cautionminvalue = -142
                    end[i] = end[i] - cautionminvalue
                if cautionmin == "CBESA":
                    if cautioncheckboxvalue[i] == "IGU - CBE":
                        cautionminvalue = 17.80
                    if cautioncheckboxvalue[i] == "SA - IGU":
                        cautionminvalue = 493.1
                    end[i] = end[i] - cautionminvalue

                end[i] = round(end[i],2)
                #EDJTJ removing negative values
                if end[i] < 0:
                    end[i] = end[i] * (-1)
                end[i] = str(end[i])
            #print(start)
            #print(end)
            end = list(map(float,end))
            for s in range(len(end)):
                for i in range(len(x)):
                    if x[i] == end[s]:
                        #print("founde")
                        #print(end[s])
                        #print(x[i])
                        founde[s] = 1
                        cautionindexend[s] = i
            #print(founde)
            for i in range(len(founde)):
                if founde[i] == 0:
                    end[i] = round(end[i],1)
                    for v in range(len(x)):
                        if x[v] == end[i]:
                            #print("founde again")
                            founde[i] = 1
                            cautionindexend[i] = v
            #print(founde)
            #print("0000000000")
            for a in range(10):
                for i in range(len(founde)):
                    if founde[i] == 0:
                        end[i] = end[i] + (a/100)
                        for v in range(len(x)):
                            if x[v] == end[i]:
                                founde[i] = 1
                                cautionindexend[i] = v
            #print(founde)
            print(start)
            print(end)
            speed = list(map(float,speed))
            lengthofcautionspot = [1] * len(cautionindexstart)
            for i in range(len(cautionindexstart)):
                lengthofcautionspot[i] = cautionindexend[i] - cautionindexstart[i]
            #for a in range(len(cautionindexstart)):
            violated = [0] * len(speed)
            
            cautionlistx = {}
            cautionlisty = {}
            for a in range(len(speed)):
                cautionspeedx = []
                cautionspeedy = [] 
                for i in range(cautionindexstart[a],(cautionindexend[a]+1)):
                    cautionspeedx.append(x[i])
                    cautionspeedy.append(y[i])
                    if y[i] > speed[a]:
                        violated[a] = 1
                        #print("caution violated")
                cautionlistx[str(a)] = cautionspeedx
                cautionlisty[str(a)] = cautionspeedy
            violatedcount = 0
            for i in range(len(speed)):
                if violated[i] == 1:
                    violatedcount = violatedcount+1
            violatedvalue = ""
            if violatedcount == 0:
                violatedvalue = "All caution spots crossed successfully with limited speed"
            if violatedcount !=0:
                violatedvalue = str(violatedcount) + " caution spots violated"
            #print(start)
            #print(end)
            #print(speed)
            #print(cautionindexstart)
            #print(cautionindexend)
            #print(lengthofcautionspot)
            #print(violated)
            
            #print(cautionspeedx)
            #print(cautionspeedy)
            #print(cautionlistx['0'])
            #print(cautionlisty['0'])
            #for i in range(len(speed)):
                #print(cautionlistx[f'{i}'])





        """
        plt.figure(figsize=(graphsizef,6))
        plt.xlabel("Distance in (KM)")
        plt.ylabel("Speed in (KMPH)")
        plt.title("Speedometer graph")
        plt.plot(x,y)
        plt.scatter(highlightx,highlighty, color = "r", marker = "o")
        #loop for annotation
        for i, label in enumerate(annot):
            plt.text(highlightx[i],highlighty[i],label)
        #fig = plt.show()
        buffer = BytesIO()
        plt.savefig(buffer, format = 'png')
        buffer.seek(0)
        image_png = buffer.getvalue()
        #chart.clear()
        chart = base64.b64encode(image_png)
        chart = chart.decode('utf-8')
        buffer.close()
        plt.clf()

        """
        #print(request.POST)
        #inputvalue = request.POST["ivalue"]
        #inputvalue = request.POST
        #return render(request, "index.html", {"minvalue" : minvalue, "today" : today, "chart" : chart, "inputvalue" : inputvalue})






        fig = go.Figure()
        scatter = go.Scatter(x=x, y=y, mode='lines', name='SPM Chart', opacity=0.8, marker_color='blue')
        fig.add_trace(scatter)
        """
        if bftindexend != 0:
            scatterbft = go.Scatter(x=bftdistance, y=bftspeed, mode = "markers", name = "Brake Feet Test", marker_color = "orange")
            fig.add_trace(scatterbft)
        """
        if startindex !=0:
            scatterbpt = go.Scatter(x=bptdistancevalues, y=bptspeedvalues, mode='markers' , name = 'Brake Power Test', marker_color='black')
            fig.add_trace(scatterbpt)
        if bftstartindex !=0:
            scatterbft = go.Scatter(x=bftdistancevalues, y=bftspeedvalues, mode = 'markers' , name = 'Brake Feel Test', marker_color='orange')
            fig.add_trace(scatterbft)
        
        scattera=[1] * len(speed)
        for i in range(len(speed)):
            if violated[i] == 1:
                violationcolor = 'red'
                cautionname = 'Caution - '+str(i+1) + '- violated'
            if violated[i] == 0:
                violationcolor = 'green'
                cautionname = 'Caution - '+str(i+1) + '- not - violated'
            scattera[i] = go.Scatter(x=cautionlistx[f'{i}'], y=cautionlisty[f'{i}'], mode ='markers', name=cautionname, marker_color=violationcolor)
            fig.add_trace(scattera[i])

        #scatter2 = go.Scatter(x=cautionspeedx, y=cautionspeedy, mode='markers', name = 'Caution', marker_color='red')
        #fig.add_trace(scatter2)
        arrow_list=[]
        for i in range(len(highlightx)):
            arrow=dict(x=highlightx[i],y=highlighty[i],text=annot[i],arrowhead = 2,
                       arrowwidth=1.5,
                       arrowcolor='rgb(255,51,0)',)
            arrow_list.append(arrow)

        fig.update_layout(annotations=arrow_list, xaxis_title="DISTANCE", yaxis_title="SPEED (KMPH)", title={'text': 'SPEEDOMETER CHART', 'y':0.9,'x':0.5,'xanchor': 'center','yanchor': 'top'})
        #fig.update_xaxes(type='category')
        #fig.update_layout(xaxis = dict(tickmode = 'array',tickvals = plotsignal, ticktext = x))
        plot_div = plot(fig, output_type='div')

        

        loading = "true"
        if startdate == enddate:
            datevalue = startdate
        if startdate != enddate:
            datevalue = "From " + str(startdate) + " To " + str(enddate)

       
        print(highlightxtime)
        
        for i in range(len(highlightxtime)):
            highlightxtime[i] = datetimesplit[highlightxtime[i]]

        highlightxtime = [0] * len(nstn)

        for i in range(len(x)):
            for a in range(len(nstn)):
                if round(x[i],1) == round(dstn[a],1):
                    highlightxtime[a] = datetimesplit[i].time()
        print(highlightxtime)

        hightime = []
        for i in range(len(highlightxtime)):
            if highlightxtime[i] == 0:
                hightime.append(i)

        
        for i in range(len(datetimesplit)):
            datetimesplit[i] = datetimesplit[i].time()

        fig2 = go.Figure()
        scatter2 = go.Scatter(x=datetimesplit, y=y, mode='lines', name='SPM Chart', opacity=0.8, marker_color='blue')
        fig2.add_trace(scatter2)
        """
        if bftindexend != 0:
            scatterbft = go.Scatter(x=bftdistance, y=bftspeed, mode = "markers", name = "Brake Feet Test", marker_color = "orange")
            fig.add_trace(scatterbft)
        """
        if startindex !=0:
            scatterbpt2 = go.Scatter(x=bptdistancevaluestime, y=bptspeedvalues, mode='markers' , name = 'Brake Power Test', marker_color='black')
            fig2.add_trace(scatterbpt2)
        if bftstartindex !=0:
            scatterbft2 = go.Scatter(x=bftdistancevaluestime, y=bftspeedvalues, mode = 'markers' , name = 'Brake Feel Test', marker_color='orange')
            fig2.add_trace(scatterbft2)

        arrow_list2=[]
        for i in range(len(highlightx)):
            if highlightxtime[i] != 0 or i==0: 
                arrow2=dict(x=highlightxtime[i],y=highlighty[i],text=annot[i],arrowhead = 2,
                           arrowwidth=1.5,
                           arrowcolor='rgb(255,51,0)',)
                arrow_list2.append(arrow2)

        fig2.update_layout(annotations=arrow_list2, xaxis_title="Time", yaxis_title="SPEED (KMPH)", title={'text': 'SPEEDOMETER CHART', 'y':0.9,'x':0.5,'xanchor': 'center','yanchor': 'top'})
       

 

        plot_div2 = plot(fig2, output_type= 'div')
     

        return render(request, "index.html", {"minvalue" : minvalue, "plot_div2": plot_div2, "avgspdwd":avgspdwd,"avgspdwod":avgspdwod, "datevalue": datevalue, "nameoflp" : nameoflp, "trainno":trainno, "locono":locono, "loading": loading, "latesignals" : latesignals, "stoppingstations" : stoppingstations,  "mpsmaxvalue": mpsmaxvalue, "mpsviolateddistance" : mpsviolateddistance, "mpsdistance" : mpsdistance, "violatedvalue" : violatedvalue, "bftvalue" : bftvalue, "bptvalue" : bptvalue, "today" : today, "mpsvalue" : mpsvalue, "plot_div" : plot_div, "sstn" : sstn, "slist" : slist, "dlist" : dlist})
"""
def indexview(request):
    return render(request,'index.html')
"""

@login_required(login_url = '/login')
def dashboardview(request):
    mydata = User.objects.filter(first_name='abc')
    superusers = User.objects.filter(is_superuser=True)
    print(mydata[0])
    print(superusers[0])
    sup = str(superusers[0])
    #ranvalue = mydata[0]
    ranvalue = str(mydata[0])
    ranvalue = int(ranvalue)
    today = date.today()
    todayvalue = str(today).split('-')
    for i in range(len(todayvalue)):
        ranvalue = ((int(todayvalue[i]))*5 + ranvalue  + (i*3) + (i*5))*5

    return render(request,'homepage.html',{"sup":sup, "ranvalue":ranvalue})

def registerview(request):
    if request.method == "POST":
        form = RegisterUserForm(request.POST)
        if form.is_valid():
            form.save()
            """username = form.cleaned_data.get('username')
                                                pwd=form.cleaned_data.get('password1')
                                                user = authenticate(username=username,password=pwd)
                                                login(request,user)"""
            return render(request,'registration/login.html')
    else:
        form = RegisterUserForm()
        mydata = User.objects.filter(first_name='abc')
        #ranvalue = mydata[0]
        ranvalue = str(mydata[0])
        ranvalue = int(ranvalue)
        today = date.today()
        todayvalue = str(today).split('-')
        for i in range(len(todayvalue)):
            ranvalue = ((int(todayvalue[i]))*5 + ranvalue  + (i*3) + (i*5))*5
    return render(request,'registration/register.html',{"form":form, "ranvalue": ranvalue})
