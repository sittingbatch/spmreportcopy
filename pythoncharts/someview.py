# HttpResponse is used to
# pass the information
# back to view
import openpyxl
#import datetime
#from matplotlib import pyplot as plt
from django.template import Context
import time
import os
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
#import numpy as np
import io, base64, uuid
from openpyxl import workbook,load_workbook
from io import BytesIO
from datetime import datetime, timedelta, timezone, tzinfo
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.template import loader
from plotly.offline import plot
import plotly.graph_objs as go



@login_required(login_url = '/login')
def something (request) :
    print(request.method)
    if request.method == "GET":
        #http = urllib3.PoolManager()

        #r = http.request('GET', 'https://cms.indianrail.gov.in/CMSREPORT/JSP/rpt/management/CrewInPRAction.do?hmode=PRNOPR&hmode=showLobby&error=S&zone=SC&div=GTL&lobbyList=SC-GTL-NRE&searchType=CREWID&crewId=&searchTypeOption=LOCO&locoNo=&hmode=showLobby&XML=%3C%3Fxml+version%3D%221.0%22+encoding%3D%22UTF-8%22%3F%3E+%3CCMSPublishXML+baseLanguage%3D%22string%22+transLanguage%3D%22string%22%3E++%3CCMSREPORT+action%3D%22REP%22+relationship%3D%22string%22++transLanguage%3D%22string%22%3E++%3Czone%3ESC%3C%2Fzone%3E+%3Cdivision%3EGTL%3C%2Fdivision%3E+%3Clobby%3ENRE%3C%2Flobby%3E+%3Cdesig1%3EALP%3C%2Fdesig1%3E+%3Cdesig2%3ESALP%3C%2Fdesig2%3E+%3Cdesig3%3ELPG%3C%2Fdesig3%3E+%3Cdesig4%3Efalse%3C%2Fdesig4%3E+%3Cdesig5%3Efalse%3C%2Fdesig5%3E+%3Cdesig6%3Efalse%3C%2Fdesig6%3E+%3Cdesig7%3Efalse%3C%2Fdesig7%3E+%3Cdesig8%3Efalse%3C%2Fdesig8%3E+%3Cdesig9%3Efalse%3C%2Fdesig9%3E+%3Cdesig10%3Efalse%3C%2Fdesig10%3E+%3Cdesig11%3Efalse%3C%2Fdesig11%3E+%3Cdesig12%3Efalse%3C%2Fdesig12%3E+%3Cdesig13%3Efalse%3C%2Fdesig13%3E+%3Cdesig14%3Efalse%3C%2Fdesig14%3E+%3Ctraction%3EALL%3C%2Ftraction%3E+%3Ccadre%3EE%27%2C%27M%27%2C%27B%3C%2Fcadre%3E+%3CcombALP%3ECOMB%3C%2FcombALP%3E+%3CdesigSelect%3EOFFICIATING%3C%2FdesigSelect%3E+%3CcontValue%3EContinuous%3C%2FcontValue%3E+%3CcontValueOption%3ESignOnOff%3C%2FcontValueOption%3E+%3CbiodataSubParts%3EDETAIL%3C%2FbiodataSubParts%3E+%3Cmonth1%3E%3C%2Fmonth1%3E+%3Cyear1%3E%3C%2Fyear1%3E+%3ClobbyselectionAs%3EHqLobby%3C%2FlobbyselectionAs%3E+%3Cdurationtype%3EFORTNIGHT%3C%2Fdurationtype%3E+%3Cactive%3EActive%3C%2Factive%3E+%3Crlevel%3ELOBBY%3C%2Frlevel%3E+%3CcurrentMidnignt%3ECURRENT%3C%2FcurrentMidnignt%3E+%3CcrewIDBaseID%3ECrewID%3C%2FcrewIDBaseID%3E+%3CcrewDesgLevel%3EGoods%3C%2FcrewDesgLevel%3E+%3CcadreFilter%3EnotCadre%3C%2FcadreFilter%3E+%3CworkingOnbehalfofCrew%3ESignOn+by+Supervisor%3C%2FworkingOnbehalfofCrew%3E+%3CfromSttn%3ESELECT%3C%2FfromSttn%3E+%3CtoSttn%3ESELECT%3C%2FtoSttn%3E+%3CstartingDate%3E%3C%2FstartingDate%3E+%3CendDate%3E%3C%2FendDate%3E+%3CpreodicCoursesVal%3EDONE%3C%2FpreodicCoursesVal%3E+%3CcrewBAStatus%3ESIGNON%3C%2FcrewBAStatus%3E+%3CcrewAvailableCheckList%3ECrewAvailableFIFO%3C%2FcrewAvailableCheckList%3E+%3CslotData%3Enot+Slot+Data%3C%2FslotData%3E+%3CtrainingSelVal%3EHistory%3C%2FtrainingSelVal%3E+%3CenegryMeterVal%3ESection%3C%2FenegryMeterVal%3E+%3ClocoTraction%3EElec.Conv%3C%2FlocoTraction%3E+%3Cspare%3EnotSpare%3C%2Fspare%3E+%3CcircularCheck%3ECurrent%3C%2FcircularCheck%3E+%3ClocoTypeWiseVal%3EGroup%3C%2FlocoTypeWiseVal%3E+%3CdutyHrsDesg%3EPILOTS%3C%2FdutyHrsDesg%3E+%3CbaReportValue%3EPilot%3C%2FbaReportValue%3E+%3Cspectacles%3ENORM%3C%2Fspectacles%3E+%3Calcohol%3EN%3C%2Falcohol%3E+%3CabnormalityType%3Enot+Nil%3C%2FabnormalityType%3E+%3CselSuburban%3ESignOn%3C%2FselSuburban%3E+%3ClocoNosearch%3E%3C%2FlocoNosearch%3E+%3CslotValueCombo%3ESlot%3C%2FslotValueCombo%3E+%3CindexValue%3E0%3C%2FindexValue%3E+%3CslotValueText%3ESlot%3C%2FslotValueText%3E+%3Cmylist%3E%3C%2Fmylist%3E+%3Cmylist1%3E%3C%2Fmylist1%3E+%3Cmylist2%3E%3C%2Fmylist2%3E+%3Cmylist3%3E%3C%2Fmylist3%3E+%3Cmylist4%3E%3C%2Fmylist4%3E+%3Cflexi1%3E%3C%2Fflexi1%3E+%3Cflexi2%3E%3C%2Fflexi2%3E+%3Cflexi3%3E%3C%2Fflexi3%3E+%3Cflexi4%3E%3C%2Fflexi4%3E+%3Cflexi5%3E%3C%2Fflexi5%3E+%3Ccomparison1%3E%3C%2Fcomparison1%3E+%3Ccomparison2%3E%3C%2Fcomparison2%3E+%3Ccomparison3%3E%3C%2Fcomparison3%3E+%3Ccomparison4%3E%3C%2Fcomparison4%3E+%3Ccomparison5%3E%3C%2Fcomparison5%3E+%3Csortlist%3E%3C%2Fsortlist%3E+%3Csortlist1%3E%3C%2Fsortlist1%3E+%3Csortlist2%3E%3C%2Fsortlist2%3E+%3Csortlist3%3E%3C%2Fsortlist3%3E+%3Croute%3E-+-+-Route-+-+-%3C%2Froute%3E+%3Croutename%3E-+-+Route-+-+-%3C%2Froutename%3E+%3CfromSttnNameRoute%3ESelect%3C%2FfromSttnNameRoute%3E+%3CtoSttnNameRoute%3ESelect%3C%2FtoSttnNameRoute%3E+%3C%2FCMSREPORT%3E++%3C%2FCMSPublishXML%3E')

        #print(r.cookie)
        """
        #url = "https://cms.indianrail.gov.in/CMSREPORT/JSP/rpt/LoginAction.do?hmode=login&isResponsive=Y&userId=NSAI&userPassword=NSAI1995"
        prurl = 'https://cms.indianrail.gov.in/CMSREPORT/JSP/rpt/management/CrewInPRAction.do?hmode=PRNOPR&hmode=showLobby&error=S&zone=SC&div=GTL&lobbyList=SC-GTL-NRE&searchType=CREWID&crewId=&searchTypeOption=LOCO&locoNo=&hmode=showLobby&XML=%3C%3Fxml+version%3D%221.0%22+encoding%3D%22UTF-8%22%3F%3E+%3CCMSPublishXML+baseLanguage%3D%22string%22+transLanguage%3D%22string%22%3E++%3CCMSREPORT+action%3D%22REP%22+relationship%3D%22string%22++transLanguage%3D%22string%22%3E++%3Czone%3ESC%3C%2Fzone%3E+%3Cdivision%3EGTL%3C%2Fdivision%3E+%3Clobby%3ENRE%3C%2Flobby%3E+%3Cdesig1%3EALP%3C%2Fdesig1%3E+%3Cdesig2%3ESALP%3C%2Fdesig2%3E+%3Cdesig3%3ELPG%3C%2Fdesig3%3E+%3Cdesig4%3Efalse%3C%2Fdesig4%3E+%3Cdesig5%3Efalse%3C%2Fdesig5%3E+%3Cdesig6%3Efalse%3C%2Fdesig6%3E+%3Cdesig7%3Efalse%3C%2Fdesig7%3E+%3Cdesig8%3Efalse%3C%2Fdesig8%3E+%3Cdesig9%3Efalse%3C%2Fdesig9%3E+%3Cdesig10%3Efalse%3C%2Fdesig10%3E+%3Cdesig11%3Efalse%3C%2Fdesig11%3E+%3Cdesig12%3Efalse%3C%2Fdesig12%3E+%3Cdesig13%3Efalse%3C%2Fdesig13%3E+%3Cdesig14%3Efalse%3C%2Fdesig14%3E+%3Ctraction%3EALL%3C%2Ftraction%3E+%3Ccadre%3EE%27%2C%27M%27%2C%27B%3C%2Fcadre%3E+%3CcombALP%3ECOMB%3C%2FcombALP%3E+%3CdesigSelect%3EOFFICIATING%3C%2FdesigSelect%3E+%3CcontValue%3EContinuous%3C%2FcontValue%3E+%3CcontValueOption%3ESignOnOff%3C%2FcontValueOption%3E+%3CbiodataSubParts%3EDETAIL%3C%2FbiodataSubParts%3E+%3Cmonth1%3E%3C%2Fmonth1%3E+%3Cyear1%3E%3C%2Fyear1%3E+%3ClobbyselectionAs%3EHqLobby%3C%2FlobbyselectionAs%3E+%3Cdurationtype%3EFORTNIGHT%3C%2Fdurationtype%3E+%3Cactive%3EActive%3C%2Factive%3E+%3Crlevel%3ELOBBY%3C%2Frlevel%3E+%3CcurrentMidnignt%3ECURRENT%3C%2FcurrentMidnignt%3E+%3CcrewIDBaseID%3ECrewID%3C%2FcrewIDBaseID%3E+%3CcrewDesgLevel%3EGoods%3C%2FcrewDesgLevel%3E+%3CcadreFilter%3EnotCadre%3C%2FcadreFilter%3E+%3CworkingOnbehalfofCrew%3ESignOn+by+Supervisor%3C%2FworkingOnbehalfofCrew%3E+%3CfromSttn%3ESELECT%3C%2FfromSttn%3E+%3CtoSttn%3ESELECT%3C%2FtoSttn%3E+%3CstartingDate%3E%3C%2FstartingDate%3E+%3CendDate%3E%3C%2FendDate%3E+%3CpreodicCoursesVal%3EDONE%3C%2FpreodicCoursesVal%3E+%3CcrewBAStatus%3ESIGNON%3C%2FcrewBAStatus%3E+%3CcrewAvailableCheckList%3ECrewAvailableFIFO%3C%2FcrewAvailableCheckList%3E+%3CslotData%3Enot+Slot+Data%3C%2FslotData%3E+%3CtrainingSelVal%3EHistory%3C%2FtrainingSelVal%3E+%3CenegryMeterVal%3ESection%3C%2FenegryMeterVal%3E+%3ClocoTraction%3EElec.Conv%3C%2FlocoTraction%3E+%3Cspare%3EnotSpare%3C%2Fspare%3E+%3CcircularCheck%3ECurrent%3C%2FcircularCheck%3E+%3ClocoTypeWiseVal%3EGroup%3C%2FlocoTypeWiseVal%3E+%3CdutyHrsDesg%3EPILOTS%3C%2FdutyHrsDesg%3E+%3CbaReportValue%3EPilot%3C%2FbaReportValue%3E+%3Cspectacles%3ENORM%3C%2Fspectacles%3E+%3Calcohol%3EN%3C%2Falcohol%3E+%3CabnormalityType%3Enot+Nil%3C%2FabnormalityType%3E+%3CselSuburban%3ESignOn%3C%2FselSuburban%3E+%3ClocoNosearch%3E%3C%2FlocoNosearch%3E+%3CslotValueCombo%3ESlot%3C%2FslotValueCombo%3E+%3CindexValue%3E0%3C%2FindexValue%3E+%3CslotValueText%3ESlot%3C%2FslotValueText%3E+%3Cmylist%3E%3C%2Fmylist%3E+%3Cmylist1%3E%3C%2Fmylist1%3E+%3Cmylist2%3E%3C%2Fmylist2%3E+%3Cmylist3%3E%3C%2Fmylist3%3E+%3Cmylist4%3E%3C%2Fmylist4%3E+%3Cflexi1%3E%3C%2Fflexi1%3E+%3Cflexi2%3E%3C%2Fflexi2%3E+%3Cflexi3%3E%3C%2Fflexi3%3E+%3Cflexi4%3E%3C%2Fflexi4%3E+%3Cflexi5%3E%3C%2Fflexi5%3E+%3Ccomparison1%3E%3C%2Fcomparison1%3E+%3Ccomparison2%3E%3C%2Fcomparison2%3E+%3Ccomparison3%3E%3C%2Fcomparison3%3E+%3Ccomparison4%3E%3C%2Fcomparison4%3E+%3Ccomparison5%3E%3C%2Fcomparison5%3E+%3Csortlist%3E%3C%2Fsortlist%3E+%3Csortlist1%3E%3C%2Fsortlist1%3E+%3Csortlist2%3E%3C%2Fsortlist2%3E+%3Csortlist3%3E%3C%2Fsortlist3%3E+%3Croute%3E-+-+-Route-+-+-%3C%2Froute%3E+%3Croutename%3E-+-+Route-+-+-%3C%2Froutename%3E+%3CfromSttnNameRoute%3ESelect%3C%2FfromSttnNameRoute%3E+%3CtoSttnNameRoute%3ESelect%3C%2FtoSttnNameRoute%3E+%3C%2FCMSREPORT%3E++%3C%2FCMSPublishXML%3E'
        #prurl = "https://cms.indianrail.gov.in/CMSREPORT/JSP/rpt/management/CrewInPRAction.do?hmode=PRNOPR&XML=%3C%3Fxml+version%3D%221.0%22+encoding%3D%22UTF-8%22%3F%3E+%3CCMSPublishXML+baseLanguage%3D%22string%22+transLanguage%3D%22string%22%3E++%3CCMSREPORT+action%3D%22REP%22+relationship%3D%22string%22++transLanguage%3D%22string%22%3E++%3Czone%3ESC%3C%2Fzone%3E+%3Cdivision%3EGTL%3C%2Fdivision%3E+%3Clobby%3ENRE%3C%2Flobby%3E+%3Cdesig1%3Efalse%3C%2Fdesig1%3E+%3Cdesig2%3ESALP%3C%2Fdesig2%3E+%3Cdesig3%3Efalse%3C%2Fdesig3%3E+%3Cdesig4%3Efalse%3C%2Fdesig4%3E+%3Cdesig5%3Efalse%3C%2Fdesig5%3E+%3Cdesig6%3Efalse%3C%2Fdesig6%3E+%3Cdesig7%3Efalse%3C%2Fdesig7%3E+%3Cdesig8%3Efalse%3C%2Fdesig8%3E+%3Cdesig9%3Efalse%3C%2Fdesig9%3E+%3Cdesig10%3Efalse%3C%2Fdesig10%3E+%3Cdesig11%3Efalse%3C%2Fdesig11%3E+%3Cdesig12%3Efalse%3C%2Fdesig12%3E+%3Cdesig13%3Efalse%3C%2Fdesig13%3E+%3Cdesig14%3Efalse%3C%2Fdesig14%3E+%3Cabnormality1%3ECOMMERCIAL%3C%2Fabnormality1%3E+%3Cabnormality2%3Efalse%3C%2Fabnormality2%3E+%3Cabnormality3%3Efalse%3C%2Fabnormality3%3E+%3Cabnormality4%3Efalse%3C%2Fabnormality4%3E+%3Cabnormality5%3Efalse%3C%2Fabnormality5%3E+%3Cabnormality6%3Efalse%3C%2Fabnormality6%3E+%3Cabnormality7%3Efalse%3C%2Fabnormality7%3E+%3Cabnormality8%3Efalse%3C%2Fabnormality8%3E+%3Cabnormality9%3Efalse%3C%2Fabnormality9%3E+%3Cabnormality10%3Efalse%3C%2Fabnormality10%3E+%3Cabnormality11%3Efalse%3C%2Fabnormality11%3E+%3Cabnormality12%3Efalse%3C%2Fabnormality12%3E+%3CstartingDate%3E%3C%2FstartingDate%3E+%3CendDate%3E%3C%2FendDate%3E+%3CmonthYearDateFormat%3E%3C%2FmonthYearDateFormat%3E+%3CmsgSrc%3ECS%3C%2FmsgSrc%3E+%3Ctraction%3EALL%3C%2Ftraction%3E+%3Ccadre%3EE'%2C'M'%2C'B%3C%2Fcadre%3E+%3CfghtCochSht%3EFghtCoch%3C%2FfghtCochSht%3E+%3Cdesignation%3EPILOT%3C%2Fdesignation%3E+%3Cactive%3EActive%3C%2Factive%3E+%3Crlevel%3ELOBBY%3C%2Frlevel%3E+%3Cdurationtype%3EFORTNIGHT%3C%2Fdurationtype%3E+%3CcombALP%3ECOMB%3C%2FcombALP%3E+%3CslotData%3Enot+Slot+Data%3C%2FslotData%3E+%3CdesigSelect%3EOFFICIATING%3C%2FdesigSelect%3E+%3CcrewAvailableCheckList%3ECrewAvailableFIFO%3C%2FcrewAvailableCheckList%3E+%3CcontValue%3EContinuous%3C%2FcontValue%3E+%3CmandatoryRequirementDueFilter%3EReft%3C%2FmandatoryRequirementDueFilter%3E+%3CsignOnOFFVal%3ESignOnVal%3C%2FsignOnOFFVal%3E+%3ClocoTraction%3EALL%3C%2FlocoTraction%3E+%3Ccont_NoncontValue%3EContinuousHQ%3C%2Fcont_NoncontValue%3E+%3CcontValueOption%3ESignOnOff%3C%2FcontValueOption%3E+%3Cspare%3Espare%3C%2Fspare%3E+%3CcrewBAStatus%3ESIGNON%3C%2FcrewBAStatus%3E+%3Cpddpad%3EHQ+crew+at+HQ%3C%2Fpddpad%3E+%3CcrewIDBaseID%3ECrewID%3C%2FcrewIDBaseID%3E+%3CcrewDesgLevel%3EGoods%3C%2FcrewDesgLevel%3E+%3CcurrentMidnignt%3ECURRENT%3C%2FcurrentMidnignt%3E+%3CabnormalityStatus%3EPN%3C%2FabnormalityStatus%3E+%3CcadreFilter%3EnotCadre%3C%2FcadreFilter%3E+%3Cyear1%3E%3C%2Fyear1%3E+%3CcrewBookingWrWorWise%3ECallBookLobbyWise%3C%2FcrewBookingWrWorWise%3E+%3Cmonth1%3E%3C%2Fmonth1%3E+%3CslotFilter%3EPrevious%3C%2FslotFilter%3E+%3CpreodicCoursesVal%3EDONE%3C%2FpreodicCoursesVal%3E+%3CdetailLevel%3ESummary%3C%2FdetailLevel%3E+%3ClocoGroupVal%3EELEC-CONV%3C%2FlocoGroupVal%3E+%3Ctime%3E4%3C%2Ftime%3E+%3ClocoTypeWiseVal%3EGroup%3C%2FlocoTypeWiseVal%3E+%3CreportGroupVal%3ELobby%3C%2FreportGroupVal%3E+%3CdfccRadio%3EALL%3C%2FdfccRadio%3E+%3CfromSttn%3Enull%3C%2FfromSttn%3E+%3CtoSttn%3Enull%3C%2FtoSttn%3E+%3CslotValueCombo%3ESlot%3C%2FslotValueCombo%3E+%3ClocoNosearch%3E%3C%2FlocoNosearch%3E+%3CmonthCombo%3EPrevious%3C%2FmonthCombo%3E+%3CmonthComboValueText%3EPrevious%3C%2FmonthComboValueText%3E+%3CindexValue%3E0%3C%2FindexValue%3E+%3CslotValueText%3ESlot%3C%2FslotValueText%3E+%3Croute%3E-+-+-Route-+-+-%3C%2Froute%3E+%3Croutename%3E-+-+Route-+-+-%3C%2Froutename%3E+%3CfromSttnNameRoute%3E%3C%2FfromSttnNameRoute%3E+%3CtoSttnNameRoute%3E%3C%2FtoSttnNameRoute%3E+%3CtrainingValue%3E-+-+Select+-+-%3C%2FtrainingValue%3E+%3C%2FCMSREPORT%3E++%3C%2FCMSPublishXML%3E&lobby=NRE&lobbyList=SC-GTL-NRE"
        lrurl = "https://cms.indianrail.gov.in/CMSREPORT/JSP/rpt/management/LRReport.do?hmode=LRReportFromToStationWise&hmode=Home&error=S&zone=SC&div=GTL&lobbyList=SC-GTL-NRE&searchType=CREWID&crewId=&searchTypeOption=LOCO&locoNo=&hmode=Home&XML=%3C%3Fxml+version%3D%221.0%22+encoding%3D%22UTF-8%22%3F%3E+%3CCMSPublishXML+baseLanguage%3D%22string%22+transLanguage%3D%22string%22%3E++%3CCMSREPORT+action%3D%22REP%22+relationship%3D%22string%22++transLanguage%3D%22string%22%3E++%3Czone%3ESC%3C%2Fzone%3E+%3Cdivision%3EGTL%3C%2Fdivision%3E+%3Clobby%3ENRE%3C%2Flobby%3E+%3Cdesig1%3Efalse%3C%2Fdesig1%3E+%3Cdesig2%3ESALP%3C%2Fdesig2%3E+%3Cdesig3%3Efalse%3C%2Fdesig3%3E+%3Cdesig4%3Efalse%3C%2Fdesig4%3E+%3Cdesig5%3Efalse%3C%2Fdesig5%3E+%3Cdesig6%3Efalse%3C%2Fdesig6%3E+%3Cdesig7%3Efalse%3C%2Fdesig7%3E+%3Cdesig8%3Efalse%3C%2Fdesig8%3E+%3Cdesig9%3Efalse%3C%2Fdesig9%3E+%3Cdesig10%3Efalse%3C%2Fdesig10%3E+%3Cdesig11%3Efalse%3C%2Fdesig11%3E+%3Cdesig12%3Efalse%3C%2Fdesig12%3E+%3Cdesig13%3Efalse%3C%2Fdesig13%3E+%3Cdesig14%3Efalse%3C%2Fdesig14%3E+%3Ctraction%3EALL%3C%2Ftraction%3E+%3Ccadre%3EE%27%2C%27M%27%2C%27B%3C%2Fcadre%3E+%3CcombALP%3ECOMB%3C%2FcombALP%3E+%3CdesigSelect%3EOFFICIATING%3C%2FdesigSelect%3E+%3CcontValue%3EContinuous%3C%2FcontValue%3E+%3CcontValueOption%3ESignOnOff%3C%2FcontValueOption%3E+%3CbiodataSubParts%3EDETAIL%3C%2FbiodataSubParts%3E+%3Cmonth1%3E%3C%2Fmonth1%3E+%3Cyear1%3E%3C%2Fyear1%3E+%3ClobbyselectionAs%3EHqLobby%3C%2FlobbyselectionAs%3E+%3Cdurationtype%3EFORTNIGHT%3C%2Fdurationtype%3E+%3Cactive%3EActive%3C%2Factive%3E+%3Crlevel%3ELOBBY%3C%2Frlevel%3E+%3CcurrentMidnignt%3ECURRENT%3C%2FcurrentMidnignt%3E+%3CcrewIDBaseID%3ECrewID%3C%2FcrewIDBaseID%3E+%3CcrewDesgLevel%3EGoods%3C%2FcrewDesgLevel%3E+%3CcadreFilter%3EnotCadre%3C%2FcadreFilter%3E+%3CworkingOnbehalfofCrew%3ESignOn+by+Supervisor%3C%2FworkingOnbehalfofCrew%3E+%3CfromSttn%3E0%3C%2FfromSttn%3E+%3CtoSttn%3E11%3C%2FtoSttn%3E+%3CstartingDate%3E%3C%2FstartingDate%3E+%3CendDate%3E%3C%2FendDate%3E+%3CpreodicCoursesVal%3EDONE%3C%2FpreodicCoursesVal%3E+%3CcrewBAStatus%3ESIGNON%3C%2FcrewBAStatus%3E+%3CcrewAvailableCheckList%3ECrewAvailableFIFO%3C%2FcrewAvailableCheckList%3E+%3CslotData%3Enot+Slot+Data%3C%2FslotData%3E+%3CtrainingSelVal%3EHistory%3C%2FtrainingSelVal%3E+%3CenegryMeterVal%3ESection%3C%2FenegryMeterVal%3E+%3ClocoTraction%3EElec.Conv%3C%2FlocoTraction%3E+%3Cspare%3EnotSpare%3C%2Fspare%3E+%3CcircularCheck%3ECurrent%3C%2FcircularCheck%3E+%3ClocoTypeWiseVal%3EGroup%3C%2FlocoTypeWiseVal%3E+%3CdutyHrsDesg%3EPILOTS%3C%2FdutyHrsDesg%3E+%3CbaReportValue%3EPilot%3C%2FbaReportValue%3E+%3Cspectacles%3ENORM%3C%2Fspectacles%3E+%3Calcohol%3EN%3C%2Falcohol%3E+%3CabnormalityType%3Enot+Nil%3C%2FabnormalityType%3E+%3CselSuburban%3ESignOn%3C%2FselSuburban%3E+%3ClocoNosearch%3E%3C%2FlocoNosearch%3E+%3CslotValueCombo%3ESlot%3C%2FslotValueCombo%3E+%3CindexValue%3E0%3C%2FindexValue%3E+%3CslotValueText%3ESlot%3C%2FslotValueText%3E+%3Cmylist%3E%3C%2Fmylist%3E+%3Cmylist1%3E%3C%2Fmylist1%3E+%3Cmylist2%3E%3C%2Fmylist2%3E+%3Cmylist3%3E%3C%2Fmylist3%3E+%3Cmylist4%3E%3C%2Fmylist4%3E+%3Cflexi1%3E%3C%2Fflexi1%3E+%3Cflexi2%3E%3C%2Fflexi2%3E+%3Cflexi3%3E%3C%2Fflexi3%3E+%3Cflexi4%3E%3C%2Fflexi4%3E+%3Cflexi5%3E%3C%2Fflexi5%3E+%3Ccomparison1%3E%3C%2Fcomparison1%3E+%3Ccomparison2%3E%3C%2Fcomparison2%3E+%3Ccomparison3%3E%3C%2Fcomparison3%3E+%3Ccomparison4%3E%3C%2Fcomparison4%3E+%3Ccomparison5%3E%3C%2Fcomparison5%3E+%3Csortlist%3E%3C%2Fsortlist%3E+%3Csortlist1%3E%3C%2Fsortlist1%3E+%3Csortlist2%3E%3C%2Fsortlist2%3E+%3Csortlist3%3E%3C%2Fsortlist3%3E+%3Croute%3E22601%3C%2Froute%3E+%3Croutename%3ENRE-TPTY-----%2822601%29%3C%2Froutename%3E+%3CfromSttnNameRoute%3ENRE%3C%2FfromSttnNameRoute%3E+%3CtoSttnNameRoute%3ERU%3C%2FtoSttnNameRoute%3E+%3C%2FCMSREPORT%3E++%3C%2FCMSPublishXML%3E"
        dueurl = "https://cms.indianrail.gov.in/CMSREPORT/JSP/rpt/crew/CrewDetail1.do?hmode=crewBiodataReports&type=OtherDetails&XML=%3C%3Fxml+version%3D%221.0%22+encoding%3D%22UTF-8%22%3F%3E+%3CCMSPublishXML+baseLanguage%3D%22string%22+transLanguage%3D%22string%22%3E++%3CCMSREPORT+action%3D%22REP%22+relationship%3D%22string%22++transLanguage%3D%22string%22%3E++%3Czone%3ESC%3C%2Fzone%3E+%3Cdivision%3EGTL%3C%2Fdivision%3E+%3Clobby%3ENRE%3C%2Flobby%3E+%3Cdesig1%3EALP%3C%2Fdesig1%3E+%3Cdesig2%3ESALP%3C%2Fdesig2%3E+%3Cdesig3%3ELPG%3C%2Fdesig3%3E+%3Cdesig4%3Efalse%3C%2Fdesig4%3E+%3Cdesig5%3Efalse%3C%2Fdesig5%3E+%3Cdesig6%3Efalse%3C%2Fdesig6%3E+%3Cdesig7%3Efalse%3C%2Fdesig7%3E+%3Cdesig8%3Efalse%3C%2Fdesig8%3E+%3Cdesig9%3Efalse%3C%2Fdesig9%3E+%3Cdesig10%3Efalse%3C%2Fdesig10%3E+%3Cdesig11%3Efalse%3C%2Fdesig11%3E+%3Cdesig12%3Efalse%3C%2Fdesig12%3E+%3Cdesig13%3Efalse%3C%2Fdesig13%3E+%3Cdesig14%3Efalse%3C%2Fdesig14%3E+%3Cabnormality1%3ECOMMERCIAL%3C%2Fabnormality1%3E+%3Cabnormality2%3Efalse%3C%2Fabnormality2%3E+%3Cabnormality3%3Efalse%3C%2Fabnormality3%3E+%3Cabnormality4%3Efalse%3C%2Fabnormality4%3E+%3Cabnormality5%3Efalse%3C%2Fabnormality5%3E+%3Cabnormality6%3Efalse%3C%2Fabnormality6%3E+%3Cabnormality7%3Efalse%3C%2Fabnormality7%3E+%3Cabnormality8%3Efalse%3C%2Fabnormality8%3E+%3Cabnormality9%3Efalse%3C%2Fabnormality9%3E+%3Cabnormality10%3Efalse%3C%2Fabnormality10%3E+%3Cabnormality11%3Efalse%3C%2Fabnormality11%3E+%3Cabnormality12%3Efalse%3C%2Fabnormality12%3E+%3CstartingDate%3E%3C%2FstartingDate%3E+%3CendDate%3E%3C%2FendDate%3E+%3CmonthYearDateFormat%3E%3C%2FmonthYearDateFormat%3E+%3CmsgSrc%3ECS%3C%2FmsgSrc%3E+%3Ctraction%3EALL%3C%2Ftraction%3E+%3Ccadre%3EE%27%2C%27M%27%2C%27B%3C%2Fcadre%3E+%3CfghtCochSht%3EFghtCoch%3C%2FfghtCochSht%3E+%3Cdesignation%3EPILOT%3C%2Fdesignation%3E+%3Cactive%3EActive%3C%2Factive%3E+%3Crlevel%3ELOBBY%3C%2Frlevel%3E+%3Cdurationtype%3EFORTNIGHT%3C%2Fdurationtype%3E+%3CcombALP%3ECOMB%3C%2FcombALP%3E+%3CslotData%3Enot+Slot+Data%3C%2FslotData%3E+%3CdesigSelect%3EOFFICIATING%3C%2FdesigSelect%3E+%3CcrewAvailableCheckList%3ECrewAvailableFIFO%3C%2FcrewAvailableCheckList%3E+%3CcontValue%3EContinuous%3C%2FcontValue%3E+%3CmandatoryRequirementDueFilter%3EReft%3C%2FmandatoryRequirementDueFilter%3E+%3CsignOnOFFVal%3ESignOnVal%3C%2FsignOnOFFVal%3E+%3ClocoTraction%3EALL%3C%2FlocoTraction%3E+%3Ccont_NoncontValue%3EContinuousHQ%3C%2Fcont_NoncontValue%3E+%3CcontValueOption%3ESignOnOff%3C%2FcontValueOption%3E+%3Cspare%3Espare%3C%2Fspare%3E+%3CcrewBAStatus%3ESIGNON%3C%2FcrewBAStatus%3E+%3Cpddpad%3EHQ+crew+at+HQ%3C%2Fpddpad%3E+%3CcrewIDBaseID%3ECrewID%3C%2FcrewIDBaseID%3E+%3CcrewDesgLevel%3EGoods%3C%2FcrewDesgLevel%3E+%3CcurrentMidnignt%3ECURRENT%3C%2FcurrentMidnignt%3E+%3CabnormalityStatus%3EPN%3C%2FabnormalityStatus%3E+%3CcadreFilter%3EnotCadre%3C%2FcadreFilter%3E+%3Cyear1%3E%3C%2Fyear1%3E+%3CcrewBookingWrWorWise%3ECallBookLobbyWise%3C%2FcrewBookingWrWorWise%3E+%3Cmonth1%3E%3C%2Fmonth1%3E+%3CslotFilter%3EPrevious%3C%2FslotFilter%3E+%3CpreodicCoursesVal%3EDONE%3C%2FpreodicCoursesVal%3E+%3CdetailLevel%3ESummary%3C%2FdetailLevel%3E+%3ClocoGroupVal%3EELEC-CONV%3C%2FlocoGroupVal%3E+%3Ctime%3E4%3C%2Ftime%3E+%3ClocoTypeWiseVal%3EGroup%3C%2FlocoTypeWiseVal%3E+%3CreportGroupVal%3ELobby%3C%2FreportGroupVal%3E+%3CdfccRadio%3EALL%3C%2FdfccRadio%3E+%3CfromSttn%3Enull%3C%2FfromSttn%3E+%3CtoSttn%3Enull%3C%2FtoSttn%3E+%3CslotValueCombo%3ESlot%3C%2FslotValueCombo%3E+%3ClocoNosearch%3E%3C%2FlocoNosearch%3E+%3CmonthCombo%3EPrevious%3C%2FmonthCombo%3E+%3CmonthComboValueText%3EPrevious%3C%2FmonthComboValueText%3E+%3CindexValue%3E0%3C%2FindexValue%3E+%3CslotValueText%3ESlot%3C%2FslotValueText%3E+%3Croute%3E-+-+-Route-+-+-%3C%2Froute%3E+%3Croutename%3E-+-+Route-+-+-%3C%2Froutename%3E+%3CfromSttnNameRoute%3E%3C%2FfromSttnNameRoute%3E+%3CtoSttnNameRoute%3E%3C%2FtoSttnNameRoute%3E+%3CtrainingValue%3E-+-+Select+-+-%3C%2FtrainingValue%3E+%3C%2FCMSREPORT%3E++%3C%2FCMSPublishXML%3E&lobby=NRE&lobbyList=SC-GTL-NRE"

        s = requests.Session()

        login_data =  { 'userId': 'NSAI', 'userPassword': 'NSAI1995'}
        s.post('https://cms.indianrail.gov.in/CMSREPORT/JSP/rpt/LoginAction.do?hmode=login&isResponsive=Y', login_data)
        r2 = s.get(prurl)
        dueget = s.get(dueurl)
        lrget = s.get(lrurl)
        print(lrget.text)
        #print(r2.text)
        soup = BeautifulSoup(r2.text, "html.parser")
        soupdue = BeautifulSoup(dueget.text, "html.parser")
        var = soupdue.find("script")
        print(var)



        prduedata = [[cell.text for cell in row("td")] for row in soup ("tr")]
        duedata = [[cell.text for cell in row("td")] for row in soupdue ("tr")]


        print(len(duedata))
        prduedata = prduedata[1]
        
        #print(len(prduedata))
        prdivvalue = int(len(prduedata)/12)
        length_to_split = [12]*(prdivvalue)

        Inputt = iter(prduedata)
        prduedata = [list(islice(Inputt, elem)) for elem in length_to_split]




                         

        """


        return render(request, "someindex.html")
    else:
        def time_diff(hr2,min2,sec2,hr1,min1,sec1):
            time_diff = (hr2-hr1) + ((min2-min1)/60) + ((sec2-sec1)/3600)
            return time_diff
        #sir here in in the variable some it is giving the formula present in the excel sheet so by just adding "data_only=True" in wb1 and wb2 we can get the value of the cell instead of formula
        
        #datas = request.POST
        #print(datas)
        #signal_sample=datas.get("signalsample")

        #signal_speed=datas.get("signalspeed")


        signal_sample = request.FILES["signalsample"]
        signal_speed = request.FILES["signalspeed"]

        #wb1 = load_workbook(filename = 'C:\wamp64\www\env_site\pythoncharts\pythoncharts\signal speed.xlsx',data_only=True)
        #wb2 = load_workbook(filename = 'C:\wamp64\www\env_site\pythoncharts\pythoncharts\signal sample.xlsx',data_only=True)
        
        wb1 = load_workbook(filename = request.FILES['signalspeed'].file,data_only=True)
        wb2 = load_workbook(filename = request.FILES['signalsample'].file,data_only=True)

        #wb1 = openpyxl.load_workbook(signal_sample)
        #wb2 = openpyxl.load_workbook(signal_speed)

        sh1 = wb1['Sheet1'] 

        sh2 = wb2['Sheet1']


        row1_ct = sh1.max_row

        col1_ct = sh1.max_column

        row2_ct = sh2.max_row

        col2_ct = sh2.max_column

        speed_list = []

        signal_list = []

        row_num_dist2 = 2

        column_num_signal_dist2 = 6

        cumu_dist1 = 2

        j = 2

        print(row1_ct)

        print(row2_ct)


        for i in range(2,row1_ct+1):
            #sir here the value that is being compared is giving some none values which cannot be compared with any of the datatype so i just added a if statement 
            
            someright = (sh1.cell(i,cumu_dist1).value)
            
            some = (sh2.cell(row_num_dist2,column_num_signal_dist2).value)
            
            if someright != None and some != None:
                if (some) == (someright):
                    
                    speed_list.append(sh1.cell(i,1).value)
                    
                    signal_list.append(sh2.cell(row_num_dist2,1).value)
                    
                    i = i+1
                    
                    row_num_dist2 = row_num_dist2 + 1
                
                elif (some) > (someright):
                    
                    i = i+1
                
                elif (some) < (someright):
                    
                    hr2 = sh1.cell(i,3).value.hour
                    min2 = sh1.cell(i,3).value.minute
                    sec2 = sh1.cell(i,3).value.second
                    
                    hr1 = sh1.cell(i-1,3).value.hour
                    min1 = sh1.cell(i-1,3).value.minute
                    sec1 = sh1.cell(i-1,3).value.second
                    
                    time_dif = time_diff(hr2,min2,sec2,hr1,min1,sec1)
                    
                    speed_dif = sh1.cell(i,1).value-sh1.cell(i-1,1).value
                    
                    dist_dif = (sh1.cell(i,2).value-sh1.cell(i-1,2).value)/1000
                    
                    distan_travelled = ((some)-(sh1.cell(i-1,cumu_dist1).value))/1000

                    #print(speed_dif)
                    #print(time_dif)
                    
                    #Sir this is to check how many decimal points are available in the variable
                    timedifcheck = str(time_dif)
                    #print(timedifcheck)
                    noofdecimal = timedifcheck[::-1].find('.')
                    #print(noofdecimal)
                    #print("{:.19f}".format(round(speed_dif, 2)))
                    
                    #Sir here we are getting an error division by zero, so i have given this if statement
                    if time_dif!=0:
                        
                        accel = speed_dif/time_dif
                    
                    x = (sh1.cell(i-1,1).value)**2
                    
                    y = (2*accel*distan_travelled)
                    
                    z = (x+y)**0.5
                    
                    speed_list.append(z)
                    
                    signal_list.append(sh2.cell(row_num_dist2,1).value)
                    
                    i = i+1
                    
                    row_num_dist2 = row_num_dist2 + 1
                
            
    #time_diff = sh1.cell(3,3).value-sh1.cell(2,3).value

    #print(time_diff)

    #time_diff_in_hours = (time_diff.hour)+((time_diff.minute)/60)+((time_diff.second)/3600)

    #print(time_diff_in_hours)
            
            
        print(speed_list)        
                
        print(signal_list)

        x = signal_list

        y = speed_list
        
        xaxis = []
        for i in range(len(signal_list)):
            xaxis.append(i)
        
        graphsize = (len(signal_list))/2
        if graphsize>100:
            graphsize = 90

        #plt.xticks(xaxis,x,rotation=90)

        """
        plt.figure(figsize=(graphsize,8))

        plt.plot(xaxis,y,color = 'green',linestyle = 'solid',linewidth = 2,marker = 'o',markerfacecolor = 'blue',markersize = 9)
        #plt.xticks(rotation=90)
        plt.xticks(xaxis, signal_list, rotation =90)
        plt.tick_params(labelsize=6)

        plt.xlabel('signal locations')

        plt.ylabel('speed graph')

        plt.title('My first graph')

        #plt.plot(x,y,color = 'green',linestyle = 'solid',linewidth = 2,marker = 'o',markerfacecolor = 'blue',markersize = 9)

        #plt.show()
        buffer = BytesIO()
        plt.savefig(buffer, format = 'png')
        buffer.seek(0)
        image_png = buffer.getvalue()
        #chart.clear()
        chart = base64.b64encode(image_png)
        chart = chart.decode('utf-8')
        buffer.close()
        plt.clf()
        """



        fig = go.Figure()
        scatter = go.Scatter(x=xaxis, y=speed_list, mode='lines', name='test', opacity=0.8, marker_color='blue')
        fig.add_trace(scatter)
        #fig.update_xaxes(type='category')
        fig.update_layout(xaxis = dict(tickmode = 'array',tickvals = xaxis, ticktext = signal_list))
        plot_div = plot(fig, output_type='div')
        #plot_div = plot([Scatter(x=signal_list, y=speed_list, mode='lines', name='test', opacity=0.8, marker_color='blue')], output_type='div')

        #return render(request, "index.html", context={'plot_div': plot_div})

        
        return render(request, "someindex.html", {"y" : y, "x" : x, 'plot_div': plot_div})
